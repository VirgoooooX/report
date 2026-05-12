"""FA Tracker analysis module — independent of db.py calculation logic.
Handles Failure Rate computation (dedup SN / Sample Size) from FA Tracker.xlsx.
"""

import datetime
import os
import re
from openpyxl import load_workbook
from app_paths import RAWDATA_DIR, ensure_runtime_dirs, iter_rawdata_files

DATA_DIR = RAWDATA_DIR

DIMENSION_FIELD = {
    'symptom': 'Failure Symptom / Failure Message',
    'location': 'Failed Location',
    'failed_test': 'Failed Test',
    'wf': 'WF',
    'config': 'Config',
}

CONFIG_ORDER = ['R1FNF', 'R2CNM', 'R3', 'R4']


def _norm_field_name(name):
    return ' '.join(str(name).split()).lower()


def _field_value(issue, field, default=''):
    if field in issue:
        return issue.get(field, default)
    wanted = _norm_field_name(field)
    for key, value in issue.items():
        if _norm_field_name(key) == wanted:
            return value
    return default


def find_fa_tracker():
    """Find the FA Tracker Excel file in rawdata/ directory."""
    ensure_runtime_dirs()
    candidates = []
    for fname, path in iter_rawdata_files():
        if 'FA Tracker' not in fname or not fname.endswith('.xlsx') or fname.startswith('~$'):
            continue
        m = re.search(r'(\d{8})', fname)
        date_key = m.group(1) if m else ''
        candidates.append((date_key, path))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[-1][1]


def read_fa_tracker(fa_path):
    """Read System TF sheet. Returns list of FA record dicts.
    Row 7 is header, data starts at row 8.
    """
    wb = load_workbook(fa_path, data_only=True)
    ws = wb['System TF']
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(7, c).value
        if v:
            headers[c] = str(v).strip().replace('\n', ' ')
    issues = []
    for r in range(8, ws.max_row + 1):
        fa_num = ws.cell(r, 1).value
        if fa_num is None:
            continue
        issue = {h: ws.cell(r, c).value for c, h in headers.items()}
        issue['_fa_num'] = fa_num
        issues.append(issue)
    wb.close()
    return issues


def read_sample_sizes(fa_path):
    """Read WF Sample Size sheet.
    Returns: { wf (str): { config (str): count (int) } }
    Row 1 header: WaterFall | Test Name | R1FNF | R2CNM | R3 | R4
    Config columns start at col 3.
    """
    wb = load_workbook(fa_path, data_only=True)
    ws = wb['WF Sample Size']
    sample_map = {}
    for r in range(2, ws.max_row + 1):
        wf_val = ws.cell(r, 1).value
        if wf_val is None:
            continue
        wf_key = str(wf_val).strip()
        sample_map[wf_key] = {}
        for ci, cfg in enumerate(CONFIG_ORDER, start=3):
            v = ws.cell(r, ci).value
            sample_map[wf_key][cfg] = int(v) if v else 0
    wb.close()
    return sample_map


def read_wf_test_names(fa_path):
    """Read WF -> test names from WF Sample Size sheet."""
    wb = load_workbook(fa_path, data_only=True)
    ws = wb['WF Sample Size']
    wf_tests = {}
    for r in range(2, ws.max_row + 1):
        wf_val = ws.cell(r, 1).value
        if wf_val is None:
            continue
        wf_key = str(wf_val).strip()
        test_text = str(ws.cell(r, 2).value or '').strip()
        if not wf_key or not test_text or test_text == '/':
            continue
        tests = [_normalize_test_name(part) for part in test_text.split('+')]
        wf_tests[wf_key] = [test for test in tests if test]
    wb.close()
    return wf_tests


def _get_dim_value(issue, dim):
    """Get the dimension value for an issue."""
    field = DIMENSION_FIELD[dim]
    val = _field_value(issue, field, '')
    return str(val).strip() if val else ''


def _failure_type(issue):
    """Returns 'spec' or 'strife' based on Failure Type field."""
    ft = str(_field_value(issue, 'Failure Type (Spec. or Strife)', '')).strip().lower()
    if 'spec' in ft:
        return 'spec'
    elif 'strife' in ft:
        return 'strife'
    return 'unknown'


def _normalize_test_name(name):
    return ' '.join(str(name or '').split()).strip()


def _normalize_config_name(name):
    return ' '.join(str(name or '').split()).strip()


def _sample_row(sample_sizes, wf):
    wf_text = str(wf or '').strip()
    if not wf_text:
        return {}
    wf_key = wf_text[2:] if wf_text.upper().startswith('WF') else wf_text
    wf_display = f'WF{wf_key}'
    return sample_sizes.get(wf_text) or sample_sizes.get(wf_key) or sample_sizes.get(wf_display) or {}


def _sample_wfs(sample_sizes):
    return [str(wf).strip() for wf in sample_sizes.keys() if str(wf).strip()]


def _wf_sample_size(sample_sizes, wf, configs=None):
    row = _sample_row(sample_sizes, wf)
    if not row:
        return 0
    if configs:
        wanted = {_normalize_config_name(config) for config in configs if _normalize_config_name(config)}
        return sum(int(value or 0) for config, value in row.items() if _normalize_config_name(config) in wanted)
    return sum(int(value or 0) for value in row.values())


def _total_sample_size(sample_sizes, wfs=None, configs=None):
    target_wfs = list(wfs) if wfs else _sample_wfs(sample_sizes)
    return sum(_wf_sample_size(sample_sizes, wf, configs) for wf in target_wfs)


def _wfs_for_failed_test(failed_test, wf_test_names):
    normalized = _normalize_test_name(failed_test)
    if not normalized or not wf_test_names:
        return []
    matched = []
    for wf, tests in wf_test_names.items():
        normalized_tests = {_normalize_test_name(test) for test in (tests or [])}
        if normalized in normalized_tests:
            matched.append(str(wf).strip())
    return matched


def _fallback_wfs_for_group(group):
    return sorted({
        str(_field_value(issue, 'WF', '')).strip()
        for issue in group
        if str(_field_value(issue, 'WF', '')).strip()
    }, key=str)


def _dimension_denominator(dim, key, sample_sizes, wf_test_names=None, group=None):
    if dim == 'wf':
        return _wf_sample_size(sample_sizes, key)
    if dim == 'config':
        return _total_sample_size(sample_sizes, configs=[key])
    if dim == 'failed_test':
        test_wfs = _wfs_for_failed_test(key, wf_test_names)
        if not test_wfs and group:
            test_wfs = _fallback_wfs_for_group(group)
        return _total_sample_size(sample_sizes, wfs=test_wfs)
    return _total_sample_size(sample_sizes)


def _date_value(issue):
    """Best-effort date extraction for the Today card."""
    for key, val in issue.items():
        key_l = str(key).lower()
        if 'date' not in key_l or 'target' in key_l:
            continue
        if isinstance(val, datetime.datetime):
            return val.date()
        if isinstance(val, datetime.date):
            return val
        if isinstance(val, str):
            text = val.strip()
            for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%m/%d/%y'):
                try:
                    return datetime.datetime.strptime(text[:10], fmt).date()
                except ValueError:
                    pass
    return None


def _summary_stats(issues, sample_sizes):
    """Compute dashboard KPI cards from FA Tracker records."""
    spec_sns = set()
    strife_sns = set()
    symptoms = set()
    wfs = set()
    configs = set()
    today = datetime.date.today()
    today_count = 0
    today_spec = 0
    today_strife = 0
    spec_issues = 0
    strife_issues = 0

    for issue in issues:
        ft = _failure_type(issue)
        sn = str(_field_value(issue, 'SN', '')).strip()
        symptom = str(_field_value(issue, 'Failure Symptom / Failure Message', '') or '').strip()
        wf = str(_field_value(issue, 'WF', '') or '').strip()
        cfg = str(_field_value(issue, 'Config', '') or '').strip()

        if symptom:
            symptoms.add(symptom)
        if wf:
            wfs.add(wf)
        if cfg:
            configs.add(cfg)

        if ft == 'spec':
            spec_issues += 1
            if sn:
                spec_sns.add(sn)
        elif ft == 'strife':
            strife_issues += 1
            if sn:
                strife_sns.add(sn)

        if _date_value(issue) == today:
            today_count += 1
            if ft == 'spec':
                today_spec += 1
            elif ft == 'strife':
                today_strife += 1

    total_sample_size = _total_sample_size(sample_sizes)
    spec_sn_count = len(spec_sns)
    strife_sn_count = len(strife_sns)

    return {
        'todayCount': today_count,
        'todaySpecCount': today_spec,
        'todayStrifeCount': today_strife,
        'totalIssues': len(issues),
        'specIssues': spec_issues,
        'strifeIssues': strife_issues,
        'specSNCount': spec_sn_count,
        'strifeSNCount': strife_sn_count,
        'uniqueSymptoms': len(symptoms),
        'uniqueWFs': len(wfs),
        'uniqueConfigs': len(configs),
        'totalSampleSize': total_sample_size,
        'specFailurePercent': round(spec_sn_count / total_sample_size * 100, 2) if total_sample_size else 0,
        'strifeFailurePercent': round(strife_sn_count / total_sample_size * 100, 2) if total_sample_size else 0,
    }


def _dim_stats(issues, sample_sizes, dim, wf_test_names=None):
    """Compute stats for a single dimension: grouped by dimension value.
    Returns sorted list of { key, specSnCount, strifeSnCount, totalSamples, display }
    """
    groups = {}
    for issue in issues:
        key = _get_dim_value(issue, dim)
        if not key:
            continue
        if key not in groups:
            groups[key] = []
        groups[key].append(issue)

    result = []
    for key, group in groups.items():
        spec_sns = set()
        strife_sns = set()
        for issue in group:
            sn = str(_field_value(issue, 'SN', '')).strip()
            ft = _failure_type(issue)
            if not sn:
                continue
            if ft == 'spec':
                spec_sns.add(sn)
            elif ft == 'strife':
                strife_sns.add(sn)

        spec_count = len(spec_sns)
        strife_count = len(strife_sns)
        total_samples = _dimension_denominator(dim, key, sample_sizes, wf_test_names, group)

        if spec_count > 0:
            display = f'{spec_count}F/{total_samples}T'
        elif strife_count > 0:
            display = f'{strife_count}SF/{total_samples}T'
        else:
            display = f'0F/{total_samples}T'

        result.append({
            'key': key,
            'specSnCount': spec_count,
            'strifeSnCount': strife_count,
            'totalSamples': total_samples,
            'display': display,
            'specRate': round(spec_count / total_samples * 100, 2) if total_samples else 0,
            'strifeRate': round(strife_count / total_samples * 100, 2) if total_samples else 0,
            'specFailureRate': round(spec_count / total_samples * 1000000) if total_samples else 0,
            'strifeFailureRate': round(strife_count / total_samples * 1000000) if total_samples else 0,
        })

    result.sort(key=lambda x: (x['specFailureRate'], x['strifeFailureRate']), reverse=True)
    return result


def _wf_label(wf, wf_names=None):
    wf_text = str(wf).strip()
    wf_display = wf_text if wf_text.upper().startswith('WF') else f'WF{wf_text}'
    wf_key = wf_text[2:] if wf_text.upper().startswith('WF') else wf_text
    name = ''
    if wf_names:
        name = (wf_names.get(wf_text, {}) or {}).get('name', '')
        if not name:
            name = (wf_names.get(wf_display, {}) or {}).get('name', '')
        if not name:
            name = (wf_names.get(wf_key, {}) or {}).get('name', '')
    return f'{wf_display} {name}'.strip()


def compute_overview(issues, sample_sizes, wf_names=None, wf_test_names=None):
    """Compute Top 10 by Symptom, Location, WF.
    Returns: { topSymptom, topLocation, topWf }
    """
    top_wf = _dim_stats(issues, sample_sizes, 'wf', wf_test_names)[:10]
    for item in top_wf:
        item['label'] = _wf_label(item['key'], wf_names)

    return {
        'summary': _summary_stats(issues, sample_sizes),
        'topSymptom': _dim_stats(issues, sample_sizes, 'symptom', wf_test_names)[:10],
        'topFailedTest': _dim_stats(issues, sample_sizes, 'failed_test', wf_test_names)[:10],
        'topWf': top_wf,
    }


def compute_cross(issues, sample_sizes, dim1, dim2, wf_test_names=None):
    """Compute cross-analysis matrix for two dimensions.
    Returns: { dim1, dim2, dim1Values, dim2Values, matrix }
    Each matrix cell: { dim1Value, dim2Value, specSnCount, strifeSnCount, totalCount, totalSamples, display, specRate, strifeRate, percentage, sns }
    """
    groups = {}
    for issue in issues:
        v1 = _get_dim_value(issue, dim1)
        v2 = _get_dim_value(issue, dim2)
        if not v1 or not v2:
            continue
        key = (v1, v2)
        if key not in groups:
            groups[key] = []
        groups[key].append(issue)

    matrix = []
    total_all = 0
    for (v1, v2), group in groups.items():
        spec_sns = set()
        strife_sns = set()
        all_sns = set()
        for issue in group:
            sn = str(_field_value(issue, 'SN', '')).strip()
            ft = _failure_type(issue)
            if not sn:
                continue
            all_sns.add(sn)
            if ft == 'spec':
                spec_sns.add(sn)
            elif ft == 'strife':
                strife_sns.add(sn)

        spec_count = len(spec_sns)
        strife_count = len(strife_sns)
        total_sn = spec_count + strife_count

        total_samples = _dimension_denominator(dim2, v2, sample_sizes, wf_test_names, group)

        total_all += total_sn

        if spec_count > 0:
            display = f'{spec_count}F/{total_samples}T'
        elif strife_count > 0:
            display = f'{strife_count}SF/{total_samples}T'
        else:
            display = f'0F/{total_samples}T'

        matrix.append({
            'dim1Value': v1,
            'dim2Value': v2,
            'specSnCount': spec_count,
            'strifeSnCount': strife_count,
            'totalCount': total_sn,
            'totalSamples': total_samples,
            'display': display,
            'specRate': round(spec_count / total_samples * 100, 2) if total_samples else 0,
            'strifeRate': round(strife_count / total_samples * 100, 2) if total_samples else 0,
            'percentage': 0,
            'sns': list(all_sns),
        })

    for m in matrix:
        m['percentage'] = round(m['totalCount'] / total_all * 100, 1) if total_all else 0

    dim1Values = sorted(set(m['dim1Value'] for m in matrix), key=str)
    dim2Values = sorted(set(m['dim2Value'] for m in matrix), key=str)

    return {
        'dim1': dim1,
        'dim2': dim2,
        'dim1Values': dim1Values,
        'dim2Values': dim2Values,
        'matrix': matrix,
    }


def compute_detail(issues, filters):
    """Filter issues by dimensions.
    filters dict keys: symptom, location, wf, config, failed_test
    Returns list of flat records for popup display.
    """
    result = []
    for issue in issues:
        match = True
        if filters.get('symptom'):
            if _get_dim_value(issue, 'symptom') != filters['symptom']:
                match = False
        if filters.get('location'):
            if _get_dim_value(issue, 'location') != filters['location']:
                match = False
        if filters.get('wf'):
            if str(_field_value(issue, 'WF', '')).strip() != filters['wf']:
                match = False
        if filters.get('config'):
            if str(_field_value(issue, 'Config', '')).strip() != filters['config']:
                match = False
        if filters.get('failed_test'):
            if str(_field_value(issue, 'Failed Test', '')).strip() != filters['failed_test']:
                match = False
        if match:
            result.append({
                'fa_num': issue.get('_fa_num', ''),
                'sn': str(_field_value(issue, 'SN', '')).strip(),
                'wf': str(_field_value(issue, 'WF', '')).strip(),
                'config': str(_field_value(issue, 'Config', '')).strip(),
                'failed_test': str(_field_value(issue, 'Failed Test', '')).strip(),
                'failure_type': str(_field_value(issue, 'Failure Type (Spec. or Strife)', '')).strip(),
                'symptom': str(_field_value(issue, 'Failure Symptom / Failure Message', '')).strip(),
                'location': str(_field_value(issue, 'Failed Location', '')).strip(),
                'fa_status': str(_field_value(issue, 'FA Status', '')).strip(),
            })
    return result
