"""
Check Item CSV Import — Daily Report Excel Generator.

核心职责：从 CSV raw data + DB 定义生成兼容 engine.py 的 Daily Report Excel。
"""

import csv
import datetime
import glob
import io
import json
import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import db
from app_paths import RAWDATA_DIR
from base_manager import get_sn_mapping_from_db, get_sn_lookup_dicts


# Mapping of filename keywords to canonical item type names.
# Keys are matched case-insensitively against the filename.
_KEYWORD_TO_TYPE: dict[str, str] = {
    "BT-OTA": "BT-OTA",
    "Charging": "Charging",
    "FACT": "FACT",
    "ISB": "ISB",
    "Touch-Cal-Post": "Touch-CAL-Post",
    "COSMETIC": "Cosmetic",
}

# Column names we need to extract from the CSV header row.
_REQUIRED_COLUMNS = [
    "SerialNumber",
    "REL Event",
    "Test Pass/Fail Status",
    "EndTime",
    "List of Failing Tests",
    "Station ID",
    "Version",
]

# Metadata row prefixes that should be skipped (rows between header and data).
_METADATA_PREFIXES = (
    "Display Name",
    "PDCA Priority",
    "Upper Limit",
    "Lower Limit",
    "Measurement Unit",
)


def identify_csv_type(filename: str) -> str | None:
    """Identify the check item type from a CSV filename by keyword matching.

    Scans the filename for known keywords (case-insensitive) and returns
    the canonical item type name. Returns None if no keyword matches.

    Args:
        filename: The CSV filename (may include path components).

    Returns:
        The canonical item type string, or None if unrecognized.
    """
    filename_lower = filename.lower()
    for keyword, item_type in _KEYWORD_TO_TYPE.items():
        if keyword.lower() in filename_lower:
            return item_type
    return None


def sort_and_filter_tfinal(records: list[dict]) -> list[dict]:
    """Sort records by end_time ascending and discard records after REL_TFINAL per SN.

    For each unique serial_number, finds the first record with rel_event == "REL_TFINAL".
    All records for that SN with end_time strictly after the REL_TFINAL event's end_time
    are discarded. SNs without a REL_TFINAL event keep all their records.

    Args:
        records: List of parsed record dicts (must have 'end_time', 'serial_number',
                 'rel_event' keys).

    Returns:
        A new list sorted by end_time ascending, with post-TFINAL records removed.
    """
    # Sort all records by end_time ascending
    sorted_records = sorted(records, key=lambda r: r.get("end_time", ""))

    # For each SN, find the end_time of the first REL_TFINAL event
    tfinal_times: dict[str, str] = {}
    for rec in sorted_records:
        sn = rec.get("serial_number", "")
        if sn not in tfinal_times and rec.get("rel_event") == "REL_TFINAL":
            tfinal_times[sn] = rec.get("end_time", "")

    # Filter: keep records where SN has no TFINAL, or end_time <= TFINAL end_time
    result = []
    for rec in sorted_records:
        sn = rec.get("serial_number", "")
        if sn not in tfinal_times:
            # No TFINAL for this SN — keep all records
            result.append(rec)
        else:
            # Keep records with end_time <= TFINAL end_time
            if rec.get("end_time", "") <= tfinal_times[sn]:
                result.append(rec)

    return result


# Events that are anomalies — they should be attributed to the preceding valid CP.
_ANOMALY_EVENTS = frozenset([
    "REL FA RETEST",
    "SEND TO FA",
    "STOP TEST",
    "RETURN TO REL",
])


def attribute_anomaly_cps(records: list[dict]) -> list[dict]:
    """Attribute anomaly events to the preceding valid CP per serial number.

    For each record, sets an `effective_cp` field:
    - If the record's rel_event is a normal CP (not an anomaly), effective_cp = rel_event.
    - If the record's rel_event is an anomaly event, effective_cp = the most recent
      preceding non-anomaly CP for that same serial number.
    - If an anomaly event has no preceding valid CP, effective_cp = rel_event itself.

    Records must already be sorted by end_time ascending.

    Args:
        records: List of parsed record dicts (sorted by end_time ascending).
                 Must have 'serial_number' and 'rel_event' keys.

    Returns:
        The same list with `effective_cp` field added/set on each record.
    """
    # Track the last valid (non-anomaly) CP per serial number
    last_valid_cp: dict[str, str] = {}

    for rec in records:
        sn = rec.get("serial_number", "")
        rel_event = rec.get("rel_event", "")

        if rel_event in _ANOMALY_EVENTS:
            # Attribute to the preceding valid CP for this SN
            if sn in last_valid_cp:
                rec["effective_cp"] = last_valid_cp[sn]
            else:
                # Edge case: no preceding valid CP — use rel_event itself
                rec["effective_cp"] = rel_event
        else:
            # Normal CP — effective_cp is itself, and update tracker
            rec["effective_cp"] = rel_event
            last_valid_cp[sn] = rel_event

    return records


def deduplicate_records(records: list[dict]) -> list[dict]:
    """Deduplicate records by keeping only the latest record per (SN, effective_cp, item) group.

    For each unique combination of (serial_number, effective_cp, item), keeps only
    the record with the latest (maximum) end_time. Since records are expected to be
    sorted by end_time ascending, the last record encountered in each group wins.

    Args:
        records: List of parsed record dicts (already sorted by end_time ascending,
                 with effective_cp set). Must have 'serial_number', 'effective_cp',
                 'item', and 'end_time' keys.

    Returns:
        A new list containing only the latest record per group.
    """
    # Use a dict keyed by (serial_number, effective_cp, item) to track the latest record
    latest: dict[tuple[str, str, str], dict] = {}

    for rec in records:
        key = (
            rec.get("serial_number", ""),
            rec.get("effective_cp", ""),
            rec.get("item", ""),
        )
        # Since records are sorted by end_time ascending, later records overwrite earlier ones.
        # But we also explicitly compare end_time to handle any ordering edge cases.
        existing = latest.get(key)
        if existing is None or rec.get("end_time", "") >= existing.get("end_time", ""):
            latest[key] = rec

    return list(latest.values())


def filter_valid_sns(records: list[dict], valid_sns: set) -> list[dict]:
    """Filter parsed records to keep only those with serial numbers in the valid set.

    Records whose serial_number is not in valid_sns are silently discarded.

    Args:
        records: List of parsed record dicts (from parse_csv_file).
        valid_sns: Set of valid serial number strings (from DB SN mapping).

    Returns:
        A new list containing only records with serial_number in valid_sns.
    """
    return [r for r in records if r.get("serial_number") in valid_sns]


def parse_csv_file(file_content: str | bytes, item_type: str) -> list[dict]:
    """Parse a check item CSV file and extract test records.

    Finds the header row by scanning for a row containing 'SerialNumber',
    then extracts data records with key fields. For FAIL records, also
    extracts parameter columns (index 13+) as a dict.

    Args:
        file_content: The CSV file content as string or bytes.
        item_type: The canonical item type (e.g., 'FACT', 'ISB').

    Returns:
        A list of dicts, each containing:
            - serial_number: str
            - rel_event: str
            - status: str ('PASS' or 'FAIL')
            - end_time: str (timestamp)
            - failing_tests: str (may be empty)
            - station_id: str
            - version: str
            - item: str (the item_type passed in)
            - test_params: dict | None (parameter name→value for FAIL records)

    Returns an empty list if no header row with 'SerialNumber' is found.
    """
    # Normalize input to string
    if isinstance(file_content, bytes):
        file_content = file_content.decode("utf-8", errors="ignore")

    reader = csv.reader(io.StringIO(file_content))

    # Step 1: Find the header row (the row containing 'SerialNumber')
    header_row = None
    header_row_idx = -1
    for idx, row in enumerate(reader):
        if "SerialNumber" in row:
            header_row = row
            header_row_idx = idx
            break

    if header_row is None:
        return []

    # Step 2: Build column index mapping for required fields
    col_indices: dict[str, int] = {}
    for col_name in _REQUIRED_COLUMNS:
        try:
            col_indices[col_name] = header_row.index(col_name)
        except ValueError:
            # Column not found — will be treated as empty string
            col_indices[col_name] = -1

    # Parameter columns start at index 13 (the 14th column, 1-indexed)
    param_start_idx = 13
    param_names = header_row[param_start_idx:] if len(header_row) > param_start_idx else []

    # Step 3: Read data rows, skipping metadata rows
    records: list[dict] = []
    for row in reader:
        # Skip empty rows
        if not row or all(cell.strip() == "" for cell in row):
            continue

        # Skip metadata rows (Display Name, PDCA Priority, etc.)
        first_cell = row[0].strip() if row else ""
        if any(first_cell.startswith(prefix) for prefix in _METADATA_PREFIXES):
            continue

        # Extract required fields
        def _get(col_name: str) -> str:
            idx = col_indices[col_name]
            if idx < 0 or idx >= len(row):
                return ""
            return row[idx].strip()

        serial_number = _get("SerialNumber")
        # Skip rows without a serial number (likely empty/malformed)
        if not serial_number:
            continue

        status = _get("Test Pass/Fail Status")
        rel_event = _get("REL Event")
        end_time = _get("EndTime")
        failing_tests = _get("List of Failing Tests")
        station_id = _get("Station ID")
        version = _get("Version")

        # For FAIL records, extract parameter columns as dict
        test_params: dict | None = None
        if status == "FAIL" and param_names:
            params = {}
            for i, param_name in enumerate(param_names):
                col_idx = param_start_idx + i
                if col_idx >= len(row):
                    break
                value = row[col_idx].strip()
                if value and param_name.strip():
                    # Try to convert to float for numeric values
                    try:
                        params[param_name.strip()] = float(value)
                    except ValueError:
                        # Keep as string if not numeric
                        params[param_name.strip()] = value
            if params:
                test_params = params

        records.append({
            "serial_number": serial_number,
            "rel_event": rel_event,
            "status": status,
            "end_time": end_time,
            "failing_tests": failing_tests,
            "station_id": station_id,
            "version": version,
            "item": item_type,
            "test_params": test_params,
        })

    return records

# Strife failure detection patterns.
# Each tuple: (compiled regex pattern, threshold value).
# If the CP name matches the pattern and the captured numeric value exceeds the threshold,
# the failure is classified as "strife_fail"; otherwise "spec_fail".
_STRIFE_PATTERNS: list[tuple[re.Pattern, int]] = [
    (re.compile(r"RANDOM DROP 1M PB_(\d+)", re.IGNORECASE), 200),
    (re.compile(r"RANDOM DROP 1M GRA_(\d+)", re.IGNORECASE), 20),
    (re.compile(r"15MM PROBE SP_(\d+)KG", re.IGNORECASE), 21),
    (re.compile(r"CLEANING_SPRAY_OP1_AFTER (\d+)HRS", re.IGNORECASE), 72),
    (re.compile(r"RBI_(\d+)CM", re.IGNORECASE), 10),
    (re.compile(r"BC_.*_(\d+)%", re.IGNORECASE), 100),
]


def _is_strife_failure(cp_name: str) -> bool:
    """Determine if a CP name indicates a strife failure based on numeric thresholds.

    Checks the CP name against known patterns. If a pattern matches and the
    captured numeric value exceeds the associated threshold, the failure is strife.

    Args:
        cp_name: The checkpoint name to evaluate.

    Returns:
        True if the CP name indicates a strife failure, False otherwise.
    """
    cp_upper = cp_name.strip().upper()
    for pattern, threshold in _STRIFE_PATTERNS:
        m = pattern.search(cp_upper)
        if m:
            value = int(m.group(1))
            if value > threshold:
                return True
    return False


def detect_strife_failures(records: list[dict], cp_schedule: dict) -> list[dict]:
    """Classify FAIL records as strife or spec failures based on CP name thresholds.

    For each record:
    - If status is not "FAIL", sets failure_type to None.
    - If status is "FAIL", checks the effective_cp against known strife patterns.
      If the CP name matches a strife pattern (numeric value exceeds threshold),
      sets failure_type to "strife_fail"; otherwise "spec_fail".

    The cp_schedule parameter provides the CP schedule from DB (wf_id → [cp_names]).
    It is available for future use (e.g., resolving CP names from schedule context)
    but the current implementation determines strife status directly from the
    effective_cp field on each record.

    Args:
        records: List of deduplicated record dicts (with effective_cp set).
                 Must have 'status' and 'effective_cp' keys.
        cp_schedule: Dict mapping wf_id to ordered list of CP names from DB.

    Returns:
        The same list with `failure_type` field added to each record:
        - "strife_fail" for strife failures
        - "spec_fail" for spec failures
        - None for PASS records
    """
    for rec in records:
        status = rec.get("status", "")
        if status != "FAIL":
            rec["failure_type"] = None
        else:
            effective_cp = rec.get("effective_cp", "")
            if _is_strife_failure(effective_cp):
                rec["failure_type"] = "strife_fail"
            else:
                rec["failure_type"] = "spec_fail"

    return records


# ---------------------------------------------------------------------------
# Excel Generation — WF Sheet Layout (openpyxl)
# Formatting matches the original generate_rel_summary.py xlsxwriter output.
# ---------------------------------------------------------------------------

# Fixed check item order for all WF sheets.
CHECK_ITEMS = ["Cosmetic", "ISB", "FACT", "BT-OTA", "Touch-CAL-Post", "Charging"]

# --- Border definitions ---
_THIN_SIDE = Side(style='thin')
_THICK_SIDE = Side(style='medium')

_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

# --- Color coding constants ---
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PASS_FONT = Font(color="006100")
SPEC_FAIL_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
STRIFE_FAIL_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
STRIFE_FAIL_FONT = Font(color="000000")

# Header styling
_CP_HEADER_FILL = PatternFill(start_color="B0B0B0", end_color="B0B0B0", fill_type="solid")
_CP_HEADER_FONT = Font(bold=True)
_CP_HEADER_BORDER = Border(
    left=_THICK_SIDE, right=_THICK_SIDE, top=_THICK_SIDE, bottom=_THICK_SIDE
)

_SUB_HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
_SUB_HEADER_FONT = Font(bold=True)
_SUB_HEADER_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THICK_SIDE, bottom=_THICK_SIDE
)
_SUB_HEADER_BORDER_RIGHT = Border(
    left=_THIN_SIDE, right=_THICK_SIDE, top=_THICK_SIDE, bottom=_THICK_SIDE
)

# SN header (Config/Unit#/S/N in sub-header row)
_SN_HEADER_BORDER = Border(
    left=_THIN_SIDE, right=_THICK_SIDE, top=_THICK_SIDE, bottom=_THICK_SIDE
)

# Number of check item columns per CP group.
_ITEMS_PER_CP = len(CHECK_ITEMS)  # 6


def _make_data_border(thick_right: bool = False, thick_bottom: bool = False) -> Border:
    """Build a data cell border with optional thick right/bottom edges."""
    return Border(
        left=_THIN_SIDE,
        right=_THICK_SIDE if thick_right else _THIN_SIDE,
        top=_THIN_SIDE,
        bottom=_THICK_SIDE if thick_bottom else _THIN_SIDE,
    )


def _is_chemical_wf(wf_num) -> bool:
    """Return True if the WF number indicates a Chemical workflow (29.x or 30).

    Chemical WFs get an extra "Chemical" column before the first CP group.

    Args:
        wf_num: WF number as int, float, or string (e.g., 29, 29.1, 30, "29.1").

    Returns:
        True if the WF is a Chemical workflow.
    """
    try:
        num = float(wf_num)
    except (TypeError, ValueError):
        return False
    # WF 29.x (29 <= num < 30) or exactly 30
    return (29 <= num < 30) or num == 30


def create_wf_sheet(wb: Workbook, wf_num, wf_name: str, cp_list: list[str],
                    sn_data: list[dict], is_chemical: bool | None = None,
                    records_by_sn: dict | None = None,
                    report_date: str | None = None):
    """Create a WF sheet with formatting matching the original xlsxwriter output.

    Layout:
    - Row 1 (openpyxl): Report Date in A1 + merged CP headers (bold, gray #B0B0B0, thick borders)
    - Row 2: Sub-headers — Config/Unit#/S/N (bold, light gray, thick right border on S/N)
             + item names per CP group (bold, light gray #D3D3D3, thick top/bottom)
    - Row 3+: Data rows with borders, alignment, color fills, config merging.

    Args:
        wb: The openpyxl Workbook to add the sheet to.
        wf_num: WF number (int, float, or string).
        wf_name: Human-readable WF name.
        cp_list: Ordered list of checkpoint names for this WF.
        sn_data: List of dicts with SN info: [{serial_number, config, unit_number}, ...].
        is_chemical: Override chemical detection. If None, auto-detects from wf_num.
        records_by_sn: Optional dict mapping serial_number to
                       {(effective_cp, item): record_dict}.
        report_date: Optional date string (YYYY-MM-DD) for the Report Date cell.

    Returns:
        The created worksheet.
    """
    # Determine chemical status
    if is_chemical is None:
        is_chemical = _is_chemical_wf(wf_num)

    # Format WF number for sheet name: integer if whole number, else decimal
    try:
        num_val = float(wf_num)
        if num_val == int(num_val):
            wf_num_str = str(int(num_val))
        else:
            wf_num_str = str(num_val)
    except (TypeError, ValueError):
        wf_num_str = str(wf_num)

    sheet_name = f"Sys WF{wf_num_str}_{wf_name}"
    # Excel sheet names cannot contain: \ / * ? : [ ]
    for ch in r'\/*?:[]':
        sheet_name = sheet_name.replace(ch, '_')
    # Excel sheet names are limited to 31 characters
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    ws = wb.create_sheet(title=sheet_name)

    # --- Row indices (1-indexed in openpyxl) ---
    header_row_idx = 2   # CP merged headers + %, Completion%, Config, Unit#, S/N
    sub_header_row_idx = 3  # Item names (Cosmetic, ISB, FACT, ...)
    report_date_row = 1  # Report Date label only

    # --- Report Date in A1 ---
    date_str = report_date or datetime.date.today().isoformat()
    cell_a1 = ws.cell(row=report_date_row, column=1, value=f"Report Date: {date_str}")
    cell_a1.font = Font(italic=True, size=9)

    # --- Fixed columns: %, Completion%, Config, Unit #, S/N (cols 1-5) ---
    # Merge each fixed header vertically across header_row and sub_header_row
    fixed_headers = ["%", "Completion%", "Config", "Unit #", "S/N"]
    for col_idx, header in enumerate(fixed_headers, start=1):
        # Merge row 2 + row 3 for each fixed column
        ws.merge_cells(
            start_row=header_row_idx, start_column=col_idx,
            end_row=sub_header_row_idx, end_column=col_idx
        )
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
        cell.font = _SUB_HEADER_FONT
        cell.fill = _SUB_HEADER_FILL
        cell.alignment = _CENTER_ALIGN
        # Use thick border on all sides for the merged fixed headers
        border = Border(
            left=_THICK_SIDE if col_idx == 1 else _THIN_SIDE,
            right=_THICK_SIDE if col_idx == 5 else _THIN_SIDE,
            top=_THICK_SIDE,
            bottom=_THICK_SIDE,
        )
        cell.border = border
        # Apply border/fill to the bottom cell of the merge as well (row 3)
        bottom_cell = ws.cell(row=sub_header_row_idx, column=col_idx)
        bottom_cell.border = border
        bottom_cell.fill = _SUB_HEADER_FILL

    # --- Chemical column (if applicable) ---
    # Normal WF: CP data starts at col 6
    # Chemical WF: col 6 = "Chemical", CP data starts at col 7
    cp_start_col = 6  # Column 6

    if is_chemical:
        # Merge Chemical header vertically across both header rows
        ws.merge_cells(
            start_row=header_row_idx, start_column=cp_start_col,
            end_row=sub_header_row_idx, end_column=cp_start_col
        )
        cell = ws.cell(row=header_row_idx, column=cp_start_col, value="Chemical")
        cell.font = _CP_HEADER_FONT
        cell.fill = _CP_HEADER_FILL
        cell.alignment = _CENTER_ALIGN
        cell.border = _CP_HEADER_BORDER
        # Apply border to bottom cell of merge
        sub_cell = ws.cell(row=sub_header_row_idx, column=cp_start_col)
        sub_cell.border = _CP_HEADER_BORDER
        sub_cell.fill = _CP_HEADER_FILL
        cp_start_col = 7  # CPs start at col 7 for chemical WFs

    # --- CP group headers (row 2) and check item sub-headers (row 3) ---
    current_col = cp_start_col
    for cp_name in cp_list:
        start_col = current_col
        end_col = current_col + _ITEMS_PER_CP - 1

        # Merge CP name across 6 columns in header row
        if _ITEMS_PER_CP > 1:
            ws.merge_cells(
                start_row=header_row_idx, start_column=start_col,
                end_row=header_row_idx, end_column=end_col
            )
        cp_cell = ws.cell(row=header_row_idx, column=start_col, value=cp_name)
        cp_cell.font = _CP_HEADER_FONT
        cp_cell.fill = _CP_HEADER_FILL
        cp_cell.alignment = _CENTER_ALIGN
        cp_cell.border = _CP_HEADER_BORDER

        # Apply thick borders to ALL cells in the merged CP range
        # (openpyxl only renders the top-left cell's border; we must set
        #  borders on every cell in the merge for the outline to display)
        for merge_col in range(start_col, end_col + 1):
            c = ws.cell(row=header_row_idx, column=merge_col)
            c.border = Border(
                left=_THICK_SIDE if merge_col == start_col else _THIN_SIDE,
                right=_THICK_SIDE if merge_col == end_col else _THIN_SIDE,
                top=_THICK_SIDE,
                bottom=_THICK_SIDE,
            )
            c.fill = _CP_HEADER_FILL

        # Write check item sub-headers in row 3
        for item_offset, item_name in enumerate(CHECK_ITEMS):
            is_last_item = (item_offset == _ITEMS_PER_CP - 1)
            sub_cell = ws.cell(row=sub_header_row_idx, column=current_col + item_offset, value=item_name)
            sub_cell.font = _SUB_HEADER_FONT
            sub_cell.fill = _SUB_HEADER_FILL
            sub_cell.alignment = _CENTER_ALIGN
            sub_cell.border = _SUB_HEADER_BORDER_RIGHT if is_last_item else _SUB_HEADER_BORDER

        current_col = end_col + 1

    # Track total columns for width setting
    total_cols = current_col - 1

    # --- Data rows (row 4+): write SN info columns ---
    for row_offset, sn_info in enumerate(sn_data):
        data_row = 4 + row_offset
        ws.cell(row=data_row, column=3, value=sn_info.get("config", ""))
        ws.cell(row=data_row, column=4, value=sn_info.get("unit_number", ""))
        ws.cell(row=data_row, column=5, value=sn_info.get("serial_number", ""))

    # --- Populate results with color coding (if records provided) ---
    if records_by_sn is not None and sn_data:
        populate_wf_data(ws, sn_data, cp_list, records_by_sn, is_chemical)

    # --- Column widths ---
    for col in range(1, 6):  # %, Completion%, Config, Unit#, S/N
        ws.column_dimensions[get_column_letter(col)].width = 15
    for col in range(6, total_cols + 1):  # Data columns
        ws.column_dimensions[get_column_letter(col)].width = 10

    # --- Freeze panes: row 3, column 6 (headers + SN columns stay visible) ---
    ws.freeze_panes = ws.cell(row=4, column=6)

    return ws


def populate_wf_data(ws, sn_data: list[dict], cp_list: list[str],
                     records_by_sn: dict, is_chemical: bool = False):
    """Populate data rows with SN results, apply color coding and full formatting.

    Matches the original xlsxwriter output:
    - All cells: thin border, center aligned, vertical center
    - Config column: merged vertically for same config, thick right border
    - Unit# column: thin border, center aligned
    - S/N column: thick right border
    - Last item in each CP group: thick right border
    - Last row in each config group: thick bottom border
    - Data cells: color fills for PASS/FAIL/Strife

    Args:
        ws: The openpyxl worksheet to populate.
        sn_data: List of dicts with SN info: [{serial_number, config, unit_number}, ...].
        cp_list: Ordered list of checkpoint names for this WF.
        records_by_sn: Dict mapping serial_number to a dict of
                       {(effective_cp, item): record_dict}.
        is_chemical: Whether this is a Chemical WF (shifts CP columns by 1).
    """
    # Fixed columns: %(1), Completion%(2), Config(3), Unit#(4), S/N(5)
    # CP data starts at col 6 (normal) or col 7 (chemical)
    cp_start_col = 6  # Column 6

    if is_chemical:
        cp_start_col = 7  # Chemical column at 6, CPs start at 7

    # Pre-compute config group boundaries (for thick bottom borders and merging)
    config_groups: list[tuple[int, int, str]] = []  # (start_row_offset, end_row_offset, config)
    if sn_data:
        current_config = sn_data[0].get("config", "")
        group_start = 0
        for i, sn_info in enumerate(sn_data):
            cfg = sn_info.get("config", "")
            if cfg != current_config:
                config_groups.append((group_start, i - 1, current_config))
                current_config = cfg
                group_start = i
        config_groups.append((group_start, len(sn_data) - 1, current_config))

    # Build a set of row offsets that are last in their config group
    last_in_config = set()
    for (_, end_offset, _) in config_groups:
        last_in_config.add(end_offset)

    # Populate data for each SN
    for row_offset, sn_info in enumerate(sn_data):
        data_row = 4 + row_offset
        sn = sn_info.get("serial_number", "")
        is_last_in_group = row_offset in last_in_config

        # --- Config column (col 1): will be merged later ---
        # --- Config column (col 3) ---
        config_cell = ws.cell(row=data_row, column=3)
        config_cell.value = sn_info.get("config", "")
        config_cell.alignment = _CENTER_ALIGN
        config_cell.border = _make_data_border(
            thick_right=True, thick_bottom=is_last_in_group
        )

        # --- Unit# column (col 4) ---
        unit_cell = ws.cell(row=data_row, column=4)
        unit_cell.value = sn_info.get("unit_number", "")
        unit_cell.alignment = _CENTER_ALIGN
        unit_cell.border = _make_data_border(
            thick_right=False, thick_bottom=is_last_in_group
        )

        # --- S/N column (col 5): thick right border ---
        sn_cell = ws.cell(row=data_row, column=5)
        sn_cell.value = sn
        sn_cell.alignment = _CENTER_ALIGN
        sn_cell.border = _make_data_border(
            thick_right=True, thick_bottom=is_last_in_group
        )

        # --- Data columns ---
        sn_records = records_by_sn.get(sn, {})
        current_col = cp_start_col
        for cp_name in cp_list:
            for item_offset, item_name in enumerate(CHECK_ITEMS):
                col = current_col + item_offset
                is_last_item = (item_offset == _ITEMS_PER_CP - 1)
                record = sn_records.get((cp_name, item_name))

                cell = ws.cell(row=data_row, column=col)
                cell.alignment = _CENTER_ALIGN
                cell.border = _make_data_border(
                    thick_right=is_last_item, thick_bottom=is_last_in_group
                )

                if record is not None:
                    status = record.get("status", "")
                    failure_type = record.get("failure_type")
                    cell.value = status

                    # Apply color coding
                    if status == "PASS":
                        cell.fill = PASS_FILL
                        cell.font = PASS_FONT
                    elif failure_type == "strife_fail":
                        cell.fill = STRIFE_FAIL_FILL
                        cell.font = STRIFE_FAIL_FONT
                    elif failure_type == "spec_fail":
                        cell.fill = SPEC_FAIL_FILL

            current_col += _ITEMS_PER_CP

    # --- Merge config cells vertically for same config (col 3) ---
    for (start_offset, end_offset, config_val) in config_groups:
        start_row = 4 + start_offset
        end_row = 4 + end_offset
        if start_row < end_row:
            ws.merge_cells(
                start_row=start_row, start_column=3,
                end_row=end_row, end_column=3
            )
            # Re-apply formatting to merged cell (openpyxl requires this)
            merged_cell = ws.cell(row=start_row, column=3)
            merged_cell.value = config_val
            merged_cell.alignment = _CENTER_ALIGN
            merged_cell.border = _make_data_border(thick_right=True, thick_bottom=True)


# ---------------------------------------------------------------------------
# Raw Record Storage
# ---------------------------------------------------------------------------


def store_raw_records(records: list[dict], summary: dict):
    """Store all parsed CSV records into the database for historical querying.

    Creates an import_batches entry with statistics, then inserts each record
    into raw_check_item_records. For FAIL records, test_params is stored as
    a JSON string; for PASS records, test_params is NULL.

    Args:
        records: List of parsed record dicts (with effective_cp set).
                 Each must have: serial_number, rel_event, effective_cp, item,
                 status, end_time, failing_tests, station_id, version,
                 source_file, test_params.
        summary: Dict with statistics: {file_count, record_count, valid_sn_count}.

    Returns:
        The import_batch_id of the created batch.
    """
    conn = db.get_conn()
    try:
        today = datetime.date.today().isoformat()
        now = datetime.datetime.now().isoformat()

        # Create import_batches entry
        cur = conn.execute(
            """INSERT INTO import_batches
               (import_date, created_at, file_count, record_count, valid_sn_count, status)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (
                today,
                now,
                summary.get('file_count', 0),
                summary.get('record_count', 0),
                summary.get('valid_sn_count', 0),
                'completed',
            ),
        )
        batch_id = cur.lastrowid

        # Insert each record into raw_check_item_records
        rows = []
        for rec in records:
            status = rec.get('status', '')
            # FAIL records: store test_params as JSON; PASS records: NULL
            test_params_value = None
            if status == 'FAIL' and rec.get('test_params'):
                test_params_value = json.dumps(rec['test_params'])

            rows.append((
                batch_id,
                today,
                rec.get('serial_number', ''),
                rec.get('rel_event', ''),
                rec.get('effective_cp', ''),
                rec.get('item', ''),
                status,
                rec.get('end_time', ''),
                rec.get('failing_tests', ''),
                rec.get('station_id', ''),
                rec.get('version', ''),
                test_params_value,
                rec.get('source_file', ''),
            ))

        if rows:
            conn.executemany(
                """INSERT INTO raw_check_item_records
                   (import_batch_id, import_date, serial_number, rel_event,
                    effective_cp, item, status, end_time, failing_tests,
                    station_id, version, test_params, source_file)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                rows,
            )

        conn.commit()
        return batch_id
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Manual Edit Preservation (回填机制)
# ---------------------------------------------------------------------------


def _get_previous_results(sn_mapping: dict, cp_schedule: dict) -> dict:
    """Retrieve previous report results for manual edit preservation.

    Priority 1: Query the most recent active report's per-SN/CP/Item data from DB.
    Priority 2: If DB has no data, parse the most recent Daily Report Excel file.

    Args:
        sn_mapping: Dict of {serial_number: info_dict} from DB.
        cp_schedule: Dict of {wf_num: [cp_name, ...]} from DB.

    Returns:
        Dict of {(serial_number, cp_name, item): status_string}.
    """
    # --- Priority 1: DB query ---
    prev_results = _get_previous_results_from_db()
    if prev_results:
        return prev_results

    # --- Priority 2: Excel fallback ---
    return _get_previous_results_from_excel()


def _get_previous_results_from_db() -> dict:
    """Query the most recent active report's check-item results from DB.

    Looks at sn_check_results table for the latest active report, falling back
    to sn_check_state_history if sn_check_results has no data.

    Returns:
        Dict of {(serial_number, cp_name, item): status_string}, or empty dict.
    """
    conn = db.get_conn()
    try:
        # Find the most recent active report
        report = conn.execute(
            """SELECT id FROM reports
               WHERE is_active = 1
               ORDER BY report_date DESC, version DESC
               LIMIT 1"""
        ).fetchone()
        if not report:
            return {}

        report_id = report['id']

        # Try sn_check_results first (legacy table, may have data)
        rows = conn.execute(
            """SELECT sn, check_item, status
               FROM sn_check_results
               WHERE report_id = ?""",
            (report_id,),
        ).fetchall()

        if rows:
            # sn_check_results has cp_idx — we need to resolve cp_name
            # Re-query with cp_name from report_cps
            rows = conn.execute(
                """SELECT scr.sn, rcp.cp_name, scr.check_item, scr.status
                   FROM sn_check_results scr
                   JOIN report_cps rcp
                     ON rcp.report_id = scr.report_id
                    AND rcp.wf_num = scr.wf_num
                    AND rcp.cp_idx = scr.cp_idx
                   WHERE scr.report_id = ?""",
                (report_id,),
            ).fetchall()
            if rows:
                result = {}
                for r in rows:
                    key = (r['sn'], r['cp_name'], r['check_item'])
                    result[key] = r['status']
                return result

        # Fallback: sn_check_state_history (lifecycle table, currently open rows)
        rows = conn.execute(
            """SELECT h.sn, rcp.cp_name, h.check_item, h.status
               FROM sn_check_state_history h
               JOIN current_cp_definitions rcp
                 ON rcp.wf_num = h.wf_num AND rcp.cp_idx = h.cp_idx
               WHERE h.closed_before_report_id IS NULL
                 AND h.last_seen_report_id = ?""",
            (report_id,),
        ).fetchall()
        if rows:
            result = {}
            for r in rows:
                key = (r['sn'], r['cp_name'], r['check_item'])
                result[key] = r['status']
            return result

        return {}
    finally:
        conn.close()


def _get_previous_results_from_excel() -> dict:
    """Parse the most recent M60 EVT Rel Daily Report Excel file from rawdata.

    Reads Row 1 for CP names (forward-filled across merged cells),
    Row 2 for item names, Row 3+ for data.

    Returns:
        Dict of {(serial_number, cp_name, item): status_string}, or empty dict.
    """
    # Find the most recent Daily Report file
    pattern = os.path.join(RAWDATA_DIR, 'M60 EVT Rel Daily Report_*.xlsx')
    report_files = glob.glob(pattern)
    if not report_files:
        # Also check subdirectories
        pattern = os.path.join(RAWDATA_DIR, '**', 'M60 EVT Rel Daily Report_*.xlsx')
        report_files = glob.glob(pattern, recursive=True)

    if not report_files:
        return {}

    # Sort by modification time, most recent first
    report_files.sort(key=os.path.getmtime, reverse=True)
    latest_file = report_files[0]

    result = {}
    try:
        wb = load_workbook(latest_file, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data = list(ws.iter_rows(values_only=True))
            if len(rows_data) < 3:
                continue

            # Row 0 (index 0): CP names (merged → forward fill)
            cp_row = list(rows_data[0])
            # Row 1 (index 1): Item names
            item_row = list(rows_data[1])

            # Forward-fill CP names (merged cells show as None after first)
            last_cp = None
            for i in range(len(cp_row)):
                if cp_row[i] is not None and str(cp_row[i]).strip():
                    val = str(cp_row[i]).strip()
                    # Skip Report Date cell and header labels
                    if 'Report Date' in val:
                        cp_row[i] = None
                        continue
                    last_cp = val
                    cp_row[i] = last_cp
                else:
                    cp_row[i] = last_cp

            # Data rows start at index 2
            # Determine SN column: look for column with "S/N" in item_row
            sn_col_idx = None
            for ci, val in enumerate(item_row):
                if val and str(val).strip().upper() in ('S/N', 'SN', 'SERIAL'):
                    sn_col_idx = ci
                    break
            # Fallback: SN is typically column 2 (0-indexed)
            if sn_col_idx is None:
                sn_col_idx = 2

            # Data columns start after the SN column
            data_start_col = sn_col_idx + 1

            for row_idx in range(2, len(rows_data)):
                row = rows_data[row_idx]
                if not row or len(row) <= sn_col_idx:
                    continue

                sn = str(row[sn_col_idx] or '').strip()
                if not sn or sn.lower() == 'nan' or sn.lower() == 's/n':
                    continue

                for col_idx in range(data_start_col, len(row)):
                    val = row[col_idx]
                    if val is None or str(val).strip() == '':
                        continue

                    # Get CP name and item name for this column
                    cp_name = cp_row[col_idx] if col_idx < len(cp_row) else None
                    item_name = item_row[col_idx] if col_idx < len(item_row) else None

                    if not cp_name or not item_name:
                        continue

                    cp_name = str(cp_name).strip()
                    item_name = str(item_name).strip()

                    # Skip non-item columns
                    if item_name not in CHECK_ITEMS:
                        continue

                    status_str = str(val).strip()
                    result[(sn, cp_name, item_name)] = status_str

        wb.close()
    except Exception:
        # If parsing fails, return empty — don't block report generation
        return {}

    return result


# ---------------------------------------------------------------------------
# Daily Report Orchestrator
# ---------------------------------------------------------------------------


def generate_daily_report(csv_files: list) -> tuple[bytes, str, dict]:
    """Orchestrate the full pipeline: read DB definitions → parse CSVs → aggregate → generate Excel.

    Processes uploaded CSV files against DB definitions to produce a Daily Report
    Excel workbook compatible with engine.py.

    Args:
        csv_files: List of file-like objects with `.filename` and `.read()` attributes
                   (e.g., Flask FileStorage objects).

    Returns:
        A tuple of (excel_bytes, filename, summary):
            - excel_bytes: The generated Excel workbook as bytes.
            - filename: The output filename (M60 EVT Rel Daily Report_YYYYMMDD.xlsx).
            - summary: Dict with statistics {file_count, record_count, valid_sn_count, warnings}.

    Raises:
        ValueError: If Base files haven't been uploaded (no SN mapping available).
    """
    # --- Step 1: Read DB definitions ---
    sn_result = get_sn_mapping_from_db()
    if sn_result is None:
        raise ValueError("请先上传 Base 文件 (SN mapping)")

    sn_mapping = sn_result['sn_mapping']
    valid_sns = set(sn_mapping.keys())

    sn_to_wf, sn_to_unit, sn_to_config = get_sn_lookup_dicts()

    # Read CP schedule and test plan/WF names from DB
    conn = db.get_conn()
    try:
        cp_defs = db.get_current_cp_definitions(conn)
        wf_names = db.get_current_wf_definitions(conn)
        test_plan = db.get_current_test_definitions(conn)
    finally:
        conn.close()

    # Build cp_schedule: {wf_id: [cp_name, ...]}
    cp_schedule = {}
    for wf_num, cp_list in cp_defs.items():
        cp_schedule[wf_num] = [cp['cp_name'] for cp in cp_list]

    # --- Step 2: Parse CSV files ---
    all_records: list[dict] = []
    warnings: list[str] = []
    file_count = 0

    for csv_file in csv_files:
        filename = csv_file.filename if hasattr(csv_file, 'filename') else str(csv_file)
        item_type = identify_csv_type(filename)

        if item_type is None:
            warnings.append(f"Unrecognized file type: {filename}")
            continue

        # Read file content
        content = csv_file.read()
        if hasattr(csv_file, 'seek'):
            csv_file.seek(0)  # Reset for potential re-read

        records = parse_csv_file(content, item_type)

        if not records:
            warnings.append(f"No SerialNumber column or no data found in: {filename}")
            continue

        # Filter valid SNs
        valid_records = filter_valid_sns(records, valid_sns)

        # Tag source file
        for rec in valid_records:
            rec['source_file'] = filename

        all_records.extend(valid_records)
        file_count += 1

    if not all_records:
        raise ValueError("所有 CSV 文件中未找到有效数据")

    valid_sn_count = len(set(r['serial_number'] for r in all_records))

    # --- Step 3: Aggregate ---
    # Sort and filter TFINAL
    sorted_records = sort_and_filter_tfinal(all_records)

    # Attribute anomaly CPs
    attributed_records = attribute_anomaly_cps(sorted_records)

    # Deduplicate (last record wins per SN/CP/Item)
    deduped_records = deduplicate_records(attributed_records)

    # Detect strife failures
    final_records = detect_strife_failures(deduped_records, cp_schedule)

    # --- Step 3.5: Merge previous results (回填机制) ---
    prev_results = _get_previous_results(sn_mapping, cp_schedule)
    if prev_results:
        # Build a lookup of existing records for fast matching
        existing_keys: dict[tuple[str, str, str], int] = {}
        for idx, rec in enumerate(final_records):
            key = (rec.get('serial_number', ''), rec.get('effective_cp', ''), rec.get('item', ''))
            existing_keys[key] = idx

        for key, prev_status in prev_results.items():
            sn, cp, item = key

            if key in existing_keys:
                # Override with previous value
                rec = final_records[existing_keys[key]]
                rec['status'] = prev_status
                # Update failure_type based on new status
                if 'FAIL' in prev_status.upper():
                    rec['failure_type'] = 'strife_fail' if _is_strife_failure(cp) else 'spec_fail'
                else:
                    rec['failure_type'] = None
            else:
                # Append as new record (manual edit that has no CSV counterpart)
                # Only add if SN is valid
                if sn in valid_sns:
                    final_records.append({
                        'serial_number': sn,
                        'effective_cp': cp,
                        'item': item,
                        'status': prev_status,
                        'failure_type': (
                            'strife_fail' if ('FAIL' in prev_status.upper() and _is_strife_failure(cp))
                            else ('spec_fail' if 'FAIL' in prev_status.upper() else None)
                        ),
                        'rel_event': cp,
                        'end_time': '',
                        'failing_tests': '',
                        'station_id': '',
                        'version': '',
                        'test_params': None,
                        'source_file': 'manual_edit',
                    })

    # --- Step 4: Generate Excel workbook ---
    wb = Workbook()
    # Remove the default sheet created by openpyxl
    if wb.active:
        wb.remove(wb.active)

    # Determine report date for the header
    all_end_times = [r.get('end_time', '') for r in all_records if r.get('end_time')]
    if all_end_times:
        latest_end_time = max(all_end_times)
        try:
            dt = datetime.datetime.fromisoformat(latest_end_time.replace('Z', '+00:00'))
        except (ValueError, AttributeError):
            try:
                for fmt in ('%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%m/%d/%Y %H:%M:%S',
                            '%Y-%m-%d %H:%M:%S.%f', '%m/%d/%Y %I:%M:%S %p'):
                    try:
                        dt = datetime.datetime.strptime(latest_end_time, fmt)
                        break
                    except ValueError:
                        continue
                else:
                    dt = datetime.datetime.now()
            except Exception:
                dt = datetime.datetime.now()
        date_str = dt.strftime('%Y%m%d')
        report_date_display = dt.strftime('%Y-%m-%d')
    else:
        dt = datetime.datetime.now()
        date_str = dt.strftime('%Y%m%d')
        report_date_display = dt.strftime('%Y-%m-%d')

    # Determine which WFs to generate sheets for (sorted numerically)
    wf_nums_sorted = sorted(cp_schedule.keys(), key=lambda x: float(x))

    for wf_num in wf_nums_sorted:
        cp_list = cp_schedule.get(wf_num, [])
        if not cp_list:
            continue

        # Get WF name
        wf_name = wf_names.get(wf_num, "")

        # Get SNs assigned to this WF
        sns_for_wf = [sn for sn, info in sn_mapping.items() if str(info['wf_id']) == str(wf_num)]

        if not sns_for_wf:
            continue

        # Sort SNs by config then unit_number (natural sort)
        def _natural_sort_key(sn):
            config = sn_to_config.get(sn, '')
            unit = sn_to_unit.get(sn, '')
            # Natural sort: split into text/number parts
            parts = [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', str(unit))]
            return (config, parts)

        sns_for_wf.sort(key=_natural_sort_key)

        # Build sn_data list
        sn_data = []
        for sn in sns_for_wf:
            sn_data.append({
                'serial_number': sn,
                'config': sn_to_config.get(sn, ''),
                'unit_number': sn_to_unit.get(sn, ''),
            })

        # Build records_by_sn lookup for this WF's SNs
        records_by_sn: dict[str, dict[tuple[str, str], dict]] = {}
        for rec in final_records:
            sn = rec.get('serial_number', '')
            if sn in set(sns_for_wf):
                if sn not in records_by_sn:
                    records_by_sn[sn] = {}
                key = (rec.get('effective_cp', ''), rec.get('item', ''))
                records_by_sn[sn][key] = rec

        # Create the WF sheet
        create_wf_sheet(wb, wf_num, wf_name, cp_list, sn_data,
                        records_by_sn=records_by_sn, report_date=report_date_display)

    # --- Step 5: Save workbook to bytes ---
    output = io.BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()

    # --- Step 6: Determine filename (date already computed in Step 4) ---
    filename = f"M60 EVT Rel Daily Report_{date_str}.xlsx"

    # --- Step 7: Build summary ---
    summary = {
        'file_count': file_count,
        'record_count': len(final_records),
        'valid_sn_count': valid_sn_count,
        'warnings': warnings,
    }

    # --- Step 8: Store raw records in DB ---
    store_raw_records(all_records, summary)

    return excel_bytes, filename, summary
