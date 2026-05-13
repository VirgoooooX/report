"""
M60 EVT REL — Data Engine
Reads Daily Report, parses all WF sheets, returns structured analysis results.
"""
from openpyxl import load_workbook
from collections import defaultdict
import re, datetime, os
import logging

try:
    from custom_rules import failure_type_for_colors, is_wf_ignored, resolve_config_alias
except Exception:
    failure_type_for_colors = None
    is_wf_ignored = lambda _wf: False
    resolve_config_alias = lambda value: value

SKIP_SHEETS = {'Sample Delivery Tracker', 'Test Schedule', 'Test Summary', 'T0 Summary'}
CHECK_NAMES = {'Cosmetic','ISB','FACT','BT-OTA','Touch-CAL-Post','Charging','Button User','User Test','Pull-out force'}
logger = logging.getLogger(__name__)

NON_CP_HEADERS = {'Comments', 'Overall Result', 'Return to Summary'}
CONFIGS = ('R1FNF', 'R2CNM', 'R3', 'R4')
ER_UNIT_CONFIG_MAP = {
    '1': 'R1FNF',
    '2': 'R2CNM',
    '3': 'R3',
    '4': 'R4',
}


class RawDataValidationError(Exception):
    """Raised when Daily Report rawdata has blocking consistency problems."""

    def __init__(self, errors):
        self.errors = errors or []
        message = _validation_error_message(self.errors)
        super().__init__(message)


def _validation_error_message(errors):
    if not errors:
        return 'Rawdata validation failed'
    first = errors[0]
    return (
        "Rawdata validation failed: "
        f"{first.get('sheet', '')} row {first.get('row', '')} "
        f"SN {first.get('sn', '')} has failure at {first.get('failed_cp', '')} "
        f"but later data at {first.get('later_cp', '')}"
    )


def is_cp_header_label(value):
    """Return True when a header cell is a real CP label, not a boundary/check column."""
    if not value or not isinstance(value, str):
        return False
    label = value.strip()
    if len(label) <= 1:
        return False
    if label in CHECK_NAMES or label in NON_CP_HEADERS:
        return False
    if re.sub(r'[^a-z0-9]+', '', label.lower()) == 't0':
        return False
    return True


def infer_config_from_unit_num(unit_num):
    """Infer config for cleaning-style sheets whose Config column is blank."""
    match = re.match(r'^ER([1-4])(?:\D|$)', str(unit_num or '').strip(), re.I)
    return ER_UNIT_CONFIG_MAP.get(match.group(1)) if match else None


def resolve_row_config(explicit_config, default_config=None, unit_num=None):
    cfg = str(explicit_config).strip() if explicit_config is not None else ''
    cfg = resolve_config_alias(cfg)
    if cfg in CONFIGS:
        return cfg
    fallback = resolve_config_alias(default_config) if default_config else ''
    if fallback in CONFIGS:
        return fallback
    return infer_config_from_unit_num(unit_num)


def has_pre_cp_data(ws, row_idx, first_cp_col):
    """Return True when S/N-adjacent baseline/T0 columns contain real data."""
    for col_idx in range(6, max(6, first_cp_col)):
        value = ws.cell(row_idx, col_idx).value
        if value is not None and str(value).strip() != '/':
            return True
    return False

def wf_sort_key(wfn):
    """自然排序键：WF1, WF2, ..., WF10, WF11, WF14.1, WF14.2"""
    try:
        parts = str(wfn).split('.')
        return tuple(int(p) if p.isdigit() else float('inf') for p in parts)
    except:
        return (float('inf'),)

# ── helpers ────────────────────────────────────────────────────────────

def normalize(s):
    return {w for w in re.sub(r'[^a-z0-9\s]',' ',s.lower()).split() if len(w)>1 and not w.isdigit()}

def wf_num(name):
    m = re.search(r'WF(\d+\.?\d*)', name)
    return m.group(1) if m else None

def parse_result(s):
    if not s or not isinstance(s, str): return None
    m = re.match(r'(\d+)(SF)?F?/(\d+)T', s.strip())
    return (int(m.group(1)), m.group(2)=='SF', int(m.group(3))) if m else None

def get_failure_type(cell):
    """Returns 'spec', 'strife', or None based on cell fill color."""
    if failure_type_for_colors:
        try:
            configured = failure_type_for_colors(cell_color_rgb(cell), font_color_rgb(cell))
            if configured:
                return configured
        except Exception:
            logger.exception("Failed to evaluate custom failure color rules")
    try:
        fg = cell.fill.fgColor
        if fg and fg.type == 'rgb':
            rgb = str(fg.rgb).upper()
            if 'FF0000' in rgb: return 'spec'
            if 'FFFF00' in rgb: return 'strife'
    except: pass
    try:
        fc = cell.font.color
        if fc and fc.type == 'rgb':
            rgb = str(fc.rgb).upper()
            if 'FF9C0006' in rgb: return 'spec'
    except: pass
    return None


def cell_color_rgb(cell):
    try:
        fg = cell.fill.fgColor
        if fg and fg.type == 'rgb':
            return str(fg.rgb).upper()
    except Exception:
        pass
    return ''


def font_color_rgb(cell):
    try:
        fc = cell.font.color
        if fc and fc.type == 'rgb':
            return str(fc.rgb).upper()
    except Exception:
        pass
    return ''


def normalize_cell_value(value):
    if value is None:
        return ''
    return str(value).strip()


def status_from_cell(cell, is_pending=False):
    if is_pending:
        return 'pending', None
    raw = normalize_cell_value(cell.value)
    if raw == '/':
        return 'skip', None
    failure_type = get_failure_type(cell)
    if failure_type == 'spec':
        return 'spec_fail', 'spec'
    if failure_type == 'strife':
        return 'strife_fail', 'strife'
    if raw:
        return 'pass', None
    return 'pending', None


def aggregate_cp_status(check_rows, cp_idx, last_real):
    rows = [r for r in check_rows if r['cp_idx'] == cp_idx]
    if cp_idx > last_real:
        return 'pending', None, 0
    if not rows:
        return 'pending', None, 0
    if all(r['status'] == 'skip' for r in rows):
        return 'skip', None, 0
    has_data = any(r['status'] not in ('skip', 'pending') for r in rows)
    if any(r['failure_type'] == 'spec' for r in rows):
        return 'spec_fail', 'spec', int(has_data)
    if any(r['failure_type'] == 'strife' for r in rows):
        return 'strife_fail', 'strife', int(has_data)
    if has_data:
        return 'pass', None, 1
    return 'pending', None, 0


def _limited_cells(cells, limit=6):
    return cells[:limit]


def _cell_info(cell):
    return {
        'cell': cell.coordinate,
        'value': normalize_cell_value(cell.value),
        'fill': cell_color_rgb(cell),
        'font': font_color_rgb(cell),
    }


def _cp_row_state(ws, row_idx, cp_idx, cp_range):
    ps, pe, cp_name = cp_range
    failure_cells = []
    data_cells = []
    raw_cells = []
    for cc in range(ps, pe + 1):
        cell = ws.cell(row_idx, cc)
        raw = normalize_cell_value(cell.value)
        if raw:
            raw_cells.append(_cell_info(cell))
        ft = get_failure_type(cell)
        if ft:
            info = _cell_info(cell)
            info['type'] = ft
            failure_cells.append(info)
        if raw and raw != '/':
            data_cells.append(_cell_info(cell))

    if failure_cells:
        state = 'fail'
    elif data_cells:
        state = 'data'
    else:
        state = 'empty_or_skip'

    return {
        'cp_idx': cp_idx,
        'cp_name': cp_name,
        'state': state,
        'failure_cells': failure_cells,
        'data_cells': data_cells,
        'raw_cells': raw_cells,
    }


def _find_rawdata_anomalies_for_sheet(ws, wf_num, ts_names, report_date='', source_file_name=''):
    errors = []
    r = 1
    while r <= ws.max_row:
        if ws.cell(r, 3).value == 'Config' and ws.cell(r, 4).value == 'Unit #':
            cp_list = []
            ls = None
            ln = ''
            for c in range(7, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and v.strip():
                    cv = v.strip()
                    if is_cp_header_label(cv):
                        if ls is not None:
                            cp_list.append((ls, c - 1, ln))
                        ls = c
                        ln = cv
            if ls is not None:
                ec = ws.max_column
                for c in range(ls, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v and isinstance(v, str) and v.strip() in ('Comments', 'Overall Result'):
                        ec = c - 1
                        break
                cp_list.append((ls, ec, ln))

            if not cp_list:
                r += 1
                continue

            cp_names = [(ps, pn) for ps, pe, pn in cp_list]
            cp_test_map = map_cps_to_tests(cp_names, ts_names)

            dr = r + 2
            default_config = None
            while dr <= ws.max_row:
                cv = ws.cell(dr, 3).value
                cv = resolve_config_alias(cv)
                if cv in CONFIGS:
                    default_config = cv
                    break
                dr += 1

            dr = r + 2
            while dr <= ws.max_row:
                dc1 = ws.cell(dr, 1).value
                dcfg = ws.cell(dr, 3).value
                if dc1 == '%' and dcfg == 'Config' and ws.cell(dr, 4).value == 'Unit #':
                    break

                sn = ws.cell(dr, 5).value
                unit_num = ws.cell(dr, 4).value
                if dcfg == 'Config' or (sn and str(sn).strip() == 'S/N') or not sn or not has_pre_cp_data(ws, dr, cp_list[0][0]):
                    dr += 1
                    continue

                row_cfg = resolve_row_config(dcfg, default_config, unit_num)
                if not row_cfg:
                    dr += 1
                    continue

                states = [_cp_row_state(ws, dr, pi, cp_range) for pi, cp_range in enumerate(cp_list)]
                for fail_idx, fail_state in enumerate(states):
                    if fail_state['state'] != 'fail':
                        continue
                    fail_test_idx = cp_test_map[fail_idx] if fail_idx < len(cp_test_map) else None
                    gap_count = 0
                    for later_idx in range(fail_idx + 1, len(states)):
                        later_state = states[later_idx]
                        later_test_idx = cp_test_map[later_idx] if later_idx < len(cp_test_map) else None
                        if later_test_idx != fail_test_idx:
                            break
                        if later_state['state'] == 'empty_or_skip':
                            gap_count += 1
                            continue
                        if later_state['state'] == 'data' and gap_count >= 3:
                            errors.append({
                                'code': 'failure_followed_by_gapped_later_data',
                                'severity': 'error',
                                'report_date': report_date,
                                'source_file': source_file_name,
                                'sheet': ws.title,
                                'wf': wf_num,
                                'row': dr,
                                'sn': str(sn).strip(),
                                'unit_num': normalize_cell_value(unit_num),
                                'config': row_cfg,
                                'failed_cp_idx': fail_idx,
                                'failed_cp': fail_state['cp_name'],
                                'later_cp_idx': later_idx,
                                'later_cp': later_state['cp_name'],
                                'failed_cells': _limited_cells(fail_state['failure_cells']),
                                'later_cells': _limited_cells(later_state['data_cells']),
                            })
                            break
                        gap_count = 0
                dr += 1
            r = dr
        else:
            r += 1
    return errors


def validate_rawdata_workbook(wb, ts_test_names, report_date='', source_file_name=''):
    """Return blocking rawdata consistency errors for a Daily Report workbook."""
    errors = []
    for name in wb.sheetnames:
        if name in SKIP_SHEETS or name.startswith('MLB'):
            continue
        wfn = wf_num(name)
        if not wfn or is_wf_ignored(wfn):
            continue
        ts_names = ts_test_names.get(wfn, [''])
        if ts_names == ['']:
            ts_names = ['(unnamed)']
        try:
            errors.extend(_find_rawdata_anomalies_for_sheet(
                wb[name], wfn, ts_names, report_date=report_date, source_file_name=source_file_name
            ))
        except Exception:
            logger.exception("Failed to validate WF sheet %s", name)
    return errors


def validate_daily_report(path, report_date=None, source_file_name=None):
    """Validate one Daily Report file and raise RawDataValidationError on blocking rawdata issues."""
    import os as _os

    source_file_name = source_file_name or _os.path.basename(path)
    wb = load_workbook(path, data_only=True)
    try:
        _, _, ts_test_names, _ = read_test_summary_from_workbook(wb)
        errors = validate_rawdata_workbook(
            wb,
            ts_test_names,
            report_date=report_date or '',
            source_file_name=source_file_name,
        )
        if errors:
            raise RawDataValidationError(errors)
        return []
    finally:
        wb.close()

# ── CP-to-Test mapping ─────────────────────────────────────────────────

def map_cps_to_tests(cp_names, ts_names):
    """
    Maps each CP to a Test Summary test index.
    cp_names: list of (start_col, name) tuples
    ts_names: list of TS test name strings
    Returns: list[int] — test index per CP, same length as cp_names
    """
    nt = len(ts_names)
    
    # Step 1: Group CPs by name similarity
    groups = []
    cg = 0
    for i, (_, cn) in enumerate(cp_names):
        if i == 0: groups.append(0); continue
        pn = cp_names[i-1][1]; pl = pn.lower(); cl = cn.lower()
        
        if cl.startswith('2nd') or pl.startswith('2nd'): cg += 1
        elif cl.startswith('3rd') or pl.startswith('3rd'): cg += 1
        elif 'margin' in cl and 'margin' not in pl: cg += 1
        elif re.search(r'(\d+)%', pn) and re.search(r'(\d+)%', cn):
            pp = int(re.search(r'(\d+)%', pn).group(1))
            cp_ct = int(re.search(r'(\d+)%', cn).group(1))
            if pp <= 100 and cp_ct >= 120: cg += 1
        elif 'ltos' in pl and 'ltos' not in cl: cg += 1
        elif 'ltos' not in pl and 'ltos' in cl: cg += 1
        elif len(normalize(pn) & normalize(cn)) == 0: cg += 1
        groups.append(cg)
    
    ng = cg + 1
    ts_words = [normalize(tn) for tn in ts_names]
    g2t = {}
    
    # Step 2: Match groups to TS tests
    for g in range(ng):
        g_cps = [cp_names[i][1] for i, gp in enumerate(groups) if gp == g]
        best_test, best_score = None, 0
        for ti, tsw in enumerate(ts_words):
            total = 0
            for cn in g_cps:
                cpw = normalize(cn)
                total += len(cpw & tsw)
                if ts_names[ti].lower() in cn.lower(): total += 10
            if total > best_score: best_score = total; best_test = ti
        if best_score >= 2: g2t[g] = best_test
    
    # Step 3: Deduplicate
    used = set()
    for g in range(ng):
        if g in g2t:
            if g2t[g] in used: del g2t[g]
            else: used.add(g2t[g])
    
    # Step 4: Forced split if groups < tests
    if ng < nt:
        cpt = len(cp_names) // nt
        if cpt > 0:
            for i in range(len(cp_names)): groups[i] = min(i // cpt, nt - 1)
            ng = nt; g2t = {g: g for g in range(nt)}
    
    # Step 5: Fill unmapped groups sequentially
    if not g2t:
        for g in range(ng): g2t[g] = min(g, nt - 1)
    else:
        sm = sorted(g2t.keys()); fa = sm[0]; lt = g2t[fa]
        for g in range(fa + 1, ng):
            if g in g2t: lt = g2t[g]
            else: lt = min(lt + 1, nt - 1); g2t[g] = lt
        lt = g2t[fa]
        for g in range(fa - 1, -1, -1):
            if g in g2t: lt = g2t[g]
            else: lt = max(lt - 1, 0); g2t[g] = lt
    
    return [min(g2t[g], nt - 1) for g in groups]


def attach_test_idx_to_cps(cp_structures, ts_test_names):
    """Attach test_idx to each CP definition using the same mapping as result parsing."""
    result = {}
    for wfn, cps in cp_structures.items():
        ts_names = ts_test_names.get(wfn, ['(unnamed)'])
        cp_names = [(cp['cp_idx'], cp['cp_name']) for cp in cps]
        cp_test_map = map_cps_to_tests(cp_names, ts_names)
        mapped = []
        for cp, test_idx in zip(cps, cp_test_map):
            item = dict(cp)
            item['test_idx'] = test_idx
            mapped.append(item)
        result[wfn] = mapped
    return result

# ── Test Summary reader ────────────────────────────────────────────────

def read_test_summary(daily_path):
    """Reads the Test Summary sheet from Daily Report.
    Returns: (ts_data, ts_qty, ts_test_names, ts_num_tests)
      ts_data: {wf: {cfg: {phase_idx: {name, result}}}}
      ts_qty: {wf: {cfg: quantity}}
      ts_test_names: {wf: [test_name, ...]}
      ts_num_tests: {wf: n}
    """
    wb = load_workbook(daily_path, data_only=True)
    result = read_test_summary_from_workbook(wb)
    wb.close()
    return result


def read_test_summary_from_workbook(wb):
    """Same as read_test_summary() but operates on an already-opened workbook."""
    ws = wb['Test Summary']

    sys_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 2).value == 'System': sys_row = r; break
    if not sys_row:
        return {}, {}, {}, {}

    config_col_map = {}
    for ps in [3, 8, 13, 18]:
        for o, c in enumerate(['R1FNF', 'R2CNM', 'R3', 'R4']):
            config_col_map[ps + o] = c

    test_cols = [7, 12, 17, 22]
    ts_data = {}
    ts_qty = {}
    ts_test_names = {}
    ts_num_tests = {}

    for r in range(14, 75):
        wf_val = ws.cell(r, 2).value
        if wf_val is None: continue
        if isinstance(wf_val, str):
            if wf_val == 'MLB': break
            continue
        if not isinstance(wf_val, (int, float)): continue
        wf_str = str(int(wf_val)) if isinstance(wf_val, float) and wf_val == int(wf_val) else str(wf_val)

        ts_qty[wf_str] = {}
        for offset, cfg in enumerate(['R1FNF', 'R2CNM', 'R3', 'R4']):
            qv = ws.cell(r, 3 + offset).value
            if qv is not None: ts_qty[wf_str][cfg] = int(qv) if isinstance(qv, (int, float)) else 0

        tests_for_wf = []
        for ti, tc in enumerate(test_cols):
            tname = ws.cell(r, tc).value
            if tname is None or not str(tname).strip(): break
            tname = str(tname).strip()
            tests_for_wf.append(tname)
            for offset, cfg in enumerate(['R1FNF', 'R2CNM', 'R3', 'R4']):
                rc = tc + 1 + offset
                val = ws.cell(r, rc).value
                if val is not None:
                    ts_data.setdefault(wf_str, {}).setdefault(cfg, {})[ti] = {'name': tname, 'result': val}

        if tests_for_wf:
            ts_test_names[wf_str] = tests_for_wf
            ts_num_tests[wf_str] = len(tests_for_wf)

    return ts_data, ts_qty, ts_test_names, ts_num_tests


def extract_cp_structure(ws, header_row=1, cp_range_start=7):
    """Extracts CP names and check items from a WF sheet header.

    CP names are in header_row (e.g. row with Config|Unit#).
    Check items are in the next row below (header_row + 1).

    Returns: [{'cp_idx': int, 'cp_name': str, 'check_items': [str]}, ...]
    """
    cp_list = []
    ls = None
    ln = ''
    # Pass 1: find CP boundaries from header_row
    for c in range(cp_range_start, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v and isinstance(v, str) and v.strip():
            cv = v.strip()
            if is_cp_header_label(cv):
                if ls is not None:
                    cp_list.append((ls, c - 1, ln))
                ls = c
                ln = cv
    if ls is not None:
        ec = ws.max_column
        for c in range(ls, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v and isinstance(v, str) and v.strip() in ('Comments', 'Overall Result'):
                ec = c - 1
                break
        cp_list.append((ls, ec, ln))

    # Pass 2: extract check items from the row BELOW header_row
    check_row = header_row + 1
    result = []
    for pi, (ps, pe, cp_name) in enumerate(cp_list):
        check_items = []
        for c in range(ps, pe + 1):
            v = ws.cell(check_row, c).value
            if v and isinstance(v, str) and v.strip():
                cv = v.strip()
                if cv in CHECK_NAMES:
                    check_items.append(cv)
        result.append({'cp_idx': pi, 'cp_name': cp_name, 'check_items': check_items})
    return result


def read_test_schedule(daily_path):
    """Reads Test Schedule sheet B/C columns and returns {wf_num: wf_name} mapping."""
    wb = load_workbook(daily_path, data_only=True)
    result = read_test_schedule_from_workbook(wb)
    wb.close()
    return result


def read_test_schedule_from_workbook(wb):
    """Same as read_test_schedule() but operates on an already-opened workbook."""
    ws = wb['Test Schedule']
    names = {}
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if b is not None and isinstance(b, (int, float)):
            wfn = str(int(b)) if isinstance(b, float) and b == int(b) else str(b)
            if c and isinstance(c, str) and c.strip():
                names[wfn] = c.strip()
    return names


def _schedule_normalize(value):
    return re.sub(r'[^a-z0-9]+', '', str(value or '').strip().lower().replace('×', 'x'))


def _schedule_percent_value(label):
    text = str(label).strip()
    if re.fullmatch(r'\d+(?:\.\d+)?', text):
        number = float(text)
        if 0 < number <= 3:
            return int(round(number * 100))
    match = re.search(r'(\d+(?:\.\d+)?)\s*%', text)
    if match:
        return int(round(float(match.group(1))))
    return None


def _schedule_cp_percent(cp_name):
    match = re.search(r'(\d+(?:\.\d+)?)\s*%', str(cp_name))
    return int(round(float(match.group(1)))) if match else None


def _schedule_label_words(label):
    words = [w.lower() for w in re.findall(r'[A-Za-z]+', str(label))]
    return [w for w in words if w not in {'h', 'hr', 'hrs', 'cyc', 'cycle', 'cycles', 'drop', 'drops', 'min', 'cm', 'k'}]


def _schedule_drop_number(label):
    text = str(label).strip().lower()
    match = re.search(r'(\d+)(?:st|nd|rd|th)?\s*drop', text)
    if match:
        return int(match.group(1))
    match = re.search(r'(\d+)\s*drops?', text)
    if match:
        return int(match.group(1))
    return None


def _schedule_profile_number(label):
    match = re.fullmatch(r'p\s*(\d+)', str(label).strip(), re.I)
    return int(match.group(1)) if match else None


def _schedule_axis_cycle(label):
    match = re.fullmatch(r'([XY])\s*-\s*(\d+)', str(label).strip(), re.I)
    if not match:
        return None
    return match.group(1).lower(), match.group(2)


def _schedule_cycle_number(label):
    match = re.fullmatch(r'cycle\s*(\d+)', str(label).strip(), re.I)
    return int(match.group(1)) if match else None


def _schedule_suffix_number(label, prefix):
    match = re.fullmatch(rf'{prefix}\s*(\d+)', str(label).strip(), re.I)
    return int(match.group(1)) if match else None


def _schedule_marker_signature(label):
    percent = _schedule_percent_value(label)
    if percent is not None:
        return ('percent', percent)

    drop_num = _schedule_drop_number(label)
    if drop_num is not None:
        return ('drop', drop_num)

    op_num = _schedule_suffix_number(label, 'op')
    if op_num is not None:
        return ('op', op_num)

    minute_num = _schedule_suffix_number(label, '')
    if minute_num is not None and 'min' in str(label).lower():
        return ('minute', minute_num)

    profile_num = _schedule_profile_number(label)
    if profile_num is not None:
        return ('profile', profile_num)

    cycle_num = _schedule_cycle_number(label)
    if cycle_num is not None:
        return ('cycle', cycle_num)

    return ('text', _schedule_normalize(label))


def _schedule_is_reset(previous_signature, current_signature):
    if not previous_signature or not current_signature:
        return False

    if previous_signature[0] == 'percent' and current_signature[0] == 'percent':
        return previous_signature[1] <= 100 and current_signature[1] > 100

    if previous_signature[0] == current_signature[0] and previous_signature[0] in {'drop', 'op', 'minute', 'profile', 'cycle'}:
        return current_signature[1] < previous_signature[1]

    return False


def _schedule_marker_candidates(label, cp_list, test_names=None):
    normalized = _schedule_normalize(label)
    if normalized in {'t0', 'end'}:
        return []

    matches = []
    percent = _schedule_percent_value(label)
    words = _schedule_label_words(label)
    if percent is not None:
        for cp in cp_list:
            if _schedule_cp_percent(cp['cp_name']) != percent:
                continue
            cp_name_lower = str(cp['cp_name']).lower()
            if words:
                if all(word in cp_name_lower for word in words):
                    matches.append(cp['test_idx'])
            else:
                matches.append(cp['test_idx'])
        if matches:
            return sorted(set(matches))
        if words:
            for cp in cp_list:
                cp_name_lower = str(cp['cp_name']).lower()
                if all(word in cp_name_lower for word in words):
                    matches.append(cp['test_idx'])
            if matches:
                return sorted(set(matches))

    for cp in cp_list:
        cp_normalized = _schedule_normalize(cp['cp_name'])
        if normalized and normalized in cp_normalized:
            matches.append(cp['test_idx'])
    if matches:
        return sorted(set(matches))

    drop_num = _schedule_drop_number(label)
    if drop_num is not None:
        token = f'drop{drop_num}'
        for cp in cp_list:
            if token in _schedule_normalize(cp['cp_name']):
                matches.append(cp['test_idx'])
        if matches:
            return sorted(set(matches))

    profile_num = _schedule_profile_number(label)
    if profile_num is not None:
        token = f'profile{profile_num}'
        for cp in cp_list:
            if token in _schedule_normalize(cp['cp_name']):
                matches.append(cp['test_idx'])
        if matches:
            return sorted(set(matches))

    axis_cycle = _schedule_axis_cycle(label)
    if axis_cycle is not None:
        axis, cycle = axis_cycle
        for cp in cp_list:
            cp_name = str(cp['cp_name']).lower()
            if axis in cp_name and cycle in cp_name:
                matches.append(cp['test_idx'])
        if matches:
            return sorted(set(matches))

    cycle_num = _schedule_cycle_number(label)
    if cycle_num is not None:
        token = f'cycling{cycle_num}'
        for cp in cp_list:
            if token in _schedule_normalize(cp['cp_name']):
                matches.append(cp['test_idx'])
        if matches:
            return sorted(set(matches))

    if '72' in str(label):
        for cp in cp_list:
            if '72' in str(cp['cp_name']):
                matches.append(cp['test_idx'])
        if matches:
            return sorted(set(matches))

    if test_names:
        label_lower = str(label).lower()
        if 'hsd' in label_lower:
            for test_idx, test_name in test_names.items():
                if 'hsd' in str(test_name).lower():
                    matches.append(test_idx)
            if matches:
                return sorted(set(matches))
        if normalized.startswith('cycle'):
            for test_idx, test_name in test_names.items():
                test_name_lower = str(test_name).lower()
                if 'battery' in test_name_lower or 'swap' in test_name_lower or 'cycl' in test_name_lower:
                    matches.append(test_idx)
            if matches:
                return sorted(set(matches))

    return []


def extract_test_schedule_segments(daily_path, ts_test_names, cps_by_wf):
    """Extract planned WF/Test/Config schedule segments from the Test Schedule sheet."""
    wb = load_workbook(daily_path, data_only=True)
    result = extract_test_schedule_segments_from_workbook(wb, ts_test_names, cps_by_wf)
    wb.close()
    return result


def extract_test_schedule_segments_from_workbook(wb, ts_test_names, cps_by_wf):
    """Same as extract_test_schedule_segments() but operates on an already-opened workbook."""
    if 'Test Schedule' not in wb.sheetnames:
        return []

    ws = wb['Test Schedule']
    header_row = None
    for row_idx in range(1, min(ws.max_row, 40) + 1):
        if str(ws.cell(row_idx, 2).value).strip() == 'WF' and str(ws.cell(row_idx, 3).value).strip() == 'Test Item':
            header_row = row_idx
            break
    if header_row is None:
        return []

    date_columns = []
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_idx).value
        if isinstance(value, datetime.datetime):
            date_columns.append((col_idx, value.date().isoformat()))

    schedule_rows = []
    current_wf = None
    current_test_item = None
    current_duration = None
    config_columns = {'R1FNF': 6, 'R2CNM': 7, 'R3': 8, 'R4': 9}

    for row_idx in range(header_row + 2, ws.max_row + 1):
        wf_value = ws.cell(row_idx, 2).value
        test_item_value = ws.cell(row_idx, 3).value
        duration_value = ws.cell(row_idx, 4).value

        if isinstance(wf_value, str) and wf_value.strip().upper() == 'MLB':
            break

        if wf_value not in (None, ''):
            current_wf = str(int(wf_value)) if isinstance(wf_value, float) and wf_value == int(wf_value) else str(wf_value).strip()
        if test_item_value not in (None, ''):
            current_test_item = str(test_item_value).strip()
        if duration_value not in (None, ''):
            current_duration = duration_value

        config = None
        allocation = None
        for config_name, col_idx in config_columns.items():
            cell_value = ws.cell(row_idx, col_idx).value
            if cell_value not in (None, ''):
                config = config_name
                allocation = cell_value
                break

        if not current_wf or not current_test_item or not config:
            continue

        markers = []
        for col_idx, date_value in date_columns:
            marker_value = ws.cell(row_idx, col_idx).value
            if marker_value not in (None, ''):
                markers.append({'date': date_value, 'label': str(marker_value).strip()})
        if not markers:
            continue

        schedule_rows.append({
            'wf_num': current_wf,
            'config': config,
            'schedule_test_item': current_test_item,
            'duration': current_duration,
            'requested': ws.cell(row_idx, 5).value,
            'allocation': allocation,
            'source_row': row_idx,
            'markers': markers,
        })

    segments = []
    for row in schedule_rows:
        wf_num = row['wf_num']
        cp_list = sorted(cps_by_wf.get(wf_num, []), key=lambda cp: cp['cp_idx'])
        ordered_tests = sorted({cp['test_idx'] for cp in cp_list if cp.get('test_idx') is not None})
        if not ordered_tests:
            names = [name for name in ts_test_names.get(wf_num, []) if name]
            ordered_tests = list(range(len(names)))
        if not ordered_tests:
            continue

        t0_marker = next((marker for marker in row['markers'] if _schedule_normalize(marker['label']) == 't0'), None)
        end_marker = next((marker for marker in row['markers'] if _schedule_normalize(marker['label']) == 'end'), None)
        if not t0_marker or not end_marker:
            continue

        test_names = {idx: name for idx, name in enumerate(ts_test_names.get(wf_num, []))}
        if len(ordered_tests) == 1:
            only_test = ordered_tests[0]
            segments.append({
                'wf_num': wf_num,
                'config': row['config'],
                'test_idx': only_test,
                'test_name': test_names.get(only_test, f'Test{only_test + 1}'),
                'schedule_test_item': row['schedule_test_item'],
                'planned_start_date': t0_marker['date'],
                'planned_end_date': end_marker['date'],
                'source_row': row['source_row'],
                'confidence': 'high',
                'inference_reason': 'single-test-row',
                'marker_labels': ['T0', 'End'],
            })
            continue

        non_boundary_markers = [
            marker for marker in row['markers']
            if _schedule_normalize(marker['label']) not in {'t0', 'end'}
        ]
        if not non_boundary_markers:
            continue

        marker_infos = []
        anchor_positions = defaultdict(list)
        reset_positions = []
        previous_signature = None
        for marker_index, marker in enumerate(non_boundary_markers):
            signature = _schedule_marker_signature(marker['label'])
            candidates = _schedule_marker_candidates(marker['label'], cp_list, test_names)
            marker_infos.append({
                'marker': marker,
                'signature': signature,
                'candidates': candidates,
            })
            if len(candidates) == 1:
                anchor_positions[candidates[0]].append(marker_index)
            if _schedule_is_reset(previous_signature, signature):
                reset_positions.append(marker_index)
            previous_signature = signature

        boundaries = []
        boundary_reasons = []
        last_boundary = -1
        for boundary_index, test_idx in enumerate(ordered_tests[:-1]):
            next_test_idx = ordered_tests[boundary_index + 1]
            next_anchors = [pos for pos in anchor_positions.get(next_test_idx, []) if pos > last_boundary]
            current_anchors = [pos for pos in anchor_positions.get(test_idx, []) if pos > last_boundary]
            usable_resets = [pos for pos in reset_positions if pos > last_boundary + 1]

            if next_anchors:
                boundary = min(next_anchors) - 1
                reason = 'next-test-anchor'
            elif current_anchors:
                boundary = max(current_anchors)
                reason = 'current-test-anchor'
            elif usable_resets:
                boundary = min(usable_resets) - 1
                reason = 'sequence-reset'
            else:
                remaining_markers = len(non_boundary_markers) - last_boundary - 1
                remaining_boundaries = len(ordered_tests) - boundary_index
                boundary = last_boundary + max(1, remaining_markers // remaining_boundaries) - 1
                reason = 'even-split-fallback'

            boundary = max(last_boundary, min(boundary, len(non_boundary_markers) - 1))
            boundaries.append(boundary)
            boundary_reasons.append(reason)
            last_boundary = boundary

        marker_ranges = []
        range_start = 0
        for boundary in boundaries:
            marker_ranges.append((range_start, boundary))
            range_start = boundary + 1
        marker_ranges.append((range_start, len(non_boundary_markers) - 1))

        for test_position, test_idx in enumerate(ordered_tests):
            range_start, range_end = marker_ranges[test_position]
            marker_group = []
            if range_start <= range_end:
                marker_group = [
                    marker_infos[index]['marker']
                    for index in range(range_start, range_end + 1)
                ]
            if not marker_group:
                continue

            marker_labels = [marker['label'] for marker in marker_group]
            if test_idx == ordered_tests[0]:
                marker_labels.insert(0, 'T0')
                planned_start = t0_marker['date']
            else:
                planned_start = marker_group[0]['date']

            if test_idx == ordered_tests[-1]:
                marker_labels.append('End')
                planned_end = end_marker['date']
            else:
                planned_end = marker_group[-1]['date']

            confidence = 'high'
            if any(reason == 'even-split-fallback' for reason in boundary_reasons):
                confidence = 'medium'

            segments.append({
                'wf_num': wf_num,
                'config': row['config'],
                'test_idx': test_idx,
                'test_name': test_names.get(test_idx, f'Test{test_idx + 1}'),
                'schedule_test_item': row['schedule_test_item'],
                'planned_start_date': planned_start,
                'planned_end_date': planned_end,
                'source_row': row['source_row'],
                'confidence': confidence,
                'inference_reason': 'boundary-inference',
                'marker_labels': marker_labels,
            })

    return segments


# ── Core WF parser ─────────────────────────────────────────────────────

def _parse_wf_sheet(ws, wfn, ts_names):
    """
    Parses one WF detail sheet.
    Returns: {config: {test_idx: {total: N, spec_fails: [SN], strife_fails: [SN]}}}
    """
    num_tests = len(ts_names)
    results = {}
    
    r = 1
    while r <= ws.max_row:
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        if c3 == 'Config' and c4 == 'Unit #':
            # Parse CPs
            cp_list = []
            ls = None; ln = ''
            for c in range(7, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and v.strip():
                    cv = v.strip()
                    if is_cp_header_label(cv):
                        if ls is not None: cp_list.append((ls, c - 1, ln))
                        ls = c; ln = cv
            if ls is not None:
                ec = ws.max_column
                for c in range(ls, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v and isinstance(v, str) and v.strip() in ('Comments', 'Overall Result'): ec = c - 1; break
                cp_list.append((ls, ec, ln))
            if not cp_list: r += 1; continue
            
            # CP → test mapping
            cp_names = [(ps, pn) for ps, pe, pn in cp_list]
            cp_test_map = map_cps_to_tests(cp_names, ts_names)
            
            # Build check item mapping per CP (row r+1 has check item names)
            cp_check_items = {}
            for ps, pe, pn in cp_list:
                items = {}
                for c in range(ps, pe + 1):
                    v = ws.cell(r + 1, c).value
                    if v and str(v).strip():
                        item_name = str(v).strip()
                        if item_name in CHECK_NAMES or len(item_name) <= 6:
                            items[c] = item_name
                if items:
                    cp_check_items[(ps, pe)] = items
            
            # Find first data config
            dr = r + 2
            config = None
            while dr <= ws.max_row:
                cv = ws.cell(dr, 3).value
                cv = resolve_config_alias(cv)
                if cv in CONFIGS: config = cv; break
                dr += 1
            
            # Init results
            if config and config not in results:
                results[config] = {
                    ti: {'total': 0, 'spec_fails': [], 'strife_fails': [], 'failure_details': []}
                    for ti in range(num_tests)
                }
            
            # Process data rows
            dr = r + 2
            while dr <= ws.max_row:
                dc1 = ws.cell(dr, 1).value
                dcfg = ws.cell(dr, 3).value
                
                if dc1 == '%' and dcfg == 'Config':
                    if ws.cell(dr, 4).value == 'Unit #': break
                    else: dr += 1; continue
                
                sn = ws.cell(dr, 5).value
                unit_num = ws.cell(dr, 4).value
                if dcfg == 'Config' or (sn and str(sn).strip() == 'S/N'): dr += 1; continue
                if not sn or not has_pre_cp_data(ws, dr, cp_list[0][0]): dr += 1; continue
                
                row_cfg = resolve_row_config(dcfg, config, unit_num)
                if not row_cfg:
                    dr += 1
                    continue
                if row_cfg not in results:
                    results[row_cfg] = {}
                    for ti in range(num_tests):
                        results[row_cfg][ti] = {
                            'total': 0, 'spec_fails': [], 'strife_fails': [],
                            'failure_details': []
                        }
                
                # Find last real CP
                last_real = None
                for pi in range(len(cp_list) - 1, -1, -1):
                    ps, pe, _ = cp_list[pi]
                    has_data = False
                    for cc in range(ps, pe + 1):
                        cv = ws.cell(dr, cc).value
                        if cv is not None and str(cv).strip() != '/': has_data = True; break
                    if has_data: last_real = pi; break
                
                if last_real is None: dr += 1; continue
                
                test_idx = cp_test_map[last_real]
                cs, ce, cp_name = cp_list[last_real]
                
                # Count unit in all tests
                for ti in range(num_tests):
                    results[row_cfg][ti]['total'] += 1
                
                # Check fills + track which check item column failed
                has_spec = False; has_strife = False
                failed_col = None
                check_items = cp_check_items.get((cs, ce), {})
                for cc in range(cs, ce + 1):
                    ft = get_failure_type(ws.cell(dr, cc))
                    if ft == 'spec':
                        has_spec = True
                        failed_col = cc
                        break
                    if ft == 'strife':
                        has_strife = True
                        if failed_col is None:
                            failed_col = cc
                
                sn_str = str(sn).strip()
                type_str = 'spec' if has_spec else ('strife' if has_strife else None)
                
                # Location = check item name (FACT/ISB/...), fallback to test name
                failed_check_item = check_items.get(failed_col, '') if failed_col else ''
                location = failed_check_item or (
                    ts_names[test_idx] if test_idx < len(ts_names) and ts_names[test_idx]
                    else f'Test{test_idx + 1}'
                )

                if has_spec:
                    results[row_cfg][test_idx]['spec_fails'].append(sn_str)
                elif has_strife:
                    results[row_cfg][test_idx]['strife_fails'].append(sn_str)

                if type_str:
                    results[row_cfg][test_idx]['failure_details'].append({
                        'sn': sn_str,
                        'type': type_str,
                        'location': location,
                        'failed_cp': cp_name,
                    })
                
                dr += 1
            r = dr
        else:
            r += 1
    return results

# ── Main entry point ────────────────────────────────────────────────────

def analyze(daily_path):
    """
    Main entry point. Parses all WF sheets in the Daily Report.
    
    Returns:
        dict: {wf_num: {config: {test_idx: {'total': N, 'spec_fails': [SN], 'strife_fails': [SN]}}}}
    """
    _, _, ts_test_names, _ = read_test_summary(daily_path)
    wb = load_workbook(daily_path)  # no data_only — need fill colors
    result = analyze_from_workbook(wb, ts_test_names)
    wb.close()
    return result


def analyze_from_workbook(wb, ts_test_names):
    """Same as analyze() but operates on an already-opened workbook and takes ts_test_names directly."""
    all_results = {}
    for name in wb.sheetnames:
        if name in SKIP_SHEETS or name.startswith('MLB'): continue
        wfn = wf_num(name)
        if not wfn: continue
        if is_wf_ignored(wfn): continue

        ts_names = ts_test_names.get(wfn, [''])
        if ts_names == ['']: ts_names = ['(unnamed)']

        try:
            wf_data = _parse_wf_sheet(wb[name], wfn, ts_names)
            if wf_data:
                all_results[wfn] = wf_data
        except Exception:
            logger.exception("Failed to parse WF sheet %s", name)

    return all_results


# ── Export helpers ──────────────────────────────────────────────────────

def result_str(f, is_sf, t):
    if f == 0: return f"0F/{t}T"
    if is_sf: return f"{f}SF/{t}T"
    return f"{f}F/{t}T"

def build_summary_table(results, ts_data=None):
    """
    Builds a flat table of results for display.
    Returns: [{wf, cfg, test_idx, test_name, raw_result, ts_result, is_match, spec, strife, total}, ...]
    """
    rows = []
    for wfn in sorted(results.keys(), key=wf_sort_key):
        for cfg in ['R1FNF','R2CNM','R3','R4']:
            if cfg not in results[wfn]: continue
            for ti, d in results[wfn][cfg].items():
                sf = len(d['spec_fails']); stf = len(d['strife_fails']); t = d['total']
                if t == 0: continue
                
                raw = result_str(sf, sf==0 and stf>0, t)
                row = {'wf': wfn, 'cfg': cfg, 'test_idx': ti, 'raw': raw,
                       'spec': sf, 'strife': stf, 'total': t}
                
                if ts_data and wfn in ts_data and cfg in ts_data[wfn] and ti in ts_data[wfn][cfg]:
                    ts_val = ts_data[wfn][cfg][ti]['result']
                    row['test_name'] = ts_data[wfn][cfg][ti]['name']
                    ts_parsed = parse_result(str(ts_val)) if not isinstance(ts_val, datetime.datetime) else None
                    row['ts'] = str(ts_val) if not isinstance(ts_val, datetime.datetime) else 'Ongoing'
                    row['is_match'] = ts_parsed and ts_parsed[0] == sf and ts_parsed[1] == (sf==0 and stf>0) and ts_parsed[2] == t
                else:
                    row['test_name'] = f'Test{ti+1}'
                    row['ts'] = '-'
                    row['is_match'] = None
                
                rows.append(row)
    return rows

def build_failure_detail(results):
    """
    Builds a per-SN failure detail list.
    Returns: [{wf, cfg, test_idx, sn, type: 'spec'|'strife'}, ...]
    """
    failures = []
    for wfn in sorted(results.keys(), key=wf_sort_key):
        for cfg in ['R1FNF','R2CNM','R3','R4']:
            if cfg not in results[wfn]: continue
            for ti, d in results[wfn][cfg].items():
                for sn in d['spec_fails']:
                    failures.append({'wf': wfn, 'cfg': cfg, 'test_idx': ti, 'sn': sn, 'type': 'spec'})
                for sn in d['strife_fails']:
                    failures.append({'wf': wfn, 'cfg': cfg, 'test_idx': ti, 'sn': sn, 'type': 'strife'})
    return failures

# ── CP Structure Extraction ────────────────────────────────────────────────

def extract_all_cp_structures(daily_path):
    """Extracts CP names and check items from all WF sheets in a Daily Report.

    Returns: {wf_num: [{'cp_idx': int, 'cp_name': str, 'check_items': [str]}, ...]}
    """
    wb = load_workbook(daily_path, data_only=True)
    result = extract_all_cp_structures_from_workbook(wb)
    wb.close()
    return result


def extract_all_cp_structures_from_workbook(wb):
    """Same as extract_all_cp_structures() but operates on an already-opened workbook."""
    all_cps = {}
    for name in wb.sheetnames:
        if name in SKIP_SHEETS or name.startswith('MLB'):
            continue
        wfn = wf_num(name)
        if not wfn:
            continue
        if is_wf_ignored(wfn):
            continue
        ws = wb[name]
        header_row = 1
        for r in range(1, ws.max_row + 1):
            c3 = ws.cell(r, 3).value
            c4 = ws.cell(r, 4).value
            if c3 == 'Config' and c4 == 'Unit #':
                header_row = r
                break
        try:
            cps = extract_cp_structure(ws, header_row)
            if cps:
                all_cps[wfn] = cps
        except Exception:
            logger.exception("Failed to extract CP structure for sheet %s", name)
    return all_cps


# ── Per-SN CP Progress Extraction ─────────────────────────────────────────

def extract_sn_progress(daily_path):
    """
    提取每个 SN 在每个 WF 中的 CP 进度信息。

    Args:
        daily_path: Daily Report Excel 文件路径

    Returns:
        dict: {wf_num: {config: [{
            'sn': str, 'unit_num': str,
            'current_cp_idx': int, 'current_cp_name': str,
            'total_cps': int, 'test_idx': int,
            'cp_results': [{'cp_idx': int, 'cp_name': str, 'status': str}, ...]
        }, ...]}}
    """
    _, _, ts_test_names, _ = read_test_summary(daily_path)
    wb = load_workbook(daily_path)  # no data_only — need fill colors
    result = extract_sn_progress_from_workbook(wb, ts_test_names)
    wb.close()
    return result


def extract_sn_progress_from_workbook(wb, ts_test_names):
    """Same as extract_sn_progress() but operates on an already-opened workbook and takes ts_test_names directly."""
    all_progress = {}

    for name in wb.sheetnames:
        if name in SKIP_SHEETS or name.startswith('MLB'):
            continue
        wfn = wf_num(name)
        if not wfn:
            continue
        if is_wf_ignored(wfn):
            continue

        ts_names = ts_test_names.get(wfn, [''])
        if ts_names == ['']:
            ts_names = ['(unnamed)']

        try:
            progress = _extract_wf_progress(wb[name], ts_names)
            if progress:
                all_progress[wfn] = progress
        except Exception:
            logger.exception("Failed to extract SN progress for sheet %s", name)

    return all_progress


def extract_sn_fact_rows(daily_path, report_id, report_date):
    """Extract all CP/check-item fact rows from a Daily Report workbook."""
    wb = load_workbook(daily_path)
    result = extract_sn_fact_rows_from_workbook(wb, report_id, report_date)
    wb.close()
    return result


def extract_sn_fact_rows_from_workbook(wb, report_id, report_date, ts_test_names=None):
    """Same as extract_sn_fact_rows() but operates on an already-opened workbook.

    If ts_test_names is None, it will be read from the workbook.
    """
    if ts_test_names is None:
        _, _, ts_test_names, _ = read_test_summary_from_workbook(wb)
    all_cp_rows = []
    all_check_rows = []

    for name in wb.sheetnames:
        if name in SKIP_SHEETS or name.startswith('MLB'):
            continue
        wfn = wf_num(name)
        if not wfn:
            continue
        if is_wf_ignored(wfn):
            continue
        ts_names = ts_test_names.get(wfn, ['(unnamed)'])
        cp_rows, check_rows = extract_wf_fact_rows(
            wb[name],
            wf_num=wfn,
            report_id=report_id,
            report_date=report_date,
            ts_names=ts_names,
        )
        all_cp_rows.extend(cp_rows)
        all_check_rows.extend(check_rows)

    return all_cp_rows, all_check_rows


def _extract_wf_progress(ws, ts_names):
    """
    解析一个 WF sheet，提取每个 SN 的 CP 进度。

    Returns: {config: [{sn, unit_num, current_cp_idx, current_cp_name,
                        total_cps, test_idx, cp_results}, ...]}
    """
    num_tests = len(ts_names)
    results = {}  # {config: [sn_entry, ...]}

    r = 1
    while r <= ws.max_row:
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        if c3 == 'Config' and c4 == 'Unit #':
            # ── 解析 CP 列范围与名称 ──
            cp_list = []
            ls = None
            ln = ''
            for c in range(7, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and v.strip():
                    cv = v.strip()
                    if is_cp_header_label(cv):
                        if ls is not None:
                            cp_list.append((ls, c - 1, ln))
                        ls = c
                        ln = cv
            if ls is not None:
                ec = ws.max_column
                for c in range(ls, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v and isinstance(v, str) and v.strip() in ('Comments', 'Overall Result'):
                        ec = c - 1
                        break
                cp_list.append((ls, ec, ln))
            if not cp_list:
                r += 1
                continue

            # CP → Test 映射
            cp_names = [(ps, pn) for ps, pe, pn in cp_list]
            cp_test_map = map_cps_to_tests(cp_names, ts_names)
            total_cps = len(cp_list)

            # ── 找到第一个数据 config ──
            dr = r + 2
            config = None
            while dr <= ws.max_row:
                cv = ws.cell(dr, 3).value
                cv = resolve_config_alias(cv)
                if cv in CONFIGS:
                    config = cv
                    break
                dr += 1

            # ── 遍历数据行 ──
            dr = r + 2
            while dr <= ws.max_row:
                dc1 = ws.cell(dr, 1).value
                dcfg = ws.cell(dr, 3).value

                if dc1 == '%' and dcfg == 'Config':
                    if ws.cell(dr, 4).value == 'Unit #':
                        break
                    else:
                        dr += 1
                        continue

                sn = ws.cell(dr, 5).value
                unit_num = ws.cell(dr, 4).value
                if dcfg == 'Config' or (sn and str(sn).strip() == 'S/N'):
                    dr += 1
                    continue
                if not sn or not has_pre_cp_data(ws, dr, cp_list[0][0]):
                    dr += 1
                    continue

                row_cfg = resolve_row_config(dcfg, config, unit_num)
                if not row_cfg:
                    dr += 1
                    continue
                unit_num_str = str(unit_num).strip() if unit_num is not None else ''

                # ── 找到 last_real CP（当前所在 CP） ──
                last_real = None
                for pi in range(total_cps - 1, -1, -1):
                    ps, pe, _ = cp_list[pi]
                    has_data = False
                    for cc in range(ps, pe + 1):
                        cv = ws.cell(dr, cc).value
                        if cv is not None and str(cv).strip() != '/':
                            has_data = True
                            break
                    if has_data:
                        last_real = pi
                        break

                if last_real is None:
                    dr += 1
                    continue

                # ── 构建所有 CP 的结果 ──
                cp_results = []
                for pi in range(total_cps):
                    ps, pe, cp_name = cp_list[pi]

                    # 检查该 CP 列范围内所有子列
                    all_slash = True
                    has_red = False
                    has_yellow = False
                    for cc in range(ps, pe + 1):
                        cv = ws.cell(dr, cc).value
                        if cv is not None and str(cv).strip() != '/':
                            all_slash = False
                        ft = get_failure_type(ws.cell(dr, cc))
                        if ft == 'spec':
                            has_red = True
                        elif ft == 'strife':
                            has_yellow = True

                    # 判定状态（优先级：pending > skip > 颜色）
                    if pi > last_real:
                        status = 'pending'
                    elif all_slash:
                        status = 'skip'
                    elif has_red:
                        status = 'spec_fail'
                    elif has_yellow:
                        status = 'strife_fail'
                    else:
                        status = 'pass'

                    cp_results.append({
                        'cp_idx': pi,
                        'cp_name': cp_name,
                        'status': status,
                    })

                # 当前 CP 信息
                current_cp_name = cp_list[last_real][2]
                test_idx = cp_test_map[last_real]

                # 只返回至少有一个非 skip 非 pending CP 的 SN
                has_real_cp = any(cr['status'] not in ('skip', 'pending') for cr in cp_results)
                if not has_real_cp:
                    dr += 1
                    continue

                sn_entry = {
                    'sn': str(sn).strip(),
                    'unit_num': unit_num_str,
                    'current_cp_idx': last_real,
                    'current_cp_name': current_cp_name,
                    'total_cps': total_cps,
                    'test_idx': test_idx,
                    'cp_results': cp_results,
                }

                results.setdefault(row_cfg, []).append(sn_entry)
                dr += 1
            r = dr
        else:
            r += 1

    return results


def extract_wf_fact_rows(ws, wf_num, report_id, report_date, ts_names):
    """Extract relational CP and check-item fact rows from one WF sheet."""
    cp_fact_rows = []
    check_fact_rows = []

    r = 1
    while r <= ws.max_row:
        if ws.cell(r, 3).value == 'Config' and ws.cell(r, 4).value == 'Unit #':
            cp_list = []
            ls = None
            ln = ''
            for c in range(7, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and v.strip():
                    cv = v.strip()
                    if is_cp_header_label(cv):
                        if ls is not None:
                            cp_list.append((ls, c - 1, ln))
                        ls = c
                        ln = cv
            if ls is not None:
                ec = ws.max_column
                for c in range(ls, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v and isinstance(v, str) and v.strip() in ('Comments', 'Overall Result'):
                        ec = c - 1
                        break
                cp_list.append((ls, ec, ln))

            if not cp_list:
                r += 1
                continue

            cp_names = [(ps, pn) for ps, pe, pn in cp_list]
            cp_test_map = map_cps_to_tests(cp_names, ts_names)

            dr = r + 2
            default_config = None
            while dr <= ws.max_row:
                cv = ws.cell(dr, 3).value
                cv = resolve_config_alias(cv)
                if cv in CONFIGS:
                    default_config = cv
                    break
                dr += 1

            dr = r + 2
            while dr <= ws.max_row:
                dc1 = ws.cell(dr, 1).value
                dcfg = ws.cell(dr, 3).value
                if dc1 == '%' and dcfg == 'Config' and ws.cell(dr, 4).value == 'Unit #':
                    break

                sn = ws.cell(dr, 5).value
                unit_num = ws.cell(dr, 4).value
                if dcfg == 'Config' or (sn and str(sn).strip() == 'S/N') or not sn or not has_pre_cp_data(ws, dr, cp_list[0][0]):
                    dr += 1
                    continue

                row_cfg = resolve_row_config(dcfg, default_config, unit_num)
                if not row_cfg:
                    dr += 1
                    continue
                unit_num_str = str(unit_num).strip() if unit_num is not None else ''
                sn_str = str(sn).strip()

                last_real = None
                for pi in range(len(cp_list) - 1, -1, -1):
                    ps, pe, _ = cp_list[pi]
                    if any(ws.cell(dr, cc).value is not None and str(ws.cell(dr, cc).value).strip() != '/' for cc in range(ps, pe + 1)):
                        last_real = pi
                        break
                if last_real is None:
                    dr += 1
                    continue

                row_checks = []
                for pi, (ps, pe, _cp_name) in enumerate(cp_list):
                    check_idx = 0
                    for cc in range(ps, pe + 1):
                        check_name = ws.cell(r + 1, cc).value
                        if not check_name or not isinstance(check_name, str) or not check_name.strip():
                            continue
                        check_name = check_name.strip()
                        cell = ws.cell(dr, cc)
                        status, failure_type = status_from_cell(cell, is_pending=pi > last_real)
                        check_row = {
                            'report_id': report_id,
                            'report_date': report_date,
                            'wf_num': wf_num,
                            'config': row_cfg,
                            'sn': sn_str,
                            'unit_num': unit_num_str,
                            'test_idx': cp_test_map[pi],
                            'cp_idx': pi,
                            'check_item_idx': check_idx,
                            'check_item': check_name,
                            'raw_value': normalize_cell_value(cell.value),
                            'normalized_value': normalize_cell_value(cell.value),
                            'status': status,
                            'failure_type': failure_type,
                            'fill_color': cell_color_rgb(cell),
                            'font_color': font_color_rgb(cell),
                            'source_row': dr,
                            'source_col': cc,
                        }
                        row_checks.append(check_row)
                        check_idx += 1

                for pi, (_ps, _pe, _cp_name) in enumerate(cp_list):
                    status, failure_type, has_data = aggregate_cp_status(row_checks, pi, last_real)
                    cp_fact_rows.append({
                        'report_id': report_id,
                        'report_date': report_date,
                        'wf_num': wf_num,
                        'config': row_cfg,
                        'sn': sn_str,
                        'unit_num': unit_num_str,
                        'test_idx': cp_test_map[pi],
                        'cp_idx': pi,
                        'status': status,
                        'failure_type': failure_type,
                        'has_data': has_data,
                        'is_current_cp': int(pi == last_real),
                    })
                check_fact_rows.extend(row_checks)
                dr += 1
            r = dr
        else:
            r += 1

    return cp_fact_rows, check_fact_rows


# ── Consolidated single-open-workbook parsing ────────────────────────────

class DailyReportParseResult:
    """Result of parsing a Daily Report workbook once."""
    __slots__ = (
        'path', 'report_date', 'source_file_name',
        'ts_data', 'ts_qty', 'ts_test_names', 'ts_num_tests',
        'wf_names', 'cp_structures', 'mapped_cps',
        'schedule_segments', 'summary_results',
        'check_rows', 'cp_rows',
        'progress_data',
        'test_names_by_wf',
    )

    def __init__(self):
        self.path = ''
        self.report_date = ''
        self.source_file_name = ''
        self.ts_data = {}
        self.ts_qty = {}
        self.ts_test_names = {}
        self.ts_num_tests = {}
        self.wf_names = {}
        self.cp_structures = {}
        self.mapped_cps = {}
        self.schedule_segments = []
        self.summary_results = {}
        self.check_rows = []
        self.cp_rows = []
        self.progress_data = {}
        self.test_names_by_wf = {}


def parse_daily_report(path, report_date=None, source_file_name=None):
    """Parse a Daily Report excel file, opening the workbook only once.

    Returns a DailyReportParseResult with all extracted data.
    """
    import os as _os

    result = DailyReportParseResult()
    result.path = path
    result.source_file_name = source_file_name or _os.path.basename(path)

    wb = load_workbook(path, data_only=True)
    try:
        # 1. Read Test Summary (needed early for ts_test_names)
        result.ts_data, result.ts_qty, result.ts_test_names, result.ts_num_tests = \
            read_test_summary_from_workbook(wb)

        # 2. Extract report date from TS data if not provided
        if not result.report_date:
            for wf_data in result.ts_data.values():
                for cfg_data in wf_data.values():
                    if cfg_data and isinstance(cfg_data, dict):
                        first_val = next(iter(cfg_data.values()), None)
                        if isinstance(first_val, dict) and first_val.get('date_raw'):
                            result.report_date = str(first_val['date_raw'])
                            break
                    if result.report_date:
                        break
                if result.report_date:
                    break

        validation_report_date = report_date or result.report_date
        validation_errors = validate_rawdata_workbook(
            wb,
            result.ts_test_names,
            report_date=validation_report_date,
            source_file_name=result.source_file_name,
        )
        if validation_errors:
            raise RawDataValidationError(validation_errors)

        # 3. Read Test Schedule → WF names
        result.wf_names = read_test_schedule_from_workbook(wb)

        # 4. Extract CP structures
        result.cp_structures = extract_all_cp_structures_from_workbook(wb)

        # 5. Map CPs to tests
        result.mapped_cps = attach_test_idx_to_cps(result.cp_structures, result.ts_test_names)

        # 6. Extract schedule segments
        result.schedule_segments = extract_test_schedule_segments_from_workbook(
            wb, result.ts_test_names, result.mapped_cps
        )

        # 7. Build test names dict for current definitions
        result.test_names_by_wf = result.ts_test_names.copy()

        # 8. Analyze (summary results for WF failure counts)
        result.summary_results = analyze_from_workbook(wb, result.ts_test_names)

        # 9. Extract SN progress (for stats, not DB writes)
        result.progress_data = extract_sn_progress_from_workbook(wb, result.ts_test_names)

        ignored = {wf for wf in set(result.wf_names) | set(result.ts_test_names) if is_wf_ignored(wf)}
        if ignored:
            for wf in ignored:
                result.ts_test_names.pop(wf, None)
                result.ts_qty.pop(wf, None)
                result.wf_names.pop(wf, None)
                result.cp_structures.pop(wf, None)
                result.mapped_cps.pop(wf, None)
                result.summary_results.pop(wf, None)
                result.progress_data.pop(wf, None)
                result.test_names_by_wf.pop(wf, None)
            result.schedule_segments = [
                segment for segment in result.schedule_segments
                if str(segment.get('wf_num', '')) not in ignored
            ]

    finally:
        wb.close()

    return result


# ── CLI self-test ──────────────────────────────────────────────────────────
if __name__ == '__main__':
    import sys, time
    path = sys.argv[1] if len(sys.argv) > 1 else None
    if not path:
        path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'rawdata', 'M60 EVT Rel Daily Report_20260508.xlsx')
    
    print(f"Analyzing: {path}")
    t0 = time.time()
    results = analyze(path)
    dt = time.time() - t0
    print(f"Done in {dt:.1f}s — {len(results)} WFs parsed")
    
    ts_data, _, _, _ = read_test_summary(path)
    rows = build_summary_table(results, ts_data)
    matched = sum(1 for r in rows if r['is_match'] is True)
    diff = sum(1 for r in rows if r['is_match'] is False)
    print(f"Summary: {len(rows)} entries, {matched} matched TS, {diff} differ")
    
    failures = build_failure_detail(results)
    spec = sum(1 for f in failures if f['type'] == 'spec')
    strife = sum(1 for f in failures if f['type'] == 'strife')
    print(f"Failures: {len(failures)} total ({spec} spec, {strife} strife)")

    # ── Test extract_sn_progress ──
    print("\n--- extract_sn_progress ---")
    t0 = time.time()
    progress = extract_sn_progress(path)
    dt = time.time() - t0
    total_entries = 0
    for wfn in sorted(progress.keys(), key=wf_sort_key):
        for cfg in sorted(progress[wfn].keys()):
            entries = progress[wfn][cfg]
            total_entries += len(entries)
    print(f"Done in {dt:.1f}s — {len(progress)} WFs, {total_entries} SN entries")
    for wfn in sorted(progress.keys(), key=wf_sort_key):
        for cfg in sorted(progress[wfn].keys()):
            entries = progress[wfn][cfg]
            print(f"  WF{wfn} {cfg}: {len(entries)} SNs")
            for entry in entries[:2]:
                cp_last = entry['cp_results'][-1]
                print(f"    SN:{entry['sn']} Unit:{entry['unit_num']} "
                      f"CP:{entry['current_cp_idx']+1}/{entry['total_cps']} "
                      f"({entry['current_cp_name']}) Test:{entry['test_idx']} "
                      f"last_cp_status:{cp_last['status']}")
