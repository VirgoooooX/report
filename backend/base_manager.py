"""
Base Data Manager — handles upload, parsing, and storage of Base definition files.

Manages:
- SN mapping (serial_number → config, wf_id, unit_number)
- WaterfallCheckpointSchedule (wf_id → CP list)
- WaterfallTestPlan (wf_id → test names)
- Test Schedule Excel (plan schedule with date markers)
"""
import csv
import json
import os
import re
import shutil
import datetime
from collections import defaultdict

from openpyxl import load_workbook

from app_paths import RAWDATA_DIR, BASE_DIR
from schedule_parser import SCHEDULE_BOUNDARY_LABELS
import db


# ---------------------------------------------------------------------------
# File type → standardized filename mapping
# ---------------------------------------------------------------------------

FILE_TYPE_NAMES = {
    'sn_mapping': 'sn_mapping.csv',
    'checkpoint_schedule': 'checkpoint_schedule.csv',
    'test_plan': 'test_plan.csv',
    'test_schedule': 'test_schedule.xlsx',
}


# ---------------------------------------------------------------------------
# File Storage and Metadata
# ---------------------------------------------------------------------------

def _get_base_dir():
    """Return the base file storage directory, creating it if needed."""
    base_dir = os.path.join(RAWDATA_DIR, 'base')
    os.makedirs(base_dir, exist_ok=True)
    return base_dir


def upload_base_file(file_storage, file_type):
    """Upload, store, and parse a Base file.

    Stores the file to rawdata/base/ with a standardized name, parses it,
    writes definitions to DB, and records metadata in base_file_meta.
    If a file of the same type already exists, replaces it.

    Args:
        file_storage: A file-like object (e.g., Flask's FileStorage) with
                      a .filename attribute and .save() method, or a path string.
        file_type: One of 'sn_mapping', 'checkpoint_schedule', 'test_plan', 'test_schedule'.

    Returns:
        dict with parsed summary information.

    Raises:
        ValueError: If file_type is not recognized.
    """
    if file_type not in FILE_TYPE_NAMES:
        raise ValueError(f"Unknown file_type: {file_type}. Must be one of {list(FILE_TYPE_NAMES.keys())}")

    standardized_name = FILE_TYPE_NAMES[file_type]
    base_dir = _get_base_dir()
    stored_path_abs = os.path.join(base_dir, standardized_name)
    # Relative path for DB storage (relative to project root)
    stored_path_rel = os.path.join('rawdata', 'base', standardized_name)

    # Determine original filename
    if hasattr(file_storage, 'filename'):
        original_filename = file_storage.filename or standardized_name
    elif isinstance(file_storage, str):
        original_filename = os.path.basename(file_storage)
    else:
        original_filename = standardized_name

    # Delete old file if it exists
    if os.path.exists(stored_path_abs):
        os.remove(stored_path_abs)

    # Save the new file
    if hasattr(file_storage, 'save'):
        # Flask FileStorage object
        file_storage.save(stored_path_abs)
    elif isinstance(file_storage, str) and os.path.isfile(file_storage):
        # Path string — copy the file
        shutil.copy2(file_storage, stored_path_abs)
    elif hasattr(file_storage, 'read'):
        # File-like object with read()
        with open(stored_path_abs, 'wb') as f:
            data = file_storage.read()
            if isinstance(data, str):
                f.write(data.encode('utf-8'))
            else:
                f.write(data)
    else:
        raise ValueError("file_storage must be a Flask FileStorage, a file path string, or a file-like object")

    # Parse the file
    parsed_summary = _parse_and_save(file_type, stored_path_abs)

    # Record metadata in base_file_meta (replace old entry for same type)
    conn = db.get_conn()
    try:
        # Delete old metadata for this file type
        conn.execute("DELETE FROM base_file_meta WHERE file_type = ?", (file_type,))
        # Insert new metadata
        conn.execute(
            """INSERT INTO base_file_meta (file_type, original_filename, stored_path, uploaded_at, parsed_summary)
               VALUES (?, ?, ?, ?, ?)""",
            (
                file_type,
                original_filename,
                stored_path_rel,
                datetime.datetime.now().isoformat(),
                json.dumps(parsed_summary),
            )
        )
        conn.commit()
    finally:
        conn.close()

    return parsed_summary


def _parse_and_save(file_type, file_path):
    """Parse a base file and save definitions to DB.

    Re-parsing semantics: every Base upload is treated as a fresh source of
    truth and fully replaces the corresponding ``current_*`` tables (the
    ``db.save_current_*`` helpers are DELETE-then-INSERT, so re-uploading
    any single file always reflects the new content immediately).

    Cross-file re-derivation: ``parse_test_schedule`` reads
    ``current_cp_definitions`` and ``current_test_definitions`` to do
    boundary inference. So when ``checkpoint_schedule`` or ``test_plan`` is
    re-uploaded, the previously persisted ``current_schedule_segments`` is
    stale even though the underlying ``test_schedule.xlsx`` on disk has not
    changed. To keep "upload one file → API reflects it" from forcing a full
    DB rebuild, we automatically re-run ``parse_test_schedule`` against the
    on-disk ``test_schedule.xlsx`` whenever a dependency is replaced.

    Args:
        file_type: The type of base file.
        file_path: Absolute path to the stored file.

    Returns:
        dict with parsed summary statistics.
    """
    if file_type == 'sn_mapping':
        result = parse_sn_mapping(file_path)
        # No direct DB write for SN mapping — it's read on-demand from the file
        return {
            'sn_count': result['sn_count'],
            'config_count': len(result['config_quantities']),
            'configs': result['config_quantities'],
        }

    elif file_type == 'checkpoint_schedule':
        result = parse_checkpoint_schedule(file_path)
        save_cp_schedule_to_db(result)
        # CP definitions feed parse_test_schedule's boundary inference; if
        # test_schedule.xlsx is already on disk, re-derive its segments now.
        _rederive_test_schedule_if_present()
        return {
            'wf_count': result['wf_count'],
            'cp_count': result['cp_count'],
        }

    elif file_type == 'test_plan':
        result = parse_test_plan(file_path)
        save_test_plan_to_db(result)
        # Test names feed parse_test_schedule's boundary inference too.
        _rederive_test_schedule_if_present()
        return {
            'wf_count': result['wf_count'],
        }

    elif file_type == 'test_schedule':
        result = parse_test_schedule(file_path)
        save_test_schedule_to_db(result)
        return {
            'segment_count': result['segment_count'],
        }

    return {}


def _rederive_test_schedule_if_present():
    """Re-run parse_test_schedule against the on-disk test_schedule.xlsx.

    Called after a dependency (``checkpoint_schedule`` / ``test_plan``) is
    re-uploaded so ``current_schedule_segments`` reflects the new CP / test
    definitions without forcing the user to re-upload the schedule file.

    Silent no-op when no Test Schedule has been uploaded yet (the standard
    initial-bootstrap order is sn_mapping → checkpoint_schedule → test_plan →
    test_schedule, and we don't want the first three uploads to fail just
    because the fourth hasn't happened yet).
    """
    standardized_name = FILE_TYPE_NAMES.get('test_schedule')
    if not standardized_name:
        return
    stored_path = os.path.join(_get_base_dir(), standardized_name)
    if not os.path.isfile(stored_path):
        return
    try:
        result = parse_test_schedule(stored_path)
        save_test_schedule_to_db(result)
    except Exception:
        # Re-derivation is best-effort; a malformed on-disk file should not
        # block the dependency upload that just succeeded. The user can
        # always re-upload the Test Schedule explicitly to get a fresh error.
        import logging
        logging.getLogger(__name__).exception(
            "Failed to re-derive test_schedule segments after dependency upload"
        )


def get_base_status():
    """Return the current status of all base files from base_file_meta.

    Returns:
        list of dicts, one per file type, with keys:
            file_type, original_filename, stored_path, uploaded_at, parsed_summary
        Only includes file types that have been uploaded.
    """
    conn = db.get_conn()
    try:
        rows = conn.execute(
            """SELECT file_type, original_filename, stored_path, uploaded_at, parsed_summary
               FROM base_file_meta
               ORDER BY file_type"""
        ).fetchall()
    finally:
        conn.close()

    result = []
    for row in rows:
        entry = dict(row)
        # Parse the JSON summary
        try:
            entry['parsed_summary'] = json.loads(entry['parsed_summary']) if entry['parsed_summary'] else {}
        except (json.JSONDecodeError, TypeError):
            entry['parsed_summary'] = {}
        result.append(entry)

    return result


# ---------------------------------------------------------------------------
# SN Mapping Parser
# ---------------------------------------------------------------------------

def parse_sn_mapping(file_path):
    """Parse SN mapping CSV and return mapping dict + per-config quantities.

    The CSV is expected to have columns:
        serial_number, config, Product, unit_number, start_date, wf_id

    Args:
        file_path: Path to the SN mapping CSV file.

    Returns:
        dict with keys:
            'sn_mapping': {sn: {'config': str, 'wf_id': str, 'unit_number': str}}
            'config_quantities': {config: count}  — number of SNs per config
            'sn_count': int — total number of unique SNs
    """
    sn_mapping = {}
    config_quantities = defaultdict(int)

    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            sn = row.get('serial_number', '').strip()
            if not sn:
                continue

            config = row.get('config', '').strip()
            wf_id = row.get('wf_id', '').strip()
            unit_number = row.get('unit_number', '').strip()

            # Only store first occurrence if duplicate SN
            if sn not in sn_mapping:
                sn_mapping[sn] = {
                    'config': config,
                    'wf_id': wf_id,
                    'unit_number': unit_number,
                }
                config_quantities[config] += 1

    return {
        'sn_mapping': sn_mapping,
        'config_quantities': dict(config_quantities),
        'sn_count': len(sn_mapping),
    }


def get_sn_mapping_from_db():
    """Load the stored SN mapping CSV from rawdata/base/ and return parsed result.

    Returns:
        Same dict as parse_sn_mapping(), or None if no SN mapping file is stored.
    """
    conn = db.get_conn()
    try:
        row = conn.execute(
            "SELECT stored_path FROM base_file_meta WHERE file_type = 'sn_mapping' ORDER BY id DESC LIMIT 1"
        ).fetchone()
    finally:
        conn.close()

    if not row or not row['stored_path']:
        return None

    stored_path = row['stored_path']
    # Handle relative paths (stored relative to project root)
    if not os.path.isabs(stored_path):
        from app_paths import BASE_DIR
        stored_path = os.path.join(BASE_DIR, stored_path)

    if not os.path.exists(stored_path):
        return None

    return parse_sn_mapping(stored_path)


def get_valid_sns():
    """Return set of valid serial numbers from stored SN mapping.

    Returns:
        set of SN strings, or empty set if no mapping available.
    """
    result = get_sn_mapping_from_db()
    if result is None:
        return set()
    return set(result['sn_mapping'].keys())


def get_sn_lookup_dicts():
    """Return lookup dicts for SN → wf_id, unit_number, config.

    Returns:
        tuple of (sn_to_wf, sn_to_unit, sn_to_config) dicts,
        or (None, None, None) if no mapping available.
    """
    result = get_sn_mapping_from_db()
    if result is None:
        return None, None, None

    sn_to_wf = {}
    sn_to_unit = {}
    sn_to_config = {}

    for sn, info in result['sn_mapping'].items():
        sn_to_wf[sn] = info['wf_id']
        sn_to_unit[sn] = info['unit_number']
        sn_to_config[sn] = info['config']

    return sn_to_wf, sn_to_unit, sn_to_config


# ---------------------------------------------------------------------------
# Checkpoint Schedule Parser
# ---------------------------------------------------------------------------

# CPs that should be excluded from the schedule.
#
# Operational CPs (FA / STOP / RETURN) are administrative steps, not test CPs.
#
# Schedule boundary labels (T0 / REL_T0 / End / TFinal / REL_TFINAL) are
# segment markers, not test CPs — they delimit the start and end of a WF on
# the Test Schedule and must NEVER appear in current_cp_definitions. The
# boundary set is sourced from SCHEDULE_BOUNDARY_LABELS (the same set used by
# engine.is_cp_header_label and schedule_parser._is_t0_marker /
# _is_end_marker), so the Plan, Actual, and daily-header layers share a
# single fact source for what counts as a boundary.
_OPERATIONAL_EXCLUDED_CPS = frozenset([
    'REL FA RETEST',
    'SEND TO FA',
    'STOP TEST',
    'RETURN TO REL',
])

# Public exclusion set retained for downstream callers/tests. Membership is
# the literal-uppercase canonical names of operational + schedule boundary
# markers; for general filtering, prefer ``_is_excluded_cp`` which handles
# separator/case variants too.
EXCLUDED_CPS = _OPERATIONAL_EXCLUDED_CPS | frozenset([
    'T0', 'REL_T0', 'End', 'TFinal', 'REL_TFINAL',
])


def _is_schedule_boundary_cp_name(cp_name):
    return re.sub(r'[^a-z0-9]+', '', str(cp_name or '').lower()) in SCHEDULE_BOUNDARY_LABELS


def _is_excluded_cp(cp_name):
    return cp_name in _OPERATIONAL_EXCLUDED_CPS or _is_schedule_boundary_cp_name(cp_name)


def _normalize_checkpoint_test_idx(raw_test_idx):
    """Convert Base checkpoint CSV numbering to internal zero-based test_idx."""
    try:
        value = int(raw_test_idx)
    except (TypeError, ValueError):
        return 0
    if value <= 0:
        return 0
    return value - 1


_DROP_ROUND_RE = re.compile(r'_(\d+)(?:ST|ND|RD|TH)\s+DROP')


def _infer_checkpoint_test_idx(raw_test_idx, test_category='', cp_name=''):
    """Infer internal test_idx from Base checkpoint numbering and category hints.

    Args:
        raw_test_idx: Raw ``wf test`` value from the Base CSV.
        test_category: Raw ``test category`` value from the Base CSV.
        cp_name: Raw ``rel event cp`` value from the Base CSV. Used to detect
            drop-round CPs (e.g. ``SIDED DROP 1M PB SEQA_2ND DROP1``) so that
            WF14/WF15 third-round drops get promoted to Margin even when their
            ``test category`` is ``SeqA`` rather than ``Margin``.
    """
    test_idx = _normalize_checkpoint_test_idx(raw_test_idx)
    category = str(test_category or '').strip().lower()

    # Some Base checkpoint exports keep WF14/WF15 drop-margin rows at wf test=2
    # even though Test Plan defines Margin as the third test. Rows that already
    # carry wf test=3, like WF6 button-margin data, should stay untouched.
    if 'margin' in category and test_idx == 1:
        return 2

    # New rule: any drop CP whose Base row puts it at wf test=2 belongs to
    # Margin, regardless of `test category`. This catches WF14/WF15 third-round
    # drops (`SeqA` category) without disturbing WF6 button-margin (different
    # cp_name shape) or non-drop CPs.
    if test_idx == 1 and _DROP_ROUND_RE.search(str(cp_name or '')):
        return 2

    return test_idx


def _schedule_boundary_label(value):
    return re.sub(r'[^a-z0-9]+', '', str(value or '').strip().lower())


def _is_schedule_t0_label(value):
    return _schedule_boundary_label(value) in {'t0', 'relt0'}


def _is_schedule_end_label(value):
    return _schedule_boundary_label(value) in {'end', 'tfinal', 'reltfinal'}


def parse_checkpoint_schedule(file_path):
    """Parse WaterfallCheckpointSchedule CSV and return {wf_id: [cp_names]}.

    The CSV is expected to have columns:
        wf_id, wf id_cp, rel event cp, wf test, ...

    Reads the CSV sorted by (wf_id, wf id_cp), groups by wf_id, extracts
    the ordered list of CP names from 'rel event cp', and filters out
    excluded CPs (REL FA RETEST, SEND TO FA, STOP TEST, RETURN TO REL, REL_TFINAL).

    Args:
        file_path: Path to the WaterfallCheckpointSchedule CSV file.

    Returns:
        dict with keys:
            'cp_schedule': {wf_id_str: [cp_name_1, cp_name_2, ...]}
            'wf_count': int — number of unique WFs
            'cp_count': int — total number of CPs (after filtering)
            'cp_details': {wf_id_str: [{'cp_name': str, 'test_idx': int}, ...]}
                — detailed info including test_idx for each CP
    """
    # Read all rows
    rows = []
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            wf_id = row.get('wf_id', '').strip()
            cp_order = row.get('wf id_cp', '').strip()
            cp_name = row.get('rel event cp', '').strip()
            test_idx = row.get('wf test', '').strip()
            test_category = row.get('test category', '').strip()

            if not wf_id or not cp_name:
                continue

            rows.append({
                'wf_id': wf_id,
                'cp_order': int(cp_order) if cp_order else 0,
                'cp_name': cp_name,
                'test_idx': _infer_checkpoint_test_idx(test_idx, test_category, cp_name),
            })

    # Sort by wf_id (numeric) then cp_order to maintain original ordering
    rows.sort(key=lambda r: (float(r['wf_id']), r['cp_order']))

    # Group by wf_id, filter excluded CPs, deduplicate while preserving order
    cp_schedule = {}
    cp_details = {}

    current_wf = None
    current_cps = []
    current_details = []
    seen_cps = set()

    for row in rows:
        wf_id = row['wf_id']

        if wf_id != current_wf:
            # Save previous WF
            if current_wf is not None and current_cps:
                cp_schedule[current_wf] = current_cps
                cp_details[current_wf] = current_details
            # Start new WF
            current_wf = wf_id
            current_cps = []
            current_details = []
            seen_cps = set()

        cp_name = row['cp_name']

        # Filter out excluded CPs (operational + schedule boundary markers).
        # T0 / REL_T0 / End / TFinal / REL_TFINAL are WF-level boundary
        # markers on the Test Schedule, not real test CPs.
        if _is_excluded_cp(cp_name):
            continue

        # Deduplicate while maintaining order
        if cp_name not in seen_cps:
            seen_cps.add(cp_name)
            current_cps.append(cp_name)
            current_details.append({
                'cp_name': cp_name,
                'test_idx': row['test_idx'],
            })

    # Don't forget the last WF
    if current_wf is not None and current_cps:
        cp_schedule[current_wf] = current_cps
        cp_details[current_wf] = current_details

    total_cps = sum(len(cps) for cps in cp_schedule.values())

    return {
        'cp_schedule': cp_schedule,
        'wf_count': len(cp_schedule),
        'cp_count': total_cps,
        'cp_details': cp_details,
    }


# ---------------------------------------------------------------------------
# Test Plan Parser
# ---------------------------------------------------------------------------

def parse_test_plan(file_path):
    """Parse WaterfallTestPlan CSV and return {wf_id: [test_names]} + wf_names.

    The CSV is expected to have columns:
        wf id, wf test_1, wf test_2, wf test_3

    Each row defines a WF with up to 3 test names. Empty test columns are
    ignored. The WF name is derived by joining non-empty test names with " + ".

    Args:
        file_path: Path to the WaterfallTestPlan CSV file.

    Returns:
        dict with keys:
            'test_plan': {wf_id_str: [test_name_1, test_name_2, ...]}
            'wf_names': {wf_id_str: wf_name}  — derived from test names joined with " + "
            'wf_count': int — number of unique WFs
    """
    test_plan = {}
    wf_names = {}

    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            wf_id = row.get('wf id', '').strip()
            if not wf_id:
                continue

            # Collect non-empty test names in order
            test_names = []
            for col in ('wf test_1', 'wf test_2', 'wf test_3'):
                name = row.get(col, '').strip()
                if name:
                    test_names.append(name)

            if not test_names:
                continue

            test_plan[wf_id] = test_names
            wf_names[wf_id] = ' + '.join(test_names)

    return {
        'test_plan': test_plan,
        'wf_names': wf_names,
        'wf_count': len(test_plan),
    }


def save_test_plan_to_db(parsed_result, conn=None):
    """Write parsed test plan to current_test_definitions and current_wf_definitions.

    Args:
        parsed_result: Output from parse_test_plan()
        conn: Optional DB connection. If None, creates and closes its own.
    """
    owns_conn = conn is None
    if conn is None:
        conn = db.get_conn()

    try:
        test_plan = parsed_result['test_plan']
        wf_names = parsed_result['wf_names']

        # Write test definitions: {wf_num: [test_name, ...]}
        # run_id=0 as placeholder for Base-sourced definitions
        db.save_current_test_definitions(conn, 0, test_plan)

        # Write WF definitions: {wf_num: wf_name}
        db.save_current_wf_definitions(conn, 0, wf_names)

        conn.commit()
    finally:
        if owns_conn:
            conn.close()


def save_cp_schedule_to_db(parsed_result, conn=None):
    """Write parsed CP schedule to current_cp_definitions table.

    Converts the parsed result into the format expected by
    db.save_current_cp_definitions: {wf_num: [{'cp_idx', 'cp_name', 'test_idx', 'check_items'}]}

    Args:
        parsed_result: Output from parse_checkpoint_schedule()
        conn: Optional DB connection. If None, creates and closes its own.
    """
    owns_conn = conn is None
    if conn is None:
        conn = db.get_conn()

    try:
        cp_details = parsed_result['cp_details']

        # Build the format expected by save_current_cp_definitions
        cps_by_wf = {}
        for wf_id, details in cp_details.items():
            cp_list = []
            for idx, detail in enumerate(details):
                cp_list.append({
                    'cp_idx': idx,
                    'cp_name': detail['cp_name'],
                    'test_idx': detail['test_idx'],
                    'check_items': [],  # Check items are not defined in the schedule
                })
            cps_by_wf[str(wf_id)] = cp_list

        # Use run_id=None since this is from Base file upload, not a report
        # We pass 0 as a placeholder run_id for Base-sourced definitions
        db.save_current_cp_definitions(conn, 0, cps_by_wf)
        conn.commit()
    finally:
        if owns_conn:
            conn.close()


# ---------------------------------------------------------------------------
# Test Schedule Parser
# ---------------------------------------------------------------------------

def parse_test_schedule(file_path):
    """Parse Test Schedule Excel and return schedule segments.

    The Excel file is expected to have a 'Test Schedule' sheet with:
    - A header row where column B = 'WF' and column C = 'Test Item'
    - Date columns (datetime values in the header row)
    - Data rows with: WF number (col B), Test Item (col C), Duration (col D),
      Requested (col E), Config allocations (R1FNF=col6, R2CNM=col7, R3=col8, R4=col9)
    - Markers at date columns (T0, End, percentage values, CP names, etc.)

    This function delegates to schedule_parser.extract_test_schedule_segments_from_workbook
    using current DB definitions (test names and CP definitions) for boundary
    inference. If DB definitions are not yet available, it falls back to a
    simplified extraction that produces one segment per schedule row.

    Args:
        file_path: Path to the Test Schedule Excel file.

    Returns:
        dict with keys:
            'segments': list of segment dicts compatible with save_current_schedule_segments
            'segment_count': int — number of segments extracted
    """
    wb = load_workbook(file_path, data_only=True)

    if 'Test Schedule' not in wb.sheetnames:
        wb.close()
        return {
            'segments': [],
            'segment_count': 0,
        }

    # Try to use the full schedule parser with DB definitions.
    conn = db.get_conn()
    try:
        ts_test_names = db.get_current_test_definitions(conn)
        cps_by_wf_raw = db.get_current_cp_definitions(conn)
    except Exception:
        ts_test_names = {}
        cps_by_wf_raw = {}
    finally:
        conn.close()

    # Convert cps_by_wf to the format expected by engine (list of dicts with cp_idx, cp_name, test_idx)
    cps_by_wf = {}
    for wf_num, cp_list in cps_by_wf_raw.items():
        cps_by_wf[wf_num] = cp_list  # Already in correct format from get_current_cp_definitions

    if ts_test_names and cps_by_wf:
        import schedule_parser
        segments = schedule_parser.extract_test_schedule_segments_from_workbook(
            wb, ts_test_names, cps_by_wf
        )
    else:
        # Fallback: simplified extraction without boundary inference
        segments = _extract_schedule_segments_simple(wb)

    wb.close()

    return {
        'segments': segments,
        'segment_count': len(segments),
    }


def _extract_schedule_segments_simple(wb):
    """Simplified schedule segment extraction without CP/test definitions.

    Extracts one segment per (WF, config) row using T0 and End markers
    as planned_start_date and planned_end_date. Uses test_idx=0 for all
    segments since we can't infer test boundaries without definitions.

    Args:
        wb: An opened openpyxl workbook with a 'Test Schedule' sheet.

    Returns:
        list of segment dicts.
    """
    ws = wb['Test Schedule']

    # Find header row
    header_row = None
    for row_idx in range(1, min(ws.max_row, 40) + 1):
        if (str(ws.cell(row_idx, 2).value).strip() == 'WF' and
                str(ws.cell(row_idx, 3).value).strip() == 'Test Item'):
            header_row = row_idx
            break

    if header_row is None:
        return []

    # Find date columns
    date_columns = []
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_idx).value
        if isinstance(value, datetime.datetime):
            date_columns.append((col_idx, value.date().isoformat()))

    # Config column mapping
    config_columns = {'R1FNF': 6, 'R2CNM': 7, 'R3': 8, 'R4': 9}

    # Parse schedule rows
    segments = []
    current_wf = None
    current_test_item = None

    for row_idx in range(header_row + 2, ws.max_row + 1):
        wf_value = ws.cell(row_idx, 2).value
        test_item_value = ws.cell(row_idx, 3).value

        # Stop at MLB marker
        if isinstance(wf_value, str) and wf_value.strip().upper() == 'MLB':
            break

        if wf_value not in (None, ''):
            current_wf = (str(int(wf_value))
                          if isinstance(wf_value, float) and wf_value == int(wf_value)
                          else str(wf_value).strip())
        if test_item_value not in (None, ''):
            current_test_item = str(test_item_value).strip()

        # Determine config for this row
        config = None
        for config_name, col_idx in config_columns.items():
            cell_value = ws.cell(row_idx, col_idx).value
            if cell_value not in (None, ''):
                config = config_name
                break

        if not current_wf or not current_test_item or not config:
            continue

        # Find T0 and End markers in date columns
        markers = []
        for col_idx, date_value in date_columns:
            marker_value = ws.cell(row_idx, col_idx).value
            if marker_value in (None, ''):
                continue
            label = str(marker_value).strip()
            if not label:
                continue
            markers.append({'date': date_value, 'label': label})

        if not markers:
            continue

        # Look for T0 and End markers
        t0_date = None
        end_date = None
        marker_labels = []

        for marker in markers:
            marker_labels.append(marker['label'])
            if _is_schedule_t0_label(marker['label']):
                t0_date = marker['date']
            elif _is_schedule_end_label(marker['label']):
                end_date = marker['date']

        if not t0_date or not end_date:
            continue

        segments.append({
            'wf_num': current_wf,
            'config': config,
            'test_idx': 0,
            'test_name': current_test_item,
            'schedule_test_item': current_test_item,
            'planned_start_date': t0_date,
            'planned_end_date': end_date,
            'source_row': row_idx,
            'confidence': 'low',
            'inference_reason': 'simple-extraction-no-definitions',
            'marker_labels': marker_labels,
        })

    return segments


def save_test_schedule_to_db(parsed_result, conn=None):
    """Write parsed test schedule segments to current_schedule_segments table.

    Args:
        parsed_result: Output from parse_test_schedule()
        conn: Optional DB connection. If None, creates and closes its own.
    """
    owns_conn = conn is None
    if conn is None:
        conn = db.get_conn()

    try:
        segments = parsed_result['segments']
        # Use run_id=0 as placeholder for Base-sourced definitions
        db.save_current_schedule_segments(conn, 0, segments)
        conn.commit()
    finally:
        if owns_conn:
            conn.close()
