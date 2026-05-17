"""Test Schedule parser.

The Test Schedule sheet is only a source for test-level planned start/end
dates. Intermediate markers are scanned transiently to infer boundaries between
tests, but are not persisted or returned as CP dates.
"""
from collections import defaultdict
import datetime
import re

from openpyxl import load_workbook


CONFIG_COLUMNS = {'R1FNF': 6, 'R2CNM': 7, 'R3': 8, 'R4': 9}

# Single source of truth for schedule boundary labels.
# A header cell whose normalized form (lowercase, non-alphanumerics stripped)
# matches one of these is a segment boundary and must NOT be treated as a CP.
# Imported by engine.is_cp_header_label (see schedule-plan-actual-split task 3.1)
# so the boundary set never drifts between the daily-header parser and the
# Test Schedule parser.
SCHEDULE_BOUNDARY_LABELS = frozenset({'t0', 'relt0', 'end', 'tfinal', 'reltfinal'})

# Strict subset of SCHEDULE_BOUNDARY_LABELS — boundaries that ALSO carry
# Daily-report check-item results (PASS/FAIL with color). Daily ingest call
# sites pass ``allow_result_boundaries=True`` to ``engine.is_cp_header_label``
# so headers in this set are treated as real CPs in the daily lifecycle and
# in the generated Excel; everything else in SCHEDULE_BOUNDARY_LABELS stays
# excluded from the result plane.
#
# Adding a label here changes the Daily ingest contract — see
# docs/plans/2026-05-17-rel-t0-daily-report-second-cut.md and
# .kiro/steering/schedule-plan-logic.md.
DAILY_RESULT_BOUNDARY_LABELS = frozenset({'relt0'})

# Subsets used by the two boundary predicates below. Defined as strict
# subsets of SCHEDULE_BOUNDARY_LABELS so the master set remains the single
# fact source.
_T0_BOUNDARY_LABELS = frozenset({'t0', 'relt0'})
_END_BOUNDARY_LABELS = SCHEDULE_BOUNDARY_LABELS - _T0_BOUNDARY_LABELS


def _normalize(value):
    return re.sub(r'[^a-z0-9]+', '', str(value or '').strip().lower())


def _is_t0_marker(label):
    return _normalize(label) in _T0_BOUNDARY_LABELS


def _is_end_marker(label):
    return _normalize(label) in _END_BOUNDARY_LABELS


def _percent_value(label):
    match = re.search(r'(\d+(?:\.\d+)?)\s*%?$', str(label).strip())
    if not match:
        return None
    value = float(match.group(1))
    if 0 < value <= 3:
        value *= 100
    return int(round(value))


def _cp_percent(cp_name):
    match = re.search(r'(\d+(?:\.\d+)?)\s*%', str(cp_name))
    return int(round(float(match.group(1)))) if match else None


def _label_words(label):
    words = [w.lower() for w in re.findall(r'[A-Za-z]+', str(label))]
    return [
        w for w in words
        if w not in {'h', 'hr', 'hrs', 'cyc', 'cycle', 'cycles', 'drop', 'drops', 'min', 'cm', 'k'}
    ]


def _drop_number(label):
    text = str(label).strip().lower()
    match = re.search(r'(\d+)\s*(?:st|nd|rd|th)?\s*drop', text)
    if match:
        return int(match.group(1))
    match = re.search(r'(\d+)\s*drops?', text)
    if match:
        return int(match.group(1))
    return None


def _profile_number(label):
    match = re.fullmatch(r'p\s*(\d+)', str(label).strip(), re.I)
    return int(match.group(1)) if match else None


def _axis_cycle(label):
    match = re.fullmatch(r'([XY])\s*-\s*(\d+)', str(label).strip(), re.I)
    if not match:
        return None
    return match.group(1).lower(), match.group(2)


def _cycle_number(label):
    match = re.fullmatch(r'cycle\s*(\d+)', str(label).strip(), re.I)
    return int(match.group(1)) if match else None


def _suffix_number(label, prefix):
    match = re.fullmatch(rf'{prefix}\s*(\d+)', str(label).strip(), re.I)
    return int(match.group(1)) if match else None


def _marker_signature(label):
    percent = _percent_value(label)
    if percent is not None:
        return ('percent', percent)

    drop_num = _drop_number(label)
    if drop_num is not None:
        return ('drop', drop_num)

    op_num = _suffix_number(label, 'op')
    if op_num is not None:
        return ('op', op_num)

    minute_num = _suffix_number(label, '')
    if minute_num is not None and 'min' in str(label).lower():
        return ('minute', minute_num)

    profile_num = _profile_number(label)
    if profile_num is not None:
        return ('profile', profile_num)

    cycle_num = _cycle_number(label)
    if cycle_num is not None:
        return ('cycle', cycle_num)

    return ('text', _normalize(label))


def _is_reset(previous_signature, current_signature):
    if not previous_signature or not current_signature:
        return False

    if previous_signature[0] == 'percent' and current_signature[0] == 'percent':
        return previous_signature[1] <= 100 and current_signature[1] > 100

    if previous_signature[0] == current_signature[0] and previous_signature[0] in {
        'drop', 'op', 'minute', 'profile', 'cycle'
    }:
        return current_signature[1] < previous_signature[1]

    return False


def _marker_candidates(label, cp_list, test_names=None):
    normalized = _normalize(label)
    if _is_t0_marker(label) or _is_end_marker(label):
        return []

    matches = []
    percent = _percent_value(label)
    words = _label_words(label)
    if percent is not None:
        for cp in cp_list:
            if _cp_percent(cp['cp_name']) != percent:
                continue
            cp_name_lower = str(cp['cp_name']).lower()
            if words:
                if all(word in cp_name_lower for word in words):
                    matches.append(cp['test_idx'])
            else:
                matches.append(cp['test_idx'])
        if matches:
            return sorted(set(matches))
        if words:
            for cp in cp_list:
                cp_name_lower = str(cp['cp_name']).lower()
                if all(word in cp_name_lower for word in words):
                    matches.append(cp['test_idx'])
            if matches:
                return sorted(set(matches))

    drop_num = _drop_number(label)
    if drop_num is not None:
        token = f'drop{drop_num}'
        for cp in cp_list:
            if token in _normalize(cp['cp_name']):
                matches.append(cp['test_idx'])
        if matches:
            return sorted(set(matches))

    for cp in cp_list:
        cp_normalized = _normalize(cp['cp_name'])
        if normalized and normalized in cp_normalized:
            matches.append(cp['test_idx'])
    if matches:
        return sorted(set(matches))

    profile_num = _profile_number(label)
    if profile_num is not None:
        token = f'profile{profile_num}'
        for cp in cp_list:
            if token in _normalize(cp['cp_name']):
                matches.append(cp['test_idx'])
        if matches:
            return sorted(set(matches))

    axis = _axis_cycle(label)
    if axis is not None:
        axis_name, cycle = axis
        for cp in cp_list:
            cp_name = str(cp['cp_name']).lower()
            if axis_name in cp_name and cycle in cp_name:
                matches.append(cp['test_idx'])
        if matches:
            return sorted(set(matches))

    cycle_num = _cycle_number(label)
    if cycle_num is not None:
        token = f'cycling{cycle_num}'
        for cp in cp_list:
            if token in _normalize(cp['cp_name']):
                matches.append(cp['test_idx'])
        if matches:
            return sorted(set(matches))

    if '72' in str(label):
        for cp in cp_list:
            if '72' in str(cp['cp_name']):
                matches.append(cp['test_idx'])
        if matches:
            return sorted(set(matches))

    if test_names:
        label_lower = str(label).lower()
        if 'hsd' in label_lower:
            for test_idx, test_name in test_names.items():
                if 'hsd' in str(test_name).lower():
                    matches.append(test_idx)
            if matches:
                return sorted(set(matches))
        if normalized.startswith('cycle'):
            for test_idx, test_name in test_names.items():
                test_name_lower = str(test_name).lower()
                if 'battery' in test_name_lower or 'swap' in test_name_lower or 'cycl' in test_name_lower:
                    matches.append(test_idx)
            if matches:
                return sorted(set(matches))

    return []


def extract_test_schedule_segments(daily_path, ts_test_names, cps_by_wf):
    """Extract planned WF/Test/Config schedule segments from the Test Schedule sheet."""
    wb = load_workbook(daily_path, data_only=True)
    try:
        return extract_test_schedule_segments_from_workbook(wb, ts_test_names, cps_by_wf)
    finally:
        wb.close()


def extract_test_schedule_segments_from_workbook(wb, ts_test_names, cps_by_wf):
    """Extract only test-level planned start/end dates from an open workbook."""
    if 'Test Schedule' not in wb.sheetnames:
        return []

    ws = wb['Test Schedule']
    header_row = None
    for row_idx in range(1, min(ws.max_row, 40) + 1):
        if str(ws.cell(row_idx, 2).value).strip() == 'WF' and str(ws.cell(row_idx, 3).value).strip() == 'Test Item':
            header_row = row_idx
            break
    if header_row is None:
        return []

    date_columns = []
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_idx).value
        if isinstance(value, datetime.datetime):
            date_columns.append((col_idx, value.date().isoformat()))

    schedule_rows = []
    current_wf = None
    current_test_item = None
    current_duration = None

    for row_idx in range(header_row + 2, ws.max_row + 1):
        wf_value = ws.cell(row_idx, 2).value
        test_item_value = ws.cell(row_idx, 3).value
        duration_value = ws.cell(row_idx, 4).value

        if isinstance(wf_value, str) and wf_value.strip().upper() == 'MLB':
            break

        if wf_value not in (None, ''):
            current_wf = str(int(wf_value)) if isinstance(wf_value, float) and wf_value == int(wf_value) else str(wf_value).strip()
        if test_item_value not in (None, ''):
            current_test_item = str(test_item_value).strip()
        if duration_value not in (None, ''):
            current_duration = duration_value

        config = None
        allocation = None
        for config_name, col_idx in CONFIG_COLUMNS.items():
            cell_value = ws.cell(row_idx, col_idx).value
            if cell_value not in (None, ''):
                config = config_name
                allocation = cell_value
                break

        if not current_wf or not current_test_item or not config:
            continue

        markers = []
        for col_idx, date_value in date_columns:
            marker_value = ws.cell(row_idx, col_idx).value
            if marker_value in (None, ''):
                continue
            label = str(marker_value).strip()
            if label:
                markers.append({'date': date_value, 'label': label})
        if not markers:
            continue

        schedule_rows.append({
            'wf_num': current_wf,
            'config': config,
            'schedule_test_item': current_test_item,
            'duration': current_duration,
            'requested': ws.cell(row_idx, 5).value,
            'allocation': allocation,
            'source_row': row_idx,
            'markers': markers,
        })

    segments = []
    for row in schedule_rows:
        wf_num = row['wf_num']
        cp_list = sorted(cps_by_wf.get(wf_num, []), key=lambda cp: cp['cp_idx'])
        ordered_tests = sorted({cp['test_idx'] for cp in cp_list if cp.get('test_idx') is not None})
        if not ordered_tests:
            names = [name for name in ts_test_names.get(wf_num, []) if name]
            ordered_tests = list(range(len(names)))
        if not ordered_tests:
            continue

        t0_marker = next((marker for marker in row['markers'] if _is_t0_marker(marker['label'])), None)
        end_marker = next((marker for marker in row['markers'] if _is_end_marker(marker['label'])), None)
        if not t0_marker or not end_marker:
            continue

        test_names = {idx: name for idx, name in enumerate(ts_test_names.get(wf_num, []))}
        if len(ordered_tests) == 1:
            only_test = ordered_tests[0]
            segments.append({
                'wf_num': wf_num,
                'config': row['config'],
                'test_idx': only_test,
                'test_name': test_names.get(only_test, f'Test{only_test + 1}'),
                'schedule_test_item': row['schedule_test_item'],
                'planned_start_date': t0_marker['date'],
                'planned_end_date': end_marker['date'],
                'source_row': row['source_row'],
                'confidence': 'high',
                'inference_reason': 'single-test-row',
                'marker_labels': [marker['label'] for marker in row['markers']],
            })
            continue

        non_boundary_markers = [
            marker for marker in row['markers']
            if not _is_t0_marker(marker['label']) and not _is_end_marker(marker['label'])
        ]
        if not non_boundary_markers:
            continue

        marker_infos = []
        anchor_positions = defaultdict(list)
        reset_positions = []
        previous_signature = None
        for marker_index, marker in enumerate(non_boundary_markers):
            signature = _marker_signature(marker['label'])
            candidates = _marker_candidates(marker['label'], cp_list, test_names)
            marker_infos.append({
                'marker': marker,
                'signature': signature,
                'candidates': candidates,
            })
            if len(candidates) == 1:
                anchor_positions[candidates[0]].append(marker_index)
            if _is_reset(previous_signature, signature):
                reset_positions.append(marker_index)
            previous_signature = signature

        boundaries = []
        boundary_reasons = []
        last_boundary = -1
        for boundary_index, test_idx in enumerate(ordered_tests[:-1]):
            next_test_idx = ordered_tests[boundary_index + 1]
            next_anchors = [pos for pos in anchor_positions.get(next_test_idx, []) if pos > last_boundary]
            current_anchors = [pos for pos in anchor_positions.get(test_idx, []) if pos > last_boundary]
            usable_resets = [pos for pos in reset_positions if pos > last_boundary + 1]

            next_start = min(next_anchors) if next_anchors else None
            bounded_current_anchors = [
                pos for pos in current_anchors
                if next_start is None or pos < next_start
            ]

            if bounded_current_anchors:
                boundary = max(bounded_current_anchors)
                reason = 'current-test-anchor'
            elif next_start is not None and next_start > last_boundary + 1:
                boundary = next_start - 1
                reason = 'next-test-anchor'
            elif usable_resets:
                boundary = min(usable_resets) - 1
                reason = 'sequence-reset'
            elif next_start is not None:
                boundary = next_start - 1
                reason = 'next-test-anchor'
            else:
                remaining_markers = len(non_boundary_markers) - last_boundary - 1
                remaining_boundaries = len(ordered_tests) - boundary_index
                boundary = last_boundary + max(1, remaining_markers // remaining_boundaries) - 1
                reason = 'even-split-fallback'

            boundary = max(last_boundary, min(boundary, len(non_boundary_markers) - 1))
            boundaries.append(boundary)
            boundary_reasons.append(reason)
            last_boundary = boundary

        marker_ranges = []
        range_start = 0
        for boundary in boundaries:
            marker_ranges.append((range_start, boundary))
            range_start = boundary + 1
        marker_ranges.append((range_start, len(non_boundary_markers) - 1))

        for test_position, test_idx in enumerate(ordered_tests):
            range_start, range_end = marker_ranges[test_position]
            marker_group = []
            if range_start <= range_end:
                marker_group = [
                    marker_infos[index]['marker']
                    for index in range(range_start, range_end + 1)
                ]
            if not marker_group:
                continue

            marker_labels = [marker['label'] for marker in marker_group]
            if test_idx == ordered_tests[0]:
                marker_labels.insert(0, 'T0')
                planned_start = t0_marker['date']
            else:
                planned_start = marker_group[0]['date']

            if test_idx == ordered_tests[-1]:
                marker_labels.append('End')
                planned_end = end_marker['date']
            else:
                planned_end = marker_group[-1]['date']

            confidence = 'high'
            if any(reason == 'even-split-fallback' for reason in boundary_reasons):
                confidence = 'medium'

            segments.append({
                'wf_num': wf_num,
                'config': row['config'],
                'test_idx': test_idx,
                'test_name': test_names.get(test_idx, f'Test{test_idx + 1}'),
                'schedule_test_item': row['schedule_test_item'],
                'planned_start_date': planned_start,
                'planned_end_date': planned_end,
                'source_row': row['source_row'],
                'confidence': confidence,
                'inference_reason': 'boundary-inference',
                'marker_labels': marker_labels,
            })

    return segments
