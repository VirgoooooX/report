"""
M60 EVT REL — Data Engine
Reads Daily Report, parses all WF sheets, returns structured analysis results.
"""
from openpyxl import load_workbook
from collections import defaultdict
import re, datetime, os

SKIP_SHEETS = {'Sample Delivery Tracker', 'Test Schedule', 'Test Summary', 'T0 Summary'}
CHECK_NAMES = {'Cosmetic','ISB','FACT','BT-OTA','Touch-CAL-Post','Charging','Button User','User Test','Pull-out force'}

def wf_sort_key(wfn):
    """自然排序键：WF1, WF2, ..., WF10, WF11, WF14.1, WF14.2"""
    try:
        parts = str(wfn).split('.')
        return tuple(int(p) if p.isdigit() else float('inf') for p in parts)
    except:
        return (float('inf'),)

# ── helpers ────────────────────────────────────────────────────────────

def normalize(s):
    return {w for w in re.sub(r'[^a-z0-9\s]',' ',s.lower()).split() if len(w)>1 and not w.isdigit()}

def wf_num(name):
    m = re.search(r'WF(\d+\.?\d*)', name)
    return m.group(1) if m else None

def parse_result(s):
    if not s or not isinstance(s, str): return None
    m = re.match(r'(\d+)(SF)?F?/(\d+)T', s.strip())
    return (int(m.group(1)), m.group(2)=='SF', int(m.group(3))) if m else None

def get_failure_type(cell):
    """Returns 'spec', 'strife', or None based on cell fill color."""
    try:
        fg = cell.fill.fgColor
        if fg and fg.type == 'rgb':
            rgb = str(fg.rgb).upper()
            if 'FF0000' in rgb: return 'spec'
            if 'FFFF00' in rgb: return 'strife'
    except: pass
    try:
        fc = cell.font.color
        if fc and fc.type == 'rgb':
            rgb = str(fc.rgb).upper()
            if 'FF9C0006' in rgb: return 'spec'
    except: pass
    return None


def cell_color_rgb(cell):
    try:
        fg = cell.fill.fgColor
        if fg and fg.type == 'rgb':
            return str(fg.rgb).upper()
    except Exception:
        pass
    return ''


def font_color_rgb(cell):
    try:
        fc = cell.font.color
        if fc and fc.type == 'rgb':
            return str(fc.rgb).upper()
    except Exception:
        pass
    return ''


def normalize_cell_value(value):
    if value is None:
        return ''
    return str(value).strip()


def status_from_cell(cell, is_pending=False):
    if is_pending:
        return 'pending', None
    raw = normalize_cell_value(cell.value)
    if raw == '/':
        return 'skip', None
    failure_type = get_failure_type(cell)
    if failure_type == 'spec':
        return 'spec_fail', 'spec'
    if failure_type == 'strife':
        return 'strife_fail', 'strife'
    if raw:
        return 'pass', None
    return 'pending', None


def aggregate_cp_status(check_rows, cp_idx, last_real):
    rows = [r for r in check_rows if r['cp_idx'] == cp_idx]
    if cp_idx > last_real:
        return 'pending', None, 0
    if not rows:
        return 'pending', None, 0
    if all(r['status'] == 'skip' for r in rows):
        return 'skip', None, 0
    has_data = any(r['status'] not in ('skip', 'pending') for r in rows)
    if any(r['failure_type'] == 'spec' for r in rows):
        return 'spec_fail', 'spec', int(has_data)
    if any(r['failure_type'] == 'strife' for r in rows):
        return 'strife_fail', 'strife', int(has_data)
    if has_data:
        return 'pass', None, 1
    return 'pending', None, 0

# ── CP-to-Test mapping ─────────────────────────────────────────────────

def map_cps_to_tests(cp_names, ts_names):
    """
    Maps each CP to a Test Summary test index.
    cp_names: list of (start_col, name) tuples
    ts_names: list of TS test name strings
    Returns: list[int] — test index per CP, same length as cp_names
    """
    nt = len(ts_names)
    
    # Step 1: Group CPs by name similarity
    groups = []
    cg = 0
    for i, (_, cn) in enumerate(cp_names):
        if i == 0: groups.append(0); continue
        pn = cp_names[i-1][1]; pl = pn.lower(); cl = cn.lower()
        
        if cl.startswith('2nd') or pl.startswith('2nd'): cg += 1
        elif cl.startswith('3rd') or pl.startswith('3rd'): cg += 1
        elif 'margin' in cl and 'margin' not in pl: cg += 1
        elif re.search(r'(\d+)%', pn) and re.search(r'(\d+)%', cn):
            pp = int(re.search(r'(\d+)%', pn).group(1))
            cp_ct = int(re.search(r'(\d+)%', cn).group(1))
            if pp <= 100 and cp_ct >= 120: cg += 1
        elif 'ltos' in pl and 'ltos' not in cl: cg += 1
        elif 'ltos' not in pl and 'ltos' in cl: cg += 1
        elif len(normalize(pn) & normalize(cn)) == 0: cg += 1
        groups.append(cg)
    
    ng = cg + 1
    ts_words = [normalize(tn) for tn in ts_names]
    g2t = {}
    
    # Step 2: Match groups to TS tests
    for g in range(ng):
        g_cps = [cp_names[i][1] for i, gp in enumerate(groups) if gp == g]
        best_test, best_score = None, 0
        for ti, tsw in enumerate(ts_words):
            total = 0
            for cn in g_cps:
                cpw = normalize(cn)
                total += len(cpw & tsw)
                if ts_names[ti].lower() in cn.lower(): total += 10
            if total > best_score: best_score = total; best_test = ti
        if best_score >= 2: g2t[g] = best_test
    
    # Step 3: Deduplicate
    used = set()
    for g in range(ng):
        if g in g2t:
            if g2t[g] in used: del g2t[g]
            else: used.add(g2t[g])
    
    # Step 4: Forced split if groups < tests
    if ng < nt:
        cpt = len(cp_names) // nt
        if cpt > 0:
            for i in range(len(cp_names)): groups[i] = min(i // cpt, nt - 1)
            ng = nt; g2t = {g: g for g in range(nt)}
    
    # Step 5: Fill unmapped groups sequentially
    if not g2t:
        for g in range(ng): g2t[g] = min(g, nt - 1)
    else:
        sm = sorted(g2t.keys()); fa = sm[0]; lt = g2t[fa]
        for g in range(fa + 1, ng):
            if g in g2t: lt = g2t[g]
            else: lt = min(lt + 1, nt - 1); g2t[g] = lt
        lt = g2t[fa]
        for g in range(fa - 1, -1, -1):
            if g in g2t: lt = g2t[g]
            else: lt = max(lt - 1, 0); g2t[g] = lt
    
    return [min(g2t[g], nt - 1) for g in groups]

# ── Test Summary reader ────────────────────────────────────────────────

def read_test_summary(daily_path):
    """Reads the Test Summary sheet from Daily Report.
    Returns: (ts_data, ts_qty, ts_test_names, ts_num_tests)
      ts_data: {wf: {cfg: {phase_idx: {name, result}}}}
      ts_qty: {wf: {cfg: quantity}}
      ts_test_names: {wf: [test_name, ...]}
      ts_num_tests: {wf: n}
    """
    wb = load_workbook(daily_path, data_only=True)
    ws = wb['Test Summary']
    
    sys_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 2).value == 'System': sys_row = r; break
    if not sys_row:
        return {}, {}, {}, {}
    
    config_col_map = {}
    for ps in [3, 8, 13, 18]:
        for o, c in enumerate(['R1FNF', 'R2CNM', 'R3', 'R4']):
            config_col_map[ps + o] = c
    
    test_cols = [7, 12, 17, 22]
    ts_data = {}
    ts_qty = {}
    ts_test_names = {}
    ts_num_tests = {}
    
    for r in range(14, 75):
        wf_val = ws.cell(r, 2).value
        if wf_val is None: continue
        if isinstance(wf_val, str):
            if wf_val == 'MLB': break
            continue
        if not isinstance(wf_val, (int, float)): continue
        wf_str = str(int(wf_val)) if isinstance(wf_val, float) and wf_val == int(wf_val) else str(wf_val)
        
        ts_qty[wf_str] = {}
        for offset, cfg in enumerate(['R1FNF', 'R2CNM', 'R3', 'R4']):
            qv = ws.cell(r, 3 + offset).value
            if qv is not None: ts_qty[wf_str][cfg] = int(qv) if isinstance(qv, (int, float)) else 0
        
        tests_for_wf = []
        for ti, tc in enumerate(test_cols):
            tname = ws.cell(r, tc).value
            if tname is None or not str(tname).strip(): break
            tname = str(tname).strip()
            tests_for_wf.append(tname)
            for offset, cfg in enumerate(['R1FNF', 'R2CNM', 'R3', 'R4']):
                rc = tc + 1 + offset
                val = ws.cell(r, rc).value
                if val is not None:
                    ts_data.setdefault(wf_str, {}).setdefault(cfg, {})[ti] = {'name': tname, 'result': val}
        
        if tests_for_wf:
            ts_test_names[wf_str] = tests_for_wf
            ts_num_tests[wf_str] = len(tests_for_wf)
    
    wb.close()
    return ts_data, ts_qty, ts_test_names, ts_num_tests


def extract_cp_structure(ws, header_row=1, cp_range_start=7):
    """Extracts CP names and check items from a WF sheet header.

    CP names are in header_row (e.g. row with Config|Unit#).
    Check items are in the next row below (header_row + 1).

    Returns: [{'cp_idx': int, 'cp_name': str, 'check_items': [str]}, ...]
    """
    cp_list = []
    ls = None
    ln = ''
    # Pass 1: find CP boundaries from header_row
    for c in range(cp_range_start, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v and isinstance(v, str) and v.strip():
            cv = v.strip()
            if cv not in CHECK_NAMES and cv not in ('Comments', 'Overall Result', 'Return to Summary') and len(cv) > 1:
                if ls is not None:
                    cp_list.append((ls, c - 1, ln))
                ls = c
                ln = cv
    if ls is not None:
        ec = ws.max_column
        for c in range(ls, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v and isinstance(v, str) and v.strip() in ('Comments', 'Overall Result'):
                ec = c - 1
                break
        cp_list.append((ls, ec, ln))

    # Pass 2: extract check items from the row BELOW header_row
    check_row = header_row + 1
    result = []
    for pi, (ps, pe, cp_name) in enumerate(cp_list):
        check_items = []
        for c in range(ps, pe + 1):
            v = ws.cell(check_row, c).value
            if v and isinstance(v, str) and v.strip():
                cv = v.strip()
                if cv in CHECK_NAMES:
                    check_items.append(cv)
        result.append({'cp_idx': pi, 'cp_name': cp_name, 'check_items': check_items})
    return result


def read_test_schedule(daily_path):
    """Reads Test Schedule sheet B/C columns and returns {wf_num: wf_name} mapping."""
    wb = load_workbook(daily_path, data_only=True)
    ws = wb['Test Schedule']
    names = {}
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if b is not None and isinstance(b, (int, float)):
            wfn = str(int(b)) if isinstance(b, float) and b == int(b) else str(b)
            if c and isinstance(c, str) and c.strip():
                names[wfn] = c.strip()
    wb.close()
    return names


# ── Core WF parser ─────────────────────────────────────────────────────

def _parse_wf_sheet(ws, wfn, ts_names):
    """
    Parses one WF detail sheet.
    Returns: {config: {test_idx: {total: N, spec_fails: [SN], strife_fails: [SN]}}}
    """
    num_tests = len(ts_names)
    results = {}
    
    r = 1
    while r <= ws.max_row:
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        if c3 == 'Config' and c4 == 'Unit #':
            # Parse CPs
            cp_list = []
            ls = None; ln = ''
            for c in range(7, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and v.strip():
                    cv = v.strip()
                    if cv not in CHECK_NAMES and cv not in ('Comments', 'Overall Result', 'Return to Summary') and len(cv) > 1:
                        if ls is not None: cp_list.append((ls, c - 1, ln))
                        ls = c; ln = cv
            if ls is not None:
                ec = ws.max_column
                for c in range(ls, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v and isinstance(v, str) and v.strip() in ('Comments', 'Overall Result'): ec = c - 1; break
                cp_list.append((ls, ec, ln))
            if not cp_list: r += 1; continue
            
            # CP → test mapping
            cp_names = [(ps, pn) for ps, pe, pn in cp_list]
            cp_test_map = map_cps_to_tests(cp_names, ts_names)
            
            # Build check item mapping per CP (row r+1 has check item names)
            cp_check_items = {}
            for ps, pe, pn in cp_list:
                items = {}
                for c in range(ps, pe + 1):
                    v = ws.cell(r + 1, c).value
                    if v and str(v).strip():
                        item_name = str(v).strip()
                        if item_name in CHECK_NAMES or len(item_name) <= 6:
                            items[c] = item_name
                if items:
                    cp_check_items[(ps, pe)] = items
            
            # Find first data config
            dr = r + 2
            config = None
            while dr <= ws.max_row:
                cv = ws.cell(dr, 3).value
                if cv in ('R1FNF', 'R2CNM', 'R3', 'R4'): config = cv; break
                dr += 1
            if not config: r += 1; continue
            
            # Init results
            if config not in results:
                results[config] = {
                    ti: {'total': 0, 'spec_fails': [], 'strife_fails': [], 'failure_details': []}
                    for ti in range(num_tests)
                }
            
            # Process data rows
            dr = r + 2
            while dr <= ws.max_row:
                dc1 = ws.cell(dr, 1).value
                dcfg = ws.cell(dr, 3).value
                
                if dc1 == '%' and dcfg == 'Config':
                    if ws.cell(dr, 4).value == 'Unit #': break
                    else: dr += 1; continue
                
                sn = ws.cell(dr, 5).value
                t0 = ws.cell(dr, 6).value
                if dcfg == 'Config' or (sn and str(sn).strip() == 'S/N'): dr += 1; continue
                if not sn or not t0: dr += 1; continue
                
                row_cfg = str(dcfg).strip() if dcfg and str(dcfg).strip() in ('R1FNF','R2CNM','R3','R4') else config
                if row_cfg not in results:
                    results[row_cfg] = {}
                    for ti in range(num_tests):
                        results[row_cfg][ti] = {
                            'total': 0, 'spec_fails': [], 'strife_fails': [],
                            'failure_details': []
                        }
                
                # Find last real CP
                last_real = None
                for pi in range(len(cp_list) - 1, -1, -1):
                    ps, pe, _ = cp_list[pi]
                    has_data = False
                    for cc in range(ps, pe + 1):
                        cv = ws.cell(dr, cc).value
                        if cv is not None and str(cv).strip() != '/': has_data = True; break
                    if has_data: last_real = pi; break
                
                if last_real is None: dr += 1; continue
                
                test_idx = cp_test_map[last_real]
                cs, ce, cp_name = cp_list[last_real]
                
                # Count unit in all tests
                for ti in range(num_tests):
                    results[row_cfg][ti]['total'] += 1
                
                # Check fills + track which check item column failed
                has_spec = False; has_strife = False
                failed_col = None
                check_items = cp_check_items.get((cs, ce), {})
                for cc in range(cs, ce + 1):
                    ft = get_failure_type(ws.cell(dr, cc))
                    if ft == 'spec':
                        has_spec = True
                        failed_col = cc
                        break
                    if ft == 'strife':
                        has_strife = True
                        if failed_col is None:
                            failed_col = cc
                
                sn_str = str(sn).strip()
                type_str = 'spec' if has_spec else ('strife' if has_strife else None)
                
                # Location = check item name (FACT/ISB/...), fallback to test name
                failed_check_item = check_items.get(failed_col, '') if failed_col else ''
                location = failed_check_item or (
                    ts_names[test_idx] if test_idx < len(ts_names) and ts_names[test_idx]
                    else f'Test{test_idx + 1}'
                )

                if has_spec:
                    results[row_cfg][test_idx]['spec_fails'].append(sn_str)
                elif has_strife:
                    results[row_cfg][test_idx]['strife_fails'].append(sn_str)

                if type_str:
                    results[row_cfg][test_idx]['failure_details'].append({
                        'sn': sn_str,
                        'type': type_str,
                        'location': location,
                        'failed_cp': cp_name,
                    })
                
                dr += 1
            r = dr
        else:
            r += 1
    return results

# ── Main entry point ────────────────────────────────────────────────────

def analyze(daily_path):
    """
    Main entry point. Parses all WF sheets in the Daily Report.
    
    Returns:
        dict: {wf_num: {config: {test_idx: {'total': N, 'spec_fails': [SN], 'strife_fails': [SN]}}}}
    """
    _, _, ts_test_names, _ = read_test_summary(daily_path)
    wb = load_workbook(daily_path)  # no data_only — need fill colors
    
    all_results = {}
    for name in wb.sheetnames:
        if name in SKIP_SHEETS or name.startswith('MLB'): continue
        wfn = wf_num(name)
        if not wfn: continue
        
        ts_names = ts_test_names.get(wfn, [''])
        if ts_names == ['']: ts_names = ['(unnamed)']
        
        try:
            wf_data = _parse_wf_sheet(wb[name], wfn, ts_names)
            if wf_data:
                all_results[wfn] = wf_data
        except Exception as e:
            pass  # Skip sheets that can't be parsed (WF34, WF33, etc.)
    
    wb.close()
    return all_results

# ── Export helpers ──────────────────────────────────────────────────────

def result_str(f, is_sf, t):
    if f == 0: return f"0F/{t}T"
    if is_sf: return f"{f}SF/{t}T"
    return f"{f}F/{t}T"

def build_summary_table(results, ts_data=None):
    """
    Builds a flat table of results for display.
    Returns: [{wf, cfg, test_idx, test_name, raw_result, ts_result, is_match, spec, strife, total}, ...]
    """
    rows = []
    for wfn in sorted(results.keys(), key=wf_sort_key):
        for cfg in ['R1FNF','R2CNM','R3','R4']:
            if cfg not in results[wfn]: continue
            for ti, d in results[wfn][cfg].items():
                sf = len(d['spec_fails']); stf = len(d['strife_fails']); t = d['total']
                if t == 0: continue
                
                raw = result_str(sf, sf==0 and stf>0, t)
                row = {'wf': wfn, 'cfg': cfg, 'test_idx': ti, 'raw': raw,
                       'spec': sf, 'strife': stf, 'total': t}
                
                if ts_data and wfn in ts_data and cfg in ts_data[wfn] and ti in ts_data[wfn][cfg]:
                    ts_val = ts_data[wfn][cfg][ti]['result']
                    row['test_name'] = ts_data[wfn][cfg][ti]['name']
                    ts_parsed = parse_result(str(ts_val)) if not isinstance(ts_val, datetime.datetime) else None
                    row['ts'] = str(ts_val) if not isinstance(ts_val, datetime.datetime) else 'Ongoing'
                    row['is_match'] = ts_parsed and ts_parsed[0] == sf and ts_parsed[1] == (sf==0 and stf>0) and ts_parsed[2] == t
                else:
                    row['test_name'] = f'Test{ti+1}'
                    row['ts'] = '-'
                    row['is_match'] = None
                
                rows.append(row)
    return rows

def build_failure_detail(results):
    """
    Builds a per-SN failure detail list.
    Returns: [{wf, cfg, test_idx, sn, type: 'spec'|'strife'}, ...]
    """
    failures = []
    for wfn in sorted(results.keys(), key=wf_sort_key):
        for cfg in ['R1FNF','R2CNM','R3','R4']:
            if cfg not in results[wfn]: continue
            for ti, d in results[wfn][cfg].items():
                for sn in d['spec_fails']:
                    failures.append({'wf': wfn, 'cfg': cfg, 'test_idx': ti, 'sn': sn, 'type': 'spec'})
                for sn in d['strife_fails']:
                    failures.append({'wf': wfn, 'cfg': cfg, 'test_idx': ti, 'sn': sn, 'type': 'strife'})
    return failures

# ── CP Structure Extraction ────────────────────────────────────────────────

def extract_all_cp_structures(daily_path):
    """Extracts CP names and check items from all WF sheets in a Daily Report.

    Returns: {wf_num: [{'cp_idx': int, 'cp_name': str, 'check_items': [str]}, ...]}
    """
    wb = load_workbook(daily_path, data_only=True)
    all_cps = {}
    for name in wb.sheetnames:
        if name in SKIP_SHEETS or name.startswith('MLB'):
            continue
        wfn = wf_num(name)
        if not wfn:
            continue
        # Find the header row (col 3 = 'Config', col 4 = 'Unit #')
        ws = wb[name]
        header_row = 1
        for r in range(1, min(ws.max_row + 1, 5)):
            c3 = ws.cell(r, 3).value
            c4 = ws.cell(r, 4).value
            if c3 == 'Config' and c4 == 'Unit #':
                header_row = r
                break
        try:
            cps = extract_cp_structure(ws, header_row)
            if cps:
                all_cps[wfn] = cps
        except Exception:
            pass
    wb.close()
    return all_cps


# ── Per-SN CP Progress Extraction ─────────────────────────────────────────

def extract_sn_progress(daily_path):
    """
    提取每个 SN 在每个 WF 中的 CP 进度信息。

    Args:
        daily_path: Daily Report Excel 文件路径

    Returns:
        dict: {wf_num: {config: [{
            'sn': str, 'unit_num': str,
            'current_cp_idx': int, 'current_cp_name': str,
            'total_cps': int, 'test_idx': int,
            'cp_results': [{'cp_idx': int, 'cp_name': str, 'status': str}, ...]
        }, ...]}}
    """
    _, _, ts_test_names, _ = read_test_summary(daily_path)
    wb = load_workbook(daily_path)  # no data_only — need fill colors

    all_progress = {}

    for name in wb.sheetnames:
        if name in SKIP_SHEETS or name.startswith('MLB'):
            continue
        wfn = wf_num(name)
        if not wfn:
            continue

        ts_names = ts_test_names.get(wfn, [''])
        if ts_names == ['']:
            ts_names = ['(unnamed)']

        try:
            progress = _extract_wf_progress(wb[name], ts_names)
            if progress:
                all_progress[wfn] = progress
        except Exception:
            pass  # 跳过无法解析的 sheet（WF34、WF33 等）

    wb.close()
    return all_progress


def extract_sn_fact_rows(daily_path, report_id, report_date):
    """Extract all CP/check-item fact rows from a Daily Report workbook."""
    _, _, ts_test_names, _ = read_test_summary(daily_path)
    wb = load_workbook(daily_path)
    all_cp_rows = []
    all_check_rows = []

    for name in wb.sheetnames:
        if name in SKIP_SHEETS or name.startswith('MLB'):
            continue
        wfn = wf_num(name)
        if not wfn:
            continue
        ts_names = ts_test_names.get(wfn, ['(unnamed)'])
        cp_rows, check_rows = extract_wf_fact_rows(
            wb[name],
            wf_num=wfn,
            report_id=report_id,
            report_date=report_date,
            ts_names=ts_names,
        )
        all_cp_rows.extend(cp_rows)
        all_check_rows.extend(check_rows)

    wb.close()
    return all_cp_rows, all_check_rows


def _extract_wf_progress(ws, ts_names):
    """
    解析一个 WF sheet，提取每个 SN 的 CP 进度。

    Returns: {config: [{sn, unit_num, current_cp_idx, current_cp_name,
                        total_cps, test_idx, cp_results}, ...]}
    """
    num_tests = len(ts_names)
    results = {}  # {config: [sn_entry, ...]}

    r = 1
    while r <= ws.max_row:
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        if c3 == 'Config' and c4 == 'Unit #':
            # ── 解析 CP 列范围与名称 ──
            cp_list = []
            ls = None
            ln = ''
            for c in range(7, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and v.strip():
                    cv = v.strip()
                    if cv not in CHECK_NAMES and cv not in ('Comments', 'Overall Result', 'Return to Summary') and len(cv) > 1:
                        if ls is not None:
                            cp_list.append((ls, c - 1, ln))
                        ls = c
                        ln = cv
            if ls is not None:
                ec = ws.max_column
                for c in range(ls, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v and isinstance(v, str) and v.strip() in ('Comments', 'Overall Result'):
                        ec = c - 1
                        break
                cp_list.append((ls, ec, ln))
            if not cp_list:
                r += 1
                continue

            # CP → Test 映射
            cp_names = [(ps, pn) for ps, pe, pn in cp_list]
            cp_test_map = map_cps_to_tests(cp_names, ts_names)
            total_cps = len(cp_list)

            # ── 找到第一个数据 config ──
            dr = r + 2
            config = None
            while dr <= ws.max_row:
                cv = ws.cell(dr, 3).value
                if cv in ('R1FNF', 'R2CNM', 'R3', 'R4'):
                    config = cv
                    break
                dr += 1
            if not config:
                r += 1
                continue

            # ── 遍历数据行 ──
            dr = r + 2
            while dr <= ws.max_row:
                dc1 = ws.cell(dr, 1).value
                dcfg = ws.cell(dr, 3).value

                if dc1 == '%' and dcfg == 'Config':
                    if ws.cell(dr, 4).value == 'Unit #':
                        break
                    else:
                        dr += 1
                        continue

                sn = ws.cell(dr, 5).value
                t0 = ws.cell(dr, 6).value
                if dcfg == 'Config' or (sn and str(sn).strip() == 'S/N'):
                    dr += 1
                    continue
                if not sn or not t0:
                    dr += 1
                    continue

                row_cfg = str(dcfg).strip() if dcfg and str(dcfg).strip() in ('R1FNF', 'R2CNM', 'R3', 'R4') else config
                unit_num = ws.cell(dr, 4).value
                unit_num_str = str(unit_num).strip() if unit_num is not None else ''

                # ── 找到 last_real CP（当前所在 CP） ──
                last_real = None
                for pi in range(total_cps - 1, -1, -1):
                    ps, pe, _ = cp_list[pi]
                    has_data = False
                    for cc in range(ps, pe + 1):
                        cv = ws.cell(dr, cc).value
                        if cv is not None and str(cv).strip() != '/':
                            has_data = True
                            break
                    if has_data:
                        last_real = pi
                        break

                if last_real is None:
                    dr += 1
                    continue

                # ── 构建所有 CP 的结果 ──
                cp_results = []
                for pi in range(total_cps):
                    ps, pe, cp_name = cp_list[pi]

                    # 检查该 CP 列范围内所有子列
                    all_slash = True
                    has_red = False
                    has_yellow = False
                    for cc in range(ps, pe + 1):
                        cv = ws.cell(dr, cc).value
                        if cv is not None and str(cv).strip() != '/':
                            all_slash = False
                        ft = get_failure_type(ws.cell(dr, cc))
                        if ft == 'spec':
                            has_red = True
                        elif ft == 'strife':
                            has_yellow = True

                    # 判定状态（优先级：pending > skip > 颜色）
                    if pi > last_real:
                        status = 'pending'
                    elif all_slash:
                        status = 'skip'
                    elif has_red:
                        status = 'spec_fail'
                    elif has_yellow:
                        status = 'strife_fail'
                    else:
                        status = 'pass'

                    cp_results.append({
                        'cp_idx': pi,
                        'cp_name': cp_name,
                        'status': status,
                    })

                # 当前 CP 信息
                current_cp_name = cp_list[last_real][2]
                test_idx = cp_test_map[last_real]

                # 只返回至少有一个非 skip 非 pending CP 的 SN
                has_real_cp = any(cr['status'] not in ('skip', 'pending') for cr in cp_results)
                if not has_real_cp:
                    dr += 1
                    continue

                sn_entry = {
                    'sn': str(sn).strip(),
                    'unit_num': unit_num_str,
                    'current_cp_idx': last_real,
                    'current_cp_name': current_cp_name,
                    'total_cps': total_cps,
                    'test_idx': test_idx,
                    'cp_results': cp_results,
                }

                results.setdefault(row_cfg, []).append(sn_entry)
                dr += 1
            r = dr
        else:
            r += 1

    return results


def extract_wf_fact_rows(ws, wf_num, report_id, report_date, ts_names):
    """Extract relational CP and check-item fact rows from one WF sheet."""
    cp_fact_rows = []
    check_fact_rows = []

    r = 1
    while r <= ws.max_row:
        if ws.cell(r, 3).value == 'Config' and ws.cell(r, 4).value == 'Unit #':
            cp_list = []
            ls = None
            ln = ''
            for c in range(7, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and v.strip():
                    cv = v.strip()
                    if cv not in CHECK_NAMES and cv not in ('Comments', 'Overall Result', 'Return to Summary') and len(cv) > 1:
                        if ls is not None:
                            cp_list.append((ls, c - 1, ln))
                        ls = c
                        ln = cv
            if ls is not None:
                ec = ws.max_column
                for c in range(ls, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v and isinstance(v, str) and v.strip() in ('Comments', 'Overall Result'):
                        ec = c - 1
                        break
                cp_list.append((ls, ec, ln))

            if not cp_list:
                r += 1
                continue

            cp_names = [(ps, pn) for ps, pe, pn in cp_list]
            cp_test_map = map_cps_to_tests(cp_names, ts_names)

            dr = r + 2
            default_config = None
            while dr <= ws.max_row:
                cv = ws.cell(dr, 3).value
                if cv in ('R1FNF', 'R2CNM', 'R3', 'R4'):
                    default_config = cv
                    break
                dr += 1
            if not default_config:
                r += 1
                continue

            dr = r + 2
            while dr <= ws.max_row:
                dc1 = ws.cell(dr, 1).value
                dcfg = ws.cell(dr, 3).value
                if dc1 == '%' and dcfg == 'Config' and ws.cell(dr, 4).value == 'Unit #':
                    break

                sn = ws.cell(dr, 5).value
                t0 = ws.cell(dr, 6).value
                if dcfg == 'Config' or (sn and str(sn).strip() == 'S/N') or not sn or not t0:
                    dr += 1
                    continue

                row_cfg = str(dcfg).strip() if dcfg and str(dcfg).strip() in ('R1FNF', 'R2CNM', 'R3', 'R4') else default_config
                unit_num = ws.cell(dr, 4).value
                unit_num_str = str(unit_num).strip() if unit_num is not None else ''
                sn_str = str(sn).strip()

                last_real = None
                for pi in range(len(cp_list) - 1, -1, -1):
                    ps, pe, _ = cp_list[pi]
                    if any(ws.cell(dr, cc).value is not None and str(ws.cell(dr, cc).value).strip() != '/' for cc in range(ps, pe + 1)):
                        last_real = pi
                        break
                if last_real is None:
                    dr += 1
                    continue

                row_checks = []
                for pi, (ps, pe, _cp_name) in enumerate(cp_list):
                    check_idx = 0
                    for cc in range(ps, pe + 1):
                        check_name = ws.cell(r + 1, cc).value
                        if not check_name or not isinstance(check_name, str) or not check_name.strip():
                            continue
                        check_name = check_name.strip()
                        cell = ws.cell(dr, cc)
                        status, failure_type = status_from_cell(cell, is_pending=pi > last_real)
                        check_row = {
                            'report_id': report_id,
                            'report_date': report_date,
                            'wf_num': wf_num,
                            'config': row_cfg,
                            'sn': sn_str,
                            'unit_num': unit_num_str,
                            'test_idx': cp_test_map[pi],
                            'cp_idx': pi,
                            'check_item_idx': check_idx,
                            'check_item': check_name,
                            'raw_value': normalize_cell_value(cell.value),
                            'normalized_value': normalize_cell_value(cell.value),
                            'status': status,
                            'failure_type': failure_type,
                            'fill_color': cell_color_rgb(cell),
                            'font_color': font_color_rgb(cell),
                            'source_row': dr,
                            'source_col': cc,
                        }
                        row_checks.append(check_row)
                        check_idx += 1

                for pi, (_ps, _pe, _cp_name) in enumerate(cp_list):
                    status, failure_type, has_data = aggregate_cp_status(row_checks, pi, last_real)
                    cp_fact_rows.append({
                        'report_id': report_id,
                        'report_date': report_date,
                        'wf_num': wf_num,
                        'config': row_cfg,
                        'sn': sn_str,
                        'unit_num': unit_num_str,
                        'test_idx': cp_test_map[pi],
                        'cp_idx': pi,
                        'status': status,
                        'failure_type': failure_type,
                        'has_data': has_data,
                        'is_current_cp': int(pi == last_real),
                    })
                check_fact_rows.extend(row_checks)
                dr += 1
            r = dr
        else:
            r += 1

    return cp_fact_rows, check_fact_rows


# ── CLI self-test ───────────────────────────────────────────────────────
if __name__ == '__main__':
    import sys, time
    path = sys.argv[1] if len(sys.argv) > 1 else None
    if not path:
        path = os.path.join(os.path.dirname(__file__), 'data', 'M60 EVT Rel Daily Report_20260508.xlsx')
    
    print(f"Analyzing: {path}")
    t0 = time.time()
    results = analyze(path)
    dt = time.time() - t0
    print(f"Done in {dt:.1f}s — {len(results)} WFs parsed")
    
    ts_data, _, _, _ = read_test_summary(path)
    rows = build_summary_table(results, ts_data)
    matched = sum(1 for r in rows if r['is_match'] is True)
    diff = sum(1 for r in rows if r['is_match'] is False)
    print(f"Summary: {len(rows)} entries, {matched} matched TS, {diff} differ")
    
    failures = build_failure_detail(results)
    spec = sum(1 for f in failures if f['type'] == 'spec')
    strife = sum(1 for f in failures if f['type'] == 'strife')
    print(f"Failures: {len(failures)} total ({spec} spec, {strife} strife)")

    # ── Test extract_sn_progress ──
    print("\n--- extract_sn_progress ---")
    t0 = time.time()
    progress = extract_sn_progress(path)
    dt = time.time() - t0
    total_entries = 0
    for wfn in sorted(progress.keys(), key=wf_sort_key):
        for cfg in sorted(progress[wfn].keys()):
            entries = progress[wfn][cfg]
            total_entries += len(entries)
    print(f"Done in {dt:.1f}s — {len(progress)} WFs, {total_entries} SN entries")
    for wfn in sorted(progress.keys(), key=wf_sort_key):
        for cfg in sorted(progress[wfn].keys()):
            entries = progress[wfn][cfg]
            print(f"  WF{wfn} {cfg}: {len(entries)} SNs")
            for entry in entries[:2]:
                cp_last = entry['cp_results'][-1]
                print(f"    SN:{entry['sn']} Unit:{entry['unit_num']} "
                      f"CP:{entry['current_cp_idx']+1}/{entry['total_cps']} "
                      f"({entry['current_cp_name']}) Test:{entry['test_idx']} "
                      f"last_cp_status:{cp_last['status']}")
