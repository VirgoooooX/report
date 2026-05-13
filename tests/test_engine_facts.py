import unittest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

import engine


class EngineFactExtractionTests(unittest.TestCase):
    def make_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'WF16.1'

        ws.cell(1, 3).value = 'Config'
        ws.cell(1, 4).value = 'Unit #'
        ws.cell(1, 5).value = 'S/N'
        ws.cell(1, 6).value = 'T0'
        ws.cell(1, 7).value = 'THC X 6'
        ws.cell(1, 9).value = 'Bottom Surface After 450Cyc'
        ws.cell(1, 11).value = 'Comments'

        ws.cell(2, 7).value = 'Cosmetic'
        ws.cell(2, 8).value = 'ISB'
        ws.cell(2, 9).value = 'FACT'
        ws.cell(2, 10).value = 'BT-OTA'

        ws.cell(3, 3).value = 'R3'
        ws.cell(3, 4).value = 'ER3-16.1-6'
        ws.cell(3, 5).value = 'D06P3Q46X0'
        ws.cell(3, 6).value = 'T0'
        ws.cell(3, 7).value = 'PASS'
        ws.cell(3, 8).value = 'PASS'
        ws.cell(3, 9).value = 'FAIL'
        ws.cell(3, 9).fill = PatternFill('solid', fgColor='FFFF0000')
        ws.cell(3, 10).value = '/'

        return ws

    def test_extract_wf_fact_rows_emits_cp_and_check_rows(self):
        ws = self.make_sheet()

        cp_rows, check_rows = engine.extract_wf_fact_rows(
            ws,
            wf_num='16.1',
            report_id=9,
            report_date='2026-05-09',
            ts_names=['THC X 6', 'Repetitive HSD PB'],
        )

        self.assertEqual(len(cp_rows), 2)
        self.assertEqual(cp_rows[1]['cp_idx'], 1)
        self.assertEqual(cp_rows[1]['status'], 'spec_fail')
        self.assertEqual(cp_rows[1]['failure_type'], 'spec')
        self.assertEqual(cp_rows[1]['is_current_cp'], 1)

        self.assertEqual(len(check_rows), 4)
        fact = [r for r in check_rows if r['check_item'] == 'FACT'][0]
        self.assertEqual(fact['status'], 'spec_fail')
        self.assertEqual(fact['failure_type'], 'spec')
        self.assertEqual(fact['source_row'], 3)
        self.assertEqual(fact['source_col'], 9)

    def make_cleaning_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'WF29.1'

        ws.cell(1, 3).value = 'Config'
        ws.cell(1, 4).value = 'Unit #'
        ws.cell(1, 5).value = 'S/N'
        ws.cell(1, 6).value = 'Chemichal'
        ws.cell(1, 7).value = 'T0'
        ws.cell(1, 13).value = '1st A 18hrs'
        ws.cell(1, 15).value = 'Comments'

        ws.cell(2, 7).value = 'Cosmetic'
        ws.cell(2, 8).value = 'ISB'
        ws.cell(2, 13).value = 'Cosmetic'
        ws.cell(2, 14).value = 'ISB'

        for offset, (unit, sn) in enumerate([
            ('ER1-29.1-1', 'SN-R1'),
            ('ER2-29.1-1', 'SN-R2'),
            ('ER3-29.1-1', 'SN-R3'),
            ('ER4-29.1-1', 'SN-R4'),
        ], start=3):
            ws.cell(offset, 4).value = unit
            ws.cell(offset, 5).value = sn
            ws.cell(offset, 7).value = 'PASS'
            ws.cell(offset, 8).value = 'PASS'
            ws.cell(offset, 13).value = 'PASS'
            ws.cell(offset, 14).value = 'PASS'

        return ws

    def test_cleaning_layout_infers_config_from_er_unit_prefix_for_fact_rows(self):
        ws = self.make_cleaning_sheet()

        cp_rows, check_rows = engine.extract_wf_fact_rows(
            ws,
            wf_num='29.1',
            report_id=9,
            report_date='2026-05-09',
            ts_names=['Cleaning Test_Spray_Option1'],
        )

        current_by_sn = {
            row['sn']: row
            for row in cp_rows
            if row['is_current_cp']
        }
        self.assertEqual(current_by_sn['SN-R1']['config'], 'R1FNF')
        self.assertEqual(current_by_sn['SN-R2']['config'], 'R2CNM')
        self.assertEqual(current_by_sn['SN-R3']['config'], 'R3')
        self.assertEqual(current_by_sn['SN-R4']['config'], 'R4')
        self.assertEqual(current_by_sn['SN-R1']['cp_idx'], 0)
        self.assertTrue(check_rows)

    def test_cleaning_layout_infers_config_from_er_unit_prefix_for_progress_rows(self):
        ws = self.make_cleaning_sheet()

        progress = engine._extract_wf_progress(ws, ['Cleaning Test_Spray_Option1'])

        self.assertEqual(len(progress['R1FNF']), 1)
        self.assertEqual(len(progress['R2CNM']), 1)
        self.assertEqual(len(progress['R3']), 1)
        self.assertEqual(len(progress['R4']), 1)
        self.assertEqual(progress['R1FNF'][0]['sn'], 'SN-R1')
        self.assertEqual(progress['R1FNF'][0]['current_cp_idx'], 0)


class RawdataValidationTests(unittest.TestCase):
    """Tests for rawdata validation gate: detect anomalous fail-gap-pass patterns."""

    def _make_wb_with_sheet(self, cp_names, sn_data):
        """Helper to build a workbook with one WF sheet.

        Args:
            cp_names: list of CP name strings (e.g. ['Drop20', 'Drop30', ..., 'Drop200'])
            sn_data: list of dicts, each with:
                - 'config', 'unit_num', 'sn'
                - 'cells': list of (value, fill_color_or_None) per CP
        Returns:
            (wb, ws)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'WF10'

        # Header row: Config | Unit # | S/N | T0 | CP1 | CP2 | ... | Comments
        ws.cell(1, 3).value = 'Config'
        ws.cell(1, 4).value = 'Unit #'
        ws.cell(1, 5).value = 'S/N'
        ws.cell(1, 6).value = 'T0'
        for i, cp_name in enumerate(cp_names):
            ws.cell(1, 7 + i).value = cp_name
        ws.cell(1, 7 + len(cp_names)).value = 'Comments'

        # Check items row (one check item per CP for simplicity)
        for i in range(len(cp_names)):
            ws.cell(2, 7 + i).value = 'Cosmetic'

        # Data rows
        for row_offset, sn_row in enumerate(sn_data):
            r = 3 + row_offset
            ws.cell(r, 3).value = sn_row['config']
            ws.cell(r, 4).value = sn_row['unit_num']
            ws.cell(r, 5).value = sn_row['sn']
            ws.cell(r, 6).value = 'T0'  # pre-CP data so row is recognized
            for i, (value, fill_color) in enumerate(sn_row['cells']):
                cell = ws.cell(r, 7 + i)
                cell.value = value
                if fill_color:
                    cell.fill = PatternFill('solid', fgColor=fill_color)

        return wb, ws

    def test_fail_gap_pass_reports_validation_error(self):
        """Fail at CP0, gap of '/' at CP1-CP8, then PASS at CP9 → must report error."""
        cp_names = [f'Drop{i * 20 + 20}' for i in range(10)]  # Drop20..Drop200
        cells = []
        # CP0: FAIL (red)
        cells.append(('FAIL', 'FFFF0000'))
        # CP1-CP8: all '/' (gap)
        for _ in range(8):
            cells.append(('/', None))
        # CP9: PASS (no color)
        cells.append(('PASS', None))

        wb, ws = self._make_wb_with_sheet(cp_names, [{
            'config': 'R3',
            'unit_num': 'ER3-10-1',
            'sn': 'D4VVLY9FWM',
            'cells': cells,
        }])

        errors = engine._find_rawdata_anomalies_for_sheet(
            ws, '10', ['Drop Test'], report_date='2026-05-12', source_file_name='test.xlsx'
        )

        self.assertGreater(len(errors), 0)
        err = errors[0]
        self.assertEqual(err['code'], 'failure_followed_by_gapped_later_data')
        self.assertEqual(err['sn'], 'D4VVLY9FWM')
        self.assertEqual(err['failed_cp'], 'Drop20')
        self.assertEqual(err['later_cp'], 'Drop200')
        self.assertEqual(err['config'], 'R3')

    def test_fail_then_adjacent_pass_no_error(self):
        """Fail at CP0, immediately PASS at CP1 (no gap) → no error, normal retest pass."""
        cp_names = ['Drop20', 'Drop40', 'Drop60']
        cells = [
            ('FAIL', 'FFFF0000'),  # CP0: fail
            ('PASS', None),         # CP1: pass (adjacent, no gap)
            ('PASS', None),         # CP2: pass
        ]

        wb, ws = self._make_wb_with_sheet(cp_names, [{
            'config': 'R3',
            'unit_num': 'ER3-10-1',
            'sn': 'SN_RETEST',
            'cells': cells,
        }])

        errors = engine._find_rawdata_anomalies_for_sheet(
            ws, '10', ['Drop Test'], report_date='2026-05-12', source_file_name='test.xlsx'
        )

        self.assertEqual(len(errors), 0)

    def test_fail_no_later_pass_no_error(self):
        """Fail at CP0, rest are '/' or empty → no error, normal fail-and-stop."""
        cp_names = ['Drop20', 'Drop40', 'Drop60', 'Drop80']
        cells = [
            ('FAIL', 'FFFF0000'),  # CP0: fail
            ('/', None),            # CP1: skip
            ('/', None),            # CP2: skip
            ('/', None),            # CP3: skip
        ]

        wb, ws = self._make_wb_with_sheet(cp_names, [{
            'config': 'R3',
            'unit_num': 'ER3-10-1',
            'sn': 'SN_STOPPED',
            'cells': cells,
        }])

        errors = engine._find_rawdata_anomalies_for_sheet(
            ws, '10', ['Drop Test'], report_date='2026-05-12', source_file_name='test.xlsx'
        )

        self.assertEqual(len(errors), 0)

    def test_fail_gap_pass_with_multiple_tests_only_flags_same_test(self):
        """Fail in test0 CP, gap, then pass in test1 CP → no error (different tests)."""
        # 4 CPs: first 2 belong to test0, last 2 belong to test1
        cp_names = ['Thermal 100Cyc', 'Thermal 200Cyc', 'HSD Drop1', 'HSD Drop2']
        cells = [
            ('FAIL', 'FFFF0000'),  # test0 CP0: fail
            ('/', None),            # test0 CP1: gap
            ('PASS', None),         # test1 CP0: pass (different test!)
            ('PASS', None),         # test1 CP1: pass
        ]

        wb, ws = self._make_wb_with_sheet(cp_names, [{
            'config': 'R3',
            'unit_num': 'ER3-10-1',
            'sn': 'SN_CROSS_TEST',
            'cells': cells,
        }])

        # Two test names → CPs should be mapped to different tests
        errors = engine._find_rawdata_anomalies_for_sheet(
            ws, '10', ['Thermal', 'HSD'], report_date='2026-05-12', source_file_name='test.xlsx'
        )

        # Should NOT flag because the later PASS is in a different test
        self.assertEqual(len(errors), 0)

    def test_validate_rawdata_workbook_aggregates_sheet_errors(self):
        """validate_rawdata_workbook should collect errors from all WF sheets."""
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Create a WF sheet with anomaly
        ws = wb.create_sheet('WF10')
        cp_names = ['Drop20', 'Drop40', 'Drop60']
        ws.cell(1, 3).value = 'Config'
        ws.cell(1, 4).value = 'Unit #'
        ws.cell(1, 5).value = 'S/N'
        ws.cell(1, 6).value = 'T0'
        for i, name in enumerate(cp_names):
            ws.cell(1, 7 + i).value = name
        ws.cell(1, 10).value = 'Comments'
        for i in range(3):
            ws.cell(2, 7 + i).value = 'Cosmetic'
        ws.cell(3, 3).value = 'R3'
        ws.cell(3, 4).value = 'ER3-10-1'
        ws.cell(3, 5).value = 'SN001'
        ws.cell(3, 6).value = 'T0'
        ws.cell(3, 7).value = 'FAIL'
        ws.cell(3, 7).fill = PatternFill('solid', fgColor='FFFF0000')
        ws.cell(3, 8).value = '/'
        ws.cell(3, 9).value = 'PASS'

        # Also add Test Summary sheet (required by validate but we test workbook-level)
        ts_test_names = {'10': ['Drop Test']}

        errors = engine.validate_rawdata_workbook(wb, ts_test_names, report_date='2026-05-12')
        self.assertGreater(len(errors), 0)
        self.assertEqual(errors[0]['wf'], '10')
        self.assertEqual(errors[0]['sn'], 'SN001')


if __name__ == '__main__':
    unittest.main()
