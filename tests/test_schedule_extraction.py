import os
import sqlite3
import tempfile
import unittest
import datetime

import db
import engine
import schedule_parser
from openpyxl import Workbook


DATA_PATH = os.path.join('rawdata', 'M60 EVT Rel Daily Report_20260509.xlsx')


def build_schedule_segments():
    _, _, ts_test_names, _ = engine.read_test_summary(DATA_PATH)
    cp_structures = engine.extract_all_cp_structures(DATA_PATH)
    # Attach test_idx via the legacy daily-only heuristic mapping. The
    # Plan/Actual split moved daily ingest to normalize_daily_cps_to_base
    # (which requires Base CP definitions in the DB), but this fixture-only
    # test deliberately exercises schedule_parser against daily-derived CPs
    # without touching the DB; we inline the few lines that engine.attach_test_idx_to_cps
    # used to provide so the test stays self-contained.
    mapped_cps = {}
    for wfn, cps in cp_structures.items():
        ts_names = ts_test_names.get(wfn, ['(unnamed)'])
        cp_names = [(cp['cp_idx'], cp['cp_name']) for cp in cps]
        cp_test_map = engine.map_cps_to_tests(cp_names, ts_names)
        mapped_cps[wfn] = [
            {**cp, 'test_idx': test_idx}
            for cp, test_idx in zip(cps, cp_test_map)
        ]
    return schedule_parser.extract_test_schedule_segments(DATA_PATH, ts_test_names, mapped_cps)


def find_segment(segments, wf_num, config, test_idx):
    for segment in segments:
        if (
            segment['wf_num'] == wf_num
            and segment['config'] == config
            and segment['test_idx'] == test_idx
        ):
            return segment
    raise AssertionError(f'missing segment wf={wf_num} cfg={config} test={test_idx}')


@unittest.skipUnless(os.path.exists(DATA_PATH), f"Data file not found: {DATA_PATH}")
class ScheduleExtractionTests(unittest.TestCase):
    def test_wf4_splits_altitude_and_rock_tumble(self):
        segments = build_schedule_segments()

        altitude = find_segment(segments, '4', 'R1FNF', 0)
        tumble = find_segment(segments, '4', 'R1FNF', 1)

        self.assertEqual(altitude['planned_start_date'], '2026-04-17')
        self.assertEqual(altitude['planned_end_date'], '2026-04-27')
        self.assertEqual(altitude['test_name'], 'Altitude')

        self.assertEqual(tumble['planned_start_date'], '2026-04-29')
        self.assertEqual(tumble['planned_end_date'], '2026-05-09')
        self.assertEqual(tumble['test_name'], 'Rock Tumble')

    def test_wf6_splits_margin_segment_after_full_scale(self):
        segments = build_schedule_segments()

        hs = find_segment(segments, '6', 'R1FNF', 0)
        main = find_segment(segments, '6', 'R1FNF', 1)
        margin = find_segment(segments, '6', 'R1FNF', 2)

        self.assertEqual(hs['planned_start_date'], '2026-04-17')
        self.assertEqual(hs['planned_end_date'], '2026-04-25')

        self.assertEqual(main['planned_start_date'], '2026-04-27')
        self.assertEqual(main['planned_end_date'], '2026-05-20')

        self.assertEqual(margin['planned_start_date'], '2026-05-23')
        self.assertEqual(margin['planned_end_date'], '2026-05-30')

    def test_wf14_2_uses_repeated_drop_sequence_to_split_main_and_margin(self):
        segments = build_schedule_segments()

        hs = find_segment(segments, '14.2', 'R1FNF', 0)
        main = find_segment(segments, '14.2', 'R1FNF', 1)
        margin = find_segment(segments, '14.2', 'R1FNF', 2)

        self.assertEqual(hs['planned_start_date'], '2026-04-17')
        self.assertEqual(hs['planned_end_date'], '2026-04-24')

        self.assertEqual(main['planned_start_date'], '2026-04-25')
        self.assertEqual(main['planned_end_date'], '2026-05-01')

        self.assertEqual(margin['planned_start_date'], '2026-05-02')
        self.assertEqual(margin['planned_end_date'], '2026-05-09')

    def test_wf39_splits_hsd_then_battery_swap_then_hsd(self):
        segments = build_schedule_segments()

        first_hsd = find_segment(segments, '39', 'R1FNF', 0)
        battery_swap = find_segment(segments, '39', 'R1FNF', 1)
        second_hsd = find_segment(segments, '39', 'R1FNF', 2)

        self.assertEqual(first_hsd['planned_start_date'], '2026-04-17')
        self.assertEqual(first_hsd['planned_end_date'], '2026-05-08')

        self.assertEqual(battery_swap['planned_start_date'], '2026-05-09')
        self.assertEqual(battery_swap['planned_end_date'], '2026-05-09')

        self.assertEqual(second_hsd['planned_start_date'], '2026-05-12')
        self.assertEqual(second_hsd['planned_end_date'], '2026-05-15')

    def test_wf16_1_splits_thc_from_repetitive_hsd(self):
        segments = build_schedule_segments()

        thc = find_segment(segments, '16.1', 'R1FNF', 0)
        hsd = find_segment(segments, '16.1', 'R1FNF', 1)

        self.assertEqual(thc['planned_start_date'], '2026-04-17')
        self.assertEqual(thc['planned_end_date'], '2026-04-24')
        self.assertEqual(thc['confidence'], 'high')

        self.assertEqual(hsd['planned_start_date'], '2026-04-25')
        self.assertEqual(hsd['planned_end_date'], '2026-05-12')
        self.assertEqual(hsd['confidence'], 'high')

    def test_wf17_splits_thc_from_random_vibration_axes(self):
        segments = build_schedule_segments()

        thc = find_segment(segments, '17', 'R1FNF', 0)
        vibration = find_segment(segments, '17', 'R1FNF', 1)

        self.assertEqual(thc['planned_start_date'], '2026-04-17')
        self.assertEqual(thc['planned_end_date'], '2026-04-24')
        self.assertEqual(thc['confidence'], 'high')

        self.assertEqual(vibration['planned_start_date'], '2026-04-27')
        self.assertEqual(vibration['planned_end_date'], '2026-05-02')
        self.assertEqual(vibration['confidence'], 'high')

    def test_wf18_splits_thc_from_squeeze_pressure_loads(self):
        segments = build_schedule_segments()

        thc = find_segment(segments, '18', 'R1FNF', 0)
        squeeze = find_segment(segments, '18', 'R1FNF', 1)

        self.assertEqual(thc['planned_start_date'], '2026-04-17')
        self.assertEqual(thc['planned_end_date'], '2026-04-24')
        self.assertEqual(thc['confidence'], 'high')

        self.assertEqual(squeeze['planned_start_date'], '2026-04-27')
        self.assertEqual(squeeze['planned_end_date'], '2026-05-13')
        self.assertEqual(squeeze['confidence'], 'high')


class ScheduleParserBoundaryUnitTests(unittest.TestCase):
    def _workbook_for_rows(self, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Test Schedule'
        ws.cell(1, 2).value = 'WF'
        ws.cell(1, 3).value = 'Test Item'
        dates = [
            '2026-04-17',
            '2026-04-24',
            '2026-04-25',
            '2026-05-01',
            '2026-05-02',
            '2026-05-09',
        ]
        for index, date_text in enumerate(dates, start=10):
            ws.cell(1, index).value = datetime.datetime.fromisoformat(date_text)
        for row_number, row in enumerate(rows, start=3):
            ws.cell(row_number, 2).value = row['wf_num']
            ws.cell(row_number, 3).value = row.get('test_item', 'Drop')
            ws.cell(row_number, 6).value = 1
            for offset, label in enumerate(row['labels'], start=10):
                ws.cell(row_number, offset).value = label
        return wb

    def _drop_segments(self, wf_num):
        wb = self._workbook_for_rows([{
            'wf_num': wf_num,
            'labels': ['T0', 'HS 72 hrs', '3rd drop', '18th drop', '3rd drop', 'End'],
        }])
        try:
            return schedule_parser.extract_test_schedule_segments_from_workbook(
                wb,
                {wf_num: ['HS', 'Main Drop', 'Margin']},
                {
                    wf_num: [
                        {'cp_idx': 0, 'cp_name': 'HS_72HRS', 'test_idx': 0},
                        {'cp_idx': 1, 'cp_name': 'SIDED_DROP_SEQB_1ST_DROP3', 'test_idx': 1},
                        {'cp_idx': 2, 'cp_name': 'SIDED_DROP_SEQB_18TH_DROP', 'test_idx': 1},
                        {'cp_idx': 3, 'cp_name': 'SIDED_DROP_SEQB_MARGIN_DROP3', 'test_idx': 2},
                    ]
                },
            )
        finally:
            wb.close()

    def test_wf14_variants_split_hs_main_margin_without_marker_details(self):
        for wf_num in ['14.1', '14.2', '14.3']:
            with self.subTest(wf_num=wf_num):
                segments = self._drop_segments(wf_num)
                self.assertEqual(
                    [(s['test_idx'], s['planned_start_date'], s['planned_end_date']) for s in segments],
                    [
                        (0, '2026-04-17', '2026-04-24'),
                        (1, '2026-04-25', '2026-05-01'),
                        (2, '2026-05-02', '2026-05-09'),
                    ],
                )
                self.assertTrue(all('marker_details' not in segment for segment in segments))

    def test_wf15_variants_split_hs_main_margin_without_marker_details(self):
        for wf_num in ['15.1', '15.2', '15.3']:
            with self.subTest(wf_num=wf_num):
                segments = self._drop_segments(wf_num)
                self.assertEqual(
                    [(s['test_idx'], s['planned_start_date'], s['planned_end_date']) for s in segments],
                    [
                        (0, '2026-04-17', '2026-04-24'),
                        (1, '2026-04-25', '2026-05-01'),
                        (2, '2026-05-02', '2026-05-09'),
                    ],
                )
                self.assertTrue(all('marker_details' not in segment for segment in segments))

    def test_wf29_variants_stay_single_t0_to_end_segment(self):
        for wf_num in ['29.1', '29.2', '29.3', '29.4']:
            with self.subTest(wf_num=wf_num):
                wb = self._workbook_for_rows([{
                    'wf_num': wf_num,
                    'labels': ['T0', '25%', '50%', '75%', '100%', 'End'],
                }])
                try:
                    segments = schedule_parser.extract_test_schedule_segments_from_workbook(
                        wb,
                        {wf_num: ['Single Test']},
                        {wf_num: [{'cp_idx': 0, 'cp_name': 'CP0', 'test_idx': 0}]},
                    )
                finally:
                    wb.close()

                self.assertEqual(len(segments), 1)
                self.assertEqual(segments[0]['planned_start_date'], '2026-04-17')
                self.assertEqual(segments[0]['planned_end_date'], '2026-05-09')
                self.assertNotIn('marker_details', segments[0])


class SchedulePersistenceTests(unittest.TestCase):
    def test_save_and_load_report_schedule_segments(self):
        fd, path = tempfile.mkstemp(suffix='.db')
        os.close(fd)
        conn = sqlite3.connect(path)
        conn.row_factory = db._dict_factory
        try:
            db.init_db(conn=conn)
            report_id = db.create_report_version(conn, '2026-05-09', 'report.xlsx')
            db.save_report_schedule_segments(conn, report_id, [
                {
                    'wf_num': '4',
                    'config': 'R1FNF',
                    'test_idx': 0,
                    'test_name': 'Altitude',
                    'schedule_test_item': 'Altitude + Rock tumble',
                    'planned_start_date': '2026-04-17',
                    'planned_end_date': '2026-04-27',
                    'source_row': 26,
                    'confidence': 'high',
                    'inference_reason': 'marker-sequence',
                    'marker_labels': ['T0', 'Op1', 'Op2', 'Op3', 'Op5'],
                }
            ])

            rows = db.get_report_schedule_segments(conn, report_id)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]['test_name'], 'Altitude')
            self.assertEqual(rows[0]['planned_end_date'], '2026-04-27')
            self.assertEqual(rows[0]['marker_labels'], ['T0', 'Op1', 'Op2', 'Op3', 'Op5'])
            self.assertNotIn('marker_details', rows[0])
        finally:
            conn.close()
            os.remove(path)

    def test_save_report_schedule_segments_replaces_previous_report_snapshot(self):
        fd, path = tempfile.mkstemp(suffix='.db')
        os.close(fd)
        conn = sqlite3.connect(path)
        conn.row_factory = db._dict_factory
        try:
            db.init_db(conn=conn)
            first_report_id = db.create_report_version(conn, '2026-05-08', 'first.xlsx')
            db.save_report_schedule_segments(conn, first_report_id, [
                {
                    'wf_num': '4',
                    'config': 'R1FNF',
                    'test_idx': 0,
                    'test_name': 'Altitude',
                    'schedule_test_item': 'Altitude',
                    'planned_start_date': '2026-04-17',
                    'planned_end_date': '2026-04-27',
                    'marker_labels': ['T0', 'End'],
                }
            ])

            second_report_id = db.create_report_version(conn, '2026-05-09', 'second.xlsx')
            db.save_report_schedule_segments(conn, second_report_id, [
                {
                    'wf_num': '5',
                    'config': 'R3',
                    'test_idx': 1,
                    'test_name': 'Thermal',
                    'schedule_test_item': 'Thermal',
                    'planned_start_date': '2026-05-01',
                    'planned_end_date': '2026-05-03',
                    'marker_labels': ['T0', 'End'],
                }
            ])

            self.assertEqual(db.get_report_schedule_segments(conn, first_report_id), [])
            latest_rows = db.get_report_schedule_segments(conn, second_report_id)
            self.assertEqual(len(latest_rows), 1)
            self.assertEqual(latest_rows[0]['wf_num'], '5')
            count = conn.execute("SELECT COUNT(*) AS c FROM report_schedule_segments").fetchone()['c']
            self.assertEqual(count, 1)
        finally:
            conn.close()
            os.remove(path)

    def test_init_db_prunes_historical_schedule_snapshots(self):
        fd, path = tempfile.mkstemp(suffix='.db')
        os.close(fd)
        conn = sqlite3.connect(path)
        conn.row_factory = db._dict_factory
        try:
            db.init_db(conn=conn)
            old_report_id = db.create_report_version(conn, '2026-05-08', 'old.xlsx')
            new_report_id = db.create_report_version(conn, '2026-05-09', 'new.xlsx')
            conn.execute(
                """INSERT INTO report_schedule_segments
                   (report_id, wf_num, config, test_idx, test_name, schedule_test_item,
                    planned_start_date, planned_end_date, confidence, inference_reason, marker_labels)
                   VALUES
                   (?, '4', 'R1FNF', 0, 'Altitude', 'Altitude', '2026-04-17', '2026-04-27', 'high', '', '[]'),
                   (?, '5', 'R3', 1, 'Thermal', 'Thermal', '2026-05-01', '2026-05-03', 'high', '', '[]')""",
                (old_report_id, new_report_id),
            )
            conn.commit()

            db.init_db(conn=conn)

            rows = conn.execute(
                "SELECT report_id, wf_num FROM report_schedule_segments ORDER BY wf_num"
            ).fetchall()
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]['report_id'], new_report_id)
            self.assertEqual(rows[0]['wf_num'], '5')
        finally:
            conn.close()
            os.remove(path)


class CurrentScheduleTests(unittest.TestCase):
    """Tests for current_schedule_segments (latest-only schedule table)."""

    def test_save_and_get_current_schedule(self):
        fd, path = tempfile.mkstemp(suffix='.db')
        os.close(fd)
        conn = sqlite3.connect(path)
        conn.row_factory = db._dict_factory
        try:
            db.init_db(conn=conn)
            report_id = db.create_report_version(conn, '2026-05-09', 'test.xlsx')

            db.save_current_schedule_segments(conn, report_id, [
                {
                    'wf_num': '4',
                    'config': 'R1FNF',
                    'test_idx': 0,
                    'test_name': 'Altitude',
                    'schedule_test_item': 'Altitude',
                    'planned_start_date': '2026-04-17',
                    'planned_end_date': '2026-04-27',
                    'marker_labels': ['T0', 'End'],
                },
                {
                    'wf_num': '6',
                    'config': 'R1FNF',
                    'test_idx': 1,
                    'test_name': 'Rock Tumble',
                    'schedule_test_item': 'Rock Tumble',
                    'planned_start_date': '2026-05-01',
                    'planned_end_date': '2026-05-05',
                    'marker_labels': ['T0', 'End'],
                },
            ])

            segments = db.get_current_schedule_segments(conn)
            self.assertEqual(len(segments), 2)
            self.assertEqual(segments[0]['wf_num'], '4')
            self.assertEqual(segments[1]['wf_num'], '6')

            filtered = db.get_current_schedule_segments(conn, wf_num='4')
            self.assertEqual(len(filtered), 1)
            self.assertEqual(filtered[0]['test_name'], 'Altitude')
            self.assertNotIn('marker_details', filtered[0])
        finally:
            conn.close()
            os.remove(path)

    def test_save_current_schedule_replaces_previous(self):
        fd, path = tempfile.mkstemp(suffix='.db')
        os.close(fd)
        conn = sqlite3.connect(path)
        conn.row_factory = db._dict_factory
        try:
            db.init_db(conn=conn)
            report_id = db.create_report_version(conn, '2026-05-08', 'first.xlsx')
            db.save_current_schedule_segments(conn, report_id, [
                {
                    'wf_num': '4',
                    'config': 'R1FNF',
                    'test_idx': 0,
                    'test_name': 'Altitude',
                    'schedule_test_item': 'Altitude',
                    'planned_start_date': '2026-04-17',
                    'planned_end_date': '2026-04-27',
                    'marker_labels': ['T0', 'End'],
                }
            ])

            second_report_id = db.create_report_version(conn, '2026-05-09', 'second.xlsx')
            db.save_current_schedule_segments(conn, second_report_id, [
                {
                    'wf_num': '5',
                    'config': 'R3',
                    'test_idx': 1,
                    'test_name': 'Thermal',
                    'schedule_test_item': 'Thermal',
                    'planned_start_date': '2026-05-01',
                    'planned_end_date': '2026-05-03',
                    'marker_labels': ['T0', 'End'],
                }
            ])

            segments = db.get_current_schedule_segments(conn)
            self.assertEqual(len(segments), 1)
            self.assertEqual(segments[0]['wf_num'], '5')
            self.assertEqual(segments[0]['updated_run_id'], second_report_id)
        finally:
            conn.close()
            os.remove(path)

    def test_get_current_schedule_empty(self):
        fd, path = tempfile.mkstemp(suffix='.db')
        os.close(fd)
        conn = sqlite3.connect(path)
        conn.row_factory = db._dict_factory
        try:
            db.init_db(conn=conn)
            segments = db.get_current_schedule_segments(conn)
            self.assertEqual(segments, [])
        finally:
            conn.close()
            os.remove(path)


if __name__ == '__main__':
    unittest.main()
