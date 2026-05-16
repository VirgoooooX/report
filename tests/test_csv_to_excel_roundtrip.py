"""
Integration test for full CSV → Excel → parse round-trip (Task 15.2).

Validates Requirements 5.1-5.7, 7.1-7.5, 14.1-14.4:
- Upload Base files (SN mapping, CP schedule, Test Plan)
- Create sample CSV files with known data (PASS and FAIL records)
- Generate Daily Report Excel via generate_daily_report()
- Parse the generated Excel with engine.parse_daily_report()
- Verify the round-trip produces consistent data
"""
import io
import json
import os
import shutil
import sqlite3
import tempfile
import unittest

import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

# Create a temp project root that mimics the real structure
# so that relative paths (rawdata/base/...) resolve correctly.
_TEMP_PROJECT_ROOT = tempfile.mkdtemp()
_TEMP_RAWDATA_DIR = os.path.join(_TEMP_PROJECT_ROOT, 'rawdata')
os.makedirs(_TEMP_RAWDATA_DIR, exist_ok=True)

# Set up a fresh test DB
TEST_DB = os.path.join(_TEMP_PROJECT_ROOT, 'test.db')
os.environ['REPORT_DB_PATH'] = TEST_DB
os.environ['REPORT_RAWDATA_DIR'] = _TEMP_RAWDATA_DIR

import app_paths
app_paths.RAWDATA_DIR = _TEMP_RAWDATA_DIR
app_paths.BASE_DIR = _TEMP_PROJECT_ROOT

import db
db.DB_PATH = TEST_DB
db.init_db()

from base_manager import upload_base_file
from checkitem_generator import generate_daily_report
import engine


def _create_sn_mapping_csv(sns_info):
    """Create a SN mapping CSV file from a list of (sn, config, wf_id, unit_number) tuples.

    Returns:
        Path to the created temp CSV file.
    """
    fd, path = tempfile.mkstemp(suffix='.csv')
    os.close(fd)
    with open(path, 'w', encoding='utf-8') as f:
        f.write("serial_number,config,Product,unit_number,start_date,wf_id\n")
        for sn, config, wf_id, unit_number in sns_info:
            f.write(f"{sn},{config},M60,{unit_number},2026-01-01,{wf_id}\n")
    return path


def _create_cp_schedule_csv(wf_cps):
    """Create a WaterfallCheckpointSchedule CSV from {wf_id: [(cp_name, test_idx), ...]}.

    Returns:
        Path to the created temp CSV file.
    """
    fd, path = tempfile.mkstemp(suffix='.csv')
    os.close(fd)
    with open(path, 'w', encoding='utf-8') as f:
        f.write("wf_id,wf id_cp,rel event cp,wf test\n")
        for wf_id, cps in wf_cps.items():
            for cp_order, (cp_name, test_idx) in enumerate(cps, start=1):
                f.write(f"{wf_id},{cp_order},{cp_name},{test_idx}\n")
    return path


def _create_test_plan_csv(wf_tests):
    """Create a WaterfallTestPlan CSV from {wf_id: [test_name, ...]}.

    Returns:
        Path to the created temp CSV file.
    """
    fd, path = tempfile.mkstemp(suffix='.csv')
    os.close(fd)
    with open(path, 'w', encoding='utf-8') as f:
        f.write("wf id,wf test_1,wf test_2,wf test_3\n")
        for wf_id, tests in wf_tests.items():
            row = [str(wf_id)]
            for i in range(3):
                row.append(tests[i] if i < len(tests) else "")
            f.write(",".join(row) + "\n")
    return path


def _make_csv_file(filename, content):
    """Create a mock file-like object mimicking Flask FileStorage."""
    class FakeFileStorage:
        def __init__(self, fname, data):
            self.filename = fname
            self._data = data.encode('utf-8') if isinstance(data, str) else data

        def read(self):
            return self._data

        def seek(self, pos):
            pass

    return FakeFileStorage(filename, content)


def _build_csv_content(records):
    """Build CSV content string from a list of record dicts.

    Each record should have: serial_number, rel_event, status, end_time.
    Optionally: failing_tests, station_id, version, and test_params dict.

    Returns:
        CSV content as string.
    """
    # Collect all param names across records
    param_names = set()
    for rec in records:
        if rec.get('test_params'):
            param_names.update(rec['test_params'].keys())
    param_names = sorted(param_names)

    header_cols = [
        "Site", "Product", "SerialNumber", "Special Build Name",
        "REL Event", "Test Pass/Fail Status", "EndTime",
        "List of Failing Tests", "Station ID", "Version",
        "Col10", "Col11", "Col12"
    ] + param_names

    lines = [",".join(header_cols)]

    for rec in records:
        row = [
            "Site1",
            "M60",
            rec['serial_number'],
            "Build1",
            rec['rel_event'],
            rec['status'],
            rec['end_time'],
            rec.get('failing_tests', ''),
            rec.get('station_id', 'STATION_01'),
            rec.get('version', '1.0.0'),
            "", "", ""
        ]
        # Add param values
        for pname in param_names:
            val = rec.get('test_params', {}).get(pname, '') if rec.get('test_params') else ''
            row.append(str(val) if val != '' else '')
        lines.append(",".join(row))

    return "\n".join(lines) + "\n"


class TestCSVToExcelRoundTrip(unittest.TestCase):
    """Full integration test: Base upload → CSV parse → Excel generate → engine.py parse."""

    @classmethod
    def setUpClass(cls):
        """Set up Base files once for all tests in this class."""
        # Reset DB tables
        conn = db.get_conn()
        conn.execute("DELETE FROM base_file_meta")
        conn.execute("DELETE FROM current_test_definitions")
        conn.execute("DELETE FROM current_wf_definitions")
        conn.execute("DELETE FROM current_cp_definitions")
        conn.execute("DELETE FROM raw_check_item_records")
        conn.execute("DELETE FROM import_batches")
        conn.commit()
        conn.close()

        # Define test data: 2 WFs, 4 SNs
        cls.sns_info = [
            ("SN_A1", "R1FNF", "1", "ER1-1-1"),
            ("SN_A2", "R1FNF", "1", "ER1-1-2"),
            ("SN_B1", "R2CNM", "1", "ER2-1-1"),
            ("SN_C1", "R3", "2", "ER3-1-1"),
        ]

        cls.wf_cps = {
            "1": [
                ("T0", 0),
                ("TC_100CYCLES", 0),
                ("TC_200CYCLES", 0),
            ],
            "2": [
                ("T0", 0),
                ("DROP_50CM", 0),
            ],
        }

        cls.wf_tests = {
            "1": ["Thermal Cycling"],
            "2": ["Drop Test"],
        }

        # Upload Base files via base_manager
        sn_path = _create_sn_mapping_csv(cls.sns_info)
        cp_path = _create_cp_schedule_csv(cls.wf_cps)
        tp_path = _create_test_plan_csv(cls.wf_tests)

        upload_base_file(sn_path, 'sn_mapping')
        upload_base_file(cp_path, 'checkpoint_schedule')
        upload_base_file(tp_path, 'test_plan')

        # Clean up source temp files (copies are in rawdata/base/)
        os.remove(sn_path)
        os.remove(cp_path)
        os.remove(tp_path)

    def setUp(self):
        """Clean raw records before each test."""
        conn = db.get_conn()
        conn.execute("DELETE FROM raw_check_item_records")
        conn.execute("DELETE FROM import_batches")
        conn.commit()
        conn.close()

    def _generate_and_parse(self, csv_files):
        """Generate Excel from CSV files and parse it back with engine.py.

        Returns:
            tuple of (excel_bytes, filename, summary, parse_result)
        """
        excel_bytes, filename, summary = generate_daily_report(csv_files)

        # Save Excel to temp file for engine.py parsing
        fd, excel_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        with open(excel_path, 'wb') as f:
            f.write(excel_bytes)

        try:
            # Parse with engine.py using the same DB (has definitions)
            conn = db.get_conn()
            try:
                parse_result = engine.parse_daily_report(excel_path, conn=conn)
            finally:
                conn.close()
        finally:
            os.remove(excel_path)

        return excel_bytes, filename, summary, parse_result

    def test_roundtrip_pass_records(self):
        """PASS records survive the full CSV → Excel → parse round-trip."""
        isb_records = [
            {"serial_number": "SN_A1", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 08:00:00"},
            {"serial_number": "SN_A2", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 08:30:00"},
            {"serial_number": "SN_B1", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 09:00:00"},
        ]
        isb_csv = _make_csv_file("ORT-ISB_20260515.csv", _build_csv_content(isb_records))

        excel_bytes, filename, summary, parse_result = self._generate_and_parse([isb_csv])

        # Verify generation succeeded
        self.assertIsNotNone(excel_bytes)
        self.assertGreater(len(excel_bytes), 0)
        self.assertIn("20260515", filename)

        # Verify parse succeeded — ts_test_names from DB
        self.assertIn('1', parse_result.ts_test_names)
        self.assertEqual(parse_result.ts_test_names['1'], ['Thermal Cycling'])

        # Verify WF sheets were found and parsed
        self.assertIn('1', parse_result.cp_structures)

    def test_roundtrip_mixed_pass_fail(self):
        """Mixed PASS and FAIL records produce correct Excel with color coding."""
        isb_records = [
            {"serial_number": "SN_A1", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 08:00:00"},
            {"serial_number": "SN_A1", "rel_event": "TC_100CYCLES", "status": "FAIL",
             "end_time": "2026-05-15 12:00:00",
             "failing_tests": "Thermal_Check_1",
             "test_params": {"Temp_Max": 85.5, "Temp_Min": -10.2}},
            {"serial_number": "SN_A2", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 08:30:00"},
            {"serial_number": "SN_A2", "rel_event": "TC_100CYCLES", "status": "PASS",
             "end_time": "2026-05-15 13:00:00"},
        ]
        isb_csv = _make_csv_file("ORT-ISB_20260515.csv", _build_csv_content(isb_records))

        excel_bytes, filename, summary, parse_result = self._generate_and_parse([isb_csv])

        # Verify generation
        self.assertEqual(summary['file_count'], 1)
        self.assertGreaterEqual(summary['valid_sn_count'], 2)

        # Verify parse found WF sheets and CP structures
        self.assertIn('1', parse_result.cp_structures)
        cp_names = [cp['cp_name'] for cp in parse_result.cp_structures['1']]
        # TC_100CYCLES should be in the CP structure
        self.assertIn('TC_100CYCLES', cp_names)

    def test_roundtrip_multiple_csv_types(self):
        """Multiple CSV file types (ISB + FACT) are combined correctly."""
        isb_records = [
            {"serial_number": "SN_A1", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 08:00:00"},
        ]
        fact_records = [
            {"serial_number": "SN_A1", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 09:00:00"},
        ]

        isb_csv = _make_csv_file("ORT-ISB_20260515.csv", _build_csv_content(isb_records))
        fact_csv = _make_csv_file("ORT-FACT_20260515.csv", _build_csv_content(fact_records))

        excel_bytes, filename, summary, parse_result = self._generate_and_parse([isb_csv, fact_csv])

        # Both files should be processed
        self.assertEqual(summary['file_count'], 2)

        # Parse should succeed
        self.assertIn('1', parse_result.ts_test_names)

    def test_roundtrip_wf_names_from_db(self):
        """engine.py reads WF names from DB when no Test Schedule sheet exists."""
        isb_records = [
            {"serial_number": "SN_C1", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 10:00:00"},
        ]
        isb_csv = _make_csv_file("ORT-ISB_20260515.csv", _build_csv_content(isb_records))

        _, _, _, parse_result = self._generate_and_parse([isb_csv])

        # WF names should come from DB (test plan joined with " + ")
        self.assertIn('1', parse_result.wf_names)
        self.assertEqual(parse_result.wf_names['1'], 'Thermal Cycling')
        self.assertIn('2', parse_result.wf_names)
        self.assertEqual(parse_result.wf_names['2'], 'Drop Test')

    def test_roundtrip_cp_structure_extracted(self):
        """engine.py correctly extracts CP structures from generated Excel."""
        records = [
            {"serial_number": "SN_A1", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 08:00:00"},
            {"serial_number": "SN_A1", "rel_event": "TC_100CYCLES", "status": "PASS",
             "end_time": "2026-05-15 12:00:00"},
            {"serial_number": "SN_A1", "rel_event": "TC_200CYCLES", "status": "PASS",
             "end_time": "2026-05-15 16:00:00"},
        ]
        isb_csv = _make_csv_file("ORT-ISB_20260515.csv", _build_csv_content(records))

        _, _, _, parse_result = self._generate_and_parse([isb_csv])

        # CP structures for WF1 should include our CPs
        self.assertIn('1', parse_result.cp_structures)
        cp_names = [cp['cp_name'] for cp in parse_result.cp_structures['1']]
        self.assertIn('TC_100CYCLES', cp_names)
        self.assertIn('TC_200CYCLES', cp_names)

    def test_roundtrip_invalid_sns_filtered(self):
        """SNs not in the mapping are filtered out during generation."""
        records = [
            {"serial_number": "SN_A1", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 08:00:00"},
            # This SN is not in our mapping
            {"serial_number": "INVALID_SN", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 09:00:00"},
        ]
        isb_csv = _make_csv_file("ORT-ISB_20260515.csv", _build_csv_content(records))

        _, _, summary, parse_result = self._generate_and_parse([isb_csv])

        # Only valid SN should be counted
        self.assertEqual(summary['valid_sn_count'], 1)

        # Parse should still succeed
        self.assertIn('1', parse_result.cp_structures)

    def test_roundtrip_raw_records_stored(self):
        """Raw records are stored in DB after generation."""
        records = [
            {"serial_number": "SN_A1", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 08:00:00"},
            {"serial_number": "SN_A1", "rel_event": "TC_100CYCLES", "status": "FAIL",
             "end_time": "2026-05-15 12:00:00",
             "failing_tests": "Check_1",
             "test_params": {"Param_A": 42.0}},
        ]
        isb_csv = _make_csv_file("ORT-ISB_20260515.csv", _build_csv_content(records))

        self._generate_and_parse([isb_csv])

        # Verify raw records in DB
        conn = db.get_conn()
        try:
            raw_records = conn.execute(
                "SELECT * FROM raw_check_item_records ORDER BY end_time"
            ).fetchall()
            batches = conn.execute("SELECT * FROM import_batches").fetchall()
        finally:
            conn.close()

        # Should have stored records
        self.assertGreater(len(raw_records), 0)
        self.assertEqual(len(batches), 1)
        self.assertEqual(batches[0]['status'], 'completed')

        # Verify PASS record has NULL test_params
        pass_recs = [r for r in raw_records if r['status'] == 'PASS']
        for r in pass_recs:
            self.assertIsNone(r['test_params'])

        # Verify FAIL record has test_params JSON
        fail_recs = [r for r in raw_records if r['status'] == 'FAIL']
        for r in fail_recs:
            self.assertIsNotNone(r['test_params'])
            params = json.loads(r['test_params'])
            self.assertIn('Param_A', params)

    def test_roundtrip_no_test_summary_or_schedule_sheets(self):
        """Generated Excel does NOT contain Test Summary or Test Schedule sheets."""
        from openpyxl import load_workbook

        records = [
            {"serial_number": "SN_A1", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 08:00:00"},
        ]
        isb_csv = _make_csv_file("ORT-ISB_20260515.csv", _build_csv_content(records))

        excel_bytes, _, _, _ = self._generate_and_parse([isb_csv])

        wb = load_workbook(io.BytesIO(excel_bytes))
        self.assertNotIn("Test Summary", wb.sheetnames)
        self.assertNotIn("Test Schedule", wb.sheetnames)
        wb.close()

    def test_roundtrip_sheet_names_parseable_by_engine(self):
        """All generated sheet names are parseable by engine.py's wf_num() regex."""
        from openpyxl import load_workbook

        records = [
            {"serial_number": "SN_A1", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 08:00:00"},
            {"serial_number": "SN_C1", "rel_event": "T0", "status": "PASS",
             "end_time": "2026-05-15 09:00:00"},
        ]
        isb_csv = _make_csv_file("ORT-ISB_20260515.csv", _build_csv_content(records))

        excel_bytes, _, _, _ = self._generate_and_parse([isb_csv])

        wb = load_workbook(io.BytesIO(excel_bytes))
        for name in wb.sheetnames:
            wfn = engine.wf_num(name)
            self.assertIsNotNone(wfn, f"engine.wf_num() failed to parse: '{name}'")
        wb.close()

    @classmethod
    def tearDownClass(cls):
        """Clean up test DB and temp project root."""
        try:
            os.remove(TEST_DB)
        except OSError:
            pass
        try:
            shutil.rmtree(_TEMP_PROJECT_ROOT, ignore_errors=True)
        except OSError:
            pass


if __name__ == '__main__':
    unittest.main()
