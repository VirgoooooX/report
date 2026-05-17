"""
Tests for base_manager.py — SN mapping parser (Task 2.1).

Validates Requirements 1.1, 1.5:
- parse_sn_mapping correctly extracts SN → {config, wf_id, unit_number}
- Computes per-config SN quantities
- Handles duplicate SNs (keeps first occurrence)
- Handles edge cases (empty file, missing columns)
"""
import csv
import datetime
import json
import os
import tempfile
import unittest

import base_manager
import db


def _write_csv(rows, fieldnames=None):
    """Helper: write rows to a temp CSV file and return its path."""
    if fieldnames is None:
        fieldnames = ['serial_number', 'config', 'Product', 'unit_number', 'start_date', 'wf_id']
    fd, path = tempfile.mkstemp(suffix='.csv')
    os.close(fd)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return path


class ParseSnMappingBasicTests(unittest.TestCase):
    """Test basic parsing of SN mapping CSV."""

    def test_parse_single_row(self):
        path = _write_csv([{
            'serial_number': 'CY21NXH72X',
            'config': 'R2CNM',
            'Product': 'B529',
            'unit_number': 'ER2-1-1',
            'start_date': '20260416',
            'wf_id': '1',
        }])
        try:
            result = base_manager.parse_sn_mapping(path)
            self.assertEqual(result['sn_count'], 1)
            self.assertIn('CY21NXH72X', result['sn_mapping'])
            entry = result['sn_mapping']['CY21NXH72X']
            self.assertEqual(entry['config'], 'R2CNM')
            self.assertEqual(entry['wf_id'], '1')
            self.assertEqual(entry['unit_number'], 'ER2-1-1')
        finally:
            os.remove(path)

    def test_parse_multiple_rows(self):
        rows = [
            {'serial_number': 'SN001', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-1-1', 'start_date': '20260416', 'wf_id': '1'},
            {'serial_number': 'SN002', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-1-2', 'start_date': '20260416', 'wf_id': '1'},
            {'serial_number': 'SN003', 'config': 'R1FNF', 'Product': 'B529', 'unit_number': 'ER1-1-1', 'start_date': '20260417', 'wf_id': '2'},
        ]
        path = _write_csv(rows)
        try:
            result = base_manager.parse_sn_mapping(path)
            self.assertEqual(result['sn_count'], 3)
            self.assertEqual(len(result['sn_mapping']), 3)
        finally:
            os.remove(path)

    def test_parse_decimal_wf_id(self):
        """WF IDs can be decimal like 14.1, 29.3."""
        rows = [
            {'serial_number': 'SN001', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-14.1-1', 'start_date': '20260416', 'wf_id': '14.1'},
            {'serial_number': 'SN002', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-29.3-2', 'start_date': '20260416', 'wf_id': '29.3'},
        ]
        path = _write_csv(rows)
        try:
            result = base_manager.parse_sn_mapping(path)
            self.assertEqual(result['sn_mapping']['SN001']['wf_id'], '14.1')
            self.assertEqual(result['sn_mapping']['SN002']['wf_id'], '29.3')
        finally:
            os.remove(path)


class ParseSnMappingConfigQuantityTests(unittest.TestCase):
    """Test per-config SN quantity computation."""

    def test_single_config_quantity(self):
        rows = [
            {'serial_number': f'SN{i:03d}', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': f'ER2-1-{i}', 'start_date': '20260416', 'wf_id': '1'}
            for i in range(1, 6)
        ]
        path = _write_csv(rows)
        try:
            result = base_manager.parse_sn_mapping(path)
            self.assertEqual(result['config_quantities'], {'R2CNM': 5})
        finally:
            os.remove(path)

    def test_multiple_config_quantities(self):
        rows = [
            {'serial_number': 'SN001', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-1-1', 'start_date': '20260416', 'wf_id': '1'},
            {'serial_number': 'SN002', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-1-2', 'start_date': '20260416', 'wf_id': '1'},
            {'serial_number': 'SN003', 'config': 'R1FNF', 'Product': 'B529', 'unit_number': 'ER1-1-1', 'start_date': '20260417', 'wf_id': '1'},
            {'serial_number': 'SN004', 'config': 'R1FNF', 'Product': 'B529', 'unit_number': 'ER1-1-2', 'start_date': '20260417', 'wf_id': '2'},
            {'serial_number': 'SN005', 'config': 'R3', 'Product': 'B529', 'unit_number': 'ER3-1-1', 'start_date': '20260422', 'wf_id': '1'},
        ]
        path = _write_csv(rows)
        try:
            result = base_manager.parse_sn_mapping(path)
            self.assertEqual(result['config_quantities']['R2CNM'], 2)
            self.assertEqual(result['config_quantities']['R1FNF'], 2)
            self.assertEqual(result['config_quantities']['R3'], 1)
        finally:
            os.remove(path)


class ParseSnMappingEdgeCaseTests(unittest.TestCase):
    """Test edge cases for SN mapping parser."""

    def test_empty_file(self):
        """Empty CSV (header only) should return empty mapping."""
        path = _write_csv([])
        try:
            result = base_manager.parse_sn_mapping(path)
            self.assertEqual(result['sn_count'], 0)
            self.assertEqual(result['sn_mapping'], {})
            self.assertEqual(result['config_quantities'], {})
        finally:
            os.remove(path)

    def test_duplicate_sn_keeps_first(self):
        """Duplicate SNs should keep the first occurrence."""
        rows = [
            {'serial_number': 'SN001', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-1-1', 'start_date': '20260416', 'wf_id': '1'},
            {'serial_number': 'SN001', 'config': 'R1FNF', 'Product': 'B529', 'unit_number': 'ER1-1-1', 'start_date': '20260417', 'wf_id': '2'},
        ]
        path = _write_csv(rows)
        try:
            result = base_manager.parse_sn_mapping(path)
            self.assertEqual(result['sn_count'], 1)
            # First occurrence wins
            self.assertEqual(result['sn_mapping']['SN001']['config'], 'R2CNM')
            self.assertEqual(result['sn_mapping']['SN001']['wf_id'], '1')
            # Only counted once in config quantities
            self.assertEqual(result['config_quantities']['R2CNM'], 1)
            self.assertNotIn('R1FNF', result['config_quantities'])
        finally:
            os.remove(path)

    def test_empty_serial_number_skipped(self):
        """Rows with empty serial_number should be skipped."""
        rows = [
            {'serial_number': '', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-1-1', 'start_date': '20260416', 'wf_id': '1'},
            {'serial_number': 'SN001', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-1-2', 'start_date': '20260416', 'wf_id': '1'},
        ]
        path = _write_csv(rows)
        try:
            result = base_manager.parse_sn_mapping(path)
            self.assertEqual(result['sn_count'], 1)
            self.assertIn('SN001', result['sn_mapping'])
        finally:
            os.remove(path)

    def test_whitespace_trimmed(self):
        """Whitespace in fields should be trimmed."""
        rows = [
            {'serial_number': ' SN001 ', 'config': ' R2CNM ', 'Product': 'B529', 'unit_number': ' ER2-1-1 ', 'start_date': '20260416', 'wf_id': ' 1 '},
        ]
        path = _write_csv(rows)
        try:
            result = base_manager.parse_sn_mapping(path)
            self.assertIn('SN001', result['sn_mapping'])
            entry = result['sn_mapping']['SN001']
            self.assertEqual(entry['config'], 'R2CNM')
            self.assertEqual(entry['wf_id'], '1')
            self.assertEqual(entry['unit_number'], 'ER2-1-1')
        finally:
            os.remove(path)

    def test_real_csv_file(self):
        """Test with the actual SN mapping CSV file if available."""
        real_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            'Checkitem data', 'Base',
            'Corn_JAWX 2026-05-06-Serial Numbers EVT FATP.csv'
        )
        if not os.path.exists(real_path):
            self.skipTest("Real CSV file not available")

        result = base_manager.parse_sn_mapping(real_path)
        # Should have a substantial number of SNs
        self.assertGreater(result['sn_count'], 100)
        # Should have multiple configs
        self.assertGreater(len(result['config_quantities']), 1)
        # All known configs should be present
        self.assertIn('R2CNM', result['config_quantities'])
        self.assertIn('R1FNF', result['config_quantities'])
        # Each SN entry should have required fields
        for sn, info in list(result['sn_mapping'].items())[:5]:
            self.assertIn('config', info)
            self.assertIn('wf_id', info)
            self.assertIn('unit_number', info)
            self.assertTrue(info['config'])
            self.assertTrue(info['wf_id'])
            self.assertTrue(info['unit_number'])


if __name__ == '__main__':
    unittest.main()


# ---------------------------------------------------------------------------
# Tests for parse_checkpoint_schedule (Task 2.3)
# Validates Requirement 1.2
# ---------------------------------------------------------------------------

CP_FIELDNAMES = ['wf_id', 'wf id_cp', 'wf id_cp total', 'wf id_cp_nci',
                 'wf id_cp total_nci', 'test category', 'rel event cp',
                 'checkin', 'wf test', 'wf test_cp', 'wf test_cp total',
                 'wf id_rel event', 'wf id_rel event_prev', 'wf id_rel event_next']


def _write_cp_csv(rows, fieldnames=None):
    """Helper: write CP schedule rows to a temp CSV file and return its path."""
    if fieldnames is None:
        fieldnames = CP_FIELDNAMES
    fd, path = tempfile.mkstemp(suffix='.csv')
    os.close(fd)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return path


def _make_cp_row(wf_id, cp_idx, cp_name, test_idx=0, category=''):
    """Helper: create a CP schedule row dict."""
    return {
        'wf_id': str(wf_id),
        'wf id_cp': str(cp_idx),
        'wf id_cp total': '',
        'wf id_cp_nci': '',
        'wf id_cp total_nci': '',
        'test category': category,
        'rel event cp': cp_name,
        'checkin': '',
        'wf test': str(test_idx),
        'wf test_cp': '',
        'wf test_cp total': '',
        'wf id_rel event': f'{wf_id}_{cp_name}',
        'wf id_rel event_prev': '',
        'wf id_rel event_next': '',
    }


class ParseCheckpointScheduleBasicTests(unittest.TestCase):
    """Test basic parsing of WaterfallCheckpointSchedule CSV."""

    def test_parse_single_wf(self):
        rows = [
            _make_cp_row(1, 0, 'REL_T0', test_idx=0, category='REL_T0'),
            _make_cp_row(1, 1, 'LTHS_100HRS', test_idx=1, category='Long Term HS'),
            _make_cp_row(1, 2, 'LTHS_200HRS', test_idx=1, category='Long Term HS'),
            _make_cp_row(1, 3, 'REL_TFINAL', test_idx=0, category='REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            self.assertEqual(result['wf_count'], 1)
            self.assertIn('1', result['cp_schedule'])
            # Schedule boundary markers (REL_T0, REL_TFINAL) should be filtered out
            self.assertEqual(result['cp_schedule']['1'], ['LTHS_100HRS', 'LTHS_200HRS'])
        finally:
            os.remove(path)

    def test_parse_multiple_wfs(self):
        rows = [
            _make_cp_row(1, 0, 'REL_T0', test_idx=0),
            _make_cp_row(1, 1, 'CP_A', test_idx=1),
            _make_cp_row(1, 2, 'REL_TFINAL', test_idx=0),
            _make_cp_row(2, 0, 'REL_T0', test_idx=0),
            _make_cp_row(2, 1, 'CP_B', test_idx=1),
            _make_cp_row(2, 2, 'CP_C', test_idx=1),
            _make_cp_row(2, 3, 'REL_TFINAL', test_idx=0),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            self.assertEqual(result['wf_count'], 2)
            self.assertEqual(result['cp_schedule']['1'], ['CP_A'])
            self.assertEqual(result['cp_schedule']['2'], ['CP_B', 'CP_C'])
        finally:
            os.remove(path)

    def test_cp_count_total(self):
        rows = [
            _make_cp_row(1, 0, 'REL_T0'),
            _make_cp_row(1, 1, 'CP_A'),
            _make_cp_row(1, 2, 'REL_TFINAL'),
            _make_cp_row(2, 0, 'REL_T0'),
            _make_cp_row(2, 1, 'CP_B'),
            _make_cp_row(2, 2, 'CP_C'),
            _make_cp_row(2, 3, 'REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            # WF1: CP_A = 1; WF2: CP_B, CP_C = 2; total = 3
            # (REL_T0 / REL_TFINAL are schedule boundaries and filtered out.)
            self.assertEqual(result['cp_count'], 3)
        finally:
            os.remove(path)

    def test_normalizes_real_checkpoint_test_numbers_to_zero_based(self):
        rows = [
            _make_cp_row(4, 0, 'REL_T0', test_idx=0),
            _make_cp_row(4, 1, 'ALT_PROFILE1', test_idx=1),
            _make_cp_row(4, 2, 'ALT_PROFILE2', test_idx=1),
            _make_cp_row(4, 3, 'RT_15MIN', test_idx=2),
            _make_cp_row(4, 4, 'REL_TFINAL', test_idx=0),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            details = result['cp_details']['4']
            # Boundary rows (REL_T0 / REL_TFINAL) are kept in cp_details
            # tagged is_boundary=1 (Batch B step 3.2). Real test CPs have
            # is_boundary=0. The non-boundary test_idx sequence is 0/0/1.
            non_boundary = [d for d in details if not d.get('is_boundary')]
            self.assertEqual(
                [detail['test_idx'] for detail in non_boundary],
                [0, 0, 1],
            )
        finally:
            os.remove(path)

    def test_uses_margin_category_when_checkpoint_test_number_repeats(self):
        rows = [
            _make_cp_row(14.2, 0, 'REL_T0', test_idx=0, category='REL_T0'),
            _make_cp_row(14.2, 1, 'HS_72HRS', test_idx=1, category='HS 65/90/72'),
            _make_cp_row(14.2, 2, 'SIDED_DROP_SEQB_1ST_DROP3', test_idx=2, category='18 Sided Drop 1m PB SeqB'),
            _make_cp_row(14.2, 3, 'SIDED_DROP_SEQB_2ND_DROP3', test_idx=2, category='18 Sided Drop 1m PB SeqB - Margin'),
            _make_cp_row(14.2, 4, 'SIDED_DROP_SEQB_3RD_DROP3', test_idx=2, category='18 Sided Drop 1m PB SeqB'),
            _make_cp_row(14.2, 5, 'REL_TFINAL', test_idx=0, category='REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            details = result['cp_details']['14.2']
            # Boundary rows (REL_T0 / REL_TFINAL) are kept in cp_details
            # tagged is_boundary=1 (Batch B step 3.2). Margin-category logic
            # applies only to non-boundary CPs.
            non_boundary = [d for d in details if not d.get('is_boundary')]
            self.assertEqual(
                [detail['test_idx'] for detail in non_boundary],
                [0, 1, 2, 1],
            )
        finally:
            os.remove(path)


class ParseCheckpointScheduleFilterTests(unittest.TestCase):
    """Test that excluded CPs are properly filtered."""

    def test_filters_rel_tfinal(self):
        rows = [
            _make_cp_row(1, 0, 'REL_T0'),
            _make_cp_row(1, 1, 'CP_A'),
            _make_cp_row(1, 2, 'REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            self.assertNotIn('REL_TFINAL', result['cp_schedule']['1'])
        finally:
            os.remove(path)

    def test_filters_rel_fa_retest(self):
        rows = [
            _make_cp_row(1, 0, 'REL_T0'),
            _make_cp_row(1, 1, 'CP_A'),
            _make_cp_row(1, 2, 'REL FA RETEST'),
            _make_cp_row(1, 3, 'REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            self.assertNotIn('REL FA RETEST', result['cp_schedule']['1'])
            self.assertEqual(result['cp_schedule']['1'], ['CP_A'])
        finally:
            os.remove(path)

    def test_filters_send_to_fa(self):
        rows = [
            _make_cp_row(1, 0, 'REL_T0'),
            _make_cp_row(1, 1, 'SEND TO FA'),
            _make_cp_row(1, 2, 'CP_A'),
            _make_cp_row(1, 3, 'REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            self.assertNotIn('SEND TO FA', result['cp_schedule']['1'])
        finally:
            os.remove(path)

    def test_filters_stop_test(self):
        rows = [
            _make_cp_row(1, 0, 'REL_T0'),
            _make_cp_row(1, 1, 'CP_A'),
            _make_cp_row(1, 2, 'STOP TEST'),
            _make_cp_row(1, 3, 'REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            self.assertNotIn('STOP TEST', result['cp_schedule']['1'])
        finally:
            os.remove(path)

    def test_filters_return_to_rel(self):
        rows = [
            _make_cp_row(1, 0, 'REL_T0'),
            _make_cp_row(1, 1, 'CP_A'),
            _make_cp_row(1, 2, 'RETURN TO REL'),
            _make_cp_row(1, 3, 'REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            self.assertNotIn('RETURN TO REL', result['cp_schedule']['1'])
        finally:
            os.remove(path)

    def test_filters_all_excluded_cps(self):
        """All excluded CPs (operational + boundary markers) should be filtered in a single WF."""
        rows = [
            _make_cp_row(1, 0, 'REL_T0'),
            _make_cp_row(1, 1, 'CP_A'),
            _make_cp_row(1, 2, 'REL FA RETEST'),
            _make_cp_row(1, 3, 'SEND TO FA'),
            _make_cp_row(1, 4, 'STOP TEST'),
            _make_cp_row(1, 5, 'RETURN TO REL'),
            _make_cp_row(1, 6, 'CP_B'),
            _make_cp_row(1, 7, 'REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            excluded = {
                'REL FA RETEST', 'SEND TO FA', 'STOP TEST', 'RETURN TO REL',
                'T0', 'REL_T0', 'End', 'TFinal', 'REL_TFINAL',
            }
            for cp in result['cp_schedule']['1']:
                self.assertNotIn(cp, excluded)
            self.assertEqual(result['cp_schedule']['1'], ['CP_A', 'CP_B'])
        finally:
            os.remove(path)


class ParseCheckpointScheduleOrderingTests(unittest.TestCase):
    """Test that CP ordering is maintained correctly."""

    def test_maintains_original_order(self):
        rows = [
            _make_cp_row(1, 0, 'REL_T0'),
            _make_cp_row(1, 1, 'ALPHA'),
            _make_cp_row(1, 2, 'BETA'),
            _make_cp_row(1, 3, 'GAMMA'),
            _make_cp_row(1, 4, 'REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            self.assertEqual(result['cp_schedule']['1'], ['ALPHA', 'BETA', 'GAMMA'])
        finally:
            os.remove(path)

    def test_order_maintained_after_filtering(self):
        """Order should be preserved even when excluded CPs are interspersed."""
        rows = [
            _make_cp_row(1, 0, 'REL_T0'),
            _make_cp_row(1, 1, 'CP_A'),
            _make_cp_row(1, 2, 'REL FA RETEST'),
            _make_cp_row(1, 3, 'CP_B'),
            _make_cp_row(1, 4, 'STOP TEST'),
            _make_cp_row(1, 5, 'CP_C'),
            _make_cp_row(1, 6, 'REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            self.assertEqual(result['cp_schedule']['1'], ['CP_A', 'CP_B', 'CP_C'])
        finally:
            os.remove(path)

    def test_unsorted_csv_still_produces_correct_order(self):
        """Even if CSV rows are not in order, output should be sorted by wf id_cp."""
        rows = [
            _make_cp_row(1, 3, 'REL_TFINAL'),
            _make_cp_row(1, 1, 'CP_B'),
            _make_cp_row(1, 0, 'REL_T0'),
            _make_cp_row(1, 2, 'CP_A'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            # Should be sorted by cp_order (wf id_cp); REL_T0/REL_TFINAL boundary
            # markers are filtered out, leaving CP_B (1) and CP_A (2).
            self.assertEqual(result['cp_schedule']['1'], ['CP_B', 'CP_A'])
        finally:
            os.remove(path)

    def test_decimal_wf_ids_sorted_numerically(self):
        """WF IDs like 14.1 should sort numerically, not lexicographically."""
        rows = [
            _make_cp_row('2', 0, 'REL_T0'),
            _make_cp_row('2', 1, 'CP_2A'),
            _make_cp_row('14.1', 0, 'REL_T0'),
            _make_cp_row('14.1', 1, 'CP_14A'),
            _make_cp_row('1', 0, 'REL_T0'),
            _make_cp_row('1', 1, 'CP_1A'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            # All 3 WFs should be present
            self.assertEqual(result['wf_count'], 3)
            self.assertIn('1', result['cp_schedule'])
            self.assertIn('2', result['cp_schedule'])
            self.assertIn('14.1', result['cp_schedule'])
        finally:
            os.remove(path)


class ParseCheckpointScheduleDetailsTests(unittest.TestCase):
    """Test cp_details output with test_idx information."""

    def test_details_include_test_idx(self):
        rows = [
            _make_cp_row(1, 0, 'REL_T0', test_idx=0),
            _make_cp_row(1, 1, 'CP_A', test_idx=1),
            _make_cp_row(1, 2, 'CP_B', test_idx=1),
            _make_cp_row(1, 3, 'CP_C', test_idx=2),
            _make_cp_row(1, 4, 'REL_TFINAL', test_idx=0),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            details = result['cp_details']['1']
            # Boundary rows (REL_T0 / REL_TFINAL) appear in cp_details with
            # is_boundary=1 (Batch B step 3.2); real CPs have is_boundary=0.
            self.assertEqual(len(details), 5)
            non_boundary = [d for d in details if not d.get('is_boundary')]
            self.assertEqual(len(non_boundary), 3)
            self.assertEqual(
                non_boundary[0],
                {'cp_name': 'CP_A', 'test_idx': 0, 'is_boundary': 0},
            )
            self.assertEqual(
                non_boundary[1],
                {'cp_name': 'CP_B', 'test_idx': 0, 'is_boundary': 0},
            )
            self.assertEqual(
                non_boundary[2],
                {'cp_name': 'CP_C', 'test_idx': 1, 'is_boundary': 0},
            )
            # Boundary rows preserve their cp_name and is_boundary tag.
            boundary = [d for d in details if d.get('is_boundary')]
            self.assertEqual(len(boundary), 2)
            self.assertEqual({d['cp_name'] for d in boundary}, {'REL_T0', 'REL_TFINAL'})
        finally:
            os.remove(path)


class ParseCheckpointScheduleBoundaryTaggingTests(unittest.TestCase):
    """Batch B step 3.2 — schedule-boundary CPs must be retained in cp_details
    with is_boundary=1, and excluded from cp_schedule (name list).

    See docs/plans/2026-05-17-rel-t0-daily-report-second-cut.md.
    """

    def test_boundary_rows_kept_in_cp_details_tagged(self):
        rows = [
            _make_cp_row(1, 0, 'REL_T0'),
            _make_cp_row(1, 1, 'CP_A'),
            _make_cp_row(1, 2, 'CP_B'),
            _make_cp_row(1, 3, 'REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            details = result['cp_details']['1']
            # All 4 rows retained in cp_details.
            self.assertEqual(
                [(d['cp_name'], d['is_boundary']) for d in details],
                [
                    ('REL_T0', 1),
                    ('CP_A', 0),
                    ('CP_B', 0),
                    ('REL_TFINAL', 1),
                ],
            )
            # cp_schedule still contains only non-boundary CPs (back-compat).
            self.assertEqual(result['cp_schedule']['1'], ['CP_A', 'CP_B'])
            # cp_count counts non-boundary only.
            self.assertEqual(result['cp_count'], 2)
        finally:
            os.remove(path)

    def test_t0_and_end_variants_all_tagged(self):
        """All five SCHEDULE_BOUNDARY_LABELS variants must be tagged."""
        rows = [
            _make_cp_row(1, 0, 'T0'),
            _make_cp_row(1, 1, 'CP_REAL'),
            _make_cp_row(1, 2, 'End'),
            _make_cp_row(2, 0, 'REL_T0'),
            _make_cp_row(2, 1, 'CP_REAL2'),
            _make_cp_row(2, 2, 'TFinal'),
            _make_cp_row(3, 0, 'REL_T0'),
            _make_cp_row(3, 1, 'CP_REAL3'),
            _make_cp_row(3, 2, 'REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            for wf_id in ('1', '2', '3'):
                details = result['cp_details'][wf_id]
                self.assertEqual(len(details), 3)
                self.assertEqual(details[0]['is_boundary'], 1)
                self.assertEqual(details[1]['is_boundary'], 0)
                self.assertEqual(details[2]['is_boundary'], 1)
        finally:
            os.remove(path)

    def test_operational_excluded_cps_dropped_not_tagged(self):
        """FA / STOP / RETURN are dropped (not tagged) — they don't belong
        in current_cp_definitions at all."""
        rows = [
            _make_cp_row(1, 0, 'REL_T0'),
            _make_cp_row(1, 1, 'CP_A'),
            _make_cp_row(1, 2, 'STOP TEST'),
            _make_cp_row(1, 3, 'REL FA RETEST'),
            _make_cp_row(1, 4, 'SEND TO FA'),
            _make_cp_row(1, 5, 'RETURN TO REL'),
            _make_cp_row(1, 6, 'CP_B'),
            _make_cp_row(1, 7, 'REL_TFINAL'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            details = result['cp_details']['1']
            cp_names = [d['cp_name'] for d in details]
            # Operational CPs are gone entirely.
            self.assertNotIn('STOP TEST', cp_names)
            self.assertNotIn('REL FA RETEST', cp_names)
            self.assertNotIn('SEND TO FA', cp_names)
            self.assertNotIn('RETURN TO REL', cp_names)
            # Boundary CPs remain, tagged.
            self.assertEqual(
                cp_names,
                ['REL_T0', 'CP_A', 'CP_B', 'REL_TFINAL'],
            )
            self.assertEqual(
                [d['is_boundary'] for d in details],
                [1, 0, 0, 1],
            )
        finally:
            os.remove(path)

    def test_save_cp_schedule_to_db_persists_boundary_tag(self):
        """Round-trip through save_cp_schedule_to_db preserves is_boundary."""
        import db as dbmod
        import tempfile

        # Use an isolated temp DB so we don't touch the real one.
        fd, db_path = tempfile.mkstemp(suffix='.db')
        os.close(fd)
        old_db_path = dbmod.DB_PATH
        dbmod.DB_PATH = db_path
        try:
            conn = dbmod.get_conn()
            dbmod.init_db(conn=conn)
            conn.commit()
            conn.close()

            rows = [
                _make_cp_row(1, 0, 'REL_T0'),
                _make_cp_row(1, 1, 'CP_A'),
                _make_cp_row(1, 2, 'REL_TFINAL'),
            ]
            csv_path = _write_cp_csv(rows)
            try:
                parsed = base_manager.parse_checkpoint_schedule(csv_path)
                base_manager.save_cp_schedule_to_db(parsed)

                conn = dbmod.get_conn()
                try:
                    persisted = dbmod.get_current_cp_definitions(conn)
                finally:
                    conn.close()
            finally:
                os.remove(csv_path)

            self.assertEqual(
                [(cp['cp_name'], cp['is_boundary']) for cp in persisted['1']],
                [
                    ('REL_T0', 1),
                    ('CP_A', 0),
                    ('REL_TFINAL', 1),
                ],
            )
        finally:
            dbmod.DB_PATH = old_db_path
            try:
                os.remove(db_path)
            except OSError:
                pass


class ParseCheckpointScheduleEdgeCaseTests(unittest.TestCase):
    """Test edge cases for CP schedule parser."""

    def test_empty_file(self):
        """Empty CSV (header only) should return empty result."""
        path = _write_cp_csv([])
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            self.assertEqual(result['wf_count'], 0)
            self.assertEqual(result['cp_count'], 0)
            self.assertEqual(result['cp_schedule'], {})
            self.assertEqual(result['cp_details'], {})
        finally:
            os.remove(path)

    def test_wf_with_only_excluded_cps(self):
        """A WF with only excluded CPs should not appear in output."""
        rows = [
            _make_cp_row(1, 0, 'REL_TFINAL'),
            _make_cp_row(1, 1, 'STOP TEST'),
        ]
        path = _write_cp_csv(rows)
        try:
            result = base_manager.parse_checkpoint_schedule(path)
            self.assertEqual(result['wf_count'], 0)
            self.assertNotIn('1', result['cp_schedule'])
        finally:
            os.remove(path)

    def test_real_csv_file(self):
        """Test with the actual WaterfallCheckpointSchedule CSV if available."""
        real_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            'Checkitem data', 'Base',
            'Corn_JAWX2026-04-17-WaterfallCheckpointScheduleEVT20260417-15-vtf2w3.csv'
        )
        if not os.path.exists(real_path):
            self.skipTest("Real CSV file not available")

        result = base_manager.parse_checkpoint_schedule(real_path)
        # Should have many WFs
        self.assertGreater(result['wf_count'], 40)
        # Should have many CPs
        self.assertGreater(result['cp_count'], 500)
        # No excluded CPs (operational + boundary markers) should be present
        excluded = {
            'REL FA RETEST', 'SEND TO FA', 'STOP TEST', 'RETURN TO REL',
            'T0', 'REL_T0', 'End', 'TFinal', 'REL_TFINAL',
        }
        for wf, cps in result['cp_schedule'].items():
            for cp in cps:
                self.assertNotIn(cp, excluded, f"Excluded CP '{cp}' found in WF {wf}")


# ---------------------------------------------------------------------------
# Tests for parse_test_plan (Task 2.5)
# Validates Requirement 1.3
# ---------------------------------------------------------------------------

TP_FIELDNAMES = ['wf id', 'wf test_1', 'wf test_2', 'wf test_3']


def _write_tp_csv(rows, fieldnames=None):
    """Helper: write Test Plan rows to a temp CSV file and return its path."""
    if fieldnames is None:
        fieldnames = TP_FIELDNAMES
    fd, path = tempfile.mkstemp(suffix='.csv')
    os.close(fd)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return path


class ParseTestPlanBasicTests(unittest.TestCase):
    """Test basic parsing of WaterfallTestPlan CSV."""

    def test_parse_single_test(self):
        """WF with only one test name."""
        rows = [{'wf id': '1', 'wf test_1': 'Long Term HS', 'wf test_2': '', 'wf test_3': ''}]
        path = _write_tp_csv(rows)
        try:
            result = base_manager.parse_test_plan(path)
            self.assertEqual(result['wf_count'], 1)
            self.assertEqual(result['test_plan']['1'], ['Long Term HS'])
            self.assertEqual(result['wf_names']['1'], 'Long Term HS')
        finally:
            os.remove(path)

    def test_parse_two_tests(self):
        """WF with two test names."""
        rows = [{'wf id': '4', 'wf test_1': 'Altitude', 'wf test_2': 'Rock Tumble', 'wf test_3': ''}]
        path = _write_tp_csv(rows)
        try:
            result = base_manager.parse_test_plan(path)
            self.assertEqual(result['test_plan']['4'], ['Altitude', 'Rock Tumble'])
            self.assertEqual(result['wf_names']['4'], 'Altitude + Rock Tumble')
        finally:
            os.remove(path)

    def test_parse_three_tests(self):
        """WF with all three test names."""
        rows = [{'wf id': '6', 'wf test_1': 'HS 65/90/72', 'wf test_2': 'Button Cycling - CW', 'wf test_3': 'Button Cycling - CW - Margin'}]
        path = _write_tp_csv(rows)
        try:
            result = base_manager.parse_test_plan(path)
            self.assertEqual(result['test_plan']['6'], ['HS 65/90/72', 'Button Cycling - CW', 'Button Cycling - CW - Margin'])
            self.assertEqual(result['wf_names']['6'], 'HS 65/90/72 + Button Cycling - CW + Button Cycling - CW - Margin')
        finally:
            os.remove(path)

    def test_parse_multiple_wfs(self):
        """Multiple WFs in one file."""
        rows = [
            {'wf id': '1', 'wf test_1': 'Long Term HS', 'wf test_2': '', 'wf test_3': ''},
            {'wf id': '4', 'wf test_1': 'Altitude', 'wf test_2': 'Rock Tumble', 'wf test_3': ''},
            {'wf id': '6', 'wf test_1': 'HS 65/90/72', 'wf test_2': 'Button Cycling - CW', 'wf test_3': 'Margin'},
        ]
        path = _write_tp_csv(rows)
        try:
            result = base_manager.parse_test_plan(path)
            self.assertEqual(result['wf_count'], 3)
            self.assertIn('1', result['test_plan'])
            self.assertIn('4', result['test_plan'])
            self.assertIn('6', result['test_plan'])
        finally:
            os.remove(path)

    def test_decimal_wf_id(self):
        """WF IDs can be decimal like 14.1."""
        rows = [{'wf id': '14.1', 'wf test_1': 'HS 65/90/72', 'wf test_2': '18 Sided Drop 1m PB SeqA', 'wf test_3': ''}]
        path = _write_tp_csv(rows)
        try:
            result = base_manager.parse_test_plan(path)
            self.assertIn('14.1', result['test_plan'])
            self.assertEqual(result['test_plan']['14.1'], ['HS 65/90/72', '18 Sided Drop 1m PB SeqA'])
            self.assertEqual(result['wf_names']['14.1'], 'HS 65/90/72 + 18 Sided Drop 1m PB SeqA')
        finally:
            os.remove(path)


class ParseTestPlanEdgeCaseTests(unittest.TestCase):
    """Test edge cases for Test Plan parser."""

    def test_empty_file(self):
        """Empty CSV (header only) should return empty result."""
        path = _write_tp_csv([])
        try:
            result = base_manager.parse_test_plan(path)
            self.assertEqual(result['wf_count'], 0)
            self.assertEqual(result['test_plan'], {})
            self.assertEqual(result['wf_names'], {})
        finally:
            os.remove(path)

    def test_empty_wf_id_skipped(self):
        """Rows with empty wf id should be skipped."""
        rows = [
            {'wf id': '', 'wf test_1': 'Test A', 'wf test_2': '', 'wf test_3': ''},
            {'wf id': '1', 'wf test_1': 'Test B', 'wf test_2': '', 'wf test_3': ''},
        ]
        path = _write_tp_csv(rows)
        try:
            result = base_manager.parse_test_plan(path)
            self.assertEqual(result['wf_count'], 1)
            self.assertIn('1', result['test_plan'])
        finally:
            os.remove(path)

    def test_all_empty_tests_skipped(self):
        """Rows where all test columns are empty should be skipped."""
        rows = [
            {'wf id': '1', 'wf test_1': '', 'wf test_2': '', 'wf test_3': ''},
            {'wf id': '2', 'wf test_1': 'Valid Test', 'wf test_2': '', 'wf test_3': ''},
        ]
        path = _write_tp_csv(rows)
        try:
            result = base_manager.parse_test_plan(path)
            self.assertEqual(result['wf_count'], 1)
            self.assertNotIn('1', result['test_plan'])
            self.assertIn('2', result['test_plan'])
        finally:
            os.remove(path)

    def test_whitespace_trimmed(self):
        """Whitespace in fields should be trimmed."""
        rows = [{'wf id': ' 1 ', 'wf test_1': ' Long Term HS ', 'wf test_2': '', 'wf test_3': ''}]
        path = _write_tp_csv(rows)
        try:
            result = base_manager.parse_test_plan(path)
            self.assertIn('1', result['test_plan'])
            self.assertEqual(result['test_plan']['1'], ['Long Term HS'])
        finally:
            os.remove(path)

    def test_real_csv_file(self):
        """Test with the actual WaterfallTestPlan CSV file if available."""
        real_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            'Checkitem data', 'Base',
            'Corn_JAWX2026-04-17-WaterfallTestPlanEVT20260417-15-qmtn81.csv'
        )
        if not os.path.exists(real_path):
            self.skipTest("Real CSV file not available")

        result = base_manager.parse_test_plan(real_path)
        # Should have many WFs (the file has 44 rows of data)
        self.assertGreater(result['wf_count'], 40)
        # Each WF should have at least one test name
        for wf_id, tests in result['test_plan'].items():
            self.assertGreater(len(tests), 0, f"WF {wf_id} has no test names")
        # WF names should be non-empty
        for wf_id, name in result['wf_names'].items():
            self.assertTrue(name, f"WF {wf_id} has empty name")
        # Verify specific known values
        self.assertEqual(result['test_plan']['1'], ['Long Term HS 65/90/500'])
        self.assertEqual(result['wf_names']['4'], 'Altitude + Rock Tumble')


class SaveTestPlanToDbTests(unittest.TestCase):
    """Test save_test_plan_to_db writes correctly to DB tables."""

    def setUp(self):
        """Create a temp DB for testing."""
        self._tmp_fd, self._tmp_path = tempfile.mkstemp(suffix='.db')
        os.close(self._tmp_fd)
        # Patch DB_PATH
        import app_paths
        self._orig_db_path = app_paths.DB_PATH
        app_paths.DB_PATH = self._tmp_path
        # Re-import db to pick up new path
        import importlib
        importlib.reload(db)
        db.init_db()

    def tearDown(self):
        """Restore original DB path and clean up."""
        import app_paths
        app_paths.DB_PATH = self._orig_db_path
        import importlib
        importlib.reload(db)
        os.unlink(self._tmp_path)

    def test_saves_test_definitions(self):
        """save_test_plan_to_db should write to current_test_definitions."""
        parsed = {
            'test_plan': {
                '1': ['Long Term HS'],
                '4': ['Altitude', 'Rock Tumble'],
            },
            'wf_names': {
                '1': 'Long Term HS',
                '4': 'Altitude + Rock Tumble',
            },
            'wf_count': 2,
        }
        base_manager.save_test_plan_to_db(parsed)

        conn = db.get_conn()
        test_defs = db.get_current_test_definitions(conn)
        self.assertEqual(test_defs['1'], ['Long Term HS'])
        self.assertEqual(test_defs['4'], ['Altitude', 'Rock Tumble'])
        conn.close()

    def test_saves_wf_definitions(self):
        """save_test_plan_to_db should write to current_wf_definitions."""
        parsed = {
            'test_plan': {
                '1': ['Long Term HS'],
                '6': ['HS 65/90/72', 'Button Cycling - CW', 'Margin'],
            },
            'wf_names': {
                '1': 'Long Term HS',
                '6': 'HS 65/90/72 + Button Cycling - CW + Margin',
            },
            'wf_count': 2,
        }
        base_manager.save_test_plan_to_db(parsed)

        conn = db.get_conn()
        wf_defs = db.get_current_wf_definitions(conn)
        self.assertEqual(wf_defs['1'], 'Long Term HS')
        self.assertEqual(wf_defs['6'], 'HS 65/90/72 + Button Cycling - CW + Margin')
        conn.close()

    def test_replaces_previous_definitions(self):
        """Calling save_test_plan_to_db again should replace previous data."""
        parsed1 = {
            'test_plan': {'1': ['Old Test']},
            'wf_names': {'1': 'Old Test'},
            'wf_count': 1,
        }
        base_manager.save_test_plan_to_db(parsed1)

        parsed2 = {
            'test_plan': {'1': ['New Test'], '2': ['Another']},
            'wf_names': {'1': 'New Test', '2': 'Another'},
            'wf_count': 2,
        }
        base_manager.save_test_plan_to_db(parsed2)

        conn = db.get_conn()
        test_defs = db.get_current_test_definitions(conn)
        wf_defs = db.get_current_wf_definitions(conn)
        # Should have new data only
        self.assertEqual(test_defs['1'], ['New Test'])
        self.assertIn('2', test_defs)
        self.assertEqual(wf_defs['1'], 'New Test')
        self.assertEqual(wf_defs['2'], 'Another')
        conn.close()


# ---------------------------------------------------------------------------
# Tests for parse_test_schedule (Task 2.7)
# Validates Requirement 1.4
# ---------------------------------------------------------------------------

from openpyxl import Workbook


def _create_test_schedule_excel(schedule_data, dates=None):
    """Helper: create a Test Schedule Excel file and return its path.

    Args:
        schedule_data: list of dicts with keys:
            wf_num, test_item, duration, config, allocation, markers
            where markers is a list of (date_index, label) tuples
        dates: list of date strings (YYYY-MM-DD) for the date columns.
            If None, uses a default set of dates.

    Returns:
        Path to the temporary Excel file.
    """
    if dates is None:
        dates = ['2026-04-16', '2026-05-01', '2026-05-15', '2026-06-01', '2026-06-15']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Test Schedule'

    # Header row (row 1): columns A-E are labels, then config columns, then dates
    header_row = 1
    ws.cell(header_row, 1, 'No.')
    ws.cell(header_row, 2, 'WF')
    ws.cell(header_row, 3, 'Test Item')
    ws.cell(header_row, 4, 'Duration')
    ws.cell(header_row, 5, 'Requested')
    ws.cell(header_row, 6, 'R1FNF')
    ws.cell(header_row, 7, 'R2CNM')
    ws.cell(header_row, 8, 'R3')
    ws.cell(header_row, 9, 'R4')

    # Date columns start at column 10
    date_start_col = 10
    for i, date_str in enumerate(dates):
        dt = datetime.datetime.strptime(date_str, '%Y-%m-%d')
        ws.cell(header_row, date_start_col + i, dt)

    # Sub-header row (row 2) - can be empty or have labels
    # Data starts at row 3 (header_row + 2)
    config_col_map = {'R1FNF': 6, 'R2CNM': 7, 'R3': 8, 'R4': 9}

    for row_offset, data in enumerate(schedule_data):
        row_idx = header_row + 2 + row_offset
        if data.get('wf_num') is not None:
            wf_val = data['wf_num']
            # Write as number if possible
            try:
                ws.cell(row_idx, 2, float(wf_val) if '.' in str(wf_val) else int(wf_val))
            except (ValueError, TypeError):
                ws.cell(row_idx, 2, str(wf_val))
        if data.get('test_item') is not None:
            ws.cell(row_idx, 3, data['test_item'])
        if data.get('duration') is not None:
            ws.cell(row_idx, 4, data['duration'])

        config = data.get('config')
        allocation = data.get('allocation', 1)
        if config and config in config_col_map:
            ws.cell(row_idx, config_col_map[config], allocation)

        for date_idx, label in data.get('markers', []):
            ws.cell(row_idx, date_start_col + date_idx, label)

    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(path)
    wb.close()
    return path


class ParseTestScheduleBasicTests(unittest.TestCase):
    """Test basic parsing of Test Schedule Excel."""

    def setUp(self):
        """Use temp file DB so no existing definitions interfere (simple extraction path)."""
        import importlib
        import app_paths
        self._db_fd, self._db_path = tempfile.mkstemp(suffix='.db')
        os.close(self._db_fd)
        os.environ['REPORT_DB_PATH'] = self._db_path
        importlib.reload(app_paths)
        importlib.reload(db)
        db.init_db()
        importlib.reload(base_manager)

    def tearDown(self):
        if 'REPORT_DB_PATH' in os.environ:
            del os.environ['REPORT_DB_PATH']
        try:
            os.remove(self._db_path)
        except OSError:
            pass

    def test_parse_single_segment(self):
        """Single WF with T0 and End markers should produce one segment."""
        schedule_data = [
            {
                'wf_num': '1',
                'test_item': 'Long Term HS',
                'duration': '1000 hrs',
                'config': 'R2CNM',
                'allocation': 5,
                'markers': [(0, 'T0'), (4, 'End')],
            },
        ]
        path = _create_test_schedule_excel(schedule_data)
        try:
            result = base_manager.parse_test_schedule(path)
            self.assertGreater(result['segment_count'], 0)
            seg = result['segments'][0]
            self.assertEqual(seg['wf_num'], '1')
            self.assertEqual(seg['config'], 'R2CNM')
            self.assertEqual(seg['planned_start_date'], '2026-04-16')
            self.assertEqual(seg['planned_end_date'], '2026-06-15')
        finally:
            os.remove(path)

    def test_parse_multiple_configs(self):
        """Multiple configs for same WF should produce multiple segments."""
        schedule_data = [
            {
                'wf_num': '1',
                'test_item': 'Long Term HS',
                'duration': '1000 hrs',
                'config': 'R2CNM',
                'allocation': 5,
                'markers': [(0, 'T0'), (4, 'End')],
            },
            {
                'wf_num': None,  # Inherited from previous row
                'test_item': None,
                'duration': None,
                'config': 'R1FNF',
                'allocation': 3,
                'markers': [(0, 'T0'), (3, 'End')],
            },
        ]
        path = _create_test_schedule_excel(schedule_data)
        try:
            result = base_manager.parse_test_schedule(path)
            self.assertEqual(result['segment_count'], 2)
            configs = {seg['config'] for seg in result['segments']}
            self.assertIn('R2CNM', configs)
            self.assertIn('R1FNF', configs)
        finally:
            os.remove(path)

    def test_parse_multiple_wfs(self):
        """Multiple WFs should produce segments for each."""
        schedule_data = [
            {
                'wf_num': '1',
                'test_item': 'Long Term HS',
                'duration': '1000 hrs',
                'config': 'R2CNM',
                'allocation': 5,
                'markers': [(0, 'T0'), (4, 'End')],
            },
            {
                'wf_num': '4',
                'test_item': 'Altitude',
                'duration': '500 hrs',
                'config': 'R2CNM',
                'allocation': 5,
                'markers': [(1, 'T0'), (3, 'End')],
            },
        ]
        path = _create_test_schedule_excel(schedule_data)
        try:
            result = base_manager.parse_test_schedule(path)
            self.assertEqual(result['segment_count'], 2)
            wf_nums = {seg['wf_num'] for seg in result['segments']}
            self.assertIn('1', wf_nums)
            self.assertIn('4', wf_nums)
        finally:
            os.remove(path)

    def test_segment_has_required_fields(self):
        """Each segment should have all required fields for save_current_schedule_segments."""
        schedule_data = [
            {
                'wf_num': '1',
                'test_item': 'Long Term HS',
                'duration': '1000 hrs',
                'config': 'R2CNM',
                'allocation': 5,
                'markers': [(0, 'T0'), (4, 'End')],
            },
        ]
        path = _create_test_schedule_excel(schedule_data)
        try:
            result = base_manager.parse_test_schedule(path)
            seg = result['segments'][0]
            # Required fields for save_current_schedule_segments
            self.assertIn('wf_num', seg)
            self.assertIn('config', seg)
            self.assertIn('test_idx', seg)
            self.assertIn('planned_start_date', seg)
            self.assertIn('planned_end_date', seg)
            self.assertIn('source_row', seg)
            self.assertIn('confidence', seg)
            self.assertIn('inference_reason', seg)
            self.assertIn('marker_labels', seg)
        finally:
            os.remove(path)


class ParseTestScheduleEdgeCaseTests(unittest.TestCase):
    """Test edge cases for Test Schedule parser."""

    def setUp(self):
        """Use temp file DB so no existing definitions interfere."""
        import importlib
        import app_paths
        self._db_fd, self._db_path = tempfile.mkstemp(suffix='.db')
        os.close(self._db_fd)
        os.environ['REPORT_DB_PATH'] = self._db_path
        importlib.reload(app_paths)
        importlib.reload(db)
        db.init_db()
        importlib.reload(base_manager)

    def tearDown(self):
        if 'REPORT_DB_PATH' in os.environ:
            del os.environ['REPORT_DB_PATH']
        try:
            os.remove(self._db_path)
        except OSError:
            pass

    def test_no_test_schedule_sheet(self):
        """Excel without 'Test Schedule' sheet should return empty result."""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Other Sheet'
        ws.cell(1, 1, 'data')

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(path)
        wb.close()

        try:
            result = base_manager.parse_test_schedule(path)
            self.assertEqual(result['segment_count'], 0)
            self.assertEqual(result['segments'], [])
        finally:
            os.remove(path)

    def test_no_header_row(self):
        """Sheet without proper header row should return empty result."""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Test Schedule'
        ws.cell(1, 1, 'Random')
        ws.cell(1, 2, 'Data')

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(path)
        wb.close()

        try:
            result = base_manager.parse_test_schedule(path)
            self.assertEqual(result['segment_count'], 0)
            self.assertEqual(result['segments'], [])
        finally:
            os.remove(path)

    def test_rows_without_markers_skipped(self):
        """Rows without any markers in date columns should be skipped."""
        schedule_data = [
            {
                'wf_num': '1',
                'test_item': 'Long Term HS',
                'duration': '1000 hrs',
                'config': 'R2CNM',
                'allocation': 5,
                'markers': [],  # No markers
            },
        ]
        path = _create_test_schedule_excel(schedule_data)
        try:
            result = base_manager.parse_test_schedule(path)
            self.assertEqual(result['segment_count'], 0)
        finally:
            os.remove(path)

    def test_rows_without_t0_or_end_skipped(self):
        """Rows with markers but no T0 or End should be skipped in simple mode."""
        schedule_data = [
            {
                'wf_num': '1',
                'test_item': 'Long Term HS',
                'duration': '1000 hrs',
                'config': 'R2CNM',
                'allocation': 5,
                'markers': [(1, '50%'), (2, '100%')],  # No T0 or End
            },
        ]
        path = _create_test_schedule_excel(schedule_data)
        try:
            result = base_manager.parse_test_schedule(path)
            self.assertEqual(result['segment_count'], 0)
        finally:
            os.remove(path)

    def test_mlb_row_stops_parsing(self):
        """MLB marker in WF column should stop parsing."""
        schedule_data = [
            {
                'wf_num': '1',
                'test_item': 'Long Term HS',
                'duration': '1000 hrs',
                'config': 'R2CNM',
                'allocation': 5,
                'markers': [(0, 'T0'), (4, 'End')],
            },
            {
                'wf_num': 'MLB',
                'test_item': 'Should not parse',
                'duration': None,
                'config': 'R2CNM',
                'allocation': 5,
                'markers': [(0, 'T0'), (4, 'End')],
            },
        ]
        path = _create_test_schedule_excel(schedule_data)
        try:
            result = base_manager.parse_test_schedule(path)
            # Only the first row should be parsed
            self.assertEqual(result['segment_count'], 1)
            self.assertEqual(result['segments'][0]['wf_num'], '1')
        finally:
            os.remove(path)


class ParseTestScheduleBasedataFirstRegressionTests(unittest.TestCase):
    """Regression tests for BaseData-first schedule extraction."""

    def setUp(self):
        import importlib
        import app_paths
        self._db_fd, self._db_path = tempfile.mkstemp(suffix='.db')
        os.close(self._db_fd)
        os.environ['REPORT_DB_PATH'] = self._db_path
        importlib.reload(app_paths)
        importlib.reload(db)
        db.init_db()
        importlib.reload(base_manager)

    def tearDown(self):
        if 'REPORT_DB_PATH' in os.environ:
            del os.environ['REPORT_DB_PATH']
        try:
            os.remove(self._db_path)
        except OSError:
            pass

    def test_full_extraction_handles_rel_t0_and_one_based_cp_test_indices(self):
        tp_path = _write_tp_csv([
            {'wf id': '4', 'wf test_1': 'Altitude', 'wf test_2': 'Rock Tumble', 'wf test_3': ''},
        ])
        cp_path = _write_cp_csv([
            _make_cp_row(4, 0, 'REL_T0', test_idx=0),
            _make_cp_row(4, 1, 'ALT_PROFILE1', test_idx=1),
            _make_cp_row(4, 2, 'ALT_PROFILE2', test_idx=1),
            _make_cp_row(4, 3, 'RT_15MIN', test_idx=2),
            _make_cp_row(4, 4, 'RT_30MIN', test_idx=2),
            _make_cp_row(4, 5, 'REL_TFINAL', test_idx=0),
        ])
        schedule_path = _create_test_schedule_excel(
            [
                {
                    'wf_num': '4',
                    'test_item': 'Altitude + Rock tumble',
                    'duration': '5 days',
                    'config': 'R1FNF',
                    'allocation': 1,
                    'markers': [
                        (0, 'REL_T0'),
                        (1, 'Op1'),
                        (2, 'Op2'),
                        (3, '15min'),
                        (4, 'End'),
                    ],
                },
            ],
            dates=['2026-04-16', '2026-04-17', '2026-04-18', '2026-04-19', '2026-04-20'],
        )
        try:
            base_manager.save_test_plan_to_db(base_manager.parse_test_plan(tp_path))
            base_manager.save_cp_schedule_to_db(base_manager.parse_checkpoint_schedule(cp_path))

            result = base_manager.parse_test_schedule(schedule_path)
            self.assertEqual(result['segment_count'], 2)

            segments = sorted(result['segments'], key=lambda seg: seg['test_idx'])
            altitude, tumble = segments

            self.assertEqual(altitude['test_idx'], 0)
            self.assertEqual(altitude['test_name'], 'Altitude')
            self.assertEqual(altitude['planned_start_date'], '2026-04-16')
            self.assertEqual(altitude['planned_end_date'], '2026-04-18')

            self.assertEqual(tumble['test_idx'], 1)
            self.assertEqual(tumble['test_name'], 'Rock Tumble')
            self.assertEqual(tumble['planned_start_date'], '2026-04-19')
            self.assertEqual(tumble['planned_end_date'], '2026-04-20')
        finally:
            os.remove(tp_path)
            os.remove(cp_path)
            os.remove(schedule_path)

    def test_full_extraction_splits_drop_margin_from_checkpoint_category(self):
        tp_path = _write_tp_csv([
            {
                'wf id': '14.2',
                'wf test_1': 'HS 65/90/72',
                'wf test_2': '18 Sided Drop 1m PB SeqB',
                'wf test_3': '18 Sided Drop 1m PB SeqB - Margin',
            },
        ])
        cp_path = _write_cp_csv([
            _make_cp_row(14.2, 0, 'REL_T0', test_idx=0, category='REL_T0'),
            _make_cp_row(14.2, 1, 'HS_72HRS', test_idx=1, category='HS 65/90/72'),
            _make_cp_row(14.2, 2, 'SIDED_DROP_SEQB_1ST_DROP3', test_idx=2, category='18 Sided Drop 1m PB SeqB'),
            _make_cp_row(14.2, 3, 'SIDED_DROP_SEQB_1ST_DROP6', test_idx=2, category='18 Sided Drop 1m PB SeqB'),
            _make_cp_row(14.2, 4, 'SIDED_DROP_SEQB_1ST_DROP18', test_idx=2, category='18 Sided Drop 1m PB SeqB'),
            _make_cp_row(14.2, 5, 'SIDED_DROP_SEQB_2ND_DROP3', test_idx=2, category='18 Sided Drop 1m PB SeqB - Margin'),
            _make_cp_row(14.2, 6, 'SIDED_DROP_SEQB_2ND_DROP6', test_idx=2, category='18 Sided Drop 1m PB SeqB - Margin'),
            _make_cp_row(14.2, 7, 'REL_TFINAL', test_idx=0, category='REL_TFINAL'),
        ])
        schedule_path = _create_test_schedule_excel(
            [
                {
                    'wf_num': '14.2',
                    'test_item': 'HS + Drop + Margin',
                    'duration': '7 days',
                    'config': 'R1FNF',
                    'allocation': 1,
                    'markers': [
                        (0, 'REL_T0'),
                        (1, 'HS 72 hrs'),
                        (2, '3rd drop'),
                        (3, '6th drop'),
                        (4, '18th drop'),
                        (5, '3rd drop'),
                        (6, '6th drop'),
                        (7, 'End'),
                    ],
                },
            ],
            dates=[
                '2026-04-17',
                '2026-04-18',
                '2026-04-19',
                '2026-04-20',
                '2026-04-21',
                '2026-04-22',
                '2026-04-23',
                '2026-04-24',
            ],
        )
        try:
            base_manager.save_test_plan_to_db(base_manager.parse_test_plan(tp_path))
            base_manager.save_cp_schedule_to_db(base_manager.parse_checkpoint_schedule(cp_path))

            result = base_manager.parse_test_schedule(schedule_path)
            self.assertEqual(result['segment_count'], 3)

            segments = sorted(result['segments'], key=lambda seg: seg['test_idx'])
            hs, drop, margin = segments

            self.assertEqual((hs['test_idx'], hs['planned_start_date'], hs['planned_end_date']), (0, '2026-04-17', '2026-04-18'))
            self.assertEqual((drop['test_idx'], drop['planned_start_date'], drop['planned_end_date']), (1, '2026-04-19', '2026-04-21'))
            self.assertEqual((margin['test_idx'], margin['planned_start_date'], margin['planned_end_date']), (2, '2026-04-22', '2026-04-24'))
            self.assertEqual(margin['marker_labels'][0], '3rd drop')
        finally:
            os.remove(tp_path)
            os.remove(cp_path)
            os.remove(schedule_path)

    def test_full_extraction_keeps_main_button_cycling_between_hs_and_margin(self):
        tp_path = _write_tp_csv([
            {
                'wf id': '6',
                'wf test_1': 'HS 65/90/72',
                'wf test_2': 'Button Cycling - CW',
                'wf test_3': 'Button Cycling - CW - Margin',
            },
        ])
        cp_path = _write_cp_csv([
            _make_cp_row(6, 0, 'REL_T0', test_idx=0, category='REL_T0'),
            _make_cp_row(6, 1, 'HS_72HRS', test_idx=1, category='HS 65/90/72'),
            _make_cp_row(6, 2, 'BC_CWCB_SHORT_10%', test_idx=2, category='Button Cycling - CW'),
            _make_cp_row(6, 3, 'BC_CWCB_LONG_50%', test_idx=2, category='Button Cycling - CW'),
            _make_cp_row(6, 4, 'BC_CWCB_SHORT_100%', test_idx=2, category='Button Cycling - CW'),
            _make_cp_row(6, 5, 'BC_CWCB_SHORT_140%', test_idx=3, category='Button Cycling - CW - Margin'),
            _make_cp_row(6, 6, 'BC_CWCB_LONG_200%', test_idx=3, category='Button Cycling - CW - Margin'),
            _make_cp_row(6, 7, 'REL_TFINAL', test_idx=0, category='REL_TFINAL'),
        ])
        schedule_path = _create_test_schedule_excel(
            [
                {
                    'wf_num': '6',
                    'test_item': 'HS 65C 90RH 72 hrs + Button Cycling CW',
                    'duration': '10 days',
                    'config': 'R1FNF',
                    'allocation': 1,
                    'markers': [
                        (0, 'T0'),
                        (1, 'HS 72 hrs'),
                        (2, '0.1'),
                        (3, '0.5'),
                        (4, '1'),
                        (5, '1.4'),
                        (6, '2'),
                        (7, 'End'),
                    ],
                },
            ],
            dates=[
                '2026-04-17',
                '2026-04-25',
                '2026-04-27',
                '2026-05-09',
                '2026-05-20',
                '2026-05-23',
                '2026-05-29',
                '2026-05-30',
            ],
        )
        try:
            base_manager.save_test_plan_to_db(base_manager.parse_test_plan(tp_path))
            base_manager.save_cp_schedule_to_db(base_manager.parse_checkpoint_schedule(cp_path))

            result = base_manager.parse_test_schedule(schedule_path)
            self.assertEqual(result['segment_count'], 3)

            segments = sorted(result['segments'], key=lambda seg: seg['test_idx'])
            hs, main, margin = segments

            self.assertEqual((hs['test_idx'], hs['planned_start_date'], hs['planned_end_date']), (0, '2026-04-17', '2026-04-25'))
            self.assertEqual((main['test_idx'], main['planned_start_date'], main['planned_end_date']), (1, '2026-04-27', '2026-05-20'))
            self.assertEqual((margin['test_idx'], margin['planned_start_date'], margin['planned_end_date']), (2, '2026-05-23', '2026-05-30'))
        finally:
            os.remove(tp_path)
            os.remove(cp_path)
            os.remove(schedule_path)

    def test_full_extraction_preserves_cleaning_markers_for_single_test_rows(self):
        tp_path = _write_tp_csv([
            {'wf id': '29.1', 'wf test_1': 'Cleaning Spray OP1', 'wf test_2': '', 'wf test_3': ''},
        ])
        cp_path = _write_cp_csv([
            _make_cp_row(29.1, 0, 'REL_T0', test_idx=0, category='REL_T0'),
            _make_cp_row(29.1, 1, 'CLEANING_SPRAY_OP1_AFTER 18HRS', test_idx=1, category='Cleaning Spray OP1'),
            _make_cp_row(29.1, 2, 'CLEANING_SPRAY_OP1_AFTER 36HRS', test_idx=1, category='Cleaning Spray OP1'),
            _make_cp_row(29.1, 3, 'REL_TFINAL', test_idx=0, category='REL_TFINAL'),
        ])
        schedule_path = _create_test_schedule_excel(
            [
                {
                    'wf_num': '29.1',
                    'test_item': 'Cleaning Spray OP1',
                    'duration': '3 days',
                    'config': 'R1FNF',
                    'allocation': 1,
                    'markers': [
                        (0, 'T0'),
                        (1, '1st A'),
                        (2, '1st B'),
                        (3, 'End'),
                    ],
                },
            ],
            dates=['2026-04-17', '2026-04-18', '2026-04-19', '2026-04-20'],
        )
        try:
            base_manager.save_test_plan_to_db(base_manager.parse_test_plan(tp_path))
            base_manager.save_cp_schedule_to_db(base_manager.parse_checkpoint_schedule(cp_path))

            result = base_manager.parse_test_schedule(schedule_path)
            self.assertEqual(result['segment_count'], 1)
            self.assertEqual(result['segments'][0]['marker_labels'], ['T0', '1st A', '1st B', 'End'])
        finally:
            os.remove(tp_path)
            os.remove(cp_path)
            os.remove(schedule_path)


class SaveTestScheduleToDbTests(unittest.TestCase):
    """Test save_test_schedule_to_db writes correctly to DB."""

    def setUp(self):
        """Set up a fresh DB for each test."""
        import importlib
        import app_paths
        self._db_fd, self._db_path = tempfile.mkstemp(suffix='.db')
        os.close(self._db_fd)
        os.environ['REPORT_DB_PATH'] = self._db_path
        importlib.reload(app_paths)
        importlib.reload(db)
        db.init_db()
        importlib.reload(base_manager)

    def tearDown(self):
        if 'REPORT_DB_PATH' in os.environ:
            del os.environ['REPORT_DB_PATH']
        try:
            os.remove(self._db_path)
        except OSError:
            pass

    def test_saves_segments_to_db(self):
        """Segments should be written to current_schedule_segments table."""
        parsed_result = {
            'segments': [
                {
                    'wf_num': '1',
                    'config': 'R2CNM',
                    'test_idx': 0,
                    'test_name': 'Long Term HS',
                    'schedule_test_item': 'Long Term HS',
                    'planned_start_date': '2026-04-16',
                    'planned_end_date': '2026-06-15',
                    'source_row': 3,
                    'confidence': 'low',
                    'inference_reason': 'simple-extraction-no-definitions',
                    'marker_labels': ['T0', 'End'],
                },
            ],
            'segment_count': 1,
        }
        base_manager.save_test_schedule_to_db(parsed_result)

        conn = db.get_conn()
        segments = db.get_current_schedule_segments(conn)
        conn.close()

        self.assertEqual(len(segments), 1)
        seg = segments[0]
        self.assertEqual(seg['wf_num'], '1')
        self.assertEqual(seg['config'], 'R2CNM')
        self.assertEqual(seg['test_idx'], 0)
        self.assertEqual(seg['planned_start_date'], '2026-04-16')
        self.assertEqual(seg['planned_end_date'], '2026-06-15')

    def test_replaces_previous_segments(self):
        """Calling save_test_schedule_to_db again should replace previous data."""
        parsed1 = {
            'segments': [
                {
                    'wf_num': '1',
                    'config': 'R2CNM',
                    'test_idx': 0,
                    'test_name': 'Old Test',
                    'schedule_test_item': 'Old Test',
                    'planned_start_date': '2026-04-01',
                    'planned_end_date': '2026-05-01',
                    'source_row': 3,
                    'confidence': 'low',
                    'inference_reason': 'simple',
                    'marker_labels': ['T0', 'End'],
                },
            ],
            'segment_count': 1,
        }
        base_manager.save_test_schedule_to_db(parsed1)

        parsed2 = {
            'segments': [
                {
                    'wf_num': '2',
                    'config': 'R1FNF',
                    'test_idx': 0,
                    'test_name': 'New Test',
                    'schedule_test_item': 'New Test',
                    'planned_start_date': '2026-05-01',
                    'planned_end_date': '2026-06-01',
                    'source_row': 4,
                    'confidence': 'high',
                    'inference_reason': 'new',
                    'marker_labels': ['T0', 'End'],
                },
            ],
            'segment_count': 1,
        }
        base_manager.save_test_schedule_to_db(parsed2)

        conn = db.get_conn()
        segments = db.get_current_schedule_segments(conn)
        conn.close()

        # Should only have the new data
        self.assertEqual(len(segments), 1)
        self.assertEqual(segments[0]['wf_num'], '2')
        self.assertEqual(segments[0]['config'], 'R1FNF')

    def test_saves_multiple_segments(self):
        """Multiple segments should all be saved."""
        parsed_result = {
            'segments': [
                {
                    'wf_num': '1',
                    'config': 'R2CNM',
                    'test_idx': 0,
                    'test_name': 'Test A',
                    'schedule_test_item': 'Test A',
                    'planned_start_date': '2026-04-16',
                    'planned_end_date': '2026-05-15',
                    'source_row': 3,
                    'confidence': 'high',
                    'inference_reason': 'single-test-row',
                    'marker_labels': ['T0', 'End'],
                },
                {
                    'wf_num': '1',
                    'config': 'R1FNF',
                    'test_idx': 0,
                    'test_name': 'Test A',
                    'schedule_test_item': 'Test A',
                    'planned_start_date': '2026-04-16',
                    'planned_end_date': '2026-06-01',
                    'source_row': 4,
                    'confidence': 'high',
                    'inference_reason': 'single-test-row',
                    'marker_labels': ['T0', 'End'],
                },
                {
                    'wf_num': '4',
                    'config': 'R2CNM',
                    'test_idx': 0,
                    'test_name': 'Altitude',
                    'schedule_test_item': 'Altitude',
                    'planned_start_date': '2026-05-01',
                    'planned_end_date': '2026-06-15',
                    'source_row': 5,
                    'confidence': 'medium',
                    'inference_reason': 'boundary-inference',
                    'marker_labels': ['T0', '50%', 'End'],
                },
            ],
            'segment_count': 3,
        }
        base_manager.save_test_schedule_to_db(parsed_result)

        conn = db.get_conn()
        segments = db.get_current_schedule_segments(conn)
        conn.close()

        self.assertEqual(len(segments), 3)

    def test_empty_segments_clears_table(self):
        """Saving empty segments should clear the table."""
        # First save some data
        parsed1 = {
            'segments': [
                {
                    'wf_num': '1',
                    'config': 'R2CNM',
                    'test_idx': 0,
                    'test_name': 'Test',
                    'schedule_test_item': 'Test',
                    'planned_start_date': '2026-04-16',
                    'planned_end_date': '2026-06-15',
                    'source_row': 3,
                    'confidence': 'low',
                    'inference_reason': 'simple',
                    'marker_labels': ['T0', 'End'],
                },
            ],
            'segment_count': 1,
        }
        base_manager.save_test_schedule_to_db(parsed1)

        # Then save empty
        parsed2 = {'segments': [], 'segment_count': 0}
        base_manager.save_test_schedule_to_db(parsed2)

        conn = db.get_conn()
        segments = db.get_current_schedule_segments(conn)
        conn.close()

        self.assertEqual(len(segments), 0)


# ---------------------------------------------------------------------------
# Tests for upload_base_file and get_base_status (Task 2.8)
# Validates Requirements 1.5, 1.6
# ---------------------------------------------------------------------------


class UploadBaseFileTests(unittest.TestCase):
    """Test file storage and metadata tracking."""

    def setUp(self):
        """Create a temp directory to use as RAWDATA_DIR."""
        import tempfile
        self._orig_rawdata_dir = base_manager.RAWDATA_DIR
        self._orig_base_dir_module = base_manager.BASE_DIR
        self._tmpdir = tempfile.mkdtemp()
        # Patch RAWDATA_DIR in base_manager module
        base_manager.RAWDATA_DIR = os.path.join(self._tmpdir, 'rawdata')
        base_manager.BASE_DIR = self._tmpdir

        # Also patch app_paths for get_sn_mapping_from_db
        import app_paths
        self._orig_app_paths_rawdata = app_paths.RAWDATA_DIR
        self._orig_app_paths_base = app_paths.BASE_DIR
        app_paths.RAWDATA_DIR = base_manager.RAWDATA_DIR
        app_paths.BASE_DIR = self._tmpdir

        # Create a temp DB
        self._db_fd, self._db_path = tempfile.mkstemp(suffix='.db')
        os.close(self._db_fd)
        self._orig_db_path = db.DB_PATH
        db.DB_PATH = self._db_path
        conn = db.get_conn()
        db.init_db(conn=conn)
        conn.close()

    def tearDown(self):
        """Restore original paths and clean up."""
        import shutil
        base_manager.RAWDATA_DIR = self._orig_rawdata_dir
        base_manager.BASE_DIR = self._orig_base_dir_module
        import app_paths
        app_paths.RAWDATA_DIR = self._orig_app_paths_rawdata
        app_paths.BASE_DIR = self._orig_app_paths_base
        db.DB_PATH = self._orig_db_path
        shutil.rmtree(self._tmpdir, ignore_errors=True)
        if os.path.exists(self._db_path):
            os.remove(self._db_path)

    def _create_sn_csv(self):
        """Create a minimal SN mapping CSV and return its path."""
        path = os.path.join(self._tmpdir, 'sn_input.csv')
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['serial_number', 'config', 'Product', 'unit_number', 'start_date', 'wf_id'])
            writer.writeheader()
            writer.writerow({'serial_number': 'SN001', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-1-1', 'start_date': '20260416', 'wf_id': '1'})
            writer.writerow({'serial_number': 'SN002', 'config': 'R1FNF', 'Product': 'B529', 'unit_number': 'ER1-1-1', 'start_date': '20260416', 'wf_id': '2'})
        return path

    def _create_cp_csv(self):
        """Create a minimal CP schedule CSV and return its path."""
        path = os.path.join(self._tmpdir, 'cp_input.csv')
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=CP_FIELDNAMES)
            writer.writeheader()
            writer.writerow(_make_cp_row(1, 0, 'REL_T0', test_idx=0))
            writer.writerow(_make_cp_row(1, 1, 'CP_A', test_idx=1))
            writer.writerow(_make_cp_row(1, 2, 'REL_TFINAL', test_idx=0))
        return path

    def _create_tp_csv(self):
        """Create a minimal Test Plan CSV and return its path."""
        path = os.path.join(self._tmpdir, 'tp_input.csv')
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=TP_FIELDNAMES)
            writer.writeheader()
            writer.writerow({'wf id': '1', 'wf test_1': 'Long Term HS', 'wf test_2': '', 'wf test_3': ''})
        return path

    def test_upload_sn_mapping_stores_file(self):
        """Uploading SN mapping should store file to rawdata/base/sn_mapping.csv."""
        src_path = self._create_sn_csv()
        result = base_manager.upload_base_file(src_path, 'sn_mapping')

        stored_path = os.path.join(base_manager.RAWDATA_DIR, 'base', 'sn_mapping.csv')
        self.assertTrue(os.path.exists(stored_path))
        self.assertEqual(result['sn_count'], 2)
        self.assertEqual(result['config_count'], 2)

    def test_upload_sn_mapping_records_metadata(self):
        """Uploading SN mapping should record metadata in base_file_meta."""
        src_path = self._create_sn_csv()
        base_manager.upload_base_file(src_path, 'sn_mapping')

        conn = db.get_conn()
        row = conn.execute("SELECT * FROM base_file_meta WHERE file_type = 'sn_mapping'").fetchone()
        conn.close()

        self.assertIsNotNone(row)
        self.assertEqual(row['file_type'], 'sn_mapping')
        self.assertEqual(row['original_filename'], 'sn_input.csv')
        self.assertEqual(row['stored_path'], os.path.join('rawdata', 'base', 'sn_mapping.csv'))
        self.assertTrue(row['uploaded_at'])
        summary = json.loads(row['parsed_summary'])
        self.assertEqual(summary['sn_count'], 2)

    def test_upload_cp_schedule_stores_and_parses(self):
        """Uploading CP schedule should store file and write to DB."""
        src_path = self._create_cp_csv()
        result = base_manager.upload_base_file(src_path, 'checkpoint_schedule')

        stored_path = os.path.join(base_manager.RAWDATA_DIR, 'base', 'checkpoint_schedule.csv')
        self.assertTrue(os.path.exists(stored_path))
        self.assertEqual(result['wf_count'], 1)
        self.assertEqual(result['cp_count'], 1)  # CP_A only (REL_T0 / REL_TFINAL are boundary markers, filtered out)

        # Verify DB was updated
        conn = db.get_conn()
        cps = db.get_current_cp_definitions(conn)
        conn.close()
        self.assertIn('1', cps)

    def test_upload_test_plan_stores_and_parses(self):
        """Uploading test plan should store file and write to DB."""
        src_path = self._create_tp_csv()
        result = base_manager.upload_base_file(src_path, 'test_plan')

        stored_path = os.path.join(base_manager.RAWDATA_DIR, 'base', 'test_plan.csv')
        self.assertTrue(os.path.exists(stored_path))
        self.assertEqual(result['wf_count'], 1)

        # Verify DB was updated
        conn = db.get_conn()
        test_defs = db.get_current_test_definitions(conn)
        wf_defs = db.get_current_wf_definitions(conn)
        conn.close()
        self.assertIn('1', test_defs)
        self.assertIn('1', wf_defs)

    def test_replacement_deletes_old_file_and_metadata(self):
        """Re-uploading same type should replace old file and metadata."""
        # First upload
        src_path1 = self._create_sn_csv()
        base_manager.upload_base_file(src_path1, 'sn_mapping')

        # Create a different SN CSV
        path2 = os.path.join(self._tmpdir, 'sn_input2.csv')
        with open(path2, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['serial_number', 'config', 'Product', 'unit_number', 'start_date', 'wf_id'])
            writer.writeheader()
            writer.writerow({'serial_number': 'SN_NEW', 'config': 'R3', 'Product': 'B529', 'unit_number': 'ER3-1-1', 'start_date': '20260420', 'wf_id': '3'})

        # Second upload (replacement)
        result = base_manager.upload_base_file(path2, 'sn_mapping')
        self.assertEqual(result['sn_count'], 1)

        # Only one metadata row should exist
        conn = db.get_conn()
        rows = conn.execute("SELECT * FROM base_file_meta WHERE file_type = 'sn_mapping'").fetchall()
        conn.close()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['original_filename'], 'sn_input2.csv')

        # File should contain new data
        stored_path = os.path.join(base_manager.RAWDATA_DIR, 'base', 'sn_mapping.csv')
        re_parsed = base_manager.parse_sn_mapping(stored_path)
        self.assertIn('SN_NEW', re_parsed['sn_mapping'])
        self.assertNotIn('SN001', re_parsed['sn_mapping'])

    def test_invalid_file_type_raises(self):
        """Unknown file_type should raise ValueError."""
        src_path = self._create_sn_csv()
        with self.assertRaises(ValueError):
            base_manager.upload_base_file(src_path, 'unknown_type')

    def test_upload_with_file_like_object(self):
        """Should handle file-like objects with read() method."""
        import io
        csv_content = "serial_number,config,Product,unit_number,start_date,wf_id\nSN001,R2CNM,B529,ER2-1-1,20260416,1\n"
        file_obj = io.BytesIO(csv_content.encode('utf-8'))
        file_obj.filename = 'test_upload.csv'

        # Simulate a file-like object without .save()
        result = base_manager.upload_base_file(file_obj, 'sn_mapping')
        self.assertEqual(result['sn_count'], 1)

        stored_path = os.path.join(base_manager.RAWDATA_DIR, 'base', 'sn_mapping.csv')
        self.assertTrue(os.path.exists(stored_path))

    def test_creates_base_directory_if_not_exists(self):
        """rawdata/base/ directory should be created automatically."""
        # Ensure it doesn't exist
        base_dir = os.path.join(base_manager.RAWDATA_DIR, 'base')
        if os.path.exists(base_dir):
            import shutil
            shutil.rmtree(base_dir)

        src_path = self._create_sn_csv()
        base_manager.upload_base_file(src_path, 'sn_mapping')

        self.assertTrue(os.path.isdir(base_dir))


class GetBaseStatusTests(unittest.TestCase):
    """Test get_base_status returns correct file status."""

    def setUp(self):
        """Create a temp directory and DB."""
        import tempfile
        self._orig_rawdata_dir = base_manager.RAWDATA_DIR
        self._orig_base_dir_module = base_manager.BASE_DIR
        self._tmpdir = tempfile.mkdtemp()
        base_manager.RAWDATA_DIR = os.path.join(self._tmpdir, 'rawdata')
        base_manager.BASE_DIR = self._tmpdir

        import app_paths
        self._orig_app_paths_rawdata = app_paths.RAWDATA_DIR
        self._orig_app_paths_base = app_paths.BASE_DIR
        app_paths.RAWDATA_DIR = base_manager.RAWDATA_DIR
        app_paths.BASE_DIR = self._tmpdir

        self._db_fd, self._db_path = tempfile.mkstemp(suffix='.db')
        os.close(self._db_fd)
        self._orig_db_path = db.DB_PATH
        db.DB_PATH = self._db_path
        conn = db.get_conn()
        db.init_db(conn=conn)
        conn.close()

    def tearDown(self):
        import shutil
        base_manager.RAWDATA_DIR = self._orig_rawdata_dir
        base_manager.BASE_DIR = self._orig_base_dir_module
        import app_paths
        app_paths.RAWDATA_DIR = self._orig_app_paths_rawdata
        app_paths.BASE_DIR = self._orig_app_paths_base
        db.DB_PATH = self._orig_db_path
        shutil.rmtree(self._tmpdir, ignore_errors=True)
        if os.path.exists(self._db_path):
            os.remove(self._db_path)

    def test_empty_status_when_no_files_uploaded(self):
        """get_base_status should return empty list when nothing uploaded."""
        status = base_manager.get_base_status()
        self.assertEqual(status, [])

    def test_status_after_upload(self):
        """get_base_status should return metadata for uploaded files."""
        # Upload a SN mapping
        path = os.path.join(self._tmpdir, 'sn.csv')
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['serial_number', 'config', 'Product', 'unit_number', 'start_date', 'wf_id'])
            writer.writeheader()
            writer.writerow({'serial_number': 'SN001', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-1-1', 'start_date': '20260416', 'wf_id': '1'})

        base_manager.upload_base_file(path, 'sn_mapping')

        status = base_manager.get_base_status()
        self.assertEqual(len(status), 1)
        self.assertEqual(status[0]['file_type'], 'sn_mapping')
        self.assertEqual(status[0]['original_filename'], 'sn.csv')
        self.assertIsInstance(status[0]['parsed_summary'], dict)
        self.assertEqual(status[0]['parsed_summary']['sn_count'], 1)

    def test_status_multiple_files(self):
        """get_base_status should return all uploaded file types."""
        # Upload SN mapping
        sn_path = os.path.join(self._tmpdir, 'sn.csv')
        with open(sn_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['serial_number', 'config', 'Product', 'unit_number', 'start_date', 'wf_id'])
            writer.writeheader()
            writer.writerow({'serial_number': 'SN001', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-1-1', 'start_date': '20260416', 'wf_id': '1'})

        base_manager.upload_base_file(sn_path, 'sn_mapping')

        # Upload test plan
        tp_path = os.path.join(self._tmpdir, 'tp.csv')
        with open(tp_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=TP_FIELDNAMES)
            writer.writeheader()
            writer.writerow({'wf id': '1', 'wf test_1': 'Long Term HS', 'wf test_2': '', 'wf test_3': ''})

        base_manager.upload_base_file(tp_path, 'test_plan')

        status = base_manager.get_base_status()
        self.assertEqual(len(status), 2)
        file_types = [s['file_type'] for s in status]
        self.assertIn('sn_mapping', file_types)
        self.assertIn('test_plan', file_types)

    def test_status_after_replacement(self):
        """After replacement, status should show only the latest file."""
        # First upload
        path1 = os.path.join(self._tmpdir, 'sn_v1.csv')
        with open(path1, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['serial_number', 'config', 'Product', 'unit_number', 'start_date', 'wf_id'])
            writer.writeheader()
            writer.writerow({'serial_number': 'SN001', 'config': 'R2CNM', 'Product': 'B529', 'unit_number': 'ER2-1-1', 'start_date': '20260416', 'wf_id': '1'})

        base_manager.upload_base_file(path1, 'sn_mapping')

        # Replacement upload
        path2 = os.path.join(self._tmpdir, 'sn_v2.csv')
        with open(path2, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['serial_number', 'config', 'Product', 'unit_number', 'start_date', 'wf_id'])
            writer.writeheader()
            writer.writerow({'serial_number': 'SN_A', 'config': 'R3', 'Product': 'B529', 'unit_number': 'ER3-1-1', 'start_date': '20260420', 'wf_id': '3'})
            writer.writerow({'serial_number': 'SN_B', 'config': 'R3', 'Product': 'B529', 'unit_number': 'ER3-1-2', 'start_date': '20260420', 'wf_id': '3'})

        base_manager.upload_base_file(path2, 'sn_mapping')

        status = base_manager.get_base_status()
        self.assertEqual(len(status), 1)
        self.assertEqual(status[0]['original_filename'], 'sn_v2.csv')
        self.assertEqual(status[0]['parsed_summary']['sn_count'], 2)


if __name__ == '__main__':
    unittest.main()
