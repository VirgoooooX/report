"""
Tests for engine.py definition priority logic (Requirements 7.1-7.5).

Verifies that parse_daily_report() correctly prioritizes definition sources:
- Test Summary sheet in Excel (backward compat) vs DB definitions
- Proper fallback and error handling when sources are missing
"""
import os
import sqlite3
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook

import db
import engine


def _temp_db():
    """Create a temporary SQLite DB with full schema."""
    fd, path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    conn = sqlite3.connect(path)
    conn.row_factory = db._dict_factory
    conn.execute("PRAGMA journal_mode=WAL")
    db.init_db(conn=conn)
    return conn, path


def _make_minimal_wf_sheet(wb, wf_num, cp_names, sn_data=None):
    """Add a minimal WF sheet to a workbook.

    Args:
        wb: openpyxl Workbook
        wf_num: str like '1' or '10'
        cp_names: list of CP name strings
        sn_data: optional list of (config, unit_num, sn) tuples for data rows
    """
    ws = wb.create_sheet(f'Sys WF{wf_num}_TestWF')

    # Header row
    ws.cell(1, 3).value = 'Config'
    ws.cell(1, 4).value = 'Unit #'
    ws.cell(1, 5).value = 'S/N'
    ws.cell(1, 6).value = 'T0'
    for i, cp_name in enumerate(cp_names):
        ws.cell(1, 7 + i).value = cp_name
    ws.cell(1, 7 + len(cp_names)).value = 'Comments'

    # Check items row
    for i in range(len(cp_names)):
        ws.cell(2, 7 + i).value = 'Cosmetic'

    # Data rows
    if sn_data:
        for row_offset, (config, unit_num, sn) in enumerate(sn_data):
            r = 3 + row_offset
            ws.cell(r, 3).value = config
            ws.cell(r, 4).value = unit_num
            ws.cell(r, 5).value = sn
            ws.cell(r, 6).value = 'T0'
            # Put a PASS value in first CP
            ws.cell(r, 7).value = 'PASS'

    return ws


def _add_test_summary_sheet(wb, wf_test_names):
    """Add a Test Summary sheet to the workbook.

    Args:
        wb: openpyxl Workbook
        wf_test_names: dict {wf_num_str: [test_name, ...]}
    """
    ws = wb.create_sheet('Test Summary')

    # System row marker (required by read_test_summary_from_workbook)
    ws.cell(1, 2).value = 'System'

    # WF data starts at row 14
    row = 14
    for wf_num_str, test_names in wf_test_names.items():
        ws.cell(row, 2).value = int(wf_num_str)
        # Quantities (cols 3-6)
        for offset in range(4):
            ws.cell(row, 3 + offset).value = 5

        # Test names in columns 7, 12, 17, 22
        test_cols = [7, 12, 17, 22]
        for ti, tname in enumerate(test_names[:4]):
            ws.cell(row, test_cols[ti]).value = tname
        row += 1


def _add_test_schedule_sheet(wb, wf_names):
    """Add a Test Schedule sheet to the workbook.

    Args:
        wb: openpyxl Workbook
        wf_names: dict {wf_num_str: wf_name}
    """
    ws = wb.create_sheet('Test Schedule')

    # Header row
    ws.cell(1, 2).value = 'WF'
    ws.cell(1, 3).value = 'Test Item'

    row = 2
    for wf_num_str, wf_name in wf_names.items():
        ws.cell(row, 2).value = int(wf_num_str)
        ws.cell(row, 3).value = wf_name
        row += 1


def _save_workbook_to_temp(wb):
    """Save workbook to a temp file and return the path."""
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(path)
    return path


def _populate_db_test_definitions(conn, test_defs):
    """Insert test definitions into current_test_definitions.

    Args:
        conn: DB connection
        test_defs: dict {wf_num: [test_name, ...]}
    """
    for wf_num, names in test_defs.items():
        for test_idx, test_name in enumerate(names):
            conn.execute(
                """INSERT INTO current_test_definitions (wf_num, test_idx, test_name, updated_run_id)
                   VALUES (?, ?, ?, ?)""",
                (str(wf_num), test_idx, test_name, 1)
            )
    conn.commit()


def _populate_db_wf_definitions(conn, wf_defs):
    """Insert WF definitions into current_wf_definitions.

    Args:
        conn: DB connection
        wf_defs: dict {wf_num: wf_name}
    """
    for wf_num, wf_name in wf_defs.items():
        conn.execute(
            """INSERT INTO current_wf_definitions (wf_num, wf_name, updated_run_id)
               VALUES (?, ?, ?)""",
            (str(wf_num), wf_name, 1)
        )
    conn.commit()


class DefinitionPriorityDBFirstTests(unittest.TestCase):
    """Requirement 7.1, 7.3: DB has definitions, no Test Summary sheet → uses DB."""

    def setUp(self):
        self.conn, self.db_path = _temp_db()
        self.excel_path = None

    def tearDown(self):
        self.conn.close()
        os.remove(self.db_path)
        if self.excel_path and os.path.exists(self.excel_path):
            os.remove(self.excel_path)

    def test_uses_db_test_definitions_when_no_test_summary(self):
        """When DB has test definitions and Excel has no Test Summary → uses DB definitions."""
        # Set up DB with test definitions
        db_test_defs = {'1': ['Drop Test', 'Vibration Test'], '2': ['Thermal Cycling']}
        _populate_db_test_definitions(self.conn, db_test_defs)
        _populate_db_wf_definitions(self.conn, {'1': 'Drop + Vibration', '2': 'Thermal'})

        # Create workbook WITHOUT Test Summary sheet
        wb = Workbook()
        wb.remove(wb.active)
        _make_minimal_wf_sheet(wb, '1', ['Drop20', 'Drop40'], [('R3', 'ER3-1-1', 'SN001')])
        _make_minimal_wf_sheet(wb, '2', ['TC100'], [('R3', 'ER3-2-1', 'SN002')])
        self.excel_path = _save_workbook_to_temp(wb)
        wb.close()

        # Parse — should use DB definitions
        result = engine.parse_daily_report(self.excel_path, conn=self.conn)

        self.assertEqual(result.ts_test_names['1'], ['Drop Test', 'Vibration Test'])
        self.assertEqual(result.ts_test_names['2'], ['Thermal Cycling'])

    def test_uses_db_wf_definitions_when_no_test_schedule(self):
        """When DB has WF definitions and Excel has no Test Schedule → uses DB definitions."""
        db_test_defs = {'1': ['Drop Test']}
        db_wf_defs = {'1': 'Random Drop PB'}
        _populate_db_test_definitions(self.conn, db_test_defs)
        _populate_db_wf_definitions(self.conn, db_wf_defs)

        # Create workbook WITHOUT Test Schedule sheet
        wb = Workbook()
        wb.remove(wb.active)
        _make_minimal_wf_sheet(wb, '1', ['Drop20'], [('R3', 'ER3-1-1', 'SN001')])
        self.excel_path = _save_workbook_to_temp(wb)
        wb.close()

        result = engine.parse_daily_report(self.excel_path, conn=self.conn)

        self.assertEqual(result.wf_names['1'], 'Random Drop PB')


class DefinitionPriorityExcelOverDBTests(unittest.TestCase):
    """Requirement 7.5: DB has definitions, Test Summary sheet exists → uses Excel."""

    def setUp(self):
        self.conn, self.db_path = _temp_db()
        self.excel_path = None

    def tearDown(self):
        self.conn.close()
        os.remove(self.db_path)
        if self.excel_path and os.path.exists(self.excel_path):
            os.remove(self.excel_path)

    def test_excel_test_summary_takes_priority_over_db(self):
        """When both DB and Test Summary exist, Test Summary wins for ts_test_names."""
        # DB has definitions
        db_test_defs = {'1': ['DB Test A', 'DB Test B']}
        _populate_db_test_definitions(self.conn, db_test_defs)
        _populate_db_wf_definitions(self.conn, {'1': 'DB WF Name'})

        # Create workbook WITH Test Summary sheet (different names)
        wb = Workbook()
        wb.remove(wb.active)
        _make_minimal_wf_sheet(wb, '1', ['Drop20', 'Drop40'], [('R3', 'ER3-1-1', 'SN001')])
        _add_test_summary_sheet(wb, {'1': ['Excel Test X', 'Excel Test Y']})
        self.excel_path = _save_workbook_to_temp(wb)
        wb.close()

        result = engine.parse_daily_report(self.excel_path, conn=self.conn)

        # Test Summary sheet should take priority
        self.assertEqual(result.ts_test_names['1'], ['Excel Test X', 'Excel Test Y'])

    def test_excel_test_schedule_takes_priority_over_db(self):
        """When both DB and Test Schedule exist, Test Schedule wins for wf_names."""
        # DB has definitions
        db_test_defs = {'1': ['Test A']}
        db_wf_defs = {'1': 'DB WF Name'}
        _populate_db_test_definitions(self.conn, db_test_defs)
        _populate_db_wf_definitions(self.conn, db_wf_defs)

        # Create workbook WITH Test Summary and Test Schedule sheets
        wb = Workbook()
        wb.remove(wb.active)
        _make_minimal_wf_sheet(wb, '1', ['Drop20'], [('R3', 'ER3-1-1', 'SN001')])
        _add_test_summary_sheet(wb, {'1': ['Test A']})
        _add_test_schedule_sheet(wb, {'1': 'Excel WF Name'})
        self.excel_path = _save_workbook_to_temp(wb)
        wb.close()

        result = engine.parse_daily_report(self.excel_path, conn=self.conn)

        # Test Schedule sheet should take priority
        self.assertEqual(result.wf_names['1'], 'Excel WF Name')


class DefinitionPriorityNoSourceErrorTests(unittest.TestCase):
    """Requirement 7.4: no DB definitions, no Test Summary sheet → raises error."""

    def setUp(self):
        self.conn, self.db_path = _temp_db()
        self.excel_path = None

    def tearDown(self):
        self.conn.close()
        os.remove(self.db_path)
        if self.excel_path and os.path.exists(self.excel_path):
            os.remove(self.excel_path)

    def test_raises_error_when_no_definitions_available(self):
        """When no Test Summary sheet AND no DB definitions → raises ValueError."""
        # DB is empty (no definitions)

        # Create workbook WITHOUT Test Summary sheet
        wb = Workbook()
        wb.remove(wb.active)
        _make_minimal_wf_sheet(wb, '1', ['Drop20'], [('R3', 'ER3-1-1', 'SN001')])
        self.excel_path = _save_workbook_to_temp(wb)
        wb.close()

        with self.assertRaises(ValueError) as ctx:
            engine.parse_daily_report(self.excel_path, conn=self.conn)

        self.assertIn('请先上传 Base 文件', str(ctx.exception))


class DefinitionPriorityLegacyExcelTests(unittest.TestCase):
    """Requirement 7.5: no DB definitions, Test Summary sheet exists → uses Excel (legacy)."""

    def setUp(self):
        self.conn, self.db_path = _temp_db()
        self.excel_path = None

    def tearDown(self):
        self.conn.close()
        os.remove(self.db_path)
        if self.excel_path and os.path.exists(self.excel_path):
            os.remove(self.excel_path)

    def test_uses_excel_test_summary_when_no_db_definitions(self):
        """Legacy behavior: no DB definitions but Test Summary exists → uses Excel."""
        # DB is empty (no definitions)

        # Create workbook WITH Test Summary sheet
        wb = Workbook()
        wb.remove(wb.active)
        _make_minimal_wf_sheet(wb, '1', ['Drop20', 'Drop40'], [('R3', 'ER3-1-1', 'SN001')])
        _add_test_summary_sheet(wb, {'1': ['Legacy Test A', 'Legacy Test B']})
        self.excel_path = _save_workbook_to_temp(wb)
        wb.close()

        result = engine.parse_daily_report(self.excel_path, conn=self.conn)

        self.assertEqual(result.ts_test_names['1'], ['Legacy Test A', 'Legacy Test B'])

    def test_uses_excel_test_schedule_when_no_db_definitions(self):
        """Legacy behavior: no DB definitions but Test Schedule exists → uses Excel."""
        # DB is empty (no definitions)

        # Create workbook WITH Test Summary and Test Schedule
        wb = Workbook()
        wb.remove(wb.active)
        _make_minimal_wf_sheet(wb, '1', ['Drop20'], [('R3', 'ER3-1-1', 'SN001')])
        _add_test_summary_sheet(wb, {'1': ['Legacy Test']})
        _add_test_schedule_sheet(wb, {'1': 'Legacy WF Name'})
        self.excel_path = _save_workbook_to_temp(wb)
        wb.close()

        result = engine.parse_daily_report(self.excel_path, conn=self.conn)

        self.assertEqual(result.wf_names['1'], 'Legacy WF Name')


if __name__ == '__main__':
    unittest.main()
