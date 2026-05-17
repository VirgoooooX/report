"""Unit tests for generate_daily_report() orchestrator and engine.py round-trip."""

import io
import re
import pytest
from unittest.mock import patch, MagicMock
from openpyxl import load_workbook


class TestGenerateDailyReport:
    """Tests for the generate_daily_report orchestrator function."""

    def _make_csv_file(self, filename, content):
        """Create a mock file-like object mimicking Flask FileStorage."""
        mock_file = MagicMock()
        mock_file.filename = filename
        if isinstance(content, str):
            content = content.encode('utf-8')
        mock_file.read.return_value = content
        return mock_file

    def _make_csv_content(self, item_type="ISB", sn="SN001", cp="T0", status="PASS",
                          end_time="2026-05-15 10:00:00"):
        """Generate minimal CSV content with a valid header and one data row."""
        header = "Site,Product,SerialNumber,Special Build Name,REL Event,Test Pass/Fail Status,EndTime,List of Failing Tests,Station ID,Version,Col10,Col11,Col12"
        data = f"Site1,Prod1,{sn},Build1,{cp},{status},{end_time},,Station1,1.0.0,,,,"
        return f"{header}\n{data}\n"

    def _mock_sn_mapping(self):
        """Return a mock SN mapping result."""
        return {
            'sn_mapping': {
                'SN001': {'config': 'R1FNF', 'wf_id': '1', 'unit_number': 'ER1-1-1'},
                'SN002': {'config': 'R1FNF', 'wf_id': '1', 'unit_number': 'ER1-1-2'},
                'SN003': {'config': 'R2CNM', 'wf_id': '2', 'unit_number': 'ER2-1-1'},
            },
            'config_quantities': {'R1FNF': 2, 'R2CNM': 1},
            'sn_count': 3,
        }

    def _mock_sn_lookup_dicts(self):
        """Return mock SN lookup dicts."""
        sn_to_wf = {'SN001': '1', 'SN002': '1', 'SN003': '2'}
        sn_to_unit = {'SN001': 'ER1-1-1', 'SN002': 'ER1-1-2', 'SN003': 'ER2-1-1'}
        sn_to_config = {'SN001': 'R1FNF', 'SN002': 'R1FNF', 'SN003': 'R2CNM'}
        return sn_to_wf, sn_to_unit, sn_to_config

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_raises_value_error_when_no_sn_mapping(self, mock_sn_mapping, mock_lookup, mock_db):
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = None

        with pytest.raises(ValueError, match="请先上传 Base 文件"):
            generate_daily_report([])

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_raises_value_error_when_no_valid_data(self, mock_sn_mapping, mock_lookup, mock_db):
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = self._mock_sn_mapping()
        mock_lookup.return_value = self._mock_sn_lookup_dicts()

        # Mock DB connection
        mock_conn = MagicMock()
        mock_db.get_conn.return_value = mock_conn
        mock_db.get_current_cp_definitions.return_value = {'1': [{'cp_idx': 0, 'cp_name': 'T0', 'test_idx': 0, 'check_items': []}]}
        mock_db.get_current_wf_definitions.return_value = {'1': 'ISB'}
        mock_db.get_current_test_definitions.return_value = {'1': ['ISB']}

        # CSV with unrecognized filename
        csv_file = self._make_csv_file("unknown_file.csv", "no data here\n")

        with pytest.raises(ValueError, match="未找到有效数据"):
            generate_daily_report([csv_file])

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_returns_excel_bytes_and_filename(self, mock_sn_mapping, mock_lookup, mock_db):
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = self._mock_sn_mapping()
        mock_lookup.return_value = self._mock_sn_lookup_dicts()

        # Mock DB connection
        mock_conn = MagicMock()
        mock_db.get_conn.return_value = mock_conn
        mock_db.get_current_cp_definitions.return_value = {
            '1': [
                {'cp_idx': 0, 'cp_name': 'T0', 'test_idx': 0, 'check_items': []},
                {'cp_idx': 1, 'cp_name': 'TC_100CYCLES', 'test_idx': 0, 'check_items': []},
            ]
        }
        mock_db.get_current_wf_definitions.return_value = {'1': 'ISB'}
        mock_db.get_current_test_definitions.return_value = {'1': ['ISB']}

        # Create CSV with valid data
        content = self._make_csv_content(item_type="ISB", sn="SN001", cp="T0",
                                         status="PASS", end_time="2026-05-15 10:00:00")
        csv_file = self._make_csv_file("ORT-ISB_data.csv", content)

        excel_bytes, filename, summary = generate_daily_report([csv_file])

        # Verify return types
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0
        assert filename == "M60 EVT Rel Daily Report_20260515.xlsx"
        assert isinstance(summary, dict)

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_filename_uses_latest_end_time(self, mock_sn_mapping, mock_lookup, mock_db):
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = self._mock_sn_mapping()
        mock_lookup.return_value = self._mock_sn_lookup_dicts()

        mock_conn = MagicMock()
        mock_db.get_conn.return_value = mock_conn
        mock_db.get_current_cp_definitions.return_value = {
            '1': [{'cp_idx': 0, 'cp_name': 'T0', 'test_idx': 0, 'check_items': []}]
        }
        mock_db.get_current_wf_definitions.return_value = {'1': 'ISB'}
        mock_db.get_current_test_definitions.return_value = {'1': ['ISB']}

        # Two CSV files with different dates
        content1 = self._make_csv_content(sn="SN001", cp="T0", status="PASS",
                                          end_time="2026-05-10 08:00:00")
        content2 = self._make_csv_content(sn="SN002", cp="T0", status="PASS",
                                          end_time="2026-05-20 16:30:00")
        csv_file1 = self._make_csv_file("ORT-ISB_old.csv", content1)
        csv_file2 = self._make_csv_file("ORT-ISB_new.csv", content2)

        _, filename, _ = generate_daily_report([csv_file1, csv_file2])

        # Should use the latest date (May 20)
        assert filename == "M60 EVT Rel Daily Report_20260520.xlsx"

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_summary_contains_statistics(self, mock_sn_mapping, mock_lookup, mock_db):
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = self._mock_sn_mapping()
        mock_lookup.return_value = self._mock_sn_lookup_dicts()

        mock_conn = MagicMock()
        mock_db.get_conn.return_value = mock_conn
        mock_db.get_current_cp_definitions.return_value = {
            '1': [{'cp_idx': 0, 'cp_name': 'T0', 'test_idx': 0, 'check_items': []}]
        }
        mock_db.get_current_wf_definitions.return_value = {'1': 'ISB'}
        mock_db.get_current_test_definitions.return_value = {'1': ['ISB']}

        content = self._make_csv_content(sn="SN001", cp="T0", status="PASS",
                                         end_time="2026-05-15 10:00:00")
        csv_file = self._make_csv_file("ORT-ISB_data.csv", content)

        _, _, summary = generate_daily_report([csv_file])

        assert 'file_count' in summary
        assert 'record_count' in summary
        assert 'valid_sn_count' in summary
        assert 'warnings' in summary
        assert summary['file_count'] == 1
        assert summary['valid_sn_count'] >= 1

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_generated_excel_has_wf_sheets(self, mock_sn_mapping, mock_lookup, mock_db):
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = self._mock_sn_mapping()
        mock_lookup.return_value = self._mock_sn_lookup_dicts()

        mock_conn = MagicMock()
        mock_db.get_conn.return_value = mock_conn
        mock_db.get_current_cp_definitions.return_value = {
            '1': [
                {'cp_idx': 0, 'cp_name': 'T0', 'test_idx': 0, 'check_items': []},
            ],
            '2': [
                {'cp_idx': 0, 'cp_name': 'T0', 'test_idx': 0, 'check_items': []},
            ],
        }
        mock_db.get_current_wf_definitions.return_value = {'1': 'ISB', '2': 'FACT + BT-OTA'}
        mock_db.get_current_test_definitions.return_value = {'1': ['ISB'], '2': ['FACT', 'BT-OTA']}

        content = self._make_csv_content(sn="SN001", cp="T0", status="PASS",
                                         end_time="2026-05-15 10:00:00")
        csv_file = self._make_csv_file("ORT-ISB_data.csv", content)

        excel_bytes, _, _ = generate_daily_report([csv_file])

        # Load the generated Excel and verify sheets
        wb = load_workbook(io.BytesIO(excel_bytes))
        sheet_names = wb.sheetnames

        # Should have WF sheets (at least WF1 since SN001 is in WF1)
        assert any("WF1" in name for name in sheet_names)

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_no_test_summary_or_schedule_sheets(self, mock_sn_mapping, mock_lookup, mock_db):
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = self._mock_sn_mapping()
        mock_lookup.return_value = self._mock_sn_lookup_dicts()

        mock_conn = MagicMock()
        mock_db.get_conn.return_value = mock_conn
        mock_db.get_current_cp_definitions.return_value = {
            '1': [{'cp_idx': 0, 'cp_name': 'T0', 'test_idx': 0, 'check_items': []}]
        }
        mock_db.get_current_wf_definitions.return_value = {'1': 'ISB'}
        mock_db.get_current_test_definitions.return_value = {'1': ['ISB']}

        content = self._make_csv_content(sn="SN001", cp="T0", status="PASS",
                                         end_time="2026-05-15 10:00:00")
        csv_file = self._make_csv_file("ORT-ISB_data.csv", content)

        excel_bytes, _, _ = generate_daily_report([csv_file])

        wb = load_workbook(io.BytesIO(excel_bytes))
        sheet_names = wb.sheetnames

        # Should NOT have Test Summary or Test Schedule sheets
        assert "Test Summary" not in sheet_names
        assert "Test Schedule" not in sheet_names

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_warnings_for_unrecognized_files(self, mock_sn_mapping, mock_lookup, mock_db):
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = self._mock_sn_mapping()
        mock_lookup.return_value = self._mock_sn_lookup_dicts()

        mock_conn = MagicMock()
        mock_db.get_conn.return_value = mock_conn
        mock_db.get_current_cp_definitions.return_value = {
            '1': [{'cp_idx': 0, 'cp_name': 'T0', 'test_idx': 0, 'check_items': []}]
        }
        mock_db.get_current_wf_definitions.return_value = {'1': 'ISB'}
        mock_db.get_current_test_definitions.return_value = {'1': ['ISB']}

        # One valid file and one unrecognized
        valid_content = self._make_csv_content(sn="SN001", cp="T0", status="PASS",
                                               end_time="2026-05-15 10:00:00")
        valid_file = self._make_csv_file("ORT-ISB_data.csv", valid_content)
        unknown_file = self._make_csv_file("random_notes.csv", "some,random,data\n")

        _, _, summary = generate_daily_report([valid_file, unknown_file])

        assert len(summary['warnings']) >= 1
        assert any("Unrecognized" in w for w in summary['warnings'])

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_multiple_csv_files_combined(self, mock_sn_mapping, mock_lookup, mock_db):
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = self._mock_sn_mapping()
        mock_lookup.return_value = self._mock_sn_lookup_dicts()

        mock_conn = MagicMock()
        mock_db.get_conn.return_value = mock_conn
        mock_db.get_current_cp_definitions.return_value = {
            '1': [{'cp_idx': 0, 'cp_name': 'T0', 'test_idx': 0, 'check_items': []}]
        }
        mock_db.get_current_wf_definitions.return_value = {'1': 'ISB + FACT'}
        mock_db.get_current_test_definitions.return_value = {'1': ['ISB', 'FACT']}

        # ISB file
        isb_content = self._make_csv_content(sn="SN001", cp="T0", status="PASS",
                                             end_time="2026-05-15 10:00:00")
        isb_file = self._make_csv_file("ORT-ISB_data.csv", isb_content)

        # FACT file
        fact_content = self._make_csv_content(sn="SN001", cp="T0", status="FAIL",
                                              end_time="2026-05-15 11:00:00")
        fact_file = self._make_csv_file("ORT-FACT_data.csv", fact_content)

        excel_bytes, _, summary = generate_daily_report([isb_file, fact_file])

        assert summary['file_count'] == 2
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0


class TestExcelGenerationIntegration:
    """Integration tests verifying generated Excel is compatible with engine.py parsing.

    Validates Requirements 5.1-5.7: sheet naming, column structure, color coding,
    and round-trip compatibility with engine.py.
    """

    def _make_csv_file(self, filename, content):
        """Create a mock file-like object mimicking Flask FileStorage."""
        mock_file = MagicMock()
        mock_file.filename = filename
        if isinstance(content, str):
            content = content.encode('utf-8')
        mock_file.read.return_value = content
        return mock_file

    def _make_csv_content(self, sn="SN001", cp="T0", status="PASS",
                          end_time="2026-05-15 10:00:00"):
        """Generate minimal CSV content with a valid header and one data row."""
        header = "Site,Product,SerialNumber,Special Build Name,REL Event,Test Pass/Fail Status,EndTime,List of Failing Tests,Station ID,Version,Col10,Col11,Col12"
        data = f"Site1,Prod1,{sn},Build1,{cp},{status},{end_time},,Station1,1.0.0,,,,"
        return f"{header}\n{data}\n"

    def _generate_test_excel(self, mock_sn_mapping, mock_lookup, mock_db,
                             cp_definitions=None, wf_definitions=None,
                             csv_files=None):
        """Helper to generate an Excel workbook with mocked DB and return bytes."""
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = {
            'sn_mapping': {
                'SN001': {'config': 'R1FNF', 'wf_id': '1', 'unit_number': 'ER1-1-1'},
                'SN002': {'config': 'R1FNF', 'wf_id': '1', 'unit_number': 'ER1-1-2'},
                'SN003': {'config': 'R2CNM', 'wf_id': '2', 'unit_number': 'ER2-1-1'},
            },
            'config_quantities': {'R1FNF': 2, 'R2CNM': 1},
            'sn_count': 3,
        }
        mock_lookup.return_value = (
            {'SN001': '1', 'SN002': '1', 'SN003': '2'},
            {'SN001': 'ER1-1-1', 'SN002': 'ER1-1-2', 'SN003': 'ER2-1-1'},
            {'SN001': 'R1FNF', 'SN002': 'R1FNF', 'SN003': 'R2CNM'},
        )

        mock_conn = MagicMock()
        mock_db.get_conn.return_value = mock_conn

        if cp_definitions is None:
            cp_definitions = {
                '1': [
                    {'cp_idx': 0, 'cp_name': 'T0', 'test_idx': 0, 'check_items': []},
                    {'cp_idx': 1, 'cp_name': 'TC_100CYCLES', 'test_idx': 0, 'check_items': []},
                ],
                '2': [
                    {'cp_idx': 0, 'cp_name': 'T0', 'test_idx': 0, 'check_items': []},
                ],
            }
        if wf_definitions is None:
            wf_definitions = {'1': 'ISB + FACT', '2': 'BT-OTA'}

        mock_db.get_current_cp_definitions.return_value = cp_definitions
        mock_db.get_current_wf_definitions.return_value = wf_definitions
        mock_db.get_current_test_definitions.return_value = {
            k: v.split(' + ') for k, v in wf_definitions.items()
        }

        if csv_files is None:
            csv_files = [
                self._make_csv_file("ORT-ISB_data.csv",
                                    self._make_csv_content(sn="SN001", cp="T0", status="PASS")),
                self._make_csv_file("ORT-FACT_data.csv",
                                    self._make_csv_content(sn="SN001", cp="TC_100CYCLES", status="FAIL",
                                                          end_time="2026-05-15 12:00:00")),
            ]

        excel_bytes, filename, summary = generate_daily_report(csv_files)
        return excel_bytes, filename, summary

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_engine_wf_num_parses_all_generated_sheet_names(self, mock_sn_mapping, mock_lookup, mock_db):
        """engine.py's wf_num() must extract WF numbers from all generated sheet names."""
        from engine import wf_num

        excel_bytes, _, _ = self._generate_test_excel(mock_sn_mapping, mock_lookup, mock_db)
        wb = load_workbook(io.BytesIO(excel_bytes))

        parsed_wf_nums = []
        for name in wb.sheetnames:
            wfn = wf_num(name)
            assert wfn is not None, f"engine.py wf_num() failed to parse sheet name: '{name}'"
            parsed_wf_nums.append(wfn)

        # Should have extracted WF numbers for all sheets
        assert '1' in parsed_wf_nums
        assert '2' in parsed_wf_nums
        wb.close()

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_engine_extract_cp_structure_reads_generated_sheets(self, mock_sn_mapping, mock_lookup, mock_db):
        """engine.py's extract_cp_structure() must read CP names from generated sheets."""
        from engine import extract_cp_structure

        excel_bytes, _, _ = self._generate_test_excel(mock_sn_mapping, mock_lookup, mock_db)
        wb = load_workbook(io.BytesIO(excel_bytes))

        # Find the WF1 sheet
        wf1_sheet = None
        for name in wb.sheetnames:
            if 'WF1' in name:
                wf1_sheet = wb[name]
                break

        assert wf1_sheet is not None, "WF1 sheet not found in generated Excel"

        # engine.py finds header_row by looking for Config/Unit # pattern
        # Our generator puts these at row 2, so CP headers are also in row 2
        cps = extract_cp_structure(wf1_sheet, header_row=2, cp_range_start=6)

        # Should find at least one CP (TC_100CYCLES — T0 is excluded by is_cp_header_label)
        assert len(cps) >= 1
        cp_names = [cp['cp_name'] for cp in cps]
        assert 'TC_100CYCLES' in cp_names

        # Each CP should have check items extracted from row 3
        for cp in cps:
            assert 'check_items' in cp
            # Check items should be from the known set
            for item in cp['check_items']:
                assert item in {'Cosmetic', 'ISB', 'FACT', 'BT-OTA', 'Touch-CAL-Post', 'Charging'}

        wb.close()

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_engine_get_failure_type_reads_color_coding(self, mock_sn_mapping, mock_lookup, mock_db):
        """engine.py's get_failure_type() must correctly identify colors from generated cells."""
        from engine import get_failure_type

        # Generate Excel with mixed PASS/FAIL results
        csv_files = [
            self._make_csv_file("ORT-ISB_data.csv",
                                self._make_csv_content(sn="SN001", cp="TC_100CYCLES",
                                                      status="PASS", end_time="2026-05-15 10:00:00")),
            self._make_csv_file("ORT-FACT_data.csv",
                                self._make_csv_content(sn="SN001", cp="TC_100CYCLES",
                                                      status="FAIL", end_time="2026-05-15 11:00:00")),
        ]

        excel_bytes, _, _ = self._generate_test_excel(
            mock_sn_mapping, mock_lookup, mock_db, csv_files=csv_files)

        # Load without data_only to preserve formatting
        wb = load_workbook(io.BytesIO(excel_bytes))

        # Find WF1 sheet
        wf1_sheet = None
        for name in wb.sheetnames:
            if 'WF1' in name:
                wf1_sheet = wb[name]
                break

        assert wf1_sheet is not None

        # Data starts at row 3. Find cells with PASS and FAIL values.
        pass_cells = []
        fail_cells = []
        for row in range(3, wf1_sheet.max_row + 1):
            for col in range(6, wf1_sheet.max_column + 1):
                cell = wf1_sheet.cell(row=row, column=col)
                if cell.value == "PASS":
                    pass_cells.append(cell)
                elif cell.value == "FAIL":
                    fail_cells.append(cell)

        # Verify PASS cells have no failure type (green is not a failure)
        for cell in pass_cells:
            ft = get_failure_type(cell)
            assert ft is None, f"PASS cell at ({cell.row},{cell.column}) should have no failure_type"

        # Verify FAIL cells have a recognized failure type (spec or strife)
        for cell in fail_cells:
            ft = get_failure_type(cell)
            assert ft in ('spec', 'strife'), (
                f"FAIL cell at ({cell.row},{cell.column}) should have 'spec' or 'strife' failure_type, got {ft}"
            )

        wb.close()

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_generated_excel_sheet_names_match_wf_pattern(self, mock_sn_mapping, mock_lookup, mock_db):
        """All generated sheet names must match the 'Sys WF{n}_{name}' pattern."""
        excel_bytes, _, _ = self._generate_test_excel(mock_sn_mapping, mock_lookup, mock_db)
        wb = load_workbook(io.BytesIO(excel_bytes))

        for name in wb.sheetnames:
            # Must match the pattern Sys WF{number}_{name}
            assert re.match(r'Sys WF\d+\.?\d*_.+', name), (
                f"Sheet name '{name}' does not match expected 'Sys WF{{n}}_{{name}}' pattern"
            )

        wb.close()

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_generated_excel_column_structure(self, mock_sn_mapping, mock_lookup, mock_db):
        """Generated Excel must have correct fixed columns and CP group structure."""
        excel_bytes, _, _ = self._generate_test_excel(mock_sn_mapping, mock_lookup, mock_db)
        wb = load_workbook(io.BytesIO(excel_bytes))

        for name in wb.sheetnames:
            ws = wb[name]
            # Row 1 should have Report Date in A1
            assert "Report Date" in str(ws.cell(row=1, column=1).value or '')

            # Row 2 should have fixed headers in columns 1-5
            assert ws.cell(row=2, column=1).value == "%"
            assert ws.cell(row=2, column=2).value == "Completion%"
            assert ws.cell(row=2, column=3).value == "Config"
            assert ws.cell(row=2, column=4).value == "Unit #"
            assert ws.cell(row=2, column=5).value == "S/N"

            # Check items should appear in row 3 starting from column 6
            check_items_found = []
            for col in range(6, ws.max_column + 1):
                val = ws.cell(row=3, column=col).value
                if val and val in {'Cosmetic', 'ISB', 'FACT', 'BT-OTA', 'Touch-CAL-Post', 'Charging'}:
                    check_items_found.append(val)

            # Should have at least one set of 6 check items
            assert len(check_items_found) >= 6, (
                f"Sheet '{name}' should have at least 6 check item sub-headers, found {len(check_items_found)}"
            )

        wb.close()

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_generated_excel_no_skip_sheets(self, mock_sn_mapping, mock_lookup, mock_db):
        """Generated Excel must not contain sheets that engine.py would skip."""
        excel_bytes, _, _ = self._generate_test_excel(mock_sn_mapping, mock_lookup, mock_db)
        wb = load_workbook(io.BytesIO(excel_bytes))

        skip_sheets = {'Sample Delivery Tracker', 'Test Schedule', 'Test Summary', 'T0 Summary'}
        for name in wb.sheetnames:
            assert name not in skip_sheets, (
                f"Generated Excel should not contain '{name}' sheet"
            )

        wb.close()



class TestGenerateDailyReportBoundaryFiltering:
    """Batch B step 3.5 — generated Daily Report Excel must include REL_T0
    columns (carries real check-item data) but exclude other schedule
    boundaries (REL_TFINAL/End/TFinal/T0).

    See docs/plans/2026-05-17-rel-t0-daily-report-second-cut.md.
    """

    def _make_csv_file(self, filename, content):
        mock_file = MagicMock()
        mock_file.filename = filename
        if isinstance(content, str):
            content = content.encode('utf-8')
        mock_file.read.return_value = content
        return mock_file

    def _make_csv_content(self, sn, cp, status="PASS",
                          end_time="2026-05-15 10:00:00"):
        header = (
            "Site,Product,SerialNumber,Special Build Name,REL Event,"
            "Test Pass/Fail Status,EndTime,List of Failing Tests,Station ID,"
            "Version,Col10,Col11,Col12"
        )
        data = (
            f"Site1,Prod1,{sn},Build1,{cp},{status},{end_time},,"
            "Station1,1.0.0,,,,"
        )
        return f"{header}\n{data}\n"

    def _mock_sn_mapping(self):
        return {
            'sn_mapping': {
                'SN001': {'config': 'R1FNF', 'wf_id': '1', 'unit_number': 'ER1-1-1'},
            },
            'config_quantities': {'R1FNF': 1},
            'sn_count': 1,
        }

    def _mock_sn_lookup_dicts(self):
        return (
            {'SN001': '1'},
            {'SN001': 'ER1-1-1'},
            {'SN001': 'R1FNF'},
        )

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_rel_t0_column_appears_in_generated_excel(
        self, mock_sn_mapping, mock_lookup, mock_db,
    ):
        """Base CPs include REL_T0 (boundary) + CP_A. Generator keeps both."""
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = self._mock_sn_mapping()
        mock_lookup.return_value = self._mock_sn_lookup_dicts()

        mock_conn = MagicMock()
        mock_db.get_conn.return_value = mock_conn
        mock_db.get_current_cp_definitions.return_value = {
            '1': [
                {'cp_idx': 0, 'cp_name': 'REL_T0', 'test_idx': 0,
                 'check_items': [], 'is_boundary': 1},
                {'cp_idx': 1, 'cp_name': 'CP_A', 'test_idx': 0,
                 'check_items': [], 'is_boundary': 0},
            ],
        }
        mock_db.get_current_wf_definitions.return_value = {'1': 'ISB'}
        mock_db.get_current_test_definitions.return_value = {'1': ['ISB']}

        content = self._make_csv_content(sn='SN001', cp='REL_T0', status='PASS')
        csv_file = self._make_csv_file('ORT-ISB.csv', content)
        excel_bytes, _, _ = generate_daily_report([csv_file])

        wb = load_workbook(io.BytesIO(excel_bytes))
        wf_sheet = next(name for name in wb.sheetnames if 'WF1' in name)
        ws = wb[wf_sheet]
        # Row 1 is the CP header row. Collect all CP labels.
        cp_labels_in_row = {
            str(c.value).strip() for c in ws[2] if c.value
        }
        assert 'REL_T0' in cp_labels_in_row
        assert 'CP_A' in cp_labels_in_row
        wb.close()

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_rel_tfinal_column_excluded_from_generated_excel(
        self, mock_sn_mapping, mock_lookup, mock_db,
    ):
        """REL_TFINAL is a boundary but NOT in DAILY_RESULT_BOUNDARY_LABELS;
        it must NOT render as a column even though it's persisted."""
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = self._mock_sn_mapping()
        mock_lookup.return_value = self._mock_sn_lookup_dicts()

        mock_conn = MagicMock()
        mock_db.get_conn.return_value = mock_conn
        mock_db.get_current_cp_definitions.return_value = {
            '1': [
                {'cp_idx': 0, 'cp_name': 'REL_T0', 'test_idx': 0,
                 'check_items': [], 'is_boundary': 1},
                {'cp_idx': 1, 'cp_name': 'CP_A', 'test_idx': 0,
                 'check_items': [], 'is_boundary': 0},
                {'cp_idx': 2, 'cp_name': 'REL_TFINAL', 'test_idx': 0,
                 'check_items': [], 'is_boundary': 1},
            ],
        }
        mock_db.get_current_wf_definitions.return_value = {'1': 'ISB'}
        mock_db.get_current_test_definitions.return_value = {'1': ['ISB']}

        content = self._make_csv_content(sn='SN001', cp='REL_T0', status='PASS')
        csv_file = self._make_csv_file('ORT-ISB.csv', content)
        excel_bytes, _, _ = generate_daily_report([csv_file])

        wb = load_workbook(io.BytesIO(excel_bytes))
        wf_sheet = next(name for name in wb.sheetnames if 'WF1' in name)
        ws = wb[wf_sheet]
        cp_labels_in_row = {
            str(c.value).strip() for c in ws[2] if c.value
        }
        assert 'REL_T0' in cp_labels_in_row
        assert 'CP_A' in cp_labels_in_row
        assert 'REL_TFINAL' not in cp_labels_in_row
        wb.close()

    @patch('checkitem_generator.db')
    @patch('checkitem_generator.get_sn_lookup_dicts')
    @patch('checkitem_generator.get_sn_mapping_from_db')
    def test_other_boundary_labels_excluded(
        self, mock_sn_mapping, mock_lookup, mock_db,
    ):
        """T0 / End / TFinal also stay out of the generated Excel."""
        from checkitem_generator import generate_daily_report

        mock_sn_mapping.return_value = self._mock_sn_mapping()
        mock_lookup.return_value = self._mock_sn_lookup_dicts()

        mock_conn = MagicMock()
        mock_db.get_conn.return_value = mock_conn
        mock_db.get_current_cp_definitions.return_value = {
            '1': [
                {'cp_idx': 0, 'cp_name': 'T0', 'test_idx': 0,
                 'check_items': [], 'is_boundary': 1},
                {'cp_idx': 1, 'cp_name': 'CP_A', 'test_idx': 0,
                 'check_items': [], 'is_boundary': 0},
                {'cp_idx': 2, 'cp_name': 'End', 'test_idx': 0,
                 'check_items': [], 'is_boundary': 1},
                {'cp_idx': 3, 'cp_name': 'TFinal', 'test_idx': 0,
                 'check_items': [], 'is_boundary': 1},
            ],
        }
        mock_db.get_current_wf_definitions.return_value = {'1': 'ISB'}
        mock_db.get_current_test_definitions.return_value = {'1': ['ISB']}

        content = self._make_csv_content(sn='SN001', cp='CP_A', status='PASS')
        csv_file = self._make_csv_file('ORT-ISB.csv', content)
        excel_bytes, _, _ = generate_daily_report([csv_file])

        wb = load_workbook(io.BytesIO(excel_bytes))
        wf_sheet = next(name for name in wb.sheetnames if 'WF1' in name)
        ws = wb[wf_sheet]
        cp_labels_in_row = {
            str(c.value).strip() for c in ws[2] if c.value
        }
        assert 'CP_A' in cp_labels_in_row
        assert 'T0' not in cp_labels_in_row
        assert 'End' not in cp_labels_in_row
        assert 'TFinal' not in cp_labels_in_row
        wb.close()
