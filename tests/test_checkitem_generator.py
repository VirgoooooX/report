"""Unit tests for checkitem_generator — file type identification."""

import pytest
from checkitem_generator import identify_csv_type


class TestIdentifyCsvType:
    """Tests for identify_csv_type function."""

    def test_bt_ota_keyword(self):
        filename = "R134-Export-ID-215213084484635-2026-04-01T00_00_00-2026-05-15T15_18_44-B529-BT-OTA-0.0.51.csv"
        assert identify_csv_type(filename) == "BT-OTA"

    def test_charging_keyword(self):
        filename = "R134-Export-ID-215213084484635-2026-04-01T00_00_00-2026-05-15T15_18_44-B529-ORT-Charging-1.1.7-rel_wo_fw_check-v1.1.0.csv"
        assert identify_csv_type(filename) == "Charging"

    def test_fact_keyword(self):
        filename = "R134-Export-ID-215213084484635-2026-04-01T00_00_00-2026-05-15T15_18_44-B529-ORT-FACT-Versions-2.csv"
        assert identify_csv_type(filename) == "FACT"

    def test_isb_keyword(self):
        filename = "R134-Export-ID-215213084484635-2026-04-01T00_00_00-2026-05-15T15_18_44-B529-ORT-ISB-1.1.3-rel.csv"
        assert identify_csv_type(filename) == "ISB"

    def test_touch_cal_post_keyword(self):
        filename = "R134-Export-ID-215213084484635-2026-04-01T00_00_00-2026-05-15T15_18_44-B529-Touch-Cal-Post-Versions-5.csv"
        assert identify_csv_type(filename) == "Touch-CAL-Post"

    def test_cosmetic_keyword(self):
        filename = "R134 Rel-Export-ID-215213084484690-2026-04-01T00_00_00-2026-05-15T15_18_44-B529-RLCS_REL-COSMETIC-Versions-2.csv"
        assert identify_csv_type(filename) == "Cosmetic"

    def test_unrecognized_file_returns_none(self):
        assert identify_csv_type("random_file.csv") is None
        assert identify_csv_type("report_2026.xlsx") is None
        assert identify_csv_type("") is None

    def test_case_insensitive_matching(self):
        assert identify_csv_type("some-bt-ota-file.csv") == "BT-OTA"
        assert identify_csv_type("CHARGING-test.csv") == "Charging"
        assert identify_csv_type("fact-results.csv") == "FACT"
        assert identify_csv_type("isb-data.csv") == "ISB"
        assert identify_csv_type("touch-cal-post-v2.csv") == "Touch-CAL-Post"
        assert identify_csv_type("cosmetic-check.csv") == "Cosmetic"

    def test_r2_variant_filenames(self):
        """Test with R2 variant filenames from the actual data folder."""
        assert identify_csv_type("R2 0515-Export-ID-215213084484212-2026-05-11T00_00_00-2026-05-15T23_59_59-B529-BT-OTA-0.0.51.csv") == "BT-OTA"
        assert identify_csv_type("R2 0515-Export-ID-215213084484212-2026-05-11T00_00_00-2026-05-15T23_59_59-B529-ORT-Charging-1.1.7-rel_wo_fw_check-v1.1.0.csv") == "Charging"
        assert identify_csv_type("R2 0515-Export-ID-215213084484212-2026-05-11T00_00_00-2026-05-15T23_59_59-B529-ORT-FACT-3.3.7.csv") == "FACT"
        assert identify_csv_type("R2 0515-Export-ID-215213084484212-2026-05-11T00_00_00-2026-05-15T23_59_59-B529-ORT-ISB-1.1.3-rel.csv") == "ISB"
        assert identify_csv_type("R2 0515-Export-ID-215213084484212-2026-05-11T00_00_00-2026-05-15T23_59_59-B529-Touch-Cal-Post-1.0.29-BTEnable-Rel_DVT.csv") == "Touch-CAL-Post"
        assert identify_csv_type("R2 Rel 0515-Export-ID-215213084484261-2026-05-11T00_00_00-2026-05-15T23_59_59-B529-RLCS_REL-COSMETIC-3.2.2511.csv") == "Cosmetic"


class TestParseCsvFile:
    """Tests for parse_csv_file function."""

    def _make_csv(self, header_row, data_rows, metadata_prefix="ORT-ISB,,,,,,,,,,,,,Parametric"):
        """Helper to build CSV content with optional metadata rows."""
        lines = []
        if metadata_prefix:
            lines.append(metadata_prefix)
        lines.append(",".join(header_row))
        # Add standard metadata rows between header and data
        lines.append("Display Name ----->,,,,,,,,,,,,,")
        lines.append("PDCA Priority ----->,,,,,,,,,,,,,0,0,0")
        lines.append("Upper Limit ----->,,,,,,,,,,,,,NA,NA,NA")
        lines.append("Lower Limit ----->,,,,,,,,,,,,,NA,NA,NA")
        lines.append("Measurement Unit ----->,,,,,,,,,,,,,NA,NA,NA")
        for row in data_rows:
            lines.append(",".join(str(v) for v in row))
        return "\n".join(lines)

    def test_returns_empty_list_when_no_serial_number_header(self):
        from checkitem_generator import parse_csv_file

        content = "col1,col2,col3\na,b,c\n"
        result = parse_csv_file(content, "FACT")
        assert result == []

    def test_detects_header_row_with_serial_number(self):
        from checkitem_generator import parse_csv_file

        header = ["Site", "Product", "SerialNumber", "Special Build Name",
                  "Special Build Description", "Unit Number", "Station ID",
                  "Test Pass/Fail Status", "StartTime", "EndTime", "Version",
                  "List of Failing Tests", "REL Event", "Param1", "Param2"]
        data = [["JAWX", "B529", "SN001", "B529-B529", "B529-B529_R2CNM", "",
                 "STATION_01", "PASS", "2026-05-12 08:00:00", "2026-05-12 08:01:00",
                 "1.0.0", "", "REL_T0", "1.5", "2.3"]]
        content = self._make_csv(header, data)
        result = parse_csv_file(content, "ISB")

        assert len(result) == 1
        assert result[0]["serial_number"] == "SN001"
        assert result[0]["rel_event"] == "REL_T0"
        assert result[0]["status"] == "PASS"
        assert result[0]["end_time"] == "2026-05-12 08:01:00"
        assert result[0]["station_id"] == "STATION_01"
        assert result[0]["version"] == "1.0.0"
        assert result[0]["item"] == "ISB"
        assert result[0]["test_params"] is None  # PASS record

    def test_extracts_params_for_fail_records(self):
        from checkitem_generator import parse_csv_file

        header = ["Site", "Product", "SerialNumber", "Special Build Name",
                  "Special Build Description", "Unit Number", "Station ID",
                  "Test Pass/Fail Status", "StartTime", "EndTime", "Version",
                  "List of Failing Tests", "REL Event", "ParamA", "ParamB"]
        data = [["JAWX", "B529", "SN002", "B529-B529", "B529-B529_R2CNM", "",
                 "STATION_02", "FAIL", "2026-05-13 10:00:00", "2026-05-13 10:01:00",
                 "2.0.0", "TestX;TestY", "CP_100%", "84.5", "-26.7"]]
        content = self._make_csv(header, data)
        result = parse_csv_file(content, "FACT")

        assert len(result) == 1
        rec = result[0]
        assert rec["status"] == "FAIL"
        assert rec["failing_tests"] == "TestX;TestY"
        assert rec["test_params"] is not None
        assert rec["test_params"]["ParamA"] == 84.5
        assert rec["test_params"]["ParamB"] == -26.7

    def test_pass_records_have_no_test_params(self):
        from checkitem_generator import parse_csv_file

        header = ["Site", "Product", "SerialNumber", "Special Build Name",
                  "Special Build Description", "Unit Number", "Station ID",
                  "Test Pass/Fail Status", "StartTime", "EndTime", "Version",
                  "List of Failing Tests", "REL Event", "ParamA"]
        data = [["JAWX", "B529", "SN003", "B529-B529", "B529-B529_R2CNM", "",
                 "STATION_01", "PASS", "2026-05-12 08:00:00", "2026-05-12 08:01:00",
                 "1.0.0", "", "REL_T0", "1.5"]]
        content = self._make_csv(header, data)
        result = parse_csv_file(content, "ISB")

        assert result[0]["test_params"] is None

    def test_skips_metadata_rows(self):
        from checkitem_generator import parse_csv_file

        header = ["Site", "Product", "SerialNumber", "Special Build Name",
                  "Special Build Description", "Unit Number", "Station ID",
                  "Test Pass/Fail Status", "StartTime", "EndTime", "Version",
                  "List of Failing Tests", "REL Event"]
        data = [["JAWX", "B529", "SN004", "B529-B529", "B529-B529_R2CNM", "",
                 "STATION_01", "PASS", "2026-05-12 08:00:00", "2026-05-12 08:01:00",
                 "1.0.0", "", "REL_T0"]]
        content = self._make_csv(header, data)
        result = parse_csv_file(content, "FACT")

        # Should only have 1 data record, not metadata rows
        assert len(result) == 1
        assert result[0]["serial_number"] == "SN004"

    def test_handles_bytes_input(self):
        from checkitem_generator import parse_csv_file

        header = ["Site", "Product", "SerialNumber", "Special Build Name",
                  "Special Build Description", "Unit Number", "Station ID",
                  "Test Pass/Fail Status", "StartTime", "EndTime", "Version",
                  "List of Failing Tests", "REL Event"]
        data = [["JAWX", "B529", "SN005", "B529-B529", "B529-B529_R2CNM", "",
                 "STATION_01", "PASS", "2026-05-12 08:00:00", "2026-05-12 08:01:00",
                 "1.0.0", "", "REL_T0"]]
        content = self._make_csv(header, data).encode("utf-8")
        result = parse_csv_file(content, "ISB")

        assert len(result) == 1
        assert result[0]["serial_number"] == "SN005"

    def test_multiple_records(self):
        from checkitem_generator import parse_csv_file

        header = ["Site", "Product", "SerialNumber", "Special Build Name",
                  "Special Build Description", "Unit Number", "Station ID",
                  "Test Pass/Fail Status", "StartTime", "EndTime", "Version",
                  "List of Failing Tests", "REL Event"]
        data = [
            ["JAWX", "B529", "SN_A", "B529-B529", "B529-B529_R2CNM", "",
             "STATION_01", "PASS", "2026-05-12 08:00:00", "2026-05-12 08:01:00",
             "1.0.0", "", "CP1"],
            ["JAWX", "B529", "SN_B", "B529-B529", "B529-B529_R2CNM", "",
             "STATION_02", "FAIL", "2026-05-12 09:00:00", "2026-05-12 09:01:00",
             "1.0.0", "FailTest1", "CP2"],
            ["JAWX", "B529", "SN_A", "B529-B529", "B529-B529_R2CNM", "",
             "STATION_01", "PASS", "2026-05-13 08:00:00", "2026-05-13 08:01:00",
             "1.0.0", "", "CP3"],
        ]
        content = self._make_csv(header, data)
        result = parse_csv_file(content, "Charging")

        assert len(result) == 3
        assert result[0]["serial_number"] == "SN_A"
        assert result[1]["serial_number"] == "SN_B"
        assert result[1]["status"] == "FAIL"
        assert result[2]["rel_event"] == "CP3"

    def test_skips_rows_without_serial_number(self):
        from checkitem_generator import parse_csv_file

        header = ["Site", "Product", "SerialNumber", "Special Build Name",
                  "Special Build Description", "Unit Number", "Station ID",
                  "Test Pass/Fail Status", "StartTime", "EndTime", "Version",
                  "List of Failing Tests", "REL Event"]
        data = [
            ["JAWX", "B529", "SN_VALID", "B529-B529", "B529-B529_R2CNM", "",
             "STATION_01", "PASS", "2026-05-12 08:00:00", "2026-05-12 08:01:00",
             "1.0.0", "", "CP1"],
            ["JAWX", "B529", "", "B529-B529", "B529-B529_R2CNM", "",
             "STATION_01", "PASS", "2026-05-12 09:00:00", "2026-05-12 09:01:00",
             "1.0.0", "", "CP2"],
        ]
        content = self._make_csv(header, data)
        result = parse_csv_file(content, "ISB")

        assert len(result) == 1
        assert result[0]["serial_number"] == "SN_VALID"

    def test_item_type_passed_through(self):
        from checkitem_generator import parse_csv_file

        header = ["Site", "Product", "SerialNumber", "Special Build Name",
                  "Special Build Description", "Unit Number", "Station ID",
                  "Test Pass/Fail Status", "StartTime", "EndTime", "Version",
                  "List of Failing Tests", "REL Event"]
        data = [["JAWX", "B529", "SN001", "B529-B529", "B529-B529_R2CNM", "",
                 "STATION_01", "PASS", "2026-05-12 08:00:00", "2026-05-12 08:01:00",
                 "1.0.0", "", "REL_T0"]]
        content = self._make_csv(header, data)

        for item_type in ["BT-OTA", "Charging", "FACT", "ISB", "Touch-CAL-Post", "Cosmetic"]:
            result = parse_csv_file(content, item_type)
            assert result[0]["item"] == item_type

    def test_non_numeric_param_values_kept_as_string(self):
        from checkitem_generator import parse_csv_file

        header = ["Site", "Product", "SerialNumber", "Special Build Name",
                  "Special Build Description", "Unit Number", "Station ID",
                  "Test Pass/Fail Status", "StartTime", "EndTime", "Version",
                  "List of Failing Tests", "REL Event", "ParamA", "ParamB"]
        data = [["JAWX", "B529", "SN006", "B529-B529", "B529-B529_R2CNM", "",
                 "STATION_01", "FAIL", "2026-05-12 08:00:00", "2026-05-12 08:01:00",
                 "1.0.0", "TestFail", "CP1", "NA", "3.14"]]
        content = self._make_csv(header, data)
        result = parse_csv_file(content, "FACT")

        assert result[0]["test_params"]["ParamA"] == "NA"
        assert result[0]["test_params"]["ParamB"] == 3.14


class TestFilterValidSns:
    """Tests for filter_valid_sns function."""

    def test_records_with_valid_sns_are_kept(self):
        from checkitem_generator import filter_valid_sns

        records = [
            {"serial_number": "SN001", "item": "FACT", "status": "PASS"},
            {"serial_number": "SN002", "item": "ISB", "status": "FAIL"},
            {"serial_number": "SN003", "item": "FACT", "status": "PASS"},
        ]
        valid_sns = {"SN001", "SN002", "SN003"}
        result = filter_valid_sns(records, valid_sns)

        assert len(result) == 3
        assert result[0]["serial_number"] == "SN001"
        assert result[1]["serial_number"] == "SN002"
        assert result[2]["serial_number"] == "SN003"

    def test_records_with_invalid_sns_are_removed(self):
        from checkitem_generator import filter_valid_sns

        records = [
            {"serial_number": "SN001", "item": "FACT", "status": "PASS"},
            {"serial_number": "INVALID_SN", "item": "ISB", "status": "FAIL"},
            {"serial_number": "SN002", "item": "FACT", "status": "PASS"},
            {"serial_number": "UNKNOWN", "item": "Charging", "status": "PASS"},
        ]
        valid_sns = {"SN001", "SN002"}
        result = filter_valid_sns(records, valid_sns)

        assert len(result) == 2
        assert result[0]["serial_number"] == "SN001"
        assert result[1]["serial_number"] == "SN002"

    def test_empty_valid_sns_returns_empty_list(self):
        from checkitem_generator import filter_valid_sns

        records = [
            {"serial_number": "SN001", "item": "FACT", "status": "PASS"},
            {"serial_number": "SN002", "item": "ISB", "status": "FAIL"},
        ]
        valid_sns = set()
        result = filter_valid_sns(records, valid_sns)

        assert result == []

    def test_empty_records_returns_empty_list(self):
        from checkitem_generator import filter_valid_sns

        records = []
        valid_sns = {"SN001", "SN002", "SN003"}
        result = filter_valid_sns(records, valid_sns)

        assert result == []

    def test_mixed_valid_and_invalid_preserves_order(self):
        from checkitem_generator import filter_valid_sns

        records = [
            {"serial_number": "A", "item": "FACT", "status": "PASS"},
            {"serial_number": "B", "item": "ISB", "status": "PASS"},
            {"serial_number": "C", "item": "FACT", "status": "FAIL"},
            {"serial_number": "D", "item": "Charging", "status": "PASS"},
        ]
        valid_sns = {"B", "D"}
        result = filter_valid_sns(records, valid_sns)

        assert len(result) == 2
        assert result[0]["serial_number"] == "B"
        assert result[1]["serial_number"] == "D"

    def test_preserves_all_record_fields(self):
        from checkitem_generator import filter_valid_sns

        records = [
            {
                "serial_number": "SN001",
                "rel_event": "REL_T0",
                "status": "FAIL",
                "end_time": "2026-05-12 08:01:00",
                "failing_tests": "TestX",
                "station_id": "STATION_01",
                "version": "1.0.0",
                "item": "FACT",
                "test_params": {"ParamA": 84.5},
            },
        ]
        valid_sns = {"SN001"}
        result = filter_valid_sns(records, valid_sns)

        assert len(result) == 1
        assert result[0] == records[0]


class TestSortAndFilterTfinal:
    """Tests for sort_and_filter_tfinal function."""

    def test_records_sorted_by_end_time_ascending(self):
        from checkitem_generator import sort_and_filter_tfinal

        records = [
            {"serial_number": "SN001", "rel_event": "CP2", "end_time": "2026-05-14 10:00:00"},
            {"serial_number": "SN001", "rel_event": "CP1", "end_time": "2026-05-12 08:00:00"},
            {"serial_number": "SN001", "rel_event": "CP3", "end_time": "2026-05-13 09:00:00"},
        ]
        result = sort_and_filter_tfinal(records)

        assert len(result) == 3
        assert result[0]["end_time"] == "2026-05-12 08:00:00"
        assert result[1]["end_time"] == "2026-05-13 09:00:00"
        assert result[2]["end_time"] == "2026-05-14 10:00:00"

    def test_records_after_rel_tfinal_are_removed(self):
        from checkitem_generator import sort_and_filter_tfinal

        records = [
            {"serial_number": "SN001", "rel_event": "CP1", "end_time": "2026-05-12 08:00:00"},
            {"serial_number": "SN001", "rel_event": "REL_TFINAL", "end_time": "2026-05-13 10:00:00"},
            {"serial_number": "SN001", "rel_event": "CP3", "end_time": "2026-05-14 12:00:00"},
        ]
        result = sort_and_filter_tfinal(records)

        assert len(result) == 2
        assert result[0]["rel_event"] == "CP1"
        assert result[1]["rel_event"] == "REL_TFINAL"

    def test_records_for_other_sns_unaffected(self):
        from checkitem_generator import sort_and_filter_tfinal

        records = [
            {"serial_number": "SN001", "rel_event": "CP1", "end_time": "2026-05-12 08:00:00"},
            {"serial_number": "SN001", "rel_event": "REL_TFINAL", "end_time": "2026-05-13 10:00:00"},
            {"serial_number": "SN001", "rel_event": "CP3", "end_time": "2026-05-14 12:00:00"},
            {"serial_number": "SN002", "rel_event": "CP1", "end_time": "2026-05-12 09:00:00"},
            {"serial_number": "SN002", "rel_event": "CP2", "end_time": "2026-05-14 15:00:00"},
        ]
        result = sort_and_filter_tfinal(records)

        # SN001: CP1 + REL_TFINAL kept, CP3 removed
        # SN002: both kept (no TFINAL)
        assert len(result) == 4
        sn1_records = [r for r in result if r["serial_number"] == "SN001"]
        sn2_records = [r for r in result if r["serial_number"] == "SN002"]
        assert len(sn1_records) == 2
        assert len(sn2_records) == 2
        assert sn1_records[0]["rel_event"] == "CP1"
        assert sn1_records[1]["rel_event"] == "REL_TFINAL"
        assert sn2_records[0]["rel_event"] == "CP1"
        assert sn2_records[1]["rel_event"] == "CP2"

    def test_sn_without_tfinal_keeps_all_records(self):
        from checkitem_generator import sort_and_filter_tfinal

        records = [
            {"serial_number": "SN001", "rel_event": "CP1", "end_time": "2026-05-12 08:00:00"},
            {"serial_number": "SN001", "rel_event": "CP2", "end_time": "2026-05-13 09:00:00"},
            {"serial_number": "SN001", "rel_event": "CP3", "end_time": "2026-05-14 10:00:00"},
        ]
        result = sort_and_filter_tfinal(records)

        assert len(result) == 3
        assert result[0]["rel_event"] == "CP1"
        assert result[1]["rel_event"] == "CP2"
        assert result[2]["rel_event"] == "CP3"

    def test_empty_records_returns_empty_list(self):
        from checkitem_generator import sort_and_filter_tfinal

        result = sort_and_filter_tfinal([])
        assert result == []

    def test_tfinal_record_itself_is_kept(self):
        from checkitem_generator import sort_and_filter_tfinal

        records = [
            {"serial_number": "SN001", "rel_event": "REL_TFINAL", "end_time": "2026-05-13 10:00:00"},
        ]
        result = sort_and_filter_tfinal(records)

        assert len(result) == 1
        assert result[0]["rel_event"] == "REL_TFINAL"

    def test_multiple_sns_with_different_tfinal_times(self):
        from checkitem_generator import sort_and_filter_tfinal

        records = [
            {"serial_number": "SN001", "rel_event": "CP1", "end_time": "2026-05-12 08:00:00"},
            {"serial_number": "SN001", "rel_event": "REL_TFINAL", "end_time": "2026-05-13 10:00:00"},
            {"serial_number": "SN001", "rel_event": "CP_AFTER", "end_time": "2026-05-14 12:00:00"},
            {"serial_number": "SN002", "rel_event": "CP1", "end_time": "2026-05-11 07:00:00"},
            {"serial_number": "SN002", "rel_event": "CP2", "end_time": "2026-05-12 09:00:00"},
            {"serial_number": "SN002", "rel_event": "REL_TFINAL", "end_time": "2026-05-15 16:00:00"},
            {"serial_number": "SN002", "rel_event": "CP_AFTER", "end_time": "2026-05-16 08:00:00"},
        ]
        result = sort_and_filter_tfinal(records)

        sn1_records = [r for r in result if r["serial_number"] == "SN001"]
        sn2_records = [r for r in result if r["serial_number"] == "SN002"]
        # SN001: CP1 + REL_TFINAL (CP_AFTER removed)
        assert len(sn1_records) == 2
        assert all(r["rel_event"] != "CP_AFTER" for r in sn1_records)
        # SN002: CP1 + CP2 + REL_TFINAL (CP_AFTER removed)
        assert len(sn2_records) == 3
        assert all(r["rel_event"] != "CP_AFTER" for r in sn2_records)


class TestAttributeAnomalyCps:
    """Tests for attribute_anomaly_cps function."""

    def test_normal_cps_get_effective_cp_equal_to_rel_event(self):
        from checkitem_generator import attribute_anomaly_cps

        records = [
            {"serial_number": "SN001", "rel_event": "REL_T0", "end_time": "2026-05-12 08:00:00"},
            {"serial_number": "SN001", "rel_event": "CP_100%", "end_time": "2026-05-13 09:00:00"},
            {"serial_number": "SN001", "rel_event": "CP_200%", "end_time": "2026-05-14 10:00:00"},
        ]
        result = attribute_anomaly_cps(records)

        assert result[0]["effective_cp"] == "REL_T0"
        assert result[1]["effective_cp"] == "CP_100%"
        assert result[2]["effective_cp"] == "CP_200%"

    def test_anomaly_events_attributed_to_preceding_valid_cp(self):
        from checkitem_generator import attribute_anomaly_cps

        records = [
            {"serial_number": "SN001", "rel_event": "REL_T0", "end_time": "2026-05-12 08:00:00"},
            {"serial_number": "SN001", "rel_event": "CP_100%", "end_time": "2026-05-13 09:00:00"},
            {"serial_number": "SN001", "rel_event": "REL FA RETEST", "end_time": "2026-05-13 10:00:00"},
            {"serial_number": "SN001", "rel_event": "SEND TO FA", "end_time": "2026-05-13 11:00:00"},
        ]
        result = attribute_anomaly_cps(records)

        assert result[0]["effective_cp"] == "REL_T0"
        assert result[1]["effective_cp"] == "CP_100%"
        # Anomaly events attributed to CP_100% (the preceding valid CP)
        assert result[2]["effective_cp"] == "CP_100%"
        assert result[3]["effective_cp"] == "CP_100%"

    def test_multiple_anomaly_events_in_sequence_all_attribute_to_same_cp(self):
        from checkitem_generator import attribute_anomaly_cps

        records = [
            {"serial_number": "SN001", "rel_event": "BC_CWCB_SHORT", "end_time": "2026-05-12 08:00:00"},
            {"serial_number": "SN001", "rel_event": "REL FA RETEST", "end_time": "2026-05-13 09:00:00"},
            {"serial_number": "SN001", "rel_event": "SEND TO FA", "end_time": "2026-05-13 10:00:00"},
            {"serial_number": "SN001", "rel_event": "STOP TEST", "end_time": "2026-05-13 11:00:00"},
            {"serial_number": "SN001", "rel_event": "RETURN TO REL", "end_time": "2026-05-13 12:00:00"},
        ]
        result = attribute_anomaly_cps(records)

        assert result[0]["effective_cp"] == "BC_CWCB_SHORT"
        # All 4 anomaly events should attribute to BC_CWCB_SHORT
        assert result[1]["effective_cp"] == "BC_CWCB_SHORT"
        assert result[2]["effective_cp"] == "BC_CWCB_SHORT"
        assert result[3]["effective_cp"] == "BC_CWCB_SHORT"
        assert result[4]["effective_cp"] == "BC_CWCB_SHORT"

    def test_anomaly_event_with_no_preceding_cp_uses_rel_event(self):
        from checkitem_generator import attribute_anomaly_cps

        records = [
            {"serial_number": "SN001", "rel_event": "REL FA RETEST", "end_time": "2026-05-12 08:00:00"},
            {"serial_number": "SN001", "rel_event": "SEND TO FA", "end_time": "2026-05-12 09:00:00"},
        ]
        result = attribute_anomaly_cps(records)

        # No preceding valid CP — each anomaly falls back to its own rel_event
        assert result[0]["effective_cp"] == "REL FA RETEST"
        assert result[1]["effective_cp"] == "SEND TO FA"

    def test_per_sn_tracking_independent(self):
        """Each SN tracks its own last valid CP independently."""
        from checkitem_generator import attribute_anomaly_cps

        records = [
            {"serial_number": "SN001", "rel_event": "CP_A", "end_time": "2026-05-12 08:00:00"},
            {"serial_number": "SN002", "rel_event": "CP_B", "end_time": "2026-05-12 08:30:00"},
            {"serial_number": "SN001", "rel_event": "REL FA RETEST", "end_time": "2026-05-12 09:00:00"},
            {"serial_number": "SN002", "rel_event": "STOP TEST", "end_time": "2026-05-12 09:30:00"},
        ]
        result = attribute_anomaly_cps(records)

        assert result[0]["effective_cp"] == "CP_A"
        assert result[1]["effective_cp"] == "CP_B"
        # SN001's anomaly attributed to CP_A (not CP_B from SN002)
        assert result[2]["effective_cp"] == "CP_A"
        # SN002's anomaly attributed to CP_B (not CP_A from SN001)
        assert result[3]["effective_cp"] == "CP_B"

    def test_anomaly_after_new_valid_cp_attributes_to_new_cp(self):
        """After a new valid CP, anomalies attribute to the new one."""
        from checkitem_generator import attribute_anomaly_cps

        records = [
            {"serial_number": "SN001", "rel_event": "CP_100%", "end_time": "2026-05-12 08:00:00"},
            {"serial_number": "SN001", "rel_event": "REL FA RETEST", "end_time": "2026-05-12 09:00:00"},
            {"serial_number": "SN001", "rel_event": "CP_200%", "end_time": "2026-05-13 08:00:00"},
            {"serial_number": "SN001", "rel_event": "SEND TO FA", "end_time": "2026-05-13 09:00:00"},
        ]
        result = attribute_anomaly_cps(records)

        assert result[0]["effective_cp"] == "CP_100%"
        assert result[1]["effective_cp"] == "CP_100%"  # attributed to CP_100%
        assert result[2]["effective_cp"] == "CP_200%"
        assert result[3]["effective_cp"] == "CP_200%"  # attributed to CP_200%

    def test_empty_records_returns_empty_list(self):
        from checkitem_generator import attribute_anomaly_cps

        result = attribute_anomaly_cps([])
        assert result == []

    def test_all_anomaly_event_types_recognized(self):
        """All four anomaly event types are correctly identified."""
        from checkitem_generator import attribute_anomaly_cps

        records = [
            {"serial_number": "SN001", "rel_event": "CP_VALID", "end_time": "2026-05-12 08:00:00"},
            {"serial_number": "SN001", "rel_event": "REL FA RETEST", "end_time": "2026-05-12 09:00:00"},
            {"serial_number": "SN001", "rel_event": "SEND TO FA", "end_time": "2026-05-12 10:00:00"},
            {"serial_number": "SN001", "rel_event": "STOP TEST", "end_time": "2026-05-12 11:00:00"},
            {"serial_number": "SN001", "rel_event": "RETURN TO REL", "end_time": "2026-05-12 12:00:00"},
        ]
        result = attribute_anomaly_cps(records)

        for i in range(1, 5):
            assert result[i]["effective_cp"] == "CP_VALID", (
                f"Anomaly event '{result[i]['rel_event']}' not attributed correctly"
            )


class TestDeduplicateRecords:
    """Tests for deduplicate_records function."""

    def test_single_record_per_group_kept_as_is(self):
        from checkitem_generator import deduplicate_records

        records = [
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "FACT", "end_time": "2026-05-12 08:00:00", "status": "PASS"},
            {"serial_number": "SN001", "effective_cp": "CP2", "item": "FACT", "end_time": "2026-05-13 09:00:00", "status": "PASS"},
            {"serial_number": "SN002", "effective_cp": "CP1", "item": "ISB", "end_time": "2026-05-12 10:00:00", "status": "FAIL"},
        ]
        result = deduplicate_records(records)

        assert len(result) == 3
        # All records should be present since each has a unique key
        keys = {(r["serial_number"], r["effective_cp"], r["item"]) for r in result}
        assert ("SN001", "CP1", "FACT") in keys
        assert ("SN001", "CP2", "FACT") in keys
        assert ("SN002", "CP1", "ISB") in keys

    def test_multiple_records_same_group_keeps_latest_end_time(self):
        from checkitem_generator import deduplicate_records

        records = [
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "FACT", "end_time": "2026-05-12 08:00:00", "status": "FAIL"},
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "FACT", "end_time": "2026-05-13 09:00:00", "status": "FAIL"},
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "FACT", "end_time": "2026-05-14 10:00:00", "status": "PASS"},
        ]
        result = deduplicate_records(records)

        assert len(result) == 1
        assert result[0]["end_time"] == "2026-05-14 10:00:00"
        assert result[0]["status"] == "PASS"

    def test_different_groups_are_independent(self):
        from checkitem_generator import deduplicate_records

        records = [
            # Group 1: SN001/CP1/FACT — two records
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "FACT", "end_time": "2026-05-12 08:00:00", "status": "FAIL"},
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "FACT", "end_time": "2026-05-13 09:00:00", "status": "PASS"},
            # Group 2: SN001/CP1/ISB — two records
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "ISB", "end_time": "2026-05-12 07:00:00", "status": "PASS"},
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "ISB", "end_time": "2026-05-14 11:00:00", "status": "FAIL"},
            # Group 3: SN002/CP1/FACT — one record
            {"serial_number": "SN002", "effective_cp": "CP1", "item": "FACT", "end_time": "2026-05-12 09:00:00", "status": "PASS"},
        ]
        result = deduplicate_records(records)

        assert len(result) == 3
        # Find each group's result
        result_map = {(r["serial_number"], r["effective_cp"], r["item"]): r for r in result}

        # Group 1: latest is 2026-05-13
        assert result_map[("SN001", "CP1", "FACT")]["end_time"] == "2026-05-13 09:00:00"
        assert result_map[("SN001", "CP1", "FACT")]["status"] == "PASS"

        # Group 2: latest is 2026-05-14
        assert result_map[("SN001", "CP1", "ISB")]["end_time"] == "2026-05-14 11:00:00"
        assert result_map[("SN001", "CP1", "ISB")]["status"] == "FAIL"

        # Group 3: only one record
        assert result_map[("SN002", "CP1", "FACT")]["end_time"] == "2026-05-12 09:00:00"
        assert result_map[("SN002", "CP1", "FACT")]["status"] == "PASS"

    def test_empty_input_returns_empty_output(self):
        from checkitem_generator import deduplicate_records

        result = deduplicate_records([])
        assert result == []

    def test_preserves_all_record_fields(self):
        from checkitem_generator import deduplicate_records

        records = [
            {
                "serial_number": "SN001",
                "effective_cp": "CP1",
                "item": "FACT",
                "end_time": "2026-05-12 08:00:00",
                "status": "FAIL",
                "rel_event": "CP1",
                "failing_tests": "TestX",
                "station_id": "STATION_01",
                "version": "2.0.0",
                "test_params": {"ParamA": 84.5},
            },
            {
                "serial_number": "SN001",
                "effective_cp": "CP1",
                "item": "FACT",
                "end_time": "2026-05-13 09:00:00",
                "status": "PASS",
                "rel_event": "CP1",
                "failing_tests": "",
                "station_id": "STATION_02",
                "version": "3.0.0",
                "test_params": None,
            },
        ]
        result = deduplicate_records(records)

        assert len(result) == 1
        rec = result[0]
        assert rec["serial_number"] == "SN001"
        assert rec["effective_cp"] == "CP1"
        assert rec["item"] == "FACT"
        assert rec["end_time"] == "2026-05-13 09:00:00"
        assert rec["status"] == "PASS"
        assert rec["rel_event"] == "CP1"
        assert rec["failing_tests"] == ""
        assert rec["station_id"] == "STATION_02"
        assert rec["version"] == "3.0.0"
        assert rec["test_params"] is None

    def test_different_items_same_sn_and_cp_are_separate_groups(self):
        from checkitem_generator import deduplicate_records

        records = [
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "FACT", "end_time": "2026-05-12 08:00:00", "status": "PASS"},
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "ISB", "end_time": "2026-05-12 08:00:00", "status": "PASS"},
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "Charging", "end_time": "2026-05-12 08:00:00", "status": "PASS"},
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "BT-OTA", "end_time": "2026-05-12 08:00:00", "status": "PASS"},
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "Touch-CAL-Post", "end_time": "2026-05-12 08:00:00", "status": "PASS"},
            {"serial_number": "SN001", "effective_cp": "CP1", "item": "Cosmetic", "end_time": "2026-05-12 08:00:00", "status": "PASS"},
        ]
        result = deduplicate_records(records)

        # All 6 items are different groups, so all should be kept
        assert len(result) == 6


class TestIsStrifeFailure:
    """Tests for _is_strife_failure helper function."""

    def test_random_drop_1m_pb_above_threshold(self):
        from checkitem_generator import _is_strife_failure

        # Threshold is 200; values > 200 are strife
        assert _is_strife_failure("RANDOM DROP 1M PB_250") is True
        assert _is_strife_failure("RANDOM DROP 1M PB_300") is True
        assert _is_strife_failure("RANDOM DROP 1M PB_201") is True

    def test_random_drop_1m_pb_at_or_below_threshold(self):
        from checkitem_generator import _is_strife_failure

        assert _is_strife_failure("RANDOM DROP 1M PB_200") is False
        assert _is_strife_failure("RANDOM DROP 1M PB_100") is False
        assert _is_strife_failure("RANDOM DROP 1M PB_50") is False

    def test_random_drop_1m_gra_above_threshold(self):
        from checkitem_generator import _is_strife_failure

        # Threshold is 20; values > 20 are strife
        assert _is_strife_failure("RANDOM DROP 1M GRA_25") is True
        assert _is_strife_failure("RANDOM DROP 1M GRA_30") is True
        assert _is_strife_failure("RANDOM DROP 1M GRA_21") is True

    def test_random_drop_1m_gra_at_or_below_threshold(self):
        from checkitem_generator import _is_strife_failure

        assert _is_strife_failure("RANDOM DROP 1M GRA_20") is False
        assert _is_strife_failure("RANDOM DROP 1M GRA_10") is False

    def test_15mm_probe_sp_above_threshold(self):
        from checkitem_generator import _is_strife_failure

        # Threshold is 21; values > 21 are strife
        assert _is_strife_failure("15MM PROBE SP_25KG") is True
        assert _is_strife_failure("15MM PROBE SP_30KG") is True
        assert _is_strife_failure("15MM PROBE SP_22KG") is True

    def test_15mm_probe_sp_at_or_below_threshold(self):
        from checkitem_generator import _is_strife_failure

        assert _is_strife_failure("15MM PROBE SP_21KG") is False
        assert _is_strife_failure("15MM PROBE SP_15KG") is False

    def test_cleaning_spray_above_threshold(self):
        from checkitem_generator import _is_strife_failure

        # Threshold is 72; values > 72 are strife
        assert _is_strife_failure("CLEANING_SPRAY_OP1_AFTER 96HRS") is True
        assert _is_strife_failure("CLEANING_SPRAY_OP1_AFTER 100HRS") is True
        assert _is_strife_failure("CLEANING_SPRAY_OP1_AFTER 73HRS") is True

    def test_cleaning_spray_at_or_below_threshold(self):
        from checkitem_generator import _is_strife_failure

        assert _is_strife_failure("CLEANING_SPRAY_OP1_AFTER 72HRS") is False
        assert _is_strife_failure("CLEANING_SPRAY_OP1_AFTER 48HRS") is False

    def test_rbi_above_threshold(self):
        from checkitem_generator import _is_strife_failure

        # Threshold is 10; values > 10 are strife
        assert _is_strife_failure("RBI_15CM") is True
        assert _is_strife_failure("RBI_20CM") is True
        assert _is_strife_failure("RBI_11CM") is True

    def test_rbi_at_or_below_threshold(self):
        from checkitem_generator import _is_strife_failure

        assert _is_strife_failure("RBI_10CM") is False
        assert _is_strife_failure("RBI_5CM") is False

    def test_bc_percentage_above_threshold(self):
        from checkitem_generator import _is_strife_failure

        # Threshold is 100; values > 100 are strife
        assert _is_strife_failure("BC_CWCB_SHORT_140%") is True
        assert _is_strife_failure("BC_CWCB_SHORT_200%") is True
        assert _is_strife_failure("BC_SOMETHING_101%") is True

    def test_bc_percentage_at_or_below_threshold(self):
        from checkitem_generator import _is_strife_failure

        assert _is_strife_failure("BC_CWCB_SHORT_100%") is False
        assert _is_strife_failure("BC_CWCB_SHORT_80%") is False
        assert _is_strife_failure("BC_SOMETHING_50%") is False

    def test_no_matching_pattern_returns_false(self):
        from checkitem_generator import _is_strife_failure

        assert _is_strife_failure("REL_T0") is False
        assert _is_strife_failure("TC_300CYCLES") is False
        assert _is_strife_failure("SOME_RANDOM_CP") is False
        assert _is_strife_failure("") is False

    def test_case_insensitive(self):
        from checkitem_generator import _is_strife_failure

        assert _is_strife_failure("random drop 1m pb_250") is True
        assert _is_strife_failure("Random Drop 1M PB_250") is True
        assert _is_strife_failure("bc_cwcb_short_140%") is True


class TestDetectStrifeFailures:
    """Tests for detect_strife_failures function."""

    def test_pass_records_get_none_failure_type(self):
        from checkitem_generator import detect_strife_failures

        records = [
            {"serial_number": "SN001", "effective_cp": "RANDOM DROP 1M PB_250", "status": "PASS", "item": "FACT"},
            {"serial_number": "SN002", "effective_cp": "BC_CWCB_SHORT_140%", "status": "PASS", "item": "ISB"},
        ]
        result = detect_strife_failures(records, {})

        assert result[0]["failure_type"] is None
        assert result[1]["failure_type"] is None

    def test_fail_at_strife_cp_gets_strife_fail(self):
        from checkitem_generator import detect_strife_failures

        records = [
            {"serial_number": "SN001", "effective_cp": "RANDOM DROP 1M PB_250", "status": "FAIL", "item": "FACT"},
            {"serial_number": "SN002", "effective_cp": "BC_CWCB_SHORT_140%", "status": "FAIL", "item": "ISB"},
            {"serial_number": "SN003", "effective_cp": "RBI_15CM", "status": "FAIL", "item": "Charging"},
        ]
        result = detect_strife_failures(records, {})

        assert result[0]["failure_type"] == "strife_fail"
        assert result[1]["failure_type"] == "strife_fail"
        assert result[2]["failure_type"] == "strife_fail"

    def test_fail_at_spec_cp_gets_spec_fail(self):
        from checkitem_generator import detect_strife_failures

        records = [
            {"serial_number": "SN001", "effective_cp": "RANDOM DROP 1M PB_100", "status": "FAIL", "item": "FACT"},
            {"serial_number": "SN002", "effective_cp": "BC_CWCB_SHORT_100%", "status": "FAIL", "item": "ISB"},
            {"serial_number": "SN003", "effective_cp": "REL_T0", "status": "FAIL", "item": "Charging"},
        ]
        result = detect_strife_failures(records, {})

        assert result[0]["failure_type"] == "spec_fail"
        assert result[1]["failure_type"] == "spec_fail"
        assert result[2]["failure_type"] == "spec_fail"

    def test_mixed_records(self):
        from checkitem_generator import detect_strife_failures

        records = [
            {"serial_number": "SN001", "effective_cp": "REL_T0", "status": "PASS", "item": "FACT"},
            {"serial_number": "SN001", "effective_cp": "BC_CWCB_SHORT_100%", "status": "FAIL", "item": "FACT"},
            {"serial_number": "SN001", "effective_cp": "BC_CWCB_SHORT_140%", "status": "FAIL", "item": "FACT"},
            {"serial_number": "SN002", "effective_cp": "RANDOM DROP 1M PB_200", "status": "FAIL", "item": "ISB"},
            {"serial_number": "SN002", "effective_cp": "RANDOM DROP 1M PB_250", "status": "FAIL", "item": "ISB"},
        ]
        result = detect_strife_failures(records, {})

        assert result[0]["failure_type"] is None          # PASS
        assert result[1]["failure_type"] == "spec_fail"   # BC at 100% (not > 100)
        assert result[2]["failure_type"] == "strife_fail" # BC at 140% (> 100)
        assert result[3]["failure_type"] == "spec_fail"   # PB at 200 (not > 200)
        assert result[4]["failure_type"] == "strife_fail" # PB at 250 (> 200)

    def test_empty_records_returns_empty_list(self):
        from checkitem_generator import detect_strife_failures

        result = detect_strife_failures([], {})
        assert result == []

    def test_records_mutated_in_place(self):
        from checkitem_generator import detect_strife_failures

        records = [
            {"serial_number": "SN001", "effective_cp": "REL_T0", "status": "FAIL", "item": "FACT"},
        ]
        result = detect_strife_failures(records, {})

        # The function mutates records in place and returns the same list
        assert result is records
        assert records[0]["failure_type"] == "spec_fail"

    def test_cp_schedule_param_accepted(self):
        """Verify the function accepts a cp_schedule dict without error."""
        from checkitem_generator import detect_strife_failures

        cp_schedule = {
            1.0: ["REL_T0", "RANDOM DROP 1M PB_100", "RANDOM DROP 1M PB_200", "RANDOM DROP 1M PB_250"],
            2.0: ["BC_CWCB_SHORT_100%", "BC_CWCB_SHORT_140%"],
        }
        records = [
            {"serial_number": "SN001", "effective_cp": "RANDOM DROP 1M PB_250", "status": "FAIL", "item": "FACT"},
        ]
        result = detect_strife_failures(records, cp_schedule)

        assert result[0]["failure_type"] == "strife_fail"

    def test_non_fail_statuses_get_none(self):
        """Records with status other than 'FAIL' (empty, unknown) get None."""
        from checkitem_generator import detect_strife_failures

        records = [
            {"serial_number": "SN001", "effective_cp": "RANDOM DROP 1M PB_250", "status": "", "item": "FACT"},
            {"serial_number": "SN002", "effective_cp": "BC_CWCB_SHORT_140%", "status": "UNKNOWN", "item": "ISB"},
        ]
        result = detect_strife_failures(records, {})

        assert result[0]["failure_type"] is None
        assert result[1]["failure_type"] is None


class TestIsChemicalWf:
    """Tests for _is_chemical_wf helper function."""

    def test_wf_29_is_chemical(self):
        from checkitem_generator import _is_chemical_wf

        assert _is_chemical_wf(29) is True
        assert _is_chemical_wf(29.0) is True

    def test_wf_29_x_is_chemical(self):
        from checkitem_generator import _is_chemical_wf

        assert _is_chemical_wf(29.1) is True
        assert _is_chemical_wf(29.5) is True
        assert _is_chemical_wf(29.9) is True

    def test_wf_30_is_chemical(self):
        from checkitem_generator import _is_chemical_wf

        assert _is_chemical_wf(30) is True
        assert _is_chemical_wf(30.0) is True

    def test_wf_30_x_is_not_chemical(self):
        from checkitem_generator import _is_chemical_wf

        assert _is_chemical_wf(30.1) is False
        assert _is_chemical_wf(30.5) is False

    def test_other_wfs_are_not_chemical(self):
        from checkitem_generator import _is_chemical_wf

        assert _is_chemical_wf(1) is False
        assert _is_chemical_wf(10) is False
        assert _is_chemical_wf(14.1) is False
        assert _is_chemical_wf(28) is False
        assert _is_chemical_wf(31) is False

    def test_string_wf_numbers(self):
        from checkitem_generator import _is_chemical_wf

        assert _is_chemical_wf("29") is True
        assert _is_chemical_wf("29.1") is True
        assert _is_chemical_wf("30") is True
        assert _is_chemical_wf("10") is False

    def test_invalid_input_returns_false(self):
        from checkitem_generator import _is_chemical_wf

        assert _is_chemical_wf(None) is False
        assert _is_chemical_wf("abc") is False
        assert _is_chemical_wf("") is False


class TestCreateWfSheet:
    """Tests for create_wf_sheet function."""

    def test_sheet_name_format_integer_wf(self):
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 10, "ISB + FACT", ["T0", "CP1"], [])

        assert ws.title == "Sys WF10_ISB + FACT"

    def test_sheet_name_format_decimal_wf(self):
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 14.1, "BT-OTA", ["T0"], [])

        assert ws.title == "Sys WF14.1_BT-OTA"

    def test_sheet_name_compatible_with_engine_regex(self):
        """Sheet name must be parseable by engine.py's WF regex."""
        import re
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 10, "ISB + FACT", ["T0"], [])

        m = re.search(r'WF(\d+\.?\d*)', ws.title)
        assert m is not None
        assert m.group(1) == "10"

    def test_sheet_name_decimal_compatible_with_engine_regex(self):
        import re
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 14.1, "BT-OTA", ["T0"], [])

        m = re.search(r'WF(\d+\.?\d*)', ws.title)
        assert m is not None
        assert m.group(1) == "14.1"

    def test_fixed_header_columns(self):
        """Row 2 should have %, Completion%, Config, Unit#, S/N in columns 1-5, merged with row 3."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 1, "Test", ["CP1"], [])

        # Row 1 has Report Date in A1
        assert "Report Date" in str(ws.cell(row=1, column=1).value)
        # Row 2 has %, Completion%, Config, Unit#, S/N (merged with row 3)
        assert ws.cell(row=2, column=1).value == "%"
        assert ws.cell(row=2, column=2).value == "Completion%"
        assert ws.cell(row=2, column=3).value == "Config"
        assert ws.cell(row=2, column=4).value == "Unit #"
        assert ws.cell(row=2, column=5).value == "S/N"

    def test_cp_group_merged_header(self):
        """Each CP name should be merged across 6 columns in row 2, starting at col 6."""
        from checkitem_generator import create_wf_sheet, CHECK_ITEMS
        from openpyxl import Workbook

        wb = Workbook()
        cp_list = ["T0", "BC_CWCB_SHORT"]
        ws = create_wf_sheet(wb, 1, "Test", cp_list, [])

        # First CP starts at column 6 (after 5 fixed headers)
        assert ws.cell(row=2, column=6).value == "T0"
        # Second CP starts at column 12 (6 + 6)
        assert ws.cell(row=2, column=12).value == "BC_CWCB_SHORT"

    def test_check_item_sub_headers(self):
        """Row 3 should have check item names under each CP group."""
        from checkitem_generator import create_wf_sheet, CHECK_ITEMS
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 1, "Test", ["T0"], [])

        # Check items start at column 6 (after 5 fixed headers: %, Completion%, Config, Unit#, S/N)
        for i, item_name in enumerate(CHECK_ITEMS):
            assert ws.cell(row=3, column=6 + i).value == item_name

    def test_check_item_order_is_correct(self):
        """Check items must be in the fixed order."""
        from checkitem_generator import CHECK_ITEMS

        assert CHECK_ITEMS == ["Cosmetic", "ISB", "FACT", "BT-OTA", "Touch-CAL-Post", "Charging"]

    def test_multiple_cps_sub_headers(self):
        """Each CP group gets its own set of 6 check item sub-headers."""
        from checkitem_generator import create_wf_sheet, CHECK_ITEMS
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 1, "Test", ["T0", "CP2", "CP3"], [])

        # CP1 sub-headers at columns 6-11
        for i, item_name in enumerate(CHECK_ITEMS):
            assert ws.cell(row=3, column=6 + i).value == item_name

        # CP2 sub-headers at columns 12-17
        for i, item_name in enumerate(CHECK_ITEMS):
            assert ws.cell(row=3, column=12 + i).value == item_name

        # CP3 sub-headers at columns 18-23
        for i, item_name in enumerate(CHECK_ITEMS):
            assert ws.cell(row=3, column=18 + i).value == item_name

    def test_chemical_wf_has_extra_column(self):
        """Chemical WFs (29.x, 30) should have a 'Chemical' column before CP groups."""
        from checkitem_generator import create_wf_sheet, CHECK_ITEMS
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 29, "Chemical Test", ["T0"], [])

        # Chemical column at position 6 (after 5 fixed headers), merged vertically
        assert ws.cell(row=2, column=6).value == "Chemical"
        # First CP starts at column 7
        assert ws.cell(row=2, column=7).value == "T0"
        # Check items start at column 7 in row 3
        for i, item_name in enumerate(CHECK_ITEMS):
            assert ws.cell(row=3, column=7 + i).value == item_name

    def test_chemical_wf_29_1(self):
        """WF 29.1 should also be treated as chemical."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 29.1, "Chemical Sub", ["T0"], [])

        assert ws.cell(row=2, column=6).value == "Chemical"
        assert ws.cell(row=2, column=7).value == "T0"

    def test_chemical_wf_30(self):
        """WF 30 should also be treated as chemical."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 30, "Chemical 30", ["T0"], [])

        assert ws.cell(row=2, column=6).value == "Chemical"
        assert ws.cell(row=2, column=7).value == "T0"

    def test_non_chemical_wf_no_extra_column(self):
        """Non-chemical WFs should NOT have a Chemical column."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 10, "Normal WF", ["T0"], [])

        # Column 6 should be the first CP, not "Chemical"
        assert ws.cell(row=2, column=6).value == "T0"

    def test_sn_data_written_to_data_rows(self):
        """SN data should be written starting from row 4."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        sn_data = [
            {"serial_number": "SN001", "config": "R2CNM", "unit_number": "ER2-1-1"},
            {"serial_number": "SN002", "config": "R2CNM", "unit_number": "ER2-1-2"},
            {"serial_number": "SN003", "config": "R1FNF", "unit_number": "ER1-1-1"},
        ]
        ws = create_wf_sheet(wb, 1, "Test", ["T0"], sn_data)

        # Row 4: first SN — Config(3), Unit#(4), S/N(5)
        assert ws.cell(row=4, column=3).value == "R2CNM"
        assert ws.cell(row=4, column=4).value == "ER2-1-1"
        assert ws.cell(row=4, column=5).value == "SN001"

        # Row 5: second SN
        assert ws.cell(row=5, column=3).value == "R2CNM"
        assert ws.cell(row=5, column=4).value == "ER2-1-2"
        assert ws.cell(row=5, column=5).value == "SN002"

        # Row 6: third SN
        assert ws.cell(row=6, column=3).value == "R1FNF"
        assert ws.cell(row=6, column=4).value == "ER1-1-1"
        assert ws.cell(row=6, column=5).value == "SN003"

    def test_empty_cp_list(self):
        """Sheet should still be created with just fixed headers if no CPs."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 1, "Empty", [], [])

        # Row 1 has Report Date
        assert "Report Date" in str(ws.cell(row=1, column=1).value)
        # Row 2 has %, Completion%, Config, Unit#, S/N
        assert ws.cell(row=2, column=1).value == "%"
        assert ws.cell(row=2, column=5).value == "S/N"
        # No CP headers beyond column 5
        assert ws.cell(row=2, column=6).value is None

    def test_empty_sn_data(self):
        """Sheet should still have headers even with no SN data."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 1, "NoData", ["T0"], [])

        assert ws.cell(row=2, column=6).value == "T0"
        # Row 4 should be empty (data starts at row 4)
        assert ws.cell(row=4, column=3).value is None

    def test_is_chemical_override(self):
        """is_chemical parameter should override auto-detection."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        # Force chemical on a non-chemical WF
        wb = Workbook()
        ws = create_wf_sheet(wb, 10, "Forced Chemical", ["T0"], [], is_chemical=True)
        assert ws.cell(row=2, column=6).value == "Chemical"

        # Force non-chemical on a chemical WF
        wb2 = Workbook()
        ws2 = create_wf_sheet(wb2, 29, "Forced Normal", ["T0"], [], is_chemical=False)
        assert ws2.cell(row=2, column=6).value == "T0"

    def test_sheet_name_truncated_to_31_chars(self):
        """Excel sheet names are limited to 31 characters."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        long_name = "A" * 50  # Very long WF name
        ws = create_wf_sheet(wb, 1, long_name, ["T0"], [])

        assert len(ws.title) <= 31

    def test_returns_worksheet(self):
        """Function should return the created worksheet."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook
        from openpyxl.worksheet.worksheet import Worksheet

        wb = Workbook()
        ws = create_wf_sheet(wb, 1, "Test", ["T0"], [])

        assert isinstance(ws, Worksheet)
        assert ws in wb.worksheets

    def test_merged_cells_for_cp_header(self):
        """CP headers should be merged across 6 columns in row 2."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        ws = create_wf_sheet(wb, 1, "Test", ["T0", "CP2"], [])

        # Check that merged cell ranges exist
        merged = [str(m) for m in ws.merged_cells.ranges]
        # T0 should be merged from F2:K2 (columns 6-11, row 2)
        assert "F2:K2" in merged
        # CP2 should be merged from L2:Q2 (columns 12-17, row 2)
        assert "L2:Q2" in merged


class TestPopulateWfData:
    """Tests for populate_wf_data function and color coding."""

    def test_pass_record_gets_green_fill_and_font(self):
        """PASS records should have green background (#C6EFCE) and green font (#006100)."""
        from checkitem_generator import (
            create_wf_sheet, populate_wf_data, PASS_FILL, PASS_FONT,
        )
        from openpyxl import Workbook

        wb = Workbook()
        sn_data = [{"serial_number": "SN001", "config": "R2CNM", "unit_number": "ER2-1-1"}]
        cp_list = ["T0"]
        records_by_sn = {
            "SN001": {
                ("T0", "Cosmetic"): {"status": "PASS", "failure_type": None},
            }
        }

        ws = create_wf_sheet(wb, 1, "Test", cp_list, sn_data, records_by_sn=records_by_sn)

        # Cosmetic is the first check item, at column 6 (after 5 fixed headers)
        cell = ws.cell(row=4, column=6)
        assert cell.value == "PASS"
        assert cell.fill == PASS_FILL
        assert cell.font == PASS_FONT

    def test_spec_fail_record_gets_red_fill(self):
        """Spec fail records should have red background (#FF0000)."""
        from checkitem_generator import (
            create_wf_sheet, populate_wf_data, SPEC_FAIL_FILL,
        )
        from openpyxl import Workbook

        wb = Workbook()
        sn_data = [{"serial_number": "SN001", "config": "R2CNM", "unit_number": "ER2-1-1"}]
        cp_list = ["T0"]
        records_by_sn = {
            "SN001": {
                ("T0", "FACT"): {"status": "FAIL", "failure_type": "spec_fail"},
            }
        }

        ws = create_wf_sheet(wb, 1, "Test", cp_list, sn_data, records_by_sn=records_by_sn)

        # FACT is the 3rd check item (index 2), at column 8
        cell = ws.cell(row=4, column=8)
        assert cell.value == "FAIL"
        assert cell.fill == SPEC_FAIL_FILL

    def test_strife_fail_record_gets_yellow_fill_and_black_font(self):
        """Strife fail records should have yellow background (#FFFF00) and black font."""
        from checkitem_generator import (
            create_wf_sheet, populate_wf_data, STRIFE_FAIL_FILL, STRIFE_FAIL_FONT,
        )
        from openpyxl import Workbook

        wb = Workbook()
        sn_data = [{"serial_number": "SN001", "config": "R2CNM", "unit_number": "ER2-1-1"}]
        cp_list = ["T0"]
        records_by_sn = {
            "SN001": {
                ("T0", "ISB"): {"status": "FAIL", "failure_type": "strife_fail"},
            }
        }

        ws = create_wf_sheet(wb, 1, "Test", cp_list, sn_data, records_by_sn=records_by_sn)

        # ISB is the 2nd check item (index 1), at column 7
        cell = ws.cell(row=4, column=7)
        assert cell.value == "FAIL"
        assert cell.fill == STRIFE_FAIL_FILL
        assert cell.font == STRIFE_FAIL_FONT

    def test_empty_cell_no_color(self):
        """Cells with no record should have no value and no fill."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        wb = Workbook()
        sn_data = [{"serial_number": "SN001", "config": "R2CNM", "unit_number": "ER2-1-1"}]
        cp_list = ["T0"]
        records_by_sn = {
            "SN001": {
                ("T0", "Cosmetic"): {"status": "PASS", "failure_type": None},
                # ISB has no record
            }
        }

        ws = create_wf_sheet(wb, 1, "Test", cp_list, sn_data, records_by_sn=records_by_sn)

        # ISB is at column 7 — should be empty
        cell = ws.cell(row=4, column=7)
        assert cell.value is None
        # Default fill (no fill applied)
        assert cell.fill == PatternFill()

    def test_multiple_sns_multiple_cps(self):
        """Multiple SNs and CPs should be populated correctly."""
        from checkitem_generator import (
            create_wf_sheet, PASS_FILL, SPEC_FAIL_FILL, STRIFE_FAIL_FILL,
        )
        from openpyxl import Workbook

        wb = Workbook()
        sn_data = [
            {"serial_number": "SN001", "config": "R2CNM", "unit_number": "ER2-1-1"},
            {"serial_number": "SN002", "config": "R2CNM", "unit_number": "ER2-1-2"},
        ]
        cp_list = ["T0", "CP2"]
        records_by_sn = {
            "SN001": {
                ("T0", "Cosmetic"): {"status": "PASS", "failure_type": None},
                ("CP2", "FACT"): {"status": "FAIL", "failure_type": "spec_fail"},
            },
            "SN002": {
                ("T0", "ISB"): {"status": "FAIL", "failure_type": "strife_fail"},
                ("CP2", "Charging"): {"status": "PASS", "failure_type": None},
            },
        }

        ws = create_wf_sheet(wb, 1, "Test", cp_list, sn_data, records_by_sn=records_by_sn)

        # SN001 row 4: T0/Cosmetic at col 6 = PASS
        assert ws.cell(row=4, column=6).value == "PASS"
        assert ws.cell(row=4, column=6).fill == PASS_FILL

        # SN001 row 4: CP2/FACT at col 12+2=14 (CP2 starts at col 12, FACT is index 2)
        assert ws.cell(row=4, column=14).value == "FAIL"
        assert ws.cell(row=4, column=14).fill == SPEC_FAIL_FILL

        # SN002 row 5: T0/ISB at col 7 = strife FAIL
        assert ws.cell(row=5, column=7).value == "FAIL"
        assert ws.cell(row=5, column=7).fill == STRIFE_FAIL_FILL

        # SN002 row 5: CP2/Charging at col 12+5=17 (Charging is index 5)
        assert ws.cell(row=5, column=17).value == "PASS"
        assert ws.cell(row=5, column=17).fill == PASS_FILL

    def test_chemical_wf_shifts_columns(self):
        """Chemical WF should shift CP data columns by 1 (Chemical column at col 6)."""
        from checkitem_generator import create_wf_sheet, PASS_FILL
        from openpyxl import Workbook

        wb = Workbook()
        sn_data = [{"serial_number": "SN001", "config": "R2CNM", "unit_number": "ER2-1-1"}]
        cp_list = ["T0"]
        records_by_sn = {
            "SN001": {
                ("T0", "Cosmetic"): {"status": "PASS", "failure_type": None},
            }
        }

        ws = create_wf_sheet(wb, 29, "Chemical", cp_list, sn_data, records_by_sn=records_by_sn)

        # Chemical WF: Chemical col at 6, CP starts at 7
        # Cosmetic (index 0) should be at column 7
        cell = ws.cell(row=4, column=7)
        assert cell.value == "PASS"
        assert cell.fill == PASS_FILL

        # Column 6 should be empty (Chemical column, user fills manually)
        assert ws.cell(row=4, column=6).value is None

    def test_sn_not_in_records_gets_empty_cells(self):
        """SNs not present in records_by_sn should have empty data cells."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        sn_data = [
            {"serial_number": "SN001", "config": "R2CNM", "unit_number": "ER2-1-1"},
            {"serial_number": "SN002", "config": "R2CNM", "unit_number": "ER2-1-2"},
        ]
        cp_list = ["T0"]
        records_by_sn = {
            "SN001": {
                ("T0", "Cosmetic"): {"status": "PASS", "failure_type": None},
            },
            # SN002 has no records
        }

        ws = create_wf_sheet(wb, 1, "Test", cp_list, sn_data, records_by_sn=records_by_sn)

        # SN001 has data
        assert ws.cell(row=4, column=6).value == "PASS"
        # SN002 row 5 should be empty
        assert ws.cell(row=5, column=6).value is None
        assert ws.cell(row=5, column=7).value is None

    def test_no_records_by_sn_skips_population(self):
        """When records_by_sn is None, no data population occurs."""
        from checkitem_generator import create_wf_sheet
        from openpyxl import Workbook

        wb = Workbook()
        sn_data = [{"serial_number": "SN001", "config": "R2CNM", "unit_number": "ER2-1-1"}]
        cp_list = ["T0"]

        ws = create_wf_sheet(wb, 1, "Test", cp_list, sn_data, records_by_sn=None)

        # Data cells should be empty (no population)
        assert ws.cell(row=4, column=6).value is None

    def test_populate_wf_data_directly(self):
        """Test calling populate_wf_data directly on an existing worksheet."""
        from checkitem_generator import (
            create_wf_sheet, populate_wf_data, PASS_FILL, PASS_FONT,
            SPEC_FAIL_FILL, STRIFE_FAIL_FILL, STRIFE_FAIL_FONT,
        )
        from openpyxl import Workbook

        wb = Workbook()
        sn_data = [
            {"serial_number": "SN001", "config": "R2CNM", "unit_number": "ER2-1-1"},
        ]
        cp_list = ["T0"]

        # Create sheet without records
        ws = create_wf_sheet(wb, 1, "Test", cp_list, sn_data)

        # Now populate directly
        records_by_sn = {
            "SN001": {
                ("T0", "Cosmetic"): {"status": "PASS", "failure_type": None},
                ("T0", "ISB"): {"status": "FAIL", "failure_type": "spec_fail"},
                ("T0", "BT-OTA"): {"status": "FAIL", "failure_type": "strife_fail"},
            }
        }
        populate_wf_data(ws, sn_data, cp_list, records_by_sn, is_chemical=False)

        # Cosmetic (col 6) = PASS green
        assert ws.cell(row=4, column=6).value == "PASS"
        assert ws.cell(row=4, column=6).fill == PASS_FILL
        assert ws.cell(row=4, column=6).font == PASS_FONT

        # ISB (col 7) = spec_fail red
        assert ws.cell(row=4, column=7).value == "FAIL"
        assert ws.cell(row=4, column=7).fill == SPEC_FAIL_FILL

        # BT-OTA (col 9) = strife_fail yellow
        assert ws.cell(row=4, column=9).value == "FAIL"
        assert ws.cell(row=4, column=9).fill == STRIFE_FAIL_FILL
        assert ws.cell(row=4, column=9).font == STRIFE_FAIL_FONT

    def test_all_six_check_items_correct_columns(self):
        """Verify all 6 check items map to correct columns for a single CP."""
        from checkitem_generator import create_wf_sheet, PASS_FILL, CHECK_ITEMS
        from openpyxl import Workbook

        wb = Workbook()
        sn_data = [{"serial_number": "SN001", "config": "R2CNM", "unit_number": "ER2-1-1"}]
        cp_list = ["T0"]
        records_by_sn = {
            "SN001": {
                ("T0", item): {"status": "PASS", "failure_type": None}
                for item in CHECK_ITEMS
            }
        }

        ws = create_wf_sheet(wb, 1, "Test", cp_list, sn_data, records_by_sn=records_by_sn)

        # All 6 items should be PASS at columns 6-11
        for i, item in enumerate(CHECK_ITEMS):
            cell = ws.cell(row=4, column=6 + i)
            assert cell.value == "PASS", f"Item {item} at col {6+i} should be PASS"
            assert cell.fill == PASS_FILL, f"Item {item} at col {6+i} should have PASS fill"

    def test_color_constants_correct_values(self):
        """Verify the color coding constants have the correct hex values."""
        from checkitem_generator import (
            PASS_FILL, PASS_FONT, SPEC_FAIL_FILL,
            STRIFE_FAIL_FILL, STRIFE_FAIL_FONT,
        )

        # PASS: green background, green font
        assert PASS_FILL.start_color.rgb == "00C6EFCE"
        assert PASS_FONT.color.rgb == "00006100"

        # Spec Fail: red background
        assert SPEC_FAIL_FILL.start_color.rgb == "00FF0000"

        # Strife Fail: yellow background, black font
        assert STRIFE_FAIL_FILL.start_color.rgb == "00FFFF00"
        assert STRIFE_FAIL_FONT.color.rgb == "00000000"
