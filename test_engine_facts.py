import unittest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

import engine


class EngineFactExtractionTests(unittest.TestCase):
    def make_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'WF16.1'

        ws.cell(1, 3).value = 'Config'
        ws.cell(1, 4).value = 'Unit #'
        ws.cell(1, 5).value = 'S/N'
        ws.cell(1, 6).value = 'T0'
        ws.cell(1, 7).value = 'THC X 6'
        ws.cell(1, 9).value = 'Bottom Surface After 450Cyc'
        ws.cell(1, 11).value = 'Comments'

        ws.cell(2, 7).value = 'Cosmetic'
        ws.cell(2, 8).value = 'ISB'
        ws.cell(2, 9).value = 'FACT'
        ws.cell(2, 10).value = 'BT-OTA'

        ws.cell(3, 3).value = 'R3'
        ws.cell(3, 4).value = 'ER3-16.1-6'
        ws.cell(3, 5).value = 'D06P3Q46X0'
        ws.cell(3, 6).value = 'T0'
        ws.cell(3, 7).value = 'PASS'
        ws.cell(3, 8).value = 'PASS'
        ws.cell(3, 9).value = 'FAIL'
        ws.cell(3, 9).fill = PatternFill('solid', fgColor='FFFF0000')
        ws.cell(3, 10).value = '/'

        return ws

    def test_extract_wf_fact_rows_emits_cp_and_check_rows(self):
        ws = self.make_sheet()

        cp_rows, check_rows = engine.extract_wf_fact_rows(
            ws,
            wf_num='16.1',
            report_id=9,
            report_date='2026-05-09',
            ts_names=['THC X 6', 'Repetitive HSD PB'],
        )

        self.assertEqual(len(cp_rows), 2)
        self.assertEqual(cp_rows[1]['cp_idx'], 1)
        self.assertEqual(cp_rows[1]['status'], 'spec_fail')
        self.assertEqual(cp_rows[1]['failure_type'], 'spec')
        self.assertEqual(cp_rows[1]['is_current_cp'], 1)

        self.assertEqual(len(check_rows), 4)
        fact = [r for r in check_rows if r['check_item'] == 'FACT'][0]
        self.assertEqual(fact['status'], 'spec_fail')
        self.assertEqual(fact['failure_type'], 'spec')
        self.assertEqual(fact['source_row'], 3)
        self.assertEqual(fact['source_col'], 9)

    def make_cleaning_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'WF29.1'

        ws.cell(1, 3).value = 'Config'
        ws.cell(1, 4).value = 'Unit #'
        ws.cell(1, 5).value = 'S/N'
        ws.cell(1, 6).value = 'Chemichal'
        ws.cell(1, 7).value = 'T0'
        ws.cell(1, 13).value = '1st A 18hrs'
        ws.cell(1, 15).value = 'Comments'

        ws.cell(2, 7).value = 'Cosmetic'
        ws.cell(2, 8).value = 'ISB'
        ws.cell(2, 13).value = 'Cosmetic'
        ws.cell(2, 14).value = 'ISB'

        for offset, (unit, sn) in enumerate([
            ('ER1-29.1-1', 'SN-R1'),
            ('ER2-29.1-1', 'SN-R2'),
            ('ER3-29.1-1', 'SN-R3'),
            ('ER4-29.1-1', 'SN-R4'),
        ], start=3):
            ws.cell(offset, 4).value = unit
            ws.cell(offset, 5).value = sn
            ws.cell(offset, 7).value = 'PASS'
            ws.cell(offset, 8).value = 'PASS'
            ws.cell(offset, 13).value = 'PASS'
            ws.cell(offset, 14).value = 'PASS'

        return ws

    def test_cleaning_layout_infers_config_from_er_unit_prefix_for_fact_rows(self):
        ws = self.make_cleaning_sheet()

        cp_rows, check_rows = engine.extract_wf_fact_rows(
            ws,
            wf_num='29.1',
            report_id=9,
            report_date='2026-05-09',
            ts_names=['Cleaning Test_Spray_Option1'],
        )

        current_by_sn = {
            row['sn']: row
            for row in cp_rows
            if row['is_current_cp']
        }
        self.assertEqual(current_by_sn['SN-R1']['config'], 'R1FNF')
        self.assertEqual(current_by_sn['SN-R2']['config'], 'R2CNM')
        self.assertEqual(current_by_sn['SN-R3']['config'], 'R3')
        self.assertEqual(current_by_sn['SN-R4']['config'], 'R4')
        self.assertEqual(current_by_sn['SN-R1']['cp_idx'], 0)
        self.assertTrue(check_rows)

    def test_cleaning_layout_infers_config_from_er_unit_prefix_for_progress_rows(self):
        ws = self.make_cleaning_sheet()

        progress = engine._extract_wf_progress(ws, ['Cleaning Test_Spray_Option1'])

        self.assertEqual(len(progress['R1FNF']), 1)
        self.assertEqual(len(progress['R2CNM']), 1)
        self.assertEqual(len(progress['R3']), 1)
        self.assertEqual(len(progress['R4']), 1)
        self.assertEqual(progress['R1FNF'][0]['sn'], 'SN-R1')
        self.assertEqual(progress['R1FNF'][0]['current_cp_idx'], 0)


if __name__ == '__main__':
    unittest.main()
