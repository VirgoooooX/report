"""
M60 EVT REL — FA Matcher
Cross-references FA Tracker records with engine analysis results.
"""
from openpyxl import load_workbook
from engine import wf_num  # noqa: F811 — keep import clean

def read_fa_tracker(fa_path):
    """
    Reads the System TF sheet from FA Tracker.
    Returns: list of FA record dicts with keys: FA#, SN, WF, Config, Failed Test, Failure Type, FA Status, ...
    """
    wb = load_workbook(fa_path, data_only=True)
    ws = wb['System TF']
    
    # Row 7 is the header
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(7, c).value
        if v: headers[c] = str(v).strip().replace('\n', ' ')
    
    fas = []
    for r in range(8, ws.max_row + 1):
        fa_num = ws.cell(r, 1).value
        if fa_num is None: continue
        fa = {h: ws.cell(r, c).value for c, h in headers.items()}
        fa['_fa_num'] = fa_num
        fas.append(fa)
    
    wb.close()
    return fas


def match(fa_records, wf_results):
    """
    Cross-references FA records with engine analysis results.
    Returns list of matched FA dicts with added fields:
      '_matched': bool — whether the SN appears in per-WF failure lists
      '_matched_type': 'spec'|'strife'|None
      '_matched_wf': str or None
    """
    matched = []
    for fa in fa_records:
        wf = fa.get('WF')
        sn = str(fa.get('SN', '')).strip()
        cfg = str(fa.get('Config', '')).strip()
        
        wf_str = str(int(wf)) if isinstance(wf, float) and wf == int(wf) else str(wf).strip() if wf else ''
        
        found = False
        found_type = None
        
        if wf_str in wf_results and cfg in wf_results[wf_str]:
            for ti, d in wf_results[wf_str][cfg].items():
                if sn in d['spec_fails']:
                    found = True; found_type = 'spec'; break
                if sn in d['strife_fails']:
                    found = True; found_type = 'strife'; break
        
        fa['_matched'] = found
        fa['_matched_type'] = found_type
        matched.append(fa)
    
    return matched


def summary(matched_fas):
    """
    Returns summary stats from a matched FA list.
    {total, matched, unmatched, by_status: {...}, by_wf: {...}, by_type: {...}}
    """
    total = len(matched_fas)
    m = sum(1 for f in matched_fas if f['_matched'])
    u = total - m
    
    by_status = {}
    by_wf = {}
    by_type = {}
    
    for fa in matched_fas:
        st = str(fa.get('FA Status', 'Unknown')).strip()
        by_status[st] = by_status.get(st, 0) + 1
        
        wf_val = fa.get('WF')
        wf_key = str(int(wf_val)) if isinstance(wf_val, float) and wf_val == int(wf_val) else str(wf_val).strip() if wf_val else '?'
        by_wf[wf_key] = by_wf.get(wf_key, 0) + 1
        
        ft = str(fa.get('Failure Type \n(Spec. or Strife)', '')).strip()
        by_type[ft] = by_type.get(ft, 0) + 1
    
    return {'total': total, 'matched': m, 'unmatched': u,
            'by_status': by_status, 'by_wf': dict(sorted(by_wf.items())),
            'by_type': by_type}


# ── CLI self-test ───────────────────────────────────────────────────────
if __name__ == '__main__':
    import os, sys
    dr = os.path.join(os.path.dirname(__file__), 'data', 'M60 EVT Rel Daily Report_20260508.xlsx')
    fa = os.path.join(os.path.dirname(__file__), 'data', 'M60 EVT REL FA Tracker 20260508.xlsx')
    
    from engine import analyze
    print("Analyzing Daily Report...")
    results = analyze(dr)
    
    print("Reading FA Tracker...")
    fas = read_fa_tracker(fa)
    matched = match(fas, results)
    s = summary(matched)
    
    print(f"\nFA Tracker: {s['total']} records")
    print(f"  Matched:   {s['matched']} ✅")
    print(f"  Unmatched: {s['unmatched']} ⚠️")
    print(f"  By status: {s['by_status']}")
